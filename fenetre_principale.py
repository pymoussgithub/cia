import os
import shutil
import subprocess
import sys
import customtkinter as ctk
from ttkbootstrap import Style
from tkinter import filedialog, messagebox
import tkinter.ttk as ttk
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None
import tkinter as tk
from PIL import Image
try:
    import pandas as pd
except ImportError:
    pd = None
import json
import math
from classes_details import open_classe_details
from ajouter_classes import open_add_class_dialog
from fenetre_prof import PersonnelManager

# Variables globales pour les compteurs du header
total_counter_label = None
without_class_counter_label = None
without_level_counter_label = None

# Variable globale pour mémoriser la position de scroll du dashboard par semaine
dashboard_scroll_positions = {}  # {week_number: scroll_position}

# Système de sauvegarde des préférences utilisateur
PREFERENCES_FILE = os.path.join(os.path.dirname(__file__), "user_preferences.json")

def set_counter_labels(total_label, without_class_label, without_level_label):
    """Définit les références globales pour les labels des compteurs."""
    global total_counter_label, without_class_counter_label, without_level_counter_label
    total_counter_label = total_label
    without_class_counter_label = without_class_label
    without_level_counter_label = without_level_label

def load_user_preferences():
    """Charge les préférences utilisateur depuis le fichier JSON."""
    try:
        if os.path.exists(PREFERENCES_FILE):
            with open(PREFERENCES_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"Erreur lors du chargement des préférences: {e}")
    return {}

def save_user_preferences(preferences):
    """Sauvegarde les préférences utilisateur dans le fichier JSON."""
    try:
        with open(PREFERENCES_FILE, 'w', encoding='utf-8') as f:
            json.dump(preferences, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"Erreur lors de la sauvegarde des préférences: {e}")

def update_header_counters(matrix_stats):
    """Met à jour les compteurs d'élèves dans le header."""
    global total_counter_label, without_class_counter_label, without_level_counter_label
    try:
        if total_counter_label:
            total_counter_label.configure(text=f"Total\n{matrix_stats['total_eleves']}")

        # Couleur pour les boutons selon les valeurs
        if without_class_counter_label:
            sans_classe_count = matrix_stats['sans_classe']
            without_class_counter_label.configure(text=f"Sans Cl.\n{sans_classe_count}")
            # Orange si > 0, vert doux si = 0
            if sans_classe_count > 0:
                without_class_counter_label.configure(fg_color="#ffd4a3", hover_color="#ffb366")
            else:
                without_class_counter_label.configure(fg_color="#86efac", hover_color="#4ade80")

        if without_level_counter_label:
            sans_niveau_count = matrix_stats['sans_niveau']
            without_level_counter_label.configure(text=f"Sans Niv.\n{sans_niveau_count}")
            # Orange si > 0, vert doux si = 0
            if sans_niveau_count > 0:
                without_level_counter_label.configure(fg_color="#ffb3ba", hover_color="#ff9999")
            else:
                without_level_counter_label.configure(fg_color="#86efac", hover_color="#4ade80")

    except Exception as e:
        print(f"Error updating counters: {e}")


def show_students_without_assignment(assignment_type):
    """
    Affiche une fenêtre popup avec la liste des élèves sans classe ou sans niveau.

    Args:
        assignment_type (str): "class" pour sans classe, "level" pour sans niveau
    """
    # Trouver le chemin du fichier matrix actuel
    matrix_path = None
    try:
        # Chercher dans le dossier actuel et ses sous-dossiers un fichier matrix.xlsx
        current_dir = os.getcwd()
        for root, dirs, files in os.walk(current_dir):
            for file in files:
                if file.lower() == 'matrix.xlsx':
                    matrix_path = os.path.join(root, file)
                    break
            if matrix_path:
                break
    except Exception as e:
        print(f"Erreur lors de la recherche du fichier matrix: {e}")

    if not matrix_path or not os.path.exists(matrix_path):
        # Utiliser une boîte de dialogue Tkinter standard
        import tkinter.messagebox as messagebox
        messagebox.showwarning("Erreur", "Fichier matrix.xlsx non trouvé")
        return

    # Analyser le fichier matrix pour trouver les élèves sans assignation
    if pd is None:
        import tkinter.messagebox as messagebox
        messagebox.showwarning("Erreur", "Pandas n'est pas disponible")
        return

    try:
        df = pd.read_excel(matrix_path)

        # Trouver les colonnes nécessaires
        stagiaire_col = None
        niveau_col = None
        classe_col = None

        for col in df.columns:
            col_lower = str(col).lower()
            if 'stagiaire' in col_lower or 'nom' in col_lower or 'élève' in col_lower or 'eleve' in col_lower:
                if not stagiaire_col:
                    stagiaire_col = col
            elif ('niveau' in col_lower or 'level' in col_lower) and not any(char.isdigit() for char in col):
                if not niveau_col:
                    niveau_col = col
            elif 'classe' in col_lower or 'class' in col_lower or 'groupe' in col_lower:
                if not classe_col:
                    classe_col = col

        if not stagiaire_col:
            import tkinter.messagebox as messagebox
            messagebox.showwarning("Erreur", "Colonne des noms d'élèves non trouvée")
            return

        # Filtrer les élèves selon le type demandé
        students_without_assignment = []

        for _, row in df.iterrows():
            nom = str(row.get(stagiaire_col, '')).strip()

            if assignment_type == "class":
                classe = str(row.get(classe_col, '')).strip() if classe_col else ''
                if nom and nom.lower() not in ['', 'nan', 'none'] and (not classe or classe.lower() in ['', 'nan', 'none']):
                    students_without_assignment.append(nom)
            elif assignment_type == "level":
                niveau = str(row.get(niveau_col, '')).strip() if niveau_col else ''
                if nom and nom.lower() not in ['', 'nan', 'none'] and (not niveau or niveau.lower() in ['', 'nan', 'none']):
                    students_without_assignment.append(nom)

        # Trier les noms par ordre alphabétique
        students_without_assignment.sort()

        # Afficher une fenêtre de chargement
        loading_popup = show_loading_popup("Analyse des élèves...")

        # Créer la fenêtre popup avec le même style que classes_details.py
        popup = ctk.CTkToplevel()
        popup.title(f"Élèves sans {'classe' if assignment_type == 'class' else 'niveau'} ({len(students_without_assignment)})")
        popup.geometry("550x500")
        popup.resizable(False, False)
        popup.configure(fg_color="white")  # Fond blanc comme dans classes_details.py

        # Centrer la fenêtre
        popup.transient()
        popup.grab_set()

        # === HEADER BLANC (comme dans classes_details.py) ===
        header_bar = ctk.CTkFrame(popup, fg_color="white", height=80, corner_radius=0)
        header_bar.pack(fill="x", side="top", padx=0, pady=0)
        header_bar.pack_propagate(False)

        header_content = ctk.CTkFrame(header_bar, fg_color="transparent")
        header_content.pack(fill="both", expand=True, padx=30)

        # Titre dans le header
        title_label = ctk.CTkLabel(
            header_content,
            text=f"Élèves sans {'classe' if assignment_type == 'class' else 'niveau'}",
            font=("Segoe UI", 18, "bold"),
            text_color="#0F172A"
        )
        title_label.pack(side="left")

        # Compteur d'élèves dans le header (à droite)
        counter_label = ctk.CTkLabel(
            header_content,
            text=f"{len(students_without_assignment)} élève(s)",
            font=("Segoe UI", 14, "bold"),
            text_color="#64748B"
        )
        counter_label.pack(side="right")

        # === CADRE GRISE PRINCIPAL (comme dans classes_details.py) ===
        main_grey_frame = ctk.CTkFrame(popup, fg_color="#F1F5F9", corner_radius=25)
        main_grey_frame.pack(fill="both", expand=True, padx=25, pady=15)

        # Titre interne et champ de recherche
        search_container = ctk.CTkFrame(main_grey_frame, fg_color="transparent")
        search_container.pack(fill="x", padx=30, pady=(25, 0))

        ctk.CTkLabel(search_container, text="RECHERCHER UN ÉLÈVE", font=("Segoe UI", 13, "bold"),
                     text_color="#64748B").pack(anchor="w", pady=(0, 10))

        # Champ de recherche
        search_var = ctk.StringVar()
        search_entry = ctk.CTkEntry(
            search_container,
            placeholder_text="Tapez le nom de l'élève...",
            textvariable=search_var,
            width=450,
            height=35,
            font=("Segoe UI", 11)
        )
        search_entry.pack(pady=(0, 20))

        # Frame scrollable pour la liste (dans le cadre gris)
        scrollable_frame = ctk.CTkScrollableFrame(
            main_grey_frame,
            fg_color="transparent",
            width=480,
            height=250
        )
        scrollable_frame.pack(pady=(0, 25), padx=25)

        # Variables pour les éléments dynamiques
        student_labels = []

        def update_student_list(*args):
            # Effacer les anciens labels
            for label in student_labels:
                label.destroy()
            student_labels.clear()

            # Filtrer les élèves selon la recherche
            search_text = search_var.get().lower().strip()
            filtered_students = [
                student for student in students_without_assignment
                if search_text in student.lower()
            ]

            # Créer les labels simples comme dans l'ancien code
            for student in filtered_students:
                student_label = ctk.CTkLabel(
                    scrollable_frame,
                    text=f"{student}",
                    font=("Segoe UI", 12),
                    text_color="#1E293B",
                    anchor="w"
                )
                student_label.pack(fill="x", padx=15, pady=3)
                student_labels.append(student_label)

            # Mettre à jour le compteur dans le header
            counter_label.configure(text=f"{len(filtered_students)} élève(s) trouvé(s)")

        # === FOOTER (comme dans classes_details.py) ===
        footer_area = ctk.CTkFrame(popup, fg_color="white", height=70)
        footer_area.pack(fill="x", side="bottom")
        footer_area.pack_propagate(False)

        # Bouton fermer centré
        close_button = ctk.CTkButton(
            footer_area,
            text="Fermer",
            fg_color="#1E293B",
            hover_color="#0F172A",
            corner_radius=10,
            font=("Segoe UI", 13, "bold"),
            height=40,
            width=150,
            command=popup.destroy
        )
        close_button.pack(side="right", padx=30)

        # Lier la fonction de mise à jour à la variable de recherche
        search_var.trace_add("write", update_student_list)

        # Afficher initialement tous les élèves
        update_student_list()

        # Focus sur le champ de recherche
        search_entry.focus()

        # Fermer la fenêtre de chargement
        if loading_popup and loading_popup.winfo_exists():
            loading_popup.destroy()

    except Exception as e:
        # Fermer la fenêtre de chargement en cas d'erreur
        if loading_popup and loading_popup.winfo_exists():
            loading_popup.destroy()
        import tkinter.messagebox as messagebox
        messagebox.showerror("Erreur", f"Erreur lors de l'analyse du fichier: {str(e)}")


def show_loading_popup(message):
    """Affiche un popup de chargement (inspiré de classes_details.py)."""
    loading_popup = ctk.CTkToplevel()
    loading_popup.title("")
    loading_popup.geometry("300x120")
    loading_popup.resizable(False, False)
    loading_popup.attributes("-topmost", True)
    loading_popup.overrideredirect(True)

    # Style moderne
    main_frame = ctk.CTkFrame(
        loading_popup,
        fg_color="#FFFFFF",
        corner_radius=15,
        border_width=1,
        border_color="#CBD5E1"
    )
    main_frame.pack(fill="both", expand=True, padx=2, pady=2)

    content_frame = ctk.CTkFrame(main_frame, fg_color="#F8FAFC", corner_radius=10)
    content_frame.pack(fill="both", expand=True, padx=8, pady=8)

    # Message
    message_label = ctk.CTkLabel(
        content_frame,
        text=message,
        font=("Segoe UI", 12),
        text_color="#1E293B"
    )
    message_label.pack(pady=(15, 10))

    # Barre de progression
    progress_bar = ctk.CTkProgressBar(
        content_frame,
        width=200,
        height=4,
        corner_radius=10,
        fg_color="#F1F5F9",
        progress_color="#3B82F6",
        mode="indeterminate"
    )
    progress_bar.pack(pady=(0, 15))
    progress_bar.start()

    return loading_popup


def save_dashboard_scroll_position(scrollable_frame, week_number=None):
    """Sauvegarde la position de scroll actuelle pour une semaine donnée."""
    global dashboard_scroll_positions
    try:
        if scrollable_frame and hasattr(scrollable_frame, '_parent_canvas'):
            canvas = scrollable_frame._parent_canvas
            yview_result = canvas.yview()

            if isinstance(yview_result, (list, tuple)) and len(yview_result) >= 1:
                current_position = float(yview_result[0])

                if week_number:
                    old_position = dashboard_scroll_positions.get(week_number, 0.0)

                    if abs(current_position - old_position) > 0.01:  # Seuil plus grand pour éviter spam
                        dashboard_scroll_positions[week_number] = current_position
    except Exception as e:
        pass  # Erreur silencieuse lors de la sauvegarde

def restore_dashboard_scroll_position(scrollable_frame, week_number=None):
    """Restaure la position de scroll sauvegardée pour une semaine donnée."""
    global dashboard_scroll_positions
    try:
        if scrollable_frame and hasattr(scrollable_frame, '_parent_canvas') and week_number:
            saved_position = dashboard_scroll_positions.get(week_number, 0.0)
            if saved_position > 0:
                scrollable_frame._parent_canvas.yview_moveto(saved_position)
    except Exception as e:
        pass  # Erreur silencieuse lors de la restauration


def get_students_info_from_matrix(matrix_path, student_names):
    """
    Récupère les informations complètes des élèves depuis matrix.xlsx

    Args:
        matrix_path (str): Chemin vers matrix.xlsx
        student_names (list): Liste des noms d'élèves à rechercher

    Returns:
        dict: {nom_eleve: {'niveau': str, 'age': str, 'ci': bool}}
    """
    if pd is None or not os.path.exists(matrix_path) or not student_names:
        # Retourner des valeurs par défaut
        return {name: {'niveau': '', 'age': 'N/A', 'ci': False} for name in student_names}

    try:
        df = pd.read_excel(matrix_path)

        # Normaliser les noms d'élèves pour la recherche
        student_names_normalized = {name.lower().strip(): name for name in student_names}

        # Trouver les colonnes nécessaires
        stagiaire_col = None
        niveau_col = None
        age_col = None
        ci_col = None

        print(f"[DEBUG MATRIX] Colonnes disponibles: {list(df.columns)}")

        for col in df.columns:
            col_lower = str(col).lower()
            if 'stagiaire' in col_lower or 'nom' in col_lower or 'élève' in col_lower or 'eleve' in col_lower:
                if not stagiaire_col:
                    stagiaire_col = col
            elif ('niveau' in col_lower or 'level' in col_lower) and not any(char.isdigit() for char in col):
                if not niveau_col:
                    niveau_col = col
            elif 'age' in col_lower or 'âge' in col_lower:
                if not age_col:
                    age_col = col
            elif 'ci' in col_lower or 'cours intensif' in col_lower or 'intensif' in col_lower or col_lower == 'cours 2':
                if not ci_col:
                    ci_col = col
                    print(f"[DEBUG MATRIX] Colonne CI trouvée: {col}")

        if not stagiaire_col:
            return {name: {'niveau': '', 'age': 'N/A', 'ci': False} for name in student_names}

        # Créer un dictionnaire des informations des élèves
        students_info = {}

        for _, row in df.iterrows():
            nom_brut = str(row.get(stagiaire_col, '')).strip()
            nom_normalized = nom_brut.lower().strip()

            # Chercher si ce nom correspond à un élève recherché
            original_name = None
            for search_name_norm, search_name_orig in student_names_normalized.items():
                if search_name_norm in nom_normalized or nom_normalized in search_name_norm:
                    original_name = search_name_orig
                    break

            if original_name:
                # Récupérer les informations
                niveau = str(row.get(niveau_col, '')).strip() if niveau_col else ''
                age = str(row.get(age_col, '')).strip() if age_col else 'N/A'

                # Déterminer si CI
                ci = False
                if ci_col:
                    ci_val = str(row.get(ci_col, '')).strip().lower()
                    ci = any(keyword in ci_val for keyword in ['oui', 'yes', 'true', '1', 'ci', 'intensif', 'cours intensif'])
                    print(f"[DEBUG MATRIX] Élève {original_name}: colonne='{ci_col}', valeur='{str(row.get(ci_col, ''))}', normalisée='{ci_val}', CI={ci}")

                students_info[original_name] = {
                    'niveau': niveau,
                    'age': age,
                    'ci': ci
                }

        # Pour les élèves non trouvés, utiliser des valeurs par défaut
        for name in student_names:
            if name not in students_info:
                students_info[name] = {'niveau': '', 'age': 'N/A', 'ci': False}

        return students_info

    except Exception as e:
        print(f"Erreur lors de la récupération des informations élèves depuis matrix: {e}")
        return {name: {'niveau': '', 'age': 'N/A', 'ci': False} for name in student_names}


def analyze_school_classes(week_folder):
    """
    Analyse tous les fichiers Excel d'écoles d'une semaine donnée.

    Args:
        week_folder (str): Chemin du dossier de la semaine (ex: "semaine_1")

    Returns:
        dict: {
            'ecole_a': list of class dicts,
            'ecole_b': list of class dicts,
            'ecole_c_cs': list of class dicts,
            'ecole_c_ci': list of class dicts,
            'ecole_morning': list of class dicts,
            'ecole_premium_cs': list of class dicts,
            'ecole_premium_ci': list of class dicts
        }
        Chaque dict de classe contient : {
            'horaire': str,  # nom de la feuille
            'intervenant': str,
            'type_intervenant': str,  # 'professeur' ou 'animateur'
            'classes': list of dicts with 'nom_classe', 'nb_eleves', 'niveau', and 'eleves'
        }
        Chaque dict de classe contient aussi : {
            'nom_classe': str,
            'nb_eleves': int,
            'niveau': str,
            'eleves': list of dicts with 'nom', 'niveau', 'age', 'ci'
        }
    """
    result = {
        'ecole_a': [],
        'ecole_b': [],
        'ecole_c_cs': [],
        'ecole_c_ci': [],
        'ecole_morning': [],
        'ecole_premium_cs': [],
        'ecole_premium_ci': []
    }

    if pd is None:
        return result

    # Mapping des fichiers Excel vers les clés d'écoles
    file_to_school_mapping = {
        'ecole_a.xlsx': 'ecole_a',
        'ecole_b.xlsx': 'ecole_b',
        'ECOLE_C_cours_standard.xlsx': 'ecole_c_cs',
        'ECOLE_C_cours_intensif.xlsx': 'ecole_c_ci',
        'MORNING.xlsx': 'ecole_morning',  # Par défaut, assigné aux cours standard
        'ECOLE_PREMIUM_cours_standard.xlsx': 'ecole_premium_cs',
        'ECOLE_PREMIUM_cours_intensifs.xlsx': 'ecole_premium_ci'      
    }

    # Fonction helper pour analyser un fichier Excel d'école
    def analyze_school_file(excel_path, school_key):
        if not os.path.exists(excel_path):
            return

        try:
            # Lire toutes les feuilles du fichier
            xl_file = pd.ExcelFile(excel_path)
            sheet_names = xl_file.sheet_names

            for sheet_name in sheet_names:
                try:
                    df = pd.read_excel(excel_path, sheet_name=sheet_name)

                    if df.empty:
                        # La feuille est vide (pas de données) mais on veut quand même afficher l'horaire
                        sheet_lower = sheet_name.lower()
                        type_intervenant = "animateur" if "animateur" in sheet_lower else "professeur"
                        horaire = clean_horaire_name(sheet_name)
                        
                        result[school_key].append({
                            'horaire': horaire or sheet_name,
                            'intervenant': "Non spécifié",
                            'type_intervenant': type_intervenant,
                            'classes': []
                        })
                        continue  # On skip le reste du traitement pour cette feuille

                    # Analyser le nom de la feuille pour déterminer le type d'intervenant et l'horaire
                    sheet_lower = sheet_name.lower()
                    type_intervenant = "animateur" if "animateur" in sheet_lower else "professeur"
                    horaire = clean_horaire_name(sheet_name)

                    # Chercher la colonne d'intervenant
                    intervenant_col = None
                    for col in df.columns:
                        col_lower = str(col).lower()
                        if any(keyword in col_lower for keyword in ['intervenant', 'professeur', 'animateur', 'enseignant']):
                            intervenant_col = col
                            break

                    # Chercher les colonnes de classes, d'élèves, de niveau et de liste d'élèves
                    classe_cols = []
                    eleve_cols = []
                    niveau_cols = []
                    liste_eleves_cols = []

                    for col in df.columns:
                        col_lower = str(col).lower()

                        # Vérifier d'abord les colonnes de liste d'élèves (plus spécifique)
                        if 'liste' in col_lower or 'élèves' in col_lower or 'eleves' in col_lower or 'noms' in col_lower:
                            liste_eleves_cols.append(col)
                        elif any(keyword in col_lower for keyword in ['classe', 'groupe', 'section']):
                            classe_cols.append(col)
                        elif any(keyword in col_lower for keyword in ['élève', 'eleve', 'effectif', 'nombre']):
                            # Éviter les colonnes qui contiennent "liste des élèves"
                            if not ('liste' in col_lower):
                                eleve_cols.append(col)
                        elif any(keyword in col_lower for keyword in ['niveau', 'level']):
                            niveau_cols.append(col)

                    # Extraire les intervenants et classes
                    intervenants = []
                    if intervenant_col:
                        intervenants = df[intervenant_col].dropna().unique().tolist()

                    classes_info = []
                    for _, row in df.iterrows():
                        classe_nom = None
                        nb_eleves = 0
                        niveau = ""
                        eleves_list = []
                        intervenant_classe = "Non spécifié"  # Intervenant spécifique à la classe

                        # Chercher le nom de la classe
                        for col in classe_cols:
                            val = str(row.get(col, '')).strip()
                            if val and val.lower() not in ['', 'nan', 'none']:
                                classe_nom = val
                                break

                        # Chercher l'intervenant pour cette classe spécifique
                        if intervenant_col:
                            intervenant_val = str(row.get(intervenant_col, '')).strip()
                            if intervenant_val and intervenant_val.lower() not in ['', 'nan', 'none']:
                                intervenant_classe = intervenant_val

                        # Chercher le nombre d'élèves
                        for col in eleve_cols:
                            val = row.get(col, 0)
                            if pd.notna(val) and isinstance(val, (int, float)):
                                nb_eleves = int(val)
                                break

                        # Chercher le niveau
                        for col in niveau_cols:
                            val = str(row.get(col, '')).strip()
                            if val and val.lower() not in ['', 'nan', 'none']:
                                niveau = val
                                break

                        # Chercher la liste des élèves
                        for col in liste_eleves_cols:
                            val = str(row.get(col, '')).strip()
                            if val and val.lower() not in ['', 'nan', 'none', 'liste des élèves...']:
                                # Parser la liste des élèves (séparés par des virgules, points-virgules, ou retours à la ligne)
                                import re
                                # Diviser par virgules, points-virgules, ou nouvelles lignes
                                eleves_raw = re.split(r'[;,|\n\r]+', val)
                                student_names = []
                                for eleve_nom in eleves_raw:
                                    eleve_nom = eleve_nom.strip()
                                    if eleve_nom and eleve_nom.lower() not in ['', 'nan', 'none']:
                                        student_names.append(eleve_nom)

                                # Récupérer les informations complètes des élèves depuis matrix.xlsx
                                matrix_path = os.path.join(week_folder, "matrix.xlsx")
                                students_info = get_students_info_from_matrix(matrix_path, student_names)

                                # Construire la liste des élèves avec leurs vraies informations
                                for eleve_nom in student_names:
                                    info = students_info.get(eleve_nom, {'niveau': niveau, 'age': 'N/A', 'ci': False})
                                    eleves_list.append({
                                        'nom': eleve_nom,
                                        'niveau': info['niveau'] or niveau,  # Utiliser le niveau de matrix, ou celui de la classe si vide
                                        'age': info['age'],
                                        'ci': info['ci']
                                    })
                                break

                        if classe_nom:
                            classes_info.append({
                                'nom_classe': classe_nom,
                                'nb_eleves': len(eleves_list),  # Utiliser le nombre réel d'élèves parsés
                                'niveau': niveau,
                                'eleves': eleves_list,
                                'intervenant': intervenant_classe  # Ajouter l'intervenant spécifique à la classe
                            })

                    # Créer l'entrée pour cette feuille/horaire                    
                    result[school_key].append({
                            'horaire': horaire or sheet_name,
                            'intervenant': intervenants[0] if intervenants else "Non spécifié",
                            'type_intervenant': type_intervenant,
                            'classes': classes_info
                        })

                except Exception as e:
                    print(f"Erreur lors de l'analyse de la feuille '{sheet_name}' dans {excel_path}: {e}")
                    continue

        except Exception as e:
            print(f"Erreur lors de l'analyse du fichier {excel_path}: {e}")

    # Analyser tous les fichiers Excel du dossier semaine
    if os.path.exists(week_folder):
        for filename in os.listdir(week_folder):
            if filename.lower().endswith('.xlsx') and filename.lower() != 'matrix.xlsx':
                file_path = os.path.join(week_folder, filename)
                school_key = file_to_school_mapping.get(filename)
                if school_key:
                    analyze_school_file(file_path, school_key)

    return result


def analyze_matrix_assignments(matrix_path):
    """
    Analyse le fichier matrix.xlsx et retourne les élèves avec leurs assignations.

    Returns:
        dict: {
            'eleves_assignes': list of dicts with 'nom', 'niveau', 'ecole', 'classe', 'horaire',
            'total_eleves': int,
            'sans_classe': int,
            'sans_niveau': int,
            'avec_classe': int,
            'par_niveau': dict {niveau: count}
        }
    """
    if pd is None or not os.path.exists(matrix_path):
        return {
            'eleves_assignes': [],
            'total_eleves': 0,
            'sans_classe': 0,
            'sans_niveau': 0,
            'avec_classe': 0,
            'par_niveau': {}
        }

    try:
        df = pd.read_excel(matrix_path)

        # Trouver les colonnes nécessaires
        stagiaire_col = None
        niveau_col = None
        ecole_col = None
        classe_col = None
        horaire_col = None

        # Fonction helper pour vérifier si une colonne doit être exclue
        def should_exclude_column(col_name):
            col_lower = str(col_name).lower()
            # Exclure les colonnes spécifiques aux cours (Cours 1, Cours 2, etc.)
            if any(f'cours {i}' in col_lower for i in range(1, 10)):
                return True
            # Exclure les colonnes avec des suffixes numériques (.1, .2, etc.)
            if any(col_lower.endswith(f'.{i}') for i in range(1, 10)):
                return True
            # Exclure les colonnes d'arrivée/départ spécifiques
            if 'arrivée' in col_lower or 'départ' in col_lower:
                return True
            return False

        for col in df.columns:
            if should_exclude_column(col):
                continue

            col_lower = str(col).lower()
            if 'stagiaire' in col_lower or 'nom' in col_lower or 'élève' in col_lower or 'eleve' in col_lower:
                if not stagiaire_col:  # Prendre la première trouvée
                    stagiaire_col = col
            elif ('niveau' in col_lower or 'level' in col_lower) and not any(char.isdigit() for char in col):
                if not niveau_col:  # Éviter les duplications comme "Niveau.1"
                    niveau_col = col
            elif 'ecole' in col_lower or 'école' in col_lower or 'school' in col_lower:
                if not ecole_col:  # Prendre la première école générale
                    ecole_col = col
            elif 'classe' in col_lower or 'class' in col_lower or 'groupe' in col_lower:
                if not classe_col:  # Prendre la première classe générale
                    classe_col = col
            elif 'horaire' in col_lower or 'horaire' in col_lower or 'heure' in col_lower or 'time' in col_lower:
                if not horaire_col and 'arrivée' not in col_lower and 'départ' not in col_lower:
                    horaire_col = col

        if not stagiaire_col:
            return {
                'eleves_assignes': [],
                'total_eleves': 0,
                'sans_classe': 0,
                'sans_niveau': 0,
                'avec_classe': 0,
                'par_niveau': {}
            }

        total_eleves = len(df)
        sans_classe = 0
        sans_niveau = 0
        avec_classe = 0
        par_niveau = {}
        eleves_assignes = []

        for _, row in df.iterrows():
            nom = str(row.get(stagiaire_col, '')).strip() if stagiaire_col else ''
            niveau = str(row.get(niveau_col, '')).strip() if niveau_col else ''
            ecole = str(row.get(ecole_col, '')).strip() if ecole_col else ''
            classe = str(row.get(classe_col, '')).strip() if classe_col else ''
            horaire = str(row.get(horaire_col, '')).strip() if horaire_col else ''

            # Vérifier si l'élève a toutes les informations nécessaires pour être assigné
            if nom and nom.lower() not in ['', 'nan', 'none'] and ecole and classe and horaire:
                eleves_assignes.append({
                    'nom': nom,
                    'niveau': niveau,
                    'ecole': ecole,
                    'classe': classe,
                    'horaire': horaire
                })

            # Statistiques des classes
            if not classe or classe.lower() in ['', 'nan', 'none']:
                sans_classe += 1
            else:
                avec_classe += 1

            # Statistiques des niveaux
            if not niveau or niveau.lower() in ['', 'nan', 'none']:
                sans_niveau += 1
            else:
                if niveau in par_niveau:
                    par_niveau[niveau] += 1
                else:
                    par_niveau[niveau] = 1

        return {
            'eleves_assignes': eleves_assignes,
            'total_eleves': total_eleves,
            'sans_classe': sans_classe,
            'sans_niveau': sans_niveau,
            'avec_classe': avec_classe,
            'par_niveau': par_niveau
        }

    except Exception as e:
        print(f"Erreur lors de l'analyse des assignations matrix: {e}")
        return {
            'eleves_assignes': [],
            'total_eleves': 0,
            'sans_classe': 0,
            'sans_niveau': 0,
            'avec_classe': 0,
            'par_niveau': {}
        }


def analyze_matrix_file(matrix_path):
    """
    Analyse le fichier matrix.xlsx et retourne les statistiques des élèves.

    Returns:
        dict: {
            'total_eleves': int,
            'sans_classe': int,
            'sans_niveau': int,
            'avec_classe': int,
            'par_niveau': dict {niveau: count}
        }
    """
    result = analyze_matrix_assignments(matrix_path)
    return {
        'total_eleves': result['total_eleves'],
        'sans_classe': result['sans_classe'],
        'sans_niveau': result['sans_niveau'],
        'avec_classe': result['avec_classe'],
        'par_niveau': result['par_niveau']
    }



def clear_all_school_files(matrix_path):
    """
    Vide tous les fichiers Excel des écoles en supprimant toutes les classes.

    Args:
        matrix_path (str): Chemin du fichier matrix.xlsx pour déterminer le dossier semaine
    """
    if load_workbook is None:
        print("openpyxl n'est pas disponible, impossible de vider les fichiers Excel")
        return

    # Déterminer le dossier semaine depuis le chemin matrix
    week_folder = os.path.dirname(matrix_path)

    # Mapping des écoles vers les fichiers Excel
    school_file_mapping = {
        'ecole_a.xlsx': 'École A',
        'ecole_b.xlsx': 'École B',
        'ECOLE_C_cours_standard.xlsx': 'École C/CS',
        'ECOLE_C_cours_intensif.xlsx': 'École C/CI',
        'MORNING.xlsx': 'Morning',
        'ECOLE_PREMIUM_cours_standard.xlsx': 'Premium/CS',
        'ECOLE_PREMIUM_cours_intensifs.xlsx': 'Premium/CI'
    }

    for excel_filename, school_name in school_file_mapping.items():
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            continue

        try:
            # Ouvrir le fichier Excel
            wb = load_workbook(excel_path)

            # Pour chaque feuille du fichier
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Identifier les lignes de données (commençant après la ligne 1 qui contient les headers)
                rows_to_delete = []
                for row_idx in range(2, sheet.max_row + 1):  # Commencer à la ligne 2
                    # Vérifier si la ligne contient des données
                    has_data = False
                    for col_idx in range(1, sheet.max_column + 1):
                        cell_value = str(sheet.cell(row=row_idx, column=col_idx).value or '').strip()
                        if cell_value and cell_value.lower() not in ['', 'nan', 'none']:
                            has_data = True
                            break

                    if has_data:
                        rows_to_delete.append(row_idx)

                # Supprimer les lignes en commençant par la fin pour éviter les décalages d'indices
                for row_idx in reversed(rows_to_delete):
                    sheet.delete_rows(row_idx)

            # Sauvegarder le fichier
            wb.save(excel_path)
            print(f"Fichier {excel_filename} vide avec succes")

        except Exception as e:
            print(f"Erreur lors du vidage de {excel_filename}: {e}")


def update_school_files_with_assignments(week_folder, matrix_assignments):
    """
    Met à jour les fichiers Excel des écoles avec les élèves assignés depuis matrix.xlsx

    Args:
        week_folder (str): Chemin du dossier de la semaine
        matrix_assignments (dict): Résultat de analyze_matrix_assignments()
    """
    if load_workbook is None:
        print("openpyxl n'est pas disponible, impossible de mettre à jour les fichiers Excel")
        return

    # Mapping des écoles vers les fichiers Excel
    school_file_mapping = {
        'A': 'ecole_a.xlsx',
        'B': 'ecole_b.xlsx',
        'C/CS': 'ECOLE_C_cours_standard.xlsx',
        'C/CI': 'ECOLE_C_cours_intensif.xlsx',
        'Morning': 'MORNING.xlsx',
        'Premium/CS': 'ECOLE_PREMIUM_cours_standard.xlsx',
        'Premium/CI': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
    }

    # Grouper les élèves par école, horaire et classe
    assignments_by_school = {}
    for eleve in matrix_assignments['eleves_assignes']:
        ecole = eleve['ecole']
        horaire = eleve['horaire']
        classe = eleve['classe']
        nom = eleve['nom']

        if ecole not in assignments_by_school:
            assignments_by_school[ecole] = {}

        if horaire not in assignments_by_school[ecole]:
            assignments_by_school[ecole][horaire] = {}

        if classe not in assignments_by_school[ecole][horaire]:
            assignments_by_school[ecole][horaire][classe] = []

        assignments_by_school[ecole][horaire][classe].append(nom)

    # Pour chaque école, mettre à jour le fichier Excel correspondant
    for ecole_key, horaires_data in assignments_by_school.items():
        if ecole_key not in school_file_mapping:
            continue

        excel_filename = school_file_mapping[ecole_key]
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            continue

        try:
            wb = load_workbook(excel_path)

            # Pour chaque horaire/feuille dans cette école
            for horaire, classes_data in horaires_data.items():
                # Chercher la feuille correspondant à cet horaire
                target_sheet = None
                for sheet_name in wb.sheetnames:
                    # Normaliser les noms pour la comparaison
                    sheet_normalized = sheet_name.lower().replace('animateur', '').replace('professeur', '').strip()
                    horaire_normalized = horaire.lower().strip()

                    if sheet_normalized == horaire_normalized or horaire_normalized in sheet_normalized:
                        target_sheet = wb[sheet_name]
                        break

                if target_sheet is None:
                    print(f"Feuille horaire '{horaire}' non trouvée dans {excel_filename}")
                    continue

                # Pour chaque classe dans cet horaire, trouver la colonne appropriée et y ajouter les élèves
                for classe_nom, eleves in classes_data.items():
                    # Chercher la ligne correspondant à cette classe
                    classe_row = None
                    classe_col = None

                    # Chercher la ligne de cette classe
                    for row_idx in range(2, target_sheet.max_row + 1):  # Commencer à la ligne 2 (après header)
                        for col_idx in range(1, target_sheet.max_column + 1):
                            cell_value = str(target_sheet.cell(row=row_idx, column=col_idx).value or '').strip()
                            if cell_value == classe_nom:
                                classe_row = row_idx
                                classe_col = col_idx
                                break
                        if classe_row:
                            break

                    if classe_row is None:
                        continue

                    # Chercher une colonne appropriée pour les élèves (liste des élèves ou élèves)
                    eleves_col = None

                    # D'abord chercher "liste des élèves" ou "liste élèves"
                    for col_idx in range(1, target_sheet.max_column + 1):
                        cell_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                        if ('liste' in cell_value and ('élèves' in cell_value or 'eleves' in cell_value)) or \
                           ('élèves' in cell_value or 'eleves' in cell_value):
                            eleves_col = col_idx
                            break

                    # Si pas trouvé, chercher "noms" ou "stagiaire"
                    if eleves_col is None:
                        for col_idx in range(1, target_sheet.max_column + 1):
                            cell_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                            if 'nom' in cell_value or 'stagiaire' in cell_value:
                                eleves_col = col_idx
                                break

                    # Si toujours pas trouvé, utiliser la colonne à droite de la classe
                    if eleves_col is None and classe_col:
                        eleves_col = classe_col + 1
                        # S'assurer que cette colonne existe
                        if eleves_col > target_sheet.max_column:
                            continue

                    if eleves_col is None:
                        print(f"Aucune colonne appropriée trouvée pour les élèves dans {excel_filename} - {horaire}")
                        continue

                    # Ajouter les élèves dans la cellule appropriée (sans doublons)
                    current_value = str(target_sheet.cell(row=classe_row, column=eleves_col).value or '').strip()

                    # Récupérer la liste actuelle des élèves
                    existing_eleves = []
                    if current_value and current_value not in ['', 'nan', 'none']:
                        existing_eleves = [e.strip() for e in current_value.split(',') if e.strip()]

                    # Ajouter seulement les élèves qui ne sont pas déjà présents (éviter les doublons)
                    new_eleves = []
                    for eleve in eleves:
                        if eleve not in existing_eleves:
                            new_eleves.append(eleve)

                    # Si il y a de nouveaux élèves à ajouter
                    if new_eleves:
                        if existing_eleves:
                            all_eleves = existing_eleves + new_eleves
                        else:
                            all_eleves = new_eleves

                        # Trier les élèves par ordre alphabétique pour une meilleure présentation
                        all_eleves_sorted = sorted(all_eleves)
                        eleves_text = ', '.join(all_eleves_sorted)

                        target_sheet.cell(row=classe_row, column=eleves_col, value=eleves_text)

            # Sauvegarder le fichier
            wb.save(excel_path)

        except Exception as e:
            print(f"Erreur lors de la mise à jour de {excel_filename}: {e}")


def clean_horaire_name(sheet_name):
    """Nettoie le nom de la feuille pour extraire le nom d'horaire de manière cohérente."""
    # Cette fonction doit être identique à celle utilisée dans fenetre_principale.py et Assignation des Niveaux.py
    sheet_lower = sheet_name.lower()
    type_intervenant = "animateur" if "animateur" in sheet_lower else "professeur"

    # Liste complète des mots à supprimer (avec variations)
    words_to_remove = [
        "animateur", "Animateur", "anim", "Anim",
        "professeur", "Professeur", "prof", "Prof"
    ]

    # Nettoyer le nom en supprimant tous les mots-clés d'intervenants
    horaire = sheet_name
    for word in words_to_remove:
        horaire = horaire.replace(word, "")

    # Nettoyer les espaces multiples et supprimer les espaces au début/fin
    horaire = " ".join(horaire.split()).strip()

    # Si le résultat est vide, utiliser le nom original
    return horaire or sheet_name

def show_loading_window(parent_app):
    """Affiche une fenêtre de chargement style moderne (Corrigé)."""
    loading_popup = ctk.CTkToplevel(parent_app)
    loading_popup.geometry("350x150")
    loading_popup.resizable(False, False)
    
    # Configuration pour supprimer les bords de fenêtre
    loading_popup.attributes("-topmost", True)
    loading_popup.overrideredirect(True)

    # Centrer précisément sur la fenêtre parente
    x = parent_app.winfo_rootx() + (parent_app.winfo_width() // 2) - 175
    y = parent_app.winfo_rooty() + (parent_app.winfo_height() // 2) - 75
    loading_popup.geometry(f"350x150+{x}+{y}")

    # --- STYLE GRAPHIQUE ---
    main_frame = ctk.CTkFrame(
        loading_popup,
        fg_color="#FFFFFF",
        corner_radius=20,
        border_width=1,
        border_color="#CBD5E1"
    )
    main_frame.pack(fill="both", expand=True)

    # Titre "Système" discret (Sans letter_spacing)
    title_label = ctk.CTkLabel(
        main_frame,
        text="SYNCHRONISATION",
        font=("Segoe UI", 10, "bold"), # On utilise Segoe UI qui est standard
        text_color="#64748B"
    )
    title_label.pack(pady=(20, 5))

    # Texte principal
    text_label = ctk.CTkLabel(
        main_frame,
        text="Chargement des élèves depuis matrix...",
        font=("Segoe UI", 13),
        text_color="#1E293B"
    )
    text_label.pack(pady=(0, 15))

    # Barre de progression animée
    progress_bar = ctk.CTkProgressBar(
        main_frame,
        width=250,
        height=4,
        corner_radius=10,
        fg_color="#F1F5F9",
        progress_color="#3B82F6",
        mode="indeterminate"
    )
    progress_bar.pack(pady=5)
    progress_bar.set(0)
    progress_bar.start()

    def close_loading():
        if loading_popup.winfo_exists():
            loading_popup.destroy()

    parent_app.after(2500, close_loading)

    return loading_popup


def open_main_window(username: str, screen_width: int, screen_height: int) -> None:
    """
    Fenêtre principale moderne affichée après la connexion.

    - Colonne de gauche : 9 boutons "Semaine 1" à "Semaine 9"
    - Partie droite : tableau de bord (dashboard) pour gérer classes, élèves, etc.
    """

    # Crée une nouvelle fenêtre principale CTk
    app = ctk.CTk()
    app.title(f"Tableau de bord - Connecté en tant que {username}")
    # Démarre transparente pour l'effet de fondu
    app.attributes("-alpha", 0.0)

    # Style global identique à la fenêtre de login
    Style("flatly")
    ctk.set_appearance_mode("light")
    app.configure(fg_color="white")

    # Occupe la quasi-totalité de l'écran (taille de base avant maximisation)
    width = int(screen_width * 0.9)
    height = int(screen_height * 0.9)
    pos_x = (screen_width - width) // 2
    pos_y = (screen_height - height) // 2
    app.geometry(f"{width}x{height}+{pos_x}+{pos_y}")
    app.minsize(1000, 650)

    # Configuration de la grille principale : colonne gauche (menu) + colonne droite (dashboard)
    app.grid_columnconfigure(0, weight=0)   # colonne menu
    app.grid_columnconfigure(1, weight=1)   # colonne dashboard
    app.grid_rowconfigure(0, weight=1)

    # --------- COLONNE GAUCHE : MENU SEMAINES ---------
    sidebar = ctk.CTkFrame(app, fg_color="#f2f6fb", corner_radius=0)
    sidebar.grid(row=0, column=0, sticky="nsw")
    sidebar.grid_rowconfigure(11, weight=1)  # espace flexible en bas

    # Logo centré au-dessus du titre
    try:
        logo_path = os.path.join(os.path.dirname(__file__), "2.png")
        if os.path.exists(logo_path):
            logo_image = Image.open(logo_path)
            # Redimensionner en gardant les proportions pour tenir dans 210x210 (3x plus grand)
            original_width, original_height = logo_image.size
            ratio = min(210 / original_width, 210 / original_height)

            # 🔧 AJUSTEMENT DE TAILLE : Modifiez ce pourcentage pour changer la taille (0.8 = 80% de la taille actuelle)
            size_percentage = 0.8  # Changez cette valeur pour ajuster la taille (0.5 = 50%, 1.0 = 100%, 1.2 = 120%, etc.)
            ratio = ratio * size_percentage

            new_width = int(original_width * ratio)
            new_height = int(original_height * ratio)
            logo_image = logo_image.resize((new_width, new_height), Image.Resampling.LANCZOS)
            logo_ctk = ctk.CTkImage(light_image=logo_image, size=(new_width, new_height))

            # 🔧 AJUSTEMENT DU CADRE : Modifiez corner_radius pour changer l'arrondi du cadre
            frame_corner_radius = 15  # Changez cette valeur pour ajuster l'arrondi (0 = carré, 20 = très arrondi, etc.)

            # Cadre autour du logo
            logo_frame = ctk.CTkFrame(
                sidebar,
                fg_color="white",
                corner_radius=frame_corner_radius,
                border_width=1,
                border_color="#e1e5e9"
            )
            logo_frame.grid(row=0, column=0, pady=(15, 5))

            logo_label = ctk.CTkLabel(
                logo_frame,
                image=logo_ctk,
                text="",
                fg_color="transparent"
            )
            logo_label.pack(padx=8, pady=8)
    except Exception as e:
        print(f"Erreur lors du chargement du logo: {e}")

    title_label = ctk.CTkLabel(
        sidebar,
        text="Semaines",
        font=("Arial", 20, "bold"),
        text_color="#1f4e79",
        anchor="center",
    )
    title_label.grid(row=1, column=0, sticky="ew", padx=18, pady=(5, 10))

    # Stockage de la semaine sélectionnée (vide au départ pour afficher le logo central)
    selected_week = ctk.StringVar(value="")

    # Charger les préférences utilisateur
    preferences = load_user_preferences()

    # Variables pour les filtres d'écoles (avec valeurs par défaut depuis les préférences)
    school_filters = {
        "A": ctk.BooleanVar(value=preferences.get("school_filters", {}).get("A", True)),
        "B": ctk.BooleanVar(value=preferences.get("school_filters", {}).get("B", True)),
        "C/CS": ctk.BooleanVar(value=preferences.get("school_filters", {}).get("C/CS", True)),
        "C/CI": ctk.BooleanVar(value=preferences.get("school_filters", {}).get("C/CI", True)),
        "Morning": ctk.BooleanVar(value=preferences.get("school_filters", {}).get("Morning", True)),
        "Premium/CS": ctk.BooleanVar(value=preferences.get("school_filters", {}).get("Premium/CS", True)),
        "Premium/CI": ctk.BooleanVar(value=preferences.get("school_filters", {}).get("Premium/CI", True))
    }

    # Stockage des boutons de semaine pour gérer la sélection visuelle
    week_buttons = {}

    # Variables pour la surveillance des modifications du fichier matrix
    matrix_last_modified = {}
    matrix_watch_job = None

    def open_actions_menu(anchor_btn):
    # 1. Configuration initiale du menu (invisible au départ avec alpha=0)
        menu = ctk.CTkToplevel(app)
        menu.overrideredirect(True)
        menu.attributes("-alpha", 0.0) # Commence invisible pour l'animation
        menu.configure(fg_color="white", border_width=1, border_color="#d0d7e2")
        
        menu_width = 200
        menu_height = 320

        # 2. FORCE la mise à jour des widgets pour obtenir les vraies coordonnées
        # C'est l'étape cruciale pour que winfo_width() ne renvoie pas '1'
        app.update_idletasks() 
        
        # 3. Calcul de la position centrée
        btn_x = anchor_btn.winfo_rootx()
        btn_y = anchor_btn.winfo_rooty()
        btn_w = anchor_btn.winfo_width()
        btn_h = anchor_btn.winfo_height()

        # Calcul X : Centre du bouton - Moitié de la largeur du menu
        pos_x = btn_x + (btn_w // 2) - (menu_width // 2)
        pos_y = btn_y + btn_h + 4 # 4 pixels d'écart sous le bouton

        menu.geometry(f"{menu_width}x{menu_height}+{pos_x}+{pos_y}")

        # --- Contenu du menu ---
        actions = ["Importer fichier matrix", "Ouvrir fichier matrix", "Assigner les niveaux\net les classes", "Professeurs", "Animateurs", "Récapitulatif", "Listes"]
        for action in actions:
            btn = ctk.CTkButton(
                menu, text=action, fg_color="#89B8E3", hover_color="#A1C9F1",
                text_color="white", corner_radius=6, height=36,
                command=lambda a=action: handle_action(a, menu)
            )
            btn.pack(fill="x", padx=8, pady=4)

        # --- Animation de fondu (Fade-in) ---
        def fade_in(alpha=0.0):
            if alpha < 1.0:
                alpha += 0.1 # Vitesse de l'animation
                menu.attributes("-alpha", alpha)
                menu.after(15, lambda: fade_in(alpha)) # Répète toutes les 15ms

        fade_in()

        # Fermeture automatique
        menu.bind("<FocusOut>", lambda e: menu.destroy())
        menu.focus_force()

    def _augment_matrix_file(matrix_path: str) -> bool:
        """
        Ajoute, dans le fichier matrix, quatre colonnes 'Niveau', 'Ecole', 'Horaire', 'Classe'
        immédiatement à droite de la colonne 'stagiaire' (recherchée dans la première ligne).
        """
        if load_workbook is None:
            messagebox.showerror(
                "Dépendance manquante",
                "La bibliothèque 'openpyxl' n'est pas installée.\n"
                "Installe-la avec : pip install openpyxl",
            )
            return False

        try:
            wb = load_workbook(matrix_path)
            ws = wb.active

            # Recherche de la colonne "stagiaire" (insensible à la casse, tolère accents)
            header_row = 1
            stagiaire_col_index = None
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=header_row, column=col).value
                if isinstance(cell_value, str) and "stagiaire" in cell_value.lower():
                    stagiaire_col_index = col
                    break

            if stagiaire_col_index is None:
                messagebox.showwarning(
                    "Colonne introuvable",
                    "Impossible de trouver une colonne contenant 'stagiaire' dans la première ligne.\n"
                    "Les colonnes 'niveau', 'ecole' et 'classe' n'ont pas été ajoutées.",
                )
                return False

            # Insère 4 colonnes juste à droite de la colonne stagiaire
            insert_at = stagiaire_col_index + 1
            ws.insert_cols(insert_at, amount=4)

            ws.cell(row=header_row, column=insert_at, value="Niveau")
            ws.cell(row=header_row, column=insert_at + 1, value="Ecole")
            ws.cell(row=header_row, column=insert_at + 2, value="Horaire")
            ws.cell(row=header_row, column=insert_at + 3, value="Classe")

            wb.save(matrix_path)
            return True
        except Exception as e:
            messagebox.showerror(
                "Erreur lors de la modification de matrix",
                f"Les colonnes n'ont pas pu être ajoutées.\n\nDétail : {e}",
            )
            return False

    def show_summary_window():
        """Affiche une fenêtre de récapitulatif moderne avec les statistiques des écoles, classes et élèves"""

        # Vérifier qu'une semaine est sélectionnée
        week_label = selected_week.get()
        if not week_label:
            messagebox.showwarning(
                "Aucune semaine sélectionnée",
                "Veuillez d'abord sélectionner une semaine.",
                parent=app
            )
            return

        week_num = week_label.split()[-1]
        week_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"semaine_{week_num}")

        # Analyser les données des écoles
        school_data = analyze_school_classes(week_folder)

        # Analyser les données du matrix
        matrix_path = os.path.join(week_folder, "matrix.xlsx")
        matrix_data = analyze_matrix_assignments(matrix_path)

        # Créer la fenêtre de récapitulatif
        summary_window = ctk.CTkToplevel(app)
        summary_window.title("Récapitulatif des données")
        summary_window.geometry("1200x800")
        summary_window.resizable(True, True)

        # Frame principal avec scroll
        main_frame = ctk.CTkScrollableFrame(summary_window, fg_color="#f0f2f5")
        main_frame.pack(fill="both", expand=True, padx=0, pady=0)

        # Header avec titre
        header_frame = ctk.CTkFrame(main_frame, fg_color="#2E3440", corner_radius=0, height=80)
        header_frame.pack(fill="x", pady=(0, 30))
        header_frame.pack_propagate(False)

        title_label = ctk.CTkLabel(
            header_frame,
            text="📊 Récapitulatif des Données",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color="white"
        )
        title_label.pack(pady=20)

        # Calculer les statistiques
        school_names = {
            'ecole_a': 'École A',
            'ecole_b': 'École B',
            'ecole_c_cs': 'École C (CS)',
            'ecole_c_ci': 'École C (CI)',
            'ecole_morning': 'Morning',
            'ecole_premium_cs': 'Premium (CS)',
            'ecole_premium_ci': 'Premium (CI)'
        }

        # Calculer les totaux
        total_students = 0
        students_with_level = 0
        students_without_level = 0
        level_counts = {}
        level_ci_counts = {}
        school_stats = {}

        for school_key, horaires in school_data.items():
            school_total = 0
            school_classes = 0
            
            for horaire in horaires:
                for classe in horaire['classes']:
                    nb_eleves = classe['nb_eleves']
                    niveau = classe['niveau']
                    is_ci = 'ci' in school_key.lower()
                    
                    total_students += nb_eleves
                    school_total += nb_eleves
                    school_classes += 1
                    
                    if niveau and niveau.strip():
                        students_with_level += nb_eleves
                        level_counts[niveau] = level_counts.get(niveau, 0) + nb_eleves
                        if is_ci:
                            level_ci_counts[niveau] = level_ci_counts.get(niveau, 0) + nb_eleves
                    else:
                        students_without_level += nb_eleves
            
            if school_total > 0:
                school_stats[school_key] = {
                    'name': school_names.get(school_key, school_key),
                    'students': school_total,
                    'classes': school_classes
                }

        # Container pour les cards du header
        cards_container = ctk.CTkFrame(main_frame, fg_color="transparent")
        cards_container.pack(fill="x", padx=30, pady=(0, 30))

        def create_stat_card(parent, title, value, icon, color, col, row):
            """Crée une card de statistique"""
            card = ctk.CTkFrame(parent, fg_color="white", corner_radius=15, border_width=0)
            card.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")
            
            # Icon et titre
            top_frame = ctk.CTkFrame(card, fg_color="transparent")
            top_frame.pack(fill="x", padx=20, pady=(20, 5))
            
            icon_label = ctk.CTkLabel(
                top_frame,
                text=icon,
                font=ctk.CTkFont(size=32),
                text_color=color
            )
            icon_label.pack(anchor="w")
            
            title_label = ctk.CTkLabel(
                top_frame,
                text=title,
                font=ctk.CTkFont(size=13),
                text_color="#64748b",
                anchor="w"
            )
            title_label.pack(anchor="w", pady=(5, 0))
            
            # Valeur
            value_label = ctk.CTkLabel(
                card,
                text=str(value),
                font=ctk.CTkFont(size=36, weight="bold"),
                text_color="#1e293b"
            )
            value_label.pack(padx=20, pady=(0, 20), anchor="w")
            
            return card

        # Configuration de la grille pour les cards
        for i in range(6):
            cards_container.grid_columnconfigure(i, weight=1, uniform="cards")

        # Première ligne de cards - Statistiques principales
        create_stat_card(cards_container, "Total Élèves", total_students, "👥", "#3b82f6", 0, 0)
        create_stat_card(cards_container, "Sans Niveau", students_without_level, "❓", "#ef4444", 1, 0)
        create_stat_card(cards_container, "Avec Niveau", students_with_level, "✅", "#10b981", 2, 0)

        # Cards pour les niveaux (max 3 par ligne)
        sorted_levels = sorted(level_counts.keys())
        for idx, niveau in enumerate(sorted_levels[:3]):
            total = level_counts[niveau]
            ci = level_ci_counts.get(niveau, 0)
            title = f"Niveau {niveau}"
            value_text = f"{total}"
            if ci > 0:
                value_text += f"\n({ci} CI)"
            
            col = idx + 3
            card = ctk.CTkFrame(cards_container, fg_color="white", corner_radius=15)
            card.grid(row=0, column=col, padx=10, pady=10, sticky="nsew")
            
            top_frame = ctk.CTkFrame(card, fg_color="transparent")
            top_frame.pack(fill="x", padx=20, pady=(20, 5))
            
            icon_label = ctk.CTkLabel(top_frame, text="🎓", font=ctk.CTkFont(size=32))
            icon_label.pack(anchor="w")
            
            title_label = ctk.CTkLabel(
                top_frame,
                text=title,
                font=ctk.CTkFont(size=13),
                text_color="#64748b"
            )
            title_label.pack(anchor="w", pady=(5, 0))
            
            value_label = ctk.CTkLabel(
                card,
                text=str(total),
                font=ctk.CTkFont(size=36, weight="bold"),
                text_color="#1e293b"
            )
            value_label.pack(padx=20, pady=(0, 5), anchor="w")
            
            if ci > 0:
                ci_label = ctk.CTkLabel(
                    card,
                    text=f"{ci} élèves CI",
                    font=ctk.CTkFont(size=11),
                    text_color="#6366f1"
                )
                ci_label.pack(padx=20, pady=(0, 15), anchor="w")
            else:
                ctk.CTkLabel(card, text="").pack(pady=(0, 15))

        # Deuxième ligne pour les niveaux restants si nécessaire
        if len(sorted_levels) > 3:
            for idx, niveau in enumerate(sorted_levels[3:]):
                total = level_counts[niveau]
                ci = level_ci_counts.get(niveau, 0)
                title = f"Niveau {niveau}"
                
                col = idx
                card = ctk.CTkFrame(cards_container, fg_color="white", corner_radius=15)
                card.grid(row=1, column=col, padx=10, pady=10, sticky="nsew")
                
                top_frame = ctk.CTkFrame(card, fg_color="transparent")
                top_frame.pack(fill="x", padx=20, pady=(20, 5))
                
                icon_label = ctk.CTkLabel(top_frame, text="🎓", font=ctk.CTkFont(size=32))
                icon_label.pack(anchor="w")
                
                title_label = ctk.CTkLabel(
                    top_frame,
                    text=title,
                    font=ctk.CTkFont(size=13),
                    text_color="#64748b"
                )
                title_label.pack(anchor="w", pady=(5, 0))
                
                value_label = ctk.CTkLabel(
                    card,
                    text=str(total),
                    font=ctk.CTkFont(size=36, weight="bold"),
                    text_color="#1e293b"
                )
                value_label.pack(padx=20, pady=(0, 5), anchor="w")
                
                if ci > 0:
                    ci_label = ctk.CTkLabel(
                        card,
                        text=f"{ci} élèves CI",
                        font=ctk.CTkFont(size=11),
                        text_color="#6366f1"
                    )
                    ci_label.pack(padx=20, pady=(0, 15), anchor="w")
                else:
                    ctk.CTkLabel(card, text="").pack(pady=(0, 15))

        # Section des écoles
        schools_label = ctk.CTkLabel(
            main_frame,
            text="🏫 Écoles",
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color="#1e293b"
        )
        schools_label.pack(padx=30, pady=(10, 20), anchor="w")

        # Container pour les cards d'écoles
        schools_container = ctk.CTkFrame(main_frame, fg_color="transparent")
        schools_container.pack(fill="x", padx=30, pady=(0, 30))

        # Configuration de la grille (3 écoles par ligne)
        for i in range(3):
            schools_container.grid_columnconfigure(i, weight=1, uniform="schools")

        # Créer les cards d'écoles
        school_icons = {
            'ecole_a': '🅰️',
            'ecole_b': '🅱️',
            'ecole_c_cs': '©️',
            'ecole_c_ci': '©️',
            'ecole_morning': '🌅',
            'ecole_premium_cs': '⭐',
            'ecole_premium_ci': '⭐'
        }

        school_colors = {
            'ecole_a': '#3b82f6',
            'ecole_b': '#8b5cf6',
            'ecole_c_cs': '#ec4899',
            'ecole_c_ci': '#f43f5e',
            'ecole_morning': '#f59e0b',
            'ecole_premium_cs': '#10b981',
            'ecole_premium_ci': '#14b8a6'
        }

        sorted_schools = sorted(school_stats.items(), key=lambda x: x[1]['students'], reverse=True)
        
        for idx, (school_key, stats) in enumerate(sorted_schools):
            row = idx // 3
            col = idx % 3
            
            card = ctk.CTkFrame(schools_container, fg_color="white", corner_radius=15)
            card.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")
            
            # Header de la card avec couleur
            color = school_colors.get(school_key, "#64748b")
            header = ctk.CTkFrame(card, fg_color=color, corner_radius=15, height=60)
            header.pack(fill="x", padx=0, pady=0)
            header.pack_propagate(False)
            
            icon = school_icons.get(school_key, '🏫')
            header_label = ctk.CTkLabel(
                header,
                text=f"{icon}  {stats['name']}",
                font=ctk.CTkFont(size=16, weight="bold"),
                text_color="white"
            )
            header_label.pack(pady=15, padx=20, anchor="w")
            
            # Stats de l'école
            stats_frame = ctk.CTkFrame(card, fg_color="transparent")
            stats_frame.pack(fill="x", padx=20, pady=20)
            
            # Nombre de classes
            classes_frame = ctk.CTkFrame(stats_frame, fg_color="#f8fafc", corner_radius=10)
            classes_frame.pack(fill="x", pady=(0, 10))
            
            classes_label = ctk.CTkLabel(
                classes_frame,
                text="📚 Classes",
                font=ctk.CTkFont(size=12),
                text_color="#64748b"
            )
            classes_label.pack(side="left", padx=15, pady=12)
            
            classes_value = ctk.CTkLabel(
                classes_frame,
                text=str(stats['classes']),
                font=ctk.CTkFont(size=20, weight="bold"),
                text_color="#1e293b"
            )
            classes_value.pack(side="right", padx=15, pady=12)
            
            # Nombre d'élèves
            students_frame = ctk.CTkFrame(stats_frame, fg_color="#f8fafc", corner_radius=10)
            students_frame.pack(fill="x")
            
            students_label = ctk.CTkLabel(
                students_frame,
                text="👥 Élèves",
                font=ctk.CTkFont(size=12),
                text_color="#64748b"
            )
            students_label.pack(side="left", padx=15, pady=12)
            
            students_value = ctk.CTkLabel(
                students_frame,
                text=str(stats['students']),
                font=ctk.CTkFont(size=20, weight="bold"),
                text_color="#1e293b"
            )
            students_value.pack(side="right", padx=15, pady=12)

        # Bouton de fermeture
        close_button = ctk.CTkButton(
            main_frame,
            text="✕ Fermer",
            command=summary_window.destroy,
            fg_color="#ef4444",
            hover_color="#dc2626",
            text_color="white",
            corner_radius=10,
            height=45,
            font=ctk.CTkFont(size=14, weight="bold")
        )
        close_button.pack(pady=(10, 30), padx=30, fill="x")

        # Centrer la fenêtre correctement
        summary_window.update_idletasks()
        window_width = summary_window.winfo_width()
        window_height = summary_window.winfo_height()
        screen_width = summary_window.winfo_screenwidth()
        screen_height = summary_window.winfo_screenheight()
        
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        # S'assurer que la fenêtre n'est pas trop haute
        if y < 20:
            y = 20
        
        summary_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    def handle_action(action, menu):
        # Ferme le petit menu d'actions
        menu.destroy()

        # Importer un fichier matrix dans le dossier de la semaine sélectionnée
        if action.startswith("Importer fichier matrix"):
            file_path = filedialog.askopenfilename(
                parent=app,
                title="Sélectionner le fichier matrix",
                filetypes=[("Fichiers Excel", "*.xlsx *.xls"), ("Tous les fichiers", "*.*")],
            )
            if not file_path:
                return

            # Afficher la fenêtre de chargement
            loading_popup = show_loading_window_for_matrix_import(app)

            # Forcer plusieurs mises à jour pour s'assurer que la fenêtre de chargement s'affiche immédiatement
            loading_popup.update()
            loading_popup.update_idletasks()
            app.update()
            app.update_idletasks()

            # Variable pour contrôler la boucle de mise à jour de l'animation
            keep_animating = [True]  # Utiliser une liste pour permettre la modification dans les fonctions imbriquées

            def keep_animation_alive():
                """Fonction récursive pour maintenir l'animation de la fenêtre de chargement"""
                if keep_animating[0] and loading_popup and loading_popup.winfo_exists():
                    loading_popup.update()
                    app.update_idletasks()
                    app.after(50, keep_animation_alive)  # Mettre à jour toutes les 50ms

            # Démarrer la boucle d'animation
            keep_animation_alive()

            def perform_matrix_import():
                try:
                    print("DÉBUT DE L'IMPORTATION MATRIX")  # Debug

                    # Chemin cible pour la semaine actuellement sélectionnée
                    dest_path = _get_matrix_path_for_selected_week()
                    print(f"Chemin destination: {dest_path}")  # Debug
                    os.makedirs(os.path.dirname(dest_path), exist_ok=True)

                    # Copie du fichier choisi vers le dossier de la semaine (écrase l'ancien si présent)
                    print("Copie du fichier...")  # Debug
                    shutil.copy2(file_path, dest_path)

                    # Ajoute les colonnes 'Niveau', 'Ecole', 'Horaire', 'Classe' après la colonne 'stagiaire'
                    print("Ajout des colonnes...")  # Debug
                    success = _augment_matrix_file(dest_path)
                    print(f"Succès de l'ajout des colonnes: {success}")  # Debug

                    if success:
                        # Vider tous les fichiers Excel des écoles (supprimer toutes les classes)
                        print("Vidage des fichiers écoles...")  # Debug
                        clear_all_school_files(dest_path)

                        # Remettre à zéro les classes assignées dans personnel.json
                        print("Remise à zéro des classes dans personnel.json...")  # Debug
                        clear_personnel_classes(os.path.dirname(dest_path))

                        # Rafraîchir l'affichage du dashboard pour qu'il soit vide
                        print("Rafraîchissement du dashboard...")  # Debug
                        week_folder = os.path.dirname(dest_path)
                        school_data = analyze_school_classes(week_folder)
                        create_classes_dashboard(content, school_data, week_folder)

                        print("Importation terminée avec succès")  # Debug

                        # Petit délai avant de fermer pour s'assurer que l'utilisateur voit la fin
                        def close_loading_window():
                            keep_animating[0] = False
                            try:
                                if loading_popup and loading_popup.winfo_exists():
                                    loading_popup.destroy()
                                    print("Fenêtre de chargement fermée")  # Debug
                            except Exception as e:
                                print(f"Erreur lors de la fermeture: {e}")  # Debug

                        app.after(500, close_loading_window)  # Délai de 500ms

                        messagebox.showinfo(
                            "Import réussi",
                            "Le fichier matrix a été importé avec succès.\n"
                            "Les colonnes 'Niveau', 'Ecole', 'Horaire' et 'Classe' ont été ajoutées.\n"
                            "Tous les fichiers des écoles ont été vidés.\n"
                            "Les classes assignées des intervenants ont été remises à zéro.",
                        )
                    else:
                        print("Échec de l'ajout des colonnes")  # Debug
                        # En cas d'échec, fermer la fenêtre de chargement
                        keep_animating[0] = False
                        try:
                            if loading_popup and loading_popup.winfo_exists():
                                loading_popup.destroy()
                        except:
                            pass

                    # Rafraîchit le statut visuel du badge MATRIX
                    update_matrix_status()

                except Exception as e:
                    print(f"ERREUR lors de l'importation: {e}")  # Debug
                    # Arrêter l'animation et fermer la fenêtre en cas d'erreur
                    keep_animating[0] = False
                    try:
                        if loading_popup and loading_popup.winfo_exists():
                            loading_popup.destroy()
                    except:
                        pass
                    messagebox.showerror("Erreur", f"Impossible d'importer le fichier matrix : {e}", parent=app)

            # Démarrer l'importation après un court délai
            app.after(10, perform_matrix_import)
            return

        # Ouvrir le fichier matrix de la semaine sélectionnée dans Excel / OpenOffice
        if action.startswith("Ouvrir fichier matrix"):
            matrix_path = _get_matrix_path_for_selected_week()
            if not os.path.exists(matrix_path):
                messagebox.showwarning(
                    "Matrix introuvable",
                    "Aucun fichier 'matrix.xlsx' n'a été trouvé pour cette semaine.\n"
                    "Vous pouvez d'abord l'importer via 'Importer fichier matrix'.",
                )
                return

            try:
                # Sur Windows, ouvre avec l'application associée (Excel, LibreOffice, etc.)
                os.startfile(matrix_path)
            except Exception as e:
                messagebox.showerror(
                    "Erreur d'ouverture",
                    f"Impossible d'ouvrir le fichier matrix.\n\nDétail : {e}",
                )
            return

        # Assigner les niveaux aux élèves - lance le script d'assignation
        if action.startswith("Assigner les niveaux"):
            matrix_path = _get_matrix_path_for_selected_week()
            if not os.path.exists(matrix_path):
                messagebox.showwarning(
                    "Matrix introuvable",
                    "Aucun fichier 'matrix.xlsx' n'a été trouvé pour cette semaine.\n"
                    "Vous devez d'abord l'importer via 'Importer fichier matrix'.",
                )
                return
            
            # Chemin absolu du script d'assignation
            script_path = os.path.join(os.path.dirname(__file__), "Assignation des Niveaux.py")
            if not os.path.exists(script_path):
                messagebox.showerror(
                    "Script introuvable",
                    f"Le script 'Assignation des Niveaux.py' est introuvable.\n"
                    f"Chemin attendu : {script_path}",
                )
                return
            
            # Afficher la fenêtre de chargement animée
            show_loading_window(app)

            try:
                # Lancer le script avec le chemin du fichier matrix en argument
                # Utiliser sys.executable pour utiliser le même interpréteur Python
                subprocess.Popen([sys.executable, script_path, matrix_path])

                # Fermer la fenêtre de chargement après un délai
                def close_loading():
                    # Cette fonction sera appelée depuis la fenêtre elle-même
                    pass

                # La fenêtre se fermera automatiquement après 1500ms
                app.after(1500, lambda: None)  # Délai pour laisser le temps à la fenêtre de se charger  # Réinitialise après 500ms
                
            except Exception as e:
                # Réinitialiser immédiatement en cas d'erreur
                app.configure(cursor="")
                app.update_idletasks()
                messagebox.showerror(
                    "Erreur de lancement",
                    f"Impossible de lancer le script d'assignation des niveaux.\n\nDétail : {e}",
                )
            return

        # Gestion des professeurs
        if action.startswith("Professeurs"):
            # Obtenir le dossier de la semaine sélectionnée
            week_label = selected_week.get()
            if not week_label:
                messagebox.showwarning(
                    "Aucune semaine sélectionnée",
                    "Veuillez d'abord sélectionner une semaine.",
                    parent=app
                )
                return

            week_num = week_label.split()[-1]
            week_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"semaine_{week_num}")

            # Fonction callback pour rafraîchir le dashboard quand les données changent
            def refresh_dashboard_callback():
                try:
                    school_data = analyze_school_classes(week_folder)
                    create_classes_dashboard(content, school_data, week_folder)
                except Exception as e:
                    print(f"Erreur lors du rafraîchissement du dashboard: {e}")

            # Ouvrir la fenêtre de gestion des professeurs
            PersonnelManager(app, week_folder, personnel_type="professeurs", data_changed_callback=refresh_dashboard_callback)
            return

        # Gestion des animateurs
        if action.startswith("Animateurs"):
            # Obtenir le dossier de la semaine sélectionnée
            week_label = selected_week.get()
            if not week_label:
                messagebox.showwarning(
                    "Aucune semaine sélectionnée",
                    "Veuillez d'abord sélectionner une semaine.",
                    parent=app
                )
                return

            week_num = week_label.split()[-1]
            week_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"semaine_{week_num}")

            # Fonction callback pour rafraîchir le dashboard quand les données changent
            def refresh_dashboard_callback():
                try:
                    school_data = analyze_school_classes(week_folder)
                    create_classes_dashboard(content, school_data, week_folder)
                except Exception as e:
                    print(f"Erreur lors du rafraîchissement du dashboard: {e}")

            # Ouvrir la fenêtre de gestion des animateurs
            PersonnelManager(app, week_folder, personnel_type="animateurs", data_changed_callback=refresh_dashboard_callback)
            return

        # Afficher le récapitulatif des statistiques
        if action == "Récapitulatif":
            show_summary_window()
            return

        # Bouton "Listes" (pas d'action pour l'instant)
        if action == "Listes":
            # Action à implémenter plus tard
            return

        # Autres actions (placeholder)

    def calculate_optimal_layout(total_items, min_per_row=2, max_per_row=4):
        """
        Calcule la disposition optimale pour un nombre d'items.
        
        Args:
            total_items: Nombre total d'items à afficher
            min_per_row: Nombre minimum d'items par ligne
            max_per_row: Nombre maximum d'items par ligne
        
        Returns:
            Nombre d'items par ligne optimal
        """
        if total_items <= min_per_row:
            return total_items
        elif total_items <= max_per_row:
            return min_per_row
        else:
            # Calculer pour avoir des lignes équilibrées
            rows = (total_items + max_per_row - 1) // max_per_row
            return min((total_items + rows - 1) // rows, max_per_row)

    def create_classes_dashboard(parent_frame, school_data, week_folder):
        """
        Crée une interface moderne pour afficher les classes avec disposition dynamique.
        Affiche toujours les cadres des écoles et horaires, même sans classes.

        Met également à jour les fichiers Excel des écoles avec les élèves assignés depuis matrix.xlsx
        """

        # 1) Ouvrir le fichier matrix.xlsx et analyser les assignations des élèves
        matrix_path = os.path.join(week_folder, "matrix.xlsx")
        matrix_assignments = analyze_matrix_assignments(matrix_path)

        # 2) Afficher seulement les élèves complètement assignés (école + horaire + classe)
        # (Affichage supprimé pour nettoyer la console)

        # 3) Mettre à jour les fichiers Excel des écoles avec les élèves assignés
        update_school_files_with_assignments(week_folder, matrix_assignments)

        # Nettoyer le contenu existant
        for widget in parent_frame.winfo_children():
            widget.destroy()
        
        # Désactiver temporairement les mises à jour automatiques pour créer tous les widgets d'un coup
        # On réactivera les mises à jour à la fin

        # Configuration du layout principal
        parent_frame.grid_columnconfigure(0, weight=1)
        parent_frame.grid_rowconfigure(0, weight=1)

        # Frame principal avec scroll
        main_scrollable = ctk.CTkScrollableFrame(
            parent_frame,
            fg_color="#f8fafc",
            corner_radius=0
        )
        main_scrollable.grid(row=0, column=0, sticky="nsew", padx=10, pady=5)
        main_scrollable.grid_columnconfigure(0, weight=1)

        # Attacher un event listener pour détecter les changements de scroll
        def on_scroll_change(event=None):
            # Essayer de déterminer le numéro de semaine depuis le week_folder
            week_number = None
            try:
                import os
                week_folder_name = os.path.basename(week_folder)
                if week_folder_name.startswith('semaine_'):
                    week_number = week_folder_name.split('_')[-1]
            except:
                pass

            save_dashboard_scroll_position(main_scrollable, week_number)

        # Utiliser after pour attacher l'event listener après que le widget soit complètement créé
        def attach_scroll_listener():
            try:
                if hasattr(main_scrollable, '_parent_canvas'):
                    canvas = main_scrollable._parent_canvas

                    # Attacher les events de scroll
                    canvas.bind("<Configure>", lambda e: on_scroll_change(e))
                    canvas.bind("<ButtonRelease-1>", lambda e: on_scroll_change(e))
                    canvas.bind("<MouseWheel>", lambda e: on_scroll_change(e))
                    canvas.bind("<B1-Motion>", lambda e: on_scroll_change(e))
                    canvas.bind("<Motion>", lambda e: on_scroll_change(e))

                    # Events sur le scrollable frame aussi
                    main_scrollable.bind("<Configure>", lambda e: on_scroll_change(e))
                    main_scrollable.bind("<MouseWheel>", lambda e: on_scroll_change(e))

                    pass  # Listeners attachés silencieusement
                else:
                    pass  # Pas de canvas, rien à faire
            except Exception as e:
                pass  # Erreur silencieuse lors de l'attachement

        parent_frame.after(100, attach_scroll_listener)

        current_row = 0

        # Mapping des écoles
        school_mapping = {
            'A': ('ecole_a', 'A', '#3b82f6', '#eff6ff'),
            'B': ('ecole_b', 'B', '#10b981', '#f0fdf4'),
            'C/CS': ('ecole_c_cs', 'C/CS', '#f59e0b', '#fffbeb'),
            'C/CI': ('ecole_c_ci', 'C/CI', '#8b5cf6', '#f3e8ff'),
            'Morning': ('ecole_morning', 'Morning', '#ef4444', '#fef2f2'),
            'Premium/CS': ('ecole_premium_cs', 'Premium/CS', '#06b6d4', '#ecfeff'),
            'Premium/CI': ('ecole_premium_ci', 'Premium/CI', '#f97316', '#fff7ed')
        }

        for school_name, (school_key, display_name, school_color, school_bg) in school_mapping.items():
            school_info = school_data.get(school_key, [])


            # Vérifier si l'école est activée dans les filtres
            if not school_filters[school_name].get():
                continue

            # Section pour l'école - affichée même sans données
            school_frame = ctk.CTkFrame(
                main_scrollable,
                fg_color="white",
                corner_radius=10,
                border_width=2,
                border_color=school_color
            )
            school_frame.grid(row=current_row, column=0, sticky="ew", pady=(4, 4), padx=5)
            school_frame.grid_columnconfigure(0, weight=1)
            school_frame.grid_rowconfigure(1, weight=1)

            # En-tête de l'école
            school_header = ctk.CTkFrame(
                school_frame,
                fg_color=school_bg,
                corner_radius=10
            )
            school_header.grid(row=0, column=0, sticky="ew", padx=10, pady=(2, 2))
            school_header.grid_columnconfigure(1, weight=1)

            school_icon = ctk.CTkLabel(
                school_header,
                text="🏫",
                font=("Arial", 16),
                text_color=school_color
            )
            school_icon.grid(row=0, column=0, padx=(3, 5), pady=1)

            school_title = ctk.CTkLabel(
                school_header,
                text=f"École {display_name}",
                font=("Arial", 14, "bold"),
                text_color=school_color
            )
            school_title.grid(row=0, column=1, sticky="w", pady=1)

            # Statistiques
            stats_frame = ctk.CTkFrame(school_header, fg_color="transparent")
            stats_frame.grid(row=0, column=2, sticky="e", padx=(10, 5))

            total_eleves = sum(len(horaire.get('classes', [])) for horaire in school_info)
            total_classes = sum(len(horaire.get('classes', [])) for horaire in school_info)

            stats_label = ctk.CTkLabel(
                stats_frame,
                text=f"👥 {total_eleves} élèves • 📚 {total_classes} classes",
                font=("Arial", 11),
                text_color=school_color
            )
            stats_label.pack()

            # Boutons d'action
            buttons_frame = ctk.CTkFrame(school_header, fg_color="transparent")
            buttons_frame.grid(row=0, column=3, sticky="e", padx=(5, 0))

            def soften_color(color, factor=0.6):
                if color.startswith('#'):
                    r = int(color[1:3], 16)
                    g = int(color[3:5], 16)
                    b = int(color[5:7], 16)
                    r = int(r + (255 - r) * (1 - factor))
                    g = int(g + (255 - g) * (1 - factor))
                    b = int(b + (255 - b) * (1 - factor))
                    return f"#{r:02x}{g:02x}{b:02x}"
                return color

            # Créer les boutons
            button_configs = [
                ("👨‍🏫 Profs", lambda sn=school_name: None),
                ("🎭 Anims", lambda sn=school_name: None),
                ("👥 Élèves", lambda sn=school_name: None),
                ("📄 Listes", lambda sn=school_name: None)
            ]

            for text, command in button_configs:
                btn = ctk.CTkButton(
                    buttons_frame,
                    text=text,
                    font=("Arial", 9, "bold"),
                    height=24,
                    width=70,
                    fg_color=soften_color(school_color, 0.5),
                    hover_color=soften_color(school_color, 0.3),
                    text_color="white",
                    command=command
                )
                btn.pack(side="left", padx=(0, 2))

            # Conteneur pour les horaires - TOUJOURS AFFICHÉ
            horaires_container = ctk.CTkFrame(
                school_frame,
                fg_color="transparent"
            )
            horaires_container.grid(row=1, column=0, sticky="nsew", padx=10, pady=(2, 10))

            # Si aucune donnée d'horaire pour cette école
            if not school_info or len(school_info) == 0:
                no_horaires_label = ctk.CTkLabel(
                    horaires_container,
                    text="📅 Aucun horaire configuré pour cette école",
                    font=("Arial", 11, "italic"),
                    text_color="#9ca3af"
                )
                no_horaires_label.pack(pady=20)
            
            # Calculer le nombre optimal d'horaires par ligne
            total_horaires = len(school_info) if school_info else 0
            if total_horaires > 0:
                horaires_per_row = calculate_optimal_layout(total_horaires, min_per_row=2, max_per_row=3)

                # Configurer les colonnes dynamiquement
                for col in range(horaires_per_row):
                    horaires_container.grid_columnconfigure(col, weight=1)

                horaire_row = 0
                horaire_col = 0

                # Fonction pour extraire l'heure de début
                def extract_start_hour(horaire_str):
                    """Extrait l'heure de début pour le tri (ex: '8h15-9h15' -> 8.25)"""
                    try:
                        # Extraire la partie avant le '-'
                        start = horaire_str.split('-')[0].strip()
                        # Séparer heures et minutes
                        if 'h' in start:
                            parts = start.split('h')
                            hours = int(parts[0])
                            minutes = int(parts[1]) if len(parts) > 1 and parts[1] else 0
                            return hours + minutes / 60.0
                        return 0
                    except:
                        return 0

                # Trier les horaires par heure de début
                sorted_school_info = sorted(school_info, key=lambda x: extract_start_hour(x.get('horaire', '')))

                # Pour chaque horaire de cette école
                for horaire_info in sorted_school_info:
                    horaire = horaire_info.get('horaire', 'Horaire inconnu')
                    intervenant = horaire_info.get('intervenant', 'Non spécifié')
                    type_intervenant = horaire_info.get('type_intervenant', 'professeur')
                    classes = horaire_info.get('classes', [])

                    # Frame pour cet horaire - TOUJOURS AFFICHÉ
                    horaire_frame = ctk.CTkFrame(
                        horaires_container,
                        fg_color="#f8fafc",
                        corner_radius=8,
                        border_width=1,
                        border_color="#e5e7eb"
                    )
                    horaire_frame.grid(
                        row=horaire_row, 
                        column=horaire_col, 
                        sticky="nsew", 
                        padx=(0, 5) if horaire_col < horaires_per_row - 1 else 0, 
                        pady=(0, 5)
                    )
                    horaire_frame.grid_columnconfigure(0, weight=1)

                    # En-tête de l'horaire
                    horaire_header = ctk.CTkFrame(
                        horaire_frame,
                        fg_color="white",
                        corner_radius=8
                    )
                    horaire_header.grid(row=0, column=0, sticky="ew", padx=8, pady=(8, 5))
                    horaire_header.grid_columnconfigure(1, weight=1)

                    # Icône selon le type d'intervenant
                    intervenant_icon = "👨‍🏫" if type_intervenant == "professeur" else "🎭"
                    icon_label = ctk.CTkLabel(
                        horaire_header,
                        text=intervenant_icon,
                        font=("Arial", 14),
                        text_color=school_color
                    )
                    icon_label.grid(row=0, column=0, padx=(5, 8), pady=5)

                    # Informations de l'horaire
                    horaire_text = f"Horaire : {horaire} : {type_intervenant.title()}"
                    horaire_label = ctk.CTkLabel(
                        horaire_header,
                        text=horaire_text,
                        font=("Arial", 11, "bold"),
                        text_color="#374151",
                        anchor="w"
                    )
                    horaire_label.grid(row=0, column=1, sticky="w", pady=5)

                    # Conteneur pour les boutons d'action (ajouter/supprimer)
                    action_buttons_frame = ctk.CTkFrame(horaire_header, fg_color="transparent")
                    action_buttons_frame.grid(row=0, column=2, padx=(10, 5), pady=5)

                    # Bouton pour ajouter une classe
                    add_class_btn = ctk.CTkButton(
                        action_buttons_frame,
                        text="➕",
                        font=("Arial", 10),
                        width=15,
                        height=15,
                        fg_color=soften_color(school_color, 0.5),
                        hover_color=soften_color(school_color, 0.3),
                        text_color="white",
                        corner_radius=12,
                        command=lambda h=horaire, sk=school_key, dn=display_name, sc=school_color: open_add_class_dialog(
                            h, sk, dn, sc, week_folder,
                            refresh_callback=lambda: create_classes_dashboard(parent_frame, analyze_school_classes(week_folder), week_folder)
                        )
                    )
                    add_class_btn.pack(side="left", padx=(0, 3))

                    # Bouton pour supprimer toutes les classes de l'horaire
                    delete_all_btn = ctk.CTkButton(
                        action_buttons_frame,
                        text="➖",
                        font=("Arial", 10),
                        width=15,
                        height=15,
                        fg_color=soften_color(school_color, 0.5),
                        hover_color=soften_color(school_color, 0.3),
                        text_color="white",
                        corner_radius=12,
                        command=lambda h=horaire, sn=display_name: confirm_delete_all_classes_from_horaire(h, sn, week_folder)
                    )
                    delete_all_btn.pack(side="left")

                    # Classes de cet horaire
                    if classes:
                        classes_container = ctk.CTkFrame(
                            horaire_frame,
                            fg_color="transparent"
                        )
                        classes_container.grid(row=1, column=0, sticky="nsew", padx=8, pady=(0, 8))

                        # Calculer le nombre optimal de classes par ligne
                        total_classes = len(classes)
                        classes_per_row = calculate_optimal_layout(total_classes, min_per_row=2, max_per_row=4)

                        # Configurer les colonnes dynamiquement
                        for col in range(classes_per_row):
                            classes_container.grid_columnconfigure(col, weight=1)

                        classe_row = 0
                        classe_col = 0

                        # Créer des cartes pour chaque classe
                        for classe_info in classes:
                            classe_nom = classe_info.get('nom_classe', 'Classe inconnue')
                            nb_eleves = classe_info.get('nb_eleves', 0)
                            niveau = classe_info.get('niveau', '')
                            eleves = classe_info.get('eleves', [])
                            intervenant_classe = classe_info.get('intervenant', 'Non spécifié')  # Intervenant spécifique à la classe

                            # Frame de la classe - TAILLE FIXE UNIFORME
                            classe_card = ctk.CTkFrame(
                                classes_container,
                                fg_color="white",
                                corner_radius=6,
                                border_width=1,
                                border_color="#e5e7eb",
                                width=160,
                                height=110
                            )
                            classe_card.grid(
                                row=classe_row,
                                column=classe_col,
                                padx=(0, 5) if classe_col < classes_per_row - 1 else 0,
                                pady=(0, 5),
                                sticky="nsew"
                            )
                            classe_card.grid_propagate(False)

                            # Fonction pour gérer l'effet hover
                            def on_enter(event, frame=classe_card, color=school_bg):
                                """Change la couleur de fond au survol"""
                                frame.configure(fg_color=color)

                            def on_leave(event, frame=classe_card):
                                """Restaure la couleur de fond d'origine"""
                                frame.configure(fg_color="white")

                            # Lier les événements de survol
                            classe_card.bind("<Enter>", on_enter)
                            classe_card.bind("<Leave>", on_leave)

                            # IMPORTANT : Propager les événements aux widgets enfants
                            def bind_hover_to_children(widget, enter_func, leave_func):
                                """Applique les événements hover à tous les enfants"""
                                widget.bind("<Enter>", enter_func)
                                widget.bind("<Leave>", leave_func)
                                for child in widget.winfo_children():
                                    bind_hover_to_children(child, enter_func, leave_func)

                            # Nom de la classe
                            classe_name_label = ctk.CTkLabel(
                                classe_card,
                                text=classe_nom,
                                font=("Arial", 11, "bold"),
                                text_color="#374151"
                            )
                            classe_name_label.pack(padx=8, pady=(5, 2))

                            # Niveau et intervenant
                            if niveau:
                                niveau_classe_label = ctk.CTkLabel(
                                    classe_card,
                                    text=f"📚 {niveau}\n👨‍🏫 {intervenant_classe}",
                                    font=("Arial", 9),
                                    text_color="#6b7280"
                                )
                                niveau_classe_label.pack(padx=8, pady=(0, 3))

                            # Nombre d'élèves avec indicateur circulaire
                            eleves_frame = ctk.CTkFrame(
                                classe_card,
                                fg_color="transparent"
                            )
                            eleves_frame.pack(padx=8, pady=(0, 5))

                            circle_frame = ctk.CTkFrame(
                                eleves_frame,
                                fg_color=school_color,
                                corner_radius=12,
                                width=24,
                                height=24
                            )
                            circle_frame.pack(side="left")
                            circle_frame.pack_propagate(False)

                            eleves_count_label = ctk.CTkLabel(
                                circle_frame,
                                text=str(nb_eleves),
                                font=("Arial", 10, "bold"),
                                text_color="white"
                            )
                            eleves_count_label.pack(expand=True)

                            eleves_text_label = ctk.CTkLabel(
                                eleves_frame,
                                text="élèves" if nb_eleves > 1 else "élève",
                                font=("Arial", 9),
                                text_color="#6b7280"
                            )
                            eleves_text_label.pack(side="left", padx=(5, 0))

                            # Appliquer l'effet hover à tous les enfants après leur création
                            bind_hover_to_children(classe_card, on_enter, on_leave)

                            # Ajouter un curseur pointeur pour indiquer que c'est cliquable
                            classe_card.configure(cursor="hand2")

                            # Rendre la carte cliquable
                            setup_classe_card_click(classe_card, classe_info, horaire, intervenant_classe, type_intervenant, school_color, display_name, week_folder)

                            # Gestion de la disposition
                            classe_col += 1
                            if classe_col >= classes_per_row:
                                classe_col = 0
                                classe_row += 1
                    else:
                        # Message si pas de classes pour cet horaire
                        no_classes_label = ctk.CTkLabel(
                            horaire_frame,
                            text="👥 Aucune classe pour cet horaire",
                            font=("Arial", 10, "italic"),
                            text_color="#9ca3af"
                        )
                        no_classes_label.grid(row=1, column=0, padx=10, pady=15)

                    # Gestion de la disposition des horaires
                    horaire_col += 1
                    if horaire_col >= horaires_per_row:
                        horaire_col = 0
                        horaire_row += 1

            current_row += 1

        # Message si aucune école n'est affichée (toutes filtrées)
        if current_row == 0:
            no_data_frame = ctk.CTkFrame(
                main_scrollable,
                fg_color="white",
                corner_radius=15,
                border_width=2,
                border_color="#e5e7eb"
            )
            no_data_frame.grid(row=0, column=0, sticky="ew", pady=20, padx=10)

            no_data_label = ctk.CTkLabel(
                no_data_frame,
                text="🔍 Aucune école sélectionnée\n\nActivez au moins un filtre d'école pour afficher les données.",
                font=("Arial", 13),
                text_color="#6b7280",
                justify="center"
            )
            no_data_label.pack(pady=30, padx=20)
        
        # Forcer une mise à jour complète à la fin pour afficher tous les widgets d'un coup
        parent_frame.update_idletasks()

        # Restaurer la position de scroll sauvegardée
        def delayed_scroll_restore():
            try:
                # Essayer de déterminer le numéro de semaine depuis le week_folder
                week_number = None
                try:
                    import os
                    week_folder_name = os.path.basename(week_folder)
                    if week_folder_name.startswith('semaine_'):
                        week_number = week_folder_name.split('_')[-1]
                except:
                    pass

                restore_dashboard_scroll_position(main_scrollable, week_number)
            except Exception as e:
                print(f"[ERROR] Erreur restauration scroll: {str(e)[:50]}")

        parent_frame.after(200, delayed_scroll_restore)


    def clear_personnel_classes(week_folder):
        """
        Remet à zéro les classes assignées de tous les intervenants dans personnel.json.
        Cette fonction est appelée lors de l'importation d'un nouveau fichier matrix.

        Args:
            week_folder (str): Dossier de la semaine
        """
        personnel_path = os.path.join(week_folder, "personnel.json")

        if not os.path.exists(personnel_path):
            return

        try:
            with open(personnel_path, "r", encoding="utf-8") as f:
                personnel_data = json.load(f)
        except Exception as e:
            print(f"Erreur lors de la lecture de personnel.json: {e}")
            return

        # Pour chaque type d'intervenant (professeurs et animateurs)
        for personnel_type in ["professeurs", "animateurs"]:
            if personnel_type not in personnel_data:
                continue

            # Pour chaque intervenant, vider sa liste de classes
            for intervenant in personnel_data[personnel_type]:
                if "classes" in intervenant:
                    intervenant["classes"] = []

        # Sauvegarder les modifications
        try:
            with open(personnel_path, "w", encoding="utf-8") as f:
                json.dump(personnel_data, f, indent=4, ensure_ascii=False)
        except Exception as e:
            print(f"Erreur lors de la sauvegarde de personnel.json: {e}")


    def update_personnel_after_class_deletion(week_folder, classes_to_remove):
        """
        Met à jour personnel.json en retirant les classes supprimées des intervenants.

        Args:
            week_folder (str): Dossier de la semaine
            classes_to_remove (list): Liste des noms de classes à retirer
        """
        personnel_path = os.path.join(week_folder, "personnel.json")

        if not os.path.exists(personnel_path):
            return

        try:
            with open(personnel_path, "r", encoding="utf-8") as f:
                personnel_data = json.load(f)
        except Exception as e:
            print(f"Erreur lors de la lecture de personnel.json: {e}")
            return

        # Pour chaque type d'intervenant (professeurs et animateurs)
        for personnel_type in ["professeurs", "animateurs"]:
            if personnel_type not in personnel_data:
                continue

            # Pour chaque intervenant
            for intervenant in personnel_data[personnel_type]:
                if "classes" not in intervenant:
                    continue

                # Retirer les classes supprimées de la liste de cet intervenant
                classes_original = intervenant["classes"]
                classes_filtrees = [c for c in classes_original if c not in classes_to_remove]

                # Mettre à jour si des classes ont été retirées
                if len(classes_filtrees) != len(classes_original):
                    intervenant["classes"] = classes_filtrees

        # Sauvegarder les modifications
        try:
            with open(personnel_path, "w", encoding="utf-8") as f:
                json.dump(personnel_data, f, indent=4, ensure_ascii=False)
        except Exception as e:
            print(f"Erreur lors de la sauvegarde de personnel.json: {e}")

    def delete_all_classes_from_horaire(week_folder, school_name, horaire):
        """
        Supprime toutes les classes d'un horaire spécifique du fichier Excel de l'école

        Args:
            week_folder (str): Dossier de la semaine
            school_name (str): Nom de l'école (ex: "A", "B", "C/CS", etc.)
            horaire (str): Horaire dont il faut supprimer toutes les classes
        """
        # Mapping des écoles vers les fichiers Excel
        school_file_mapping = {
            'A': 'ecole_a.xlsx',
            'B': 'ecole_b.xlsx',
            'C/CS': 'ECOLE_C_cours_standard.xlsx',
            'C/CI': 'ECOLE_C_cours_intensif.xlsx',
            'Morning': 'MORNING.xlsx',
            'Premium/CS': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'Premium/CI': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        if school_name not in school_file_mapping:
            raise ValueError(f"École inconnue: {school_name}")

        excel_filename = school_file_mapping[school_name]
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Fichier non trouvé: {excel_path}")

        # Vérifier que le fichier n'est pas ouvert par une autre application
        try:
            with open(excel_path, 'r+b') as f:
                pass
        except PermissionError:
            raise PermissionError(f"Le fichier {excel_filename} est ouvert dans une autre application. Veuillez le fermer avant de supprimer les classes.")

        try:
            # Ouvrir le fichier Excel
            wb = load_workbook(excel_path)
            target_sheet = None

            # Chercher la feuille correspondant à l'horaire
            for sheet_name in wb.sheetnames:
                # Normaliser les noms pour la comparaison
                sheet_normalized = sheet_name.lower().replace('animateur', '').replace('professeur', '').strip()
                horaire_normalized = horaire.lower().strip()

                if sheet_normalized == horaire_normalized or horaire_normalized in sheet_normalized:
                    target_sheet = wb[sheet_name]
                    break

            if target_sheet is None:
                raise ValueError(f"Feuille horaire '{horaire}' non trouvée dans {excel_filename}")

            # Créer une liste de tous les élèves qui sont dans ces classes qui vont être supprimées
            all_students = []
            classes_to_remove = []
            rows_to_delete = []

            # Chercher la colonne "Classe" pour collecter les noms des classes
            classe_col = None
            for col_idx in range(1, target_sheet.max_column + 1):
                header_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                if 'classe' in header_value or 'class' in header_value:
                    classe_col = col_idx
                    break

            for row_idx in range(2, target_sheet.max_row + 1):
                # Vérifier si la ligne contient des données (au moins une cellule non vide)
                has_data = False
                for col_idx in range(1, target_sheet.max_column + 1):
                    cell_value = str(target_sheet.cell(row=row_idx, column=col_idx).value or '').strip()
                    if cell_value:
                        has_data = True
                        break

                if has_data:
                    rows_to_delete.append(row_idx)

                    # Collecter le nom de la classe
                    if classe_col:
                        classe_nom = str(target_sheet.cell(row=row_idx, column=classe_col).value or '').strip()
                        if classe_nom:
                            classes_to_remove.append(classe_nom)

                    # Collecter les élèves de cette classe (colonne 5: "Liste des élèves")
                    students_cell = str(target_sheet.cell(row=row_idx, column=5).value or '').strip()
                    if students_cell:
                        # Séparer les élèves par virgule et nettoyer les espaces
                        students_in_class = [student.strip() for student in students_cell.split(',') if student.strip()]
                        all_students.extend(students_in_class)

            # Ouvrir le fichier matrix.xlsx et supprimer les informations pour chaque élève
            matrix_path = os.path.join(week_folder, 'matrix.xlsx')
            if os.path.exists(matrix_path):
                try:
                    matrix_wb = load_workbook(matrix_path)
                    matrix_sheet = matrix_wb.active  # Sheet1

                    # Pour chaque élève dans la liste, trouver sa ligne dans matrix.xlsx et supprimer Ecole, Horaire, Classe
                    students_updated = 0
                    for student_name in all_students:
                        # Chercher l'élève par nom (colonne 1)
                        for row_idx in range(2, matrix_sheet.max_row + 1):  # Commencer à la ligne 2 (après headers)
                            cell_value = str(matrix_sheet.cell(row=row_idx, column=1).value or '').strip()
                            if cell_value.upper() == student_name.upper():  # Comparaison insensible à la casse
                                # Supprimer les valeurs des colonnes Ecole (3), Horaire (4), Classe (5)
                                matrix_sheet.cell(row=row_idx, column=3).value = None  # Ecole
                                matrix_sheet.cell(row=row_idx, column=4).value = None  # Horaire
                                matrix_sheet.cell(row=row_idx, column=5).value = None  # Classe
                                students_updated += 1
                                break  # Passer au prochain élève une fois trouvé

                        # Sauvegarder le fichier matrix.xlsx
                        matrix_wb.save(matrix_path)

                except Exception as e:
                    print(f"Erreur lors de la mise à jour de matrix.xlsx: {e}")
            else:
                print(f"Fichier matrix.xlsx non trouvé dans {week_folder}")

            # Supprimer les lignes en commençant par la fin pour éviter les décalages d'indices
            for row_idx in reversed(rows_to_delete):
                target_sheet.delete_rows(row_idx)

            # Sauvegarder le fichier
            wb.save(excel_path)

            # Mettre à jour personnel.json pour retirer les classes supprimées
            if classes_to_remove:
                update_personnel_after_class_deletion(week_folder, classes_to_remove)

            return len(rows_to_delete)  # Retourner le nombre de classes supprimées

        except Exception as e:
            # Re-lever l'exception avec un message plus clair
            if "File is not a zip file" in str(e):
                raise Exception(f"Le fichier {excel_filename} est corrompu ou n'est pas un fichier Excel valide.")
            else:
                raise

    def delete_class_from_excel(week_folder, school_name, horaire, classe_nom):
        """
        Supprime une classe spécifique du fichier Excel de l'école

        Args:
            week_folder (str): Dossier de la semaine
            school_name (str): Nom de l'école (ex: "A", "B", "C/CS", etc.)
            horaire (str): Horaire de la classe
            classe_nom (str): Nom de la classe à supprimer
        """
        # Mapping des écoles vers les fichiers Excel
        school_file_mapping = {
            'A': 'ecole_a.xlsx',
            'B': 'ecole_b.xlsx',
            'C/CS': 'ECOLE_C_cours_standard.xlsx',
            'C/CI': 'ECOLE_C_cours_intensif.xlsx',
            'Morning': 'MORNING.xlsx',
            'Premium/CS': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'Premium/CI': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        if school_name not in school_file_mapping:
            raise ValueError(f"École inconnue: {school_name}")

        excel_filename = school_file_mapping[school_name]
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Fichier non trouvé: {excel_path}")

        # Vérifier que le fichier n'est pas ouvert par une autre application
        try:
            with open(excel_path, 'r+b') as f:
                pass
        except PermissionError:
            raise PermissionError(f"Le fichier {excel_filename} est ouvert dans une autre application. Veuillez le fermer avant de supprimer une classe.")

        try:
            # Ouvrir le fichier Excel
            wb = load_workbook(excel_path)
            target_sheet = None

            # Chercher la feuille correspondant à l'horaire
            for sheet_name in wb.sheetnames:
                # Normaliser les noms pour la comparaison
                sheet_normalized = sheet_name.lower().replace('animateur', '').replace('professeur', '').strip()
                horaire_normalized = horaire.lower().strip()

                if sheet_normalized == horaire_normalized or horaire_normalized in sheet_normalized:
                    target_sheet = wb[sheet_name]
                    break

            if target_sheet is None:
                raise ValueError(f"Feuille horaire '{horaire}' non trouvée dans {excel_filename}")

            # Chercher la colonne "Classe" spécifiquement
            classe_col = None
            for col_idx in range(1, target_sheet.max_column + 1):
                header_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                if 'classe' in header_value or 'class' in header_value:
                    classe_col = col_idx
                    break

            # Chercher et supprimer la ligne de la classe
            row_to_delete = None
            for row_idx in range(2, target_sheet.max_row + 1):  # Commencer à la ligne 2 (après header)
                for col_idx in range(1, target_sheet.max_column + 1):
                    cell_value = str(target_sheet.cell(row=row_idx, column=col_idx).value or '').strip()
                    if cell_value == classe_nom:
                        row_to_delete = row_idx
                        break
                if row_to_delete:
                    break

            if row_to_delete is None:
                raise ValueError(f"Classe '{classe_nom}' non trouvée dans la feuille '{horaire}'")

            # RÉCUPÉRER LES ÉLÈVES AVANT SUPPRESSION pour mettre à jour le matrix
            eleves_a_supprimer = []
            eleves_col = None

            # Trouver la colonne des élèves (priorité à "liste des élèves")
            # 1ère priorité : colonne contenant "liste" ET "élèves"
            for col_idx in range(1, target_sheet.max_column + 1):
                cell_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                if 'liste' in cell_value and ('élèves' in cell_value or 'eleves' in cell_value):
                    eleves_col = col_idx
                    break

            # 2ème priorité : colonne contenant juste "élèves" ou "eleves"
            if eleves_col is None:
                for col_idx in range(1, target_sheet.max_column + 1):
                    cell_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                    if ('élèves' in cell_value or 'eleves' in cell_value) and 'liste' not in cell_value:
                        eleves_col = col_idx
                        break

            # Récupérer les élèves de la classe à supprimer
            if eleves_col:
                eleves_value = str(target_sheet.cell(row=row_to_delete, column=eleves_col).value or '').strip()
                if eleves_value and eleves_value not in ['', 'nan', 'none']:
                    # Diviser par virgule et nettoyer
                    eleves_a_supprimer = [nom.strip() for nom in eleves_value.split(',') if nom.strip()]
            else:
                # Essayer la colonne 5 comme fallback (comme dans delete_all_classes_from_horaire)
                eleves_value = str(target_sheet.cell(row=row_to_delete, column=5).value or '').strip()
                if eleves_value and eleves_value not in ['', 'nan', 'none']:
                    eleves_a_supprimer = [nom.strip() for nom in eleves_value.split(',') if nom.strip()]

            # Supprimer la ligne
            target_sheet.delete_rows(row_to_delete)

            # Sauvegarder le fichier
            wb.save(excel_path)

            # METTRE À JOUR LE MATRIX : retirer l'école, la classe et l'horaire des élèves supprimés
            if eleves_a_supprimer:
                update_matrix_after_class_deletion(week_folder, eleves_a_supprimer, school_name, horaire, classe_nom)

            # METTRE À JOUR LE PERSONNEL : retirer la classe supprimée des intervenants
            update_personnel_after_class_deletion(week_folder, [classe_nom])

        except Exception as e:
            # Re-lever l'exception avec un message plus clair
            if "File is not a zip file" in str(e):
                raise Exception(f"Le fichier {excel_filename} est corrompu ou n'est pas un fichier Excel valide.")
            else:
                raise

    def update_matrix_after_class_deletion(week_folder, eleves_a_supprimer, school_name, horaire, classe_nom=None):
        """
        Met à jour le fichier matrix.xlsx pour retirer l'école, la classe et l'horaire des élèves supprimés
        """
        matrix_path = os.path.join(week_folder, "matrix.xlsx")

        if not os.path.exists(matrix_path):
            return

        if not load_workbook:
            return

        try:
            wb = load_workbook(matrix_path)
            ws = wb.active

            # Identifier les colonnes importantes
            stagiaire_col = None
            colonnes_assignation = []

            for col in range(1, ws.max_column + 1):
                col_name = str(ws.cell(row=1, column=col).value or '').strip()
                col_name_lower = col_name.lower()

                if 'stagiaire' in col_name_lower or 'nom' in col_name_lower or 'élève' in col_name_lower or 'eleve' in col_name_lower:
                    stagiaire_col = col
                elif any(keyword in col_name_lower for keyword in ['ecole', 'école', 'school', 'classe', 'class', 'groupe', 'horaire', 'horaire', 'heure', 'time']):
                    if not any(exclure in col_name_lower for exclure in ['cours 1', 'cours 2', 'cours 3', 'cours 4', '.1', '.2', 'arrivée', 'départ']):
                        colonnes_assignation.append(col)

            if not stagiaire_col:
                print("ERREUR: Aucune colonne eleves trouvee dans le matrix")
                return

            # Normaliser les noms d'élèves à supprimer
            eleves_normalises = {eleve.lower().strip() for eleve in eleves_a_supprimer}

            # Parcourir toutes les lignes pour trouver les élèves à mettre à jour
            updated_count = 0
            last_updated_row = None  # Pour la vérification

            for row_idx in range(2, ws.max_row + 1):
                eleve_nom_brut = str(ws.cell(row=row_idx, column=stagiaire_col).value or '').strip()
                eleve_nom_normalise = eleve_nom_brut.lower().strip()

                # Chercher une correspondance
                correspondance_trouvee = False
                for eleve_cible in eleves_normalises:
                    if eleve_cible in eleve_nom_normalise or eleve_nom_normalise in eleve_cible:
                        correspondance_trouvee = True
                        break

                if correspondance_trouvee:
                    last_updated_row = row_idx  # Sauvegarder la ligne pour vérification

                    # Vider TOUTES les colonnes d'assignation pour cet élève
                    for col in colonnes_assignation:
                        old_value = ws.cell(row=row_idx, column=col).value
                        if old_value is not None:
                            # Utiliser la syntaxe correcte pour modifier la cellule
                            cell = ws.cell(row=row_idx, column=col)
                            cell.value = None

                    updated_count += 1

            # Sauvegarder le fichier matrix
            wb.save(matrix_path)

        except Exception as e:
            print(f"ERREUR lors de la mise a jour du matrix: {e}")
            import traceback
            traceback.print_exc()

    def confirm_delete_all_classes_from_horaire(horaire, school_name, week_folder):
        """Affiche une boîte de dialogue de confirmation pour supprimer toutes les classes d'un horaire"""
        # Boîte de dialogue de confirmation
        result = messagebox.askyesno(
            "Confirmer la suppression massive",
            f"Êtes-vous sûr de vouloir supprimer TOUTES les classes de l'horaire '{horaire}' ?\n\n"
            f"École: {school_name}\n"
            f"Horaire: {horaire}\n\n"
            f"ATTENTION: Cette action supprimera toutes les classes de cet horaire du fichier Excel.\n"
            f"Cette action est irréversible.",
            parent=app
        )

        if result:  # L'utilisateur a confirmé
            # Afficher la fenêtre de chargement
            loading_popup = show_loading_window_for_deletion(app)

            # Forcer plusieurs mises à jour pour s'assurer que la fenêtre de chargement s'affiche immédiatement
            loading_popup.update()
            loading_popup.update_idletasks()
            app.update()
            app.update_idletasks()

            # Variable pour contrôler la boucle de mise à jour de l'animation
            keep_animating = [True]  # Utiliser une liste pour permettre la modification dans les fonctions imbriquées

            def keep_animation_alive():
                """Fonction récursive pour maintenir l'animation de la fenêtre de chargement"""
                if keep_animating[0] and loading_popup and loading_popup.winfo_exists():
                    loading_popup.update()
                    app.update_idletasks()
                    app.after(50, keep_animation_alive)  # Mettre à jour toutes les 50ms

            # Démarrer la boucle d'animation
            keep_animation_alive()

            def perform_mass_deletion():
                try:
                    deleted_count = delete_all_classes_from_horaire(week_folder, school_name, horaire)
                    # Rafraîchir l'affichage
                    school_data = analyze_school_classes(week_folder)
                    create_classes_dashboard(content, school_data, week_folder)

                    # Arrêter l'animation et fermer la fenêtre
                    keep_animating[0] = False
                    try:
                        if loading_popup and loading_popup.winfo_exists():
                            loading_popup.destroy()
                    except:
                        pass

                    messagebox.showinfo(
                        "Suppression réussie",
                        f"Toutes les classes de l'horaire '{horaire}' ont été supprimées.\n"
                        f"Nombre de classes supprimées: {deleted_count}",
                        parent=app
                    )
                except Exception as e:
                    # Arrêter l'animation et fermer la fenêtre en cas d'erreur
                    keep_animating[0] = False
                    try:
                        if loading_popup and loading_popup.winfo_exists():
                            loading_popup.destroy()
                    except:
                        pass
                    messagebox.showerror("Erreur", f"Impossible de supprimer les classes : {e}", parent=app)

            # Démarrer la suppression après un court délai
            app.after(10, perform_mass_deletion)

    def confirm_delete_class(menu_window, classe_info, horaire, school_name, week_folder):
        """Affiche une boîte de dialogue de confirmation pour supprimer la classe"""
        # Fermer le menu contextuel de manière sécurisée
        try:
            if menu_window.winfo_exists():
                menu_window.destroy()
        except:
            pass  # Ignore les erreurs si la fenêtre est déjà détruite

        classe_nom = classe_info.get('nom_classe', 'Classe inconnue')

        # Boîte de dialogue de confirmation
        result = messagebox.askyesno(
            "Confirmer la suppression",
            f"Êtes-vous sûr de vouloir supprimer la classe '{classe_nom}' ?\n\n"
            f"Cette action supprimera la classe du fichier Excel de l'école {school_name}.\n"
            f"Cette action est irréversible.",
            parent=app
        )

        if result:  # L'utilisateur a confirmé
            # Afficher la fenêtre de chargement
            loading_popup = show_loading_window_for_deletion(app)

            # Forcer plusieurs mises à jour pour s'assurer que la fenêtre de chargement s'affiche immédiatement
            loading_popup.update()
            loading_popup.update_idletasks()
            app.update()
            app.update_idletasks()

            # Variable pour contrôler la boucle de mise à jour de l'animation
            keep_animating = [True]  # Utiliser une liste pour permettre la modification dans les fonctions imbriquées

            def keep_animation_alive():
                """Fonction récursive pour maintenir l'animation de la fenêtre de chargement"""
                if keep_animating[0] and loading_popup and loading_popup.winfo_exists():
                    loading_popup.update()
                    app.update_idletasks()
                    app.after(50, keep_animation_alive)  # Mettre à jour toutes les 50ms

            # Démarrer la boucle d'animation
            keep_animation_alive()

            def perform_deletion():
                try:
                    delete_class_from_excel(week_folder, school_name, horaire, classe_nom)
                    # Rafraîchir l'affichage
                    school_data = analyze_school_classes(week_folder)
                    create_classes_dashboard(content, school_data, week_folder)

                    # Arrêter l'animation et fermer la fenêtre
                    keep_animating[0] = False
                    try:
                        if loading_popup and loading_popup.winfo_exists():
                            loading_popup.destroy()
                    except:
                        pass

                    messagebox.showinfo("Suppression réussie", f"La classe '{classe_nom}' a été supprimée.", parent=app)
                except Exception as e:
                    # Arrêter l'animation et fermer la fenêtre en cas d'erreur
                    keep_animating[0] = False
                    try:
                        if loading_popup and loading_popup.winfo_exists():
                            loading_popup.destroy()
                    except:
                        pass
                    messagebox.showerror("Erreur", f"Impossible de supprimer la classe : {e}", parent=app)

            # Démarrer la suppression après un court délai
            app.after(10, perform_deletion)

    def setup_classe_card_click(classe_card, classe_info, horaire, intervenant, type_intervenant, school_color, school_name, week_folder):
        """Configure le clic sur une carte de classe"""
        def on_click(event):
            # Callback pour rafraîchir le dashboard
            def refresh_dashboard_callback():
                try:
                    school_data = analyze_school_classes(week_folder)
                    create_classes_dashboard(content, school_data, week_folder)
                except Exception as e:
                    print(f"Erreur lors du rafraîchissement du dashboard: {e}")

            open_classe_details(classe_info, horaire, intervenant, type_intervenant, school_color, school_name, week_folder, refresh_dashboard_callback)

        def on_right_click(event):
            """Menu contextuel pour supprimer la classe"""
            # Créer un menu contextuel
            context_menu = ctk.CTkToplevel(app)
            context_menu.title("")
            context_menu.geometry("250x80")
            context_menu.resizable(False, False)
            context_menu.attributes("-topmost", True)
            context_menu.overrideredirect(True)  # Pas de barre de titre

            # Positionner le menu près du curseur
            x, y = event.x_root, event.y_root
            context_menu.geometry(f"+{x}+{y}")

            # Variable pour suivre si le menu est encore valide
            menu_destroyed = False

            # Bouton de suppression
            def safe_confirm_delete():
                if not menu_destroyed and context_menu.winfo_exists():
                    confirm_delete_class(context_menu, classe_info, horaire, school_name, week_folder)

            delete_btn = ctk.CTkButton(
                context_menu,
                text="🗑️ Supprimer la classe",
                fg_color="#dc2626",
                hover_color="#b91c1c",
                text_color="white",
                command=safe_confirm_delete
            )
            delete_btn.pack(fill="both", expand=True, padx=10, pady=10)

            # Fermer le menu si on clique ailleurs
            def close_menu(event=None):
                nonlocal menu_destroyed
                if not menu_destroyed and context_menu.winfo_exists():
                    menu_destroyed = True
                    try:
                        context_menu.destroy()
                    except:
                        pass  # Ignore les erreurs de destruction

            # Gérer la fermeture propre
            context_menu.bind("<FocusOut>", close_menu)

            # Fermer aussi sur Escape ou clic ailleurs
            def on_key_press(event):
                if event.keysym == 'Escape':
                    close_menu()

            context_menu.bind("<Key>", on_key_press)
            context_menu.bind("<Button-1>", lambda e: close_menu() if not delete_btn.winfo_containing(e.x_root, e.y_root) else None)

            # Délai avant de donner le focus pour éviter les conflits
            def focus_menu():
                if not menu_destroyed and context_menu.winfo_exists():
                    try:
                        context_menu.focus_force()
                    except:
                        pass

            context_menu.after(10, focus_menu)

        classe_card.bind("<Button-1>", on_click)
        classe_card.bind("<Button-3>", on_right_click)  # Clic droit

        # Propager le clic aux widgets enfants
        for child in classe_card.winfo_children():
            bind_click_to_children(child, on_click)
            # Propager aussi le clic droit
            child.bind("<Button-3>", on_right_click)

    def bind_click_to_children(widget, click_func):
        """Propage l'événement de clic à tous les enfants"""
        widget.bind("<Button-1>", click_func)
        for child in widget.winfo_children():
            bind_click_to_children(child, click_func)

    # Callback pour changer le contenu du dashboard
    def show_loading_window_for_week_selection(parent_app):
        """Affiche une fenêtre de chargement lors de la sélection d'une semaine."""
        loading_popup = ctk.CTkToplevel(parent_app)
        loading_popup.geometry("350x150")
        loading_popup.resizable(False, False)

        # Configuration pour supprimer les bords de fenêtre
        loading_popup.attributes("-topmost", True)
        loading_popup.overrideredirect(True)

        # Centrer précisément sur la fenêtre parente
        x = parent_app.winfo_rootx() + (parent_app.winfo_width() // 2) - 175
        y = parent_app.winfo_rooty() + (parent_app.winfo_height() // 2) - 75
        loading_popup.geometry(f"350x150+{x}+{y}")

        # --- STYLE GRAPHIQUE ---
        main_frame = ctk.CTkFrame(
            loading_popup,
            fg_color="#FFFFFF",
            corner_radius=20,
            border_width=1,
            border_color="#CBD5E1"
        )
        main_frame.pack(fill="both", expand=True)

        # Titre "Système" discret
        title_label = ctk.CTkLabel(
            main_frame,
            text="CHARGEMENT",
            font=("Segoe UI", 10, "bold"),
            text_color="#64748B"
        )
        title_label.pack(pady=(20, 5))

        # Texte principal
        text_label = ctk.CTkLabel(
            main_frame,
            text="Chargement des données de la semaine...",
            font=("Segoe UI", 13),
            text_color="#1E293B"
        )
        text_label.pack(pady=(0, 15))

        # Petite icône de chargement fixe
        loading_icon = ctk.CTkLabel(
            main_frame,
            text="⏳",
            font=("Arial", 24),
            text_color="#3B82F6"
        )
        loading_icon.pack(pady=(5, 0))

        # La fenêtre sera fermée manuellement après le chargement complet
        # Pas de fermeture automatique ici

        return loading_popup

    def show_loading_window_for_deletion(parent_app):
        """Affiche une fenêtre de chargement lors de la suppression d'une classe."""
        loading_popup = ctk.CTkToplevel(parent_app)
        loading_popup.geometry("350x150")
        loading_popup.resizable(False, False)

        # Configuration pour supprimer les bords de fenêtre
        loading_popup.attributes("-topmost", True)
        loading_popup.overrideredirect(True)

        # Centrer précisément sur la fenêtre parente
        x = parent_app.winfo_rootx() + (parent_app.winfo_width() // 2) - 175
        y = parent_app.winfo_rooty() + (parent_app.winfo_height() // 2) - 75
        loading_popup.geometry(f"350x150+{x}+{y}")

        # --- STYLE GRAPHIQUE ---
        main_frame = ctk.CTkFrame(
            loading_popup,
            fg_color="#FFFFFF",
            corner_radius=20,
            border_width=1,
            border_color="#CBD5E1"
        )
        main_frame.pack(fill="both", expand=True)

        # Titre "Système" discret
        title_label = ctk.CTkLabel(
            main_frame,
            text="MISE À JOUR",
            font=("Segoe UI", 10, "bold"),
            text_color="#64748B"
        )
        title_label.pack(pady=(20, 5))

        # Texte principal
        text_label = ctk.CTkLabel(
            main_frame,
            text="Mise à jour de la base matrix et des écoles...",
            font=("Segoe UI", 13),
            text_color="#1E293B"
        )
        text_label.pack(pady=(0, 15))

        # Petite icône de chargement fixe
        loading_icon = ctk.CTkLabel(
            main_frame,
            text="⏳",
            font=("Arial", 24),
            text_color="#3B82F6"
        )
        loading_icon.pack(pady=(5, 0))

        # La fenêtre sera fermée manuellement après la suppression complète
        # Pas de fermeture automatique ici

        return loading_popup

    def show_loading_window_for_school_filter(school_name, is_activation=True):
        """Affiche une fenêtre de chargement lors du filtrage des écoles."""
        loading_popup = ctk.CTkToplevel(app)
        loading_popup.geometry("400x150")
        loading_popup.resizable(False, False)

        # Configuration pour supprimer les bords de fenêtre
        loading_popup.attributes("-topmost", True)
        loading_popup.overrideredirect(True)

        # Centrer précisément sur la fenêtre parente
        x = app.winfo_rootx() + (app.winfo_width() // 2) - 200
        y = app.winfo_rooty() + (app.winfo_height() // 2) - 75
        loading_popup.geometry(f"400x150+{x}+{y}")

        # --- STYLE GRAPHIQUE ---
        main_frame = ctk.CTkFrame(
            loading_popup,
            fg_color="#FFFFFF",
            corner_radius=20,
            border_width=1,
            border_color="#CBD5E1"
        )
        main_frame.pack(fill="both", expand=True)

        # Titre selon l'action
        title_text = "CHARGEMENT" if is_activation else "MISE À JOUR"
        title_label = ctk.CTkLabel(
            main_frame,
            text=title_text,
            font=("Segoe UI", 10, "bold"),
            text_color="#64748B"
        )
        title_label.pack(pady=(20, 5))

        # Texte principal selon l'action
        if is_activation:
            text_content = f"Chargement des classes et des élèves de l'école {school_name}..."
        else:
            text_content = "Mise à jour de l'affichage des écoles..."

        text_label = ctk.CTkLabel(
            main_frame,
            text=text_content,
            font=("Segoe UI", 12),
            text_color="#1E293B",
            wraplength=350
        )
        text_label.pack(pady=(0, 15))

        # Petite icône de chargement fixe
        loading_icon = ctk.CTkLabel(
            main_frame,
            text="⏳",
            font=("Arial", 24),
            text_color="#3B82F6"
        )
        loading_icon.pack(pady=(5, 0))

        # La fenêtre sera fermée manuellement après le chargement complet
        # Pas de fermeture automatique ici

        return loading_popup

    def show_loading_window_for_matrix_import(parent_app):
        """Affiche une fenêtre de chargement lors de l'importation du fichier matrix."""
        loading_popup = ctk.CTkToplevel(parent_app)
        loading_popup.geometry("400x150")
        loading_popup.resizable(False, False)

        # Configuration pour supprimer les bords de fenêtre
        loading_popup.attributes("-topmost", True)
        loading_popup.overrideredirect(True)

        # Centrer précisément sur la fenêtre parente
        x = parent_app.winfo_rootx() + (parent_app.winfo_width() // 2) - 200
        y = parent_app.winfo_rooty() + (parent_app.winfo_height() // 2) - 75
        loading_popup.geometry(f"400x150+{x}+{y}")

        # --- STYLE GRAPHIQUE ---
        main_frame = ctk.CTkFrame(
            loading_popup,
            fg_color="#FFFFFF",
            corner_radius=20,
            border_width=1,
            border_color="#CBD5E1"
        )
        main_frame.pack(fill="both", expand=True)

        # Titre
        title_label = ctk.CTkLabel(
            main_frame,
            text="IMPORTATION",
            font=("Segoe UI", 10, "bold"),
            text_color="#64748B"
        )
        title_label.pack(pady=(20, 5))

        # Texte principal
        text_label = ctk.CTkLabel(
            main_frame,
            text="Importation du fichier matrix et préparation des écoles...",
            font=("Segoe UI", 12),
            text_color="#1E293B",
            wraplength=350
        )
        text_label.pack(pady=(0, 15))

        # Petite icône de chargement fixe
        loading_icon = ctk.CTkLabel(
            main_frame,
            text="⏳",
            font=("Arial", 24),
            text_color="#3B82F6"
        )
        loading_icon.pack(pady=(5, 0))

        # La fenêtre sera fermée manuellement après l'importation complète
        # Pas de fermeture automatique ici

        return loading_popup

    def on_week_selected(week_label: str):
        # Afficher la fenêtre de chargement
        loading_popup = show_loading_window_for_week_selection(app)

        # Forcer plusieurs mises à jour pour s'assurer que la fenêtre de chargement s'affiche immédiatement
        loading_popup.update()
        loading_popup.update_idletasks()
        app.update()
        app.update_idletasks()

        # Variable pour contrôler la boucle de mise à jour de l'animation
        keep_animating = [True]  # Utiliser une liste pour permettre la modification dans les fonctions imbriquées
        
        def keep_animation_alive():
            """Fonction récursive pour maintenir l'animation de la barre de progression"""
            if keep_animating[0] and loading_popup and loading_popup.winfo_exists():
                loading_popup.update()
                app.update_idletasks()
                app.after(50, keep_animation_alive)  # Mettre à jour toutes les 50ms
        
        # Démarrer la boucle d'animation
        keep_animation_alive()
        
        # Démarrer le traitement après un très court délai pour permettre à la fenêtre de chargement de s'afficher
        # On divise le traitement en étapes pour permettre à la barre de progression de rester animée
        def step1_initial_setup():
            """Étape 1 : Configuration initiale"""
            try:
                selected_week.set(week_label)
                stop_matrix_watch()
                
                # Changer la couleur des boutons
                for btn_text, btn in week_buttons.items():
                    if btn_text == week_label:
                        btn.configure(fg_color="#10b981", hover_color="#059669")
                    else:
                        btn.configure(fg_color="#89B8E3", hover_color="#A1C9F1")
                
                # Cacher le logo et afficher le header
                logo_center.grid_remove()
                header_frame.grid(row=0, column=0, sticky="ew", padx=20, pady=(16, 8))
                update_matrix_status()
                
                # Passer à l'étape suivante
                app.after(10, step2_analyze_data)
            except Exception as e:
                print(f"Erreur lors de l'étape 1: {e}")
                keep_animating[0] = False
                try:
                    if loading_popup and loading_popup.winfo_exists():
                        loading_popup.destroy()
                except:
                    pass
        
        def step2_analyze_data():
            """Étape 2 : Analyser les données"""
            try:
                # Analyser le fichier matrix
                matrix_path = _get_matrix_path_for_selected_week()
                matrix_stats = analyze_matrix_file(matrix_path)
                update_header_counters(matrix_stats)
                
                # Passer à l'étape suivante
                app.after(10, step3_analyze_schools)
            except Exception as e:
                print(f"Erreur lors de l'étape 2: {e}")
                keep_animating[0] = False
                try:
                    if loading_popup and loading_popup.winfo_exists():
                        loading_popup.destroy()
                except:
                    pass
        
        def step3_analyze_schools():
            """Étape 3 : Analyser les écoles"""
            try:
                # Analyser les données des écoles
                week_num = week_label.split()[-1]
                week_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"semaine_{week_num}")
                school_data = analyze_school_classes(week_folder)
                
                # Passer à l'étape suivante
                app.after(10, lambda: step4_create_dashboard(week_folder, school_data))
            except Exception as e:
                print(f"Erreur lors de l'étape 3: {e}")
                keep_animating[0] = False
                try:
                    if loading_popup and loading_popup.winfo_exists():
                        loading_popup.destroy()
                except:
                    pass
        
        def step4_create_dashboard(week_folder, school_data):
            """Étape 4 : Créer le dashboard"""
            try:
                # Créer l'interface des classes
                content.grid(row=1, column=0, sticky="nsew", padx=20, pady=(0, 20))
                content.update_idletasks()
                create_classes_dashboard(content, school_data, week_folder)
                
                # Passer à l'étape finale
                app.after(10, step5_finalize)
            except Exception as e:
                print(f"Erreur lors de l'étape 4: {e}")
                keep_animating[0] = False
                try:
                    if loading_popup and loading_popup.winfo_exists():
                        loading_popup.destroy()
                except:
                    pass
        
        def step5_finalize():
            """Étape 5 : Finalisation"""
            try:
                # Arrêter la boucle d'animation
                keep_animating[0] = False
                
                # Forcer la mise à jour complète de l'interface
                app.update_idletasks()
                app.update()
                
                # Démarrer la surveillance du fichier matrix
                check_matrix_modifications()
                
                # Fermer la fenêtre de chargement
                try:
                    if loading_popup and loading_popup.winfo_exists():
                        loading_popup.destroy()
                except:
                    pass
            except Exception as e:
                print(f"Erreur lors de l'étape 5: {e}")
                keep_animating[0] = False
                try:
                    if loading_popup and loading_popup.winfo_exists():
                        loading_popup.destroy()
                except:
                    pass

        # Démarrer le traitement après un très court délai (10ms) pour permettre à la fenêtre de chargement de s'afficher
        app.after(10, step1_initial_setup)

    # Création des 9 boutons de semaine
    for i in range(1, 10):
        week_text = f"Semaine {i}"
        btn = ctk.CTkButton(
            sidebar,
            text=week_text,
            width=170,
            height=40,
            fg_color="#89B8E3",
            hover_color="#A1C9F1",
            text_color="white",
            font=("Arial", 14, "bold"),
            corner_radius=8,
            command=lambda w=week_text: on_week_selected(w),
        )
        btn.grid(row=i+1, column=0, padx=16, pady=4, sticky="ew")
        week_buttons[week_text] = btn


    # Fonction de fermeture propre
    def on_app_closing():
        """Gère la fermeture propre de l'application."""
        stop_matrix_watch()
        app.destroy()

    # Bouton déconnexion / quitter
    logout_btn = ctk.CTkButton(
        sidebar,
        text="Quitter",
        width=170,
        height=36,
        fg_color="#e57373",
        hover_color="#ef9a9a",
        text_color="white",
        font=("Arial", 13, "bold"),
        corner_radius=8,
        command=on_app_closing,
    )
    logout_btn.grid(row=12, column=0, padx=16, pady=(10, 18), sticky="ew")

    # Gérer la fermeture par la croix de la fenêtre
    app.protocol("WM_DELETE_WINDOW", on_app_closing)

    # --------- PARTIE DROITE : DASHBOARD ---------
    main_area = ctk.CTkFrame(app, fg_color="white", corner_radius=0)
    main_area.grid(row=0, column=1, sticky="nsew")
    main_area.grid_rowconfigure(1, weight=1)
    main_area.grid_columnconfigure(0, weight=1)

    # En-tête du dashboard (masqué au départ)
    header_frame = ctk.CTkFrame(main_area, fg_color="#ffffff")
    header_frame.grid_remove()  # Masqué au départ
    header_frame.grid_columnconfigure(0, weight=0)  # compteurs d'élèves
    header_frame.grid_columnconfigure(1, weight=1)  # boutons de filtres d'école
    header_frame.grid_columnconfigure(2, weight=0)  # bouton actions
    header_frame.grid_columnconfigure(3, weight=0)  # statut matrix
    header_frame.grid_rowconfigure(0, weight=0)  # ligne titre et boutons

    # Compteurs d'élèves (remplace le titre de semaine)
    counters_frame = ctk.CTkFrame(
        header_frame,
        fg_color="transparent",
        corner_radius=8,
        height=36
    )
    counters_frame.grid(row=0, column=0, sticky="ew", padx=(0, 10))

    # Cadre pour les boutons de filtres d'école
    central_frame = ctk.CTkFrame(
        header_frame,
        fg_color="white",
        corner_radius=8,
        height=40
    )
    central_frame.grid(row=0, column=1, sticky="ew", padx=10)
    counters_frame.pack_propagate(False)
    counters_frame.grid_columnconfigure(0, weight=1)
    counters_frame.grid_columnconfigure(1, weight=1)
    counters_frame.grid_columnconfigure(2, weight=1)

    # --- Style 1 : Badges "Outline & Soft Fill" ---

    # Configuration des styles pour les boutons
    FONT_BUTTON = ("Arial", 11, "bold")  # Police pour les boutons en gras

    def create_metric_button(parent, title, value, button_color, hover_color=None, column=None, assignment_type=None):
        # Bouton cliquable seulement pour "sans classe" et "sans niveau" (pas pour "Total")
        def on_button_click():
            if assignment_type:
                show_students_without_assignment(assignment_type)

        # Ne définir une commande que si ce n'est pas le bouton Total
        command = on_button_click if assignment_type else None

        # Pour le bouton Total, utiliser la même couleur pour hover (pas d'effet hover)
        if hover_color is None:
            hover_color = button_color

        button = ctk.CTkButton(
            parent,
            text=f"{title}\n{value}",
            font=FONT_BUTTON,
            fg_color=button_color,
            hover_color=hover_color,
            text_color="white",
            width=80,
            height=40,
            corner_radius=8,
            command=command
        )
        button.grid(row=0, column=column, padx=5, pady=5)

        return button

    # --- Mise en œuvre ---

    # 1. Total (Couleur douce de l'école A)
    total_counter_button = create_metric_button(
        counters_frame, "Total", "1,254", "#93c5fd", None, 0
    )

    # 2. Sans Classe (Orange pastel)
    without_class_counter_button = create_metric_button(
        counters_frame, "Sans Cl.", "12", "#ffd4a3", "#ffb366", 1, "class"
    )

    # 3. Sans Niveau (Rouge pastel)
    without_level_counter_button = create_metric_button(
        counters_frame, "Sans Niv.", "05", "#ffd4a3", "#ffb366", 2, "level"
    )

    # Initialiser les références globales des compteurs
    set_counter_labels(total_counter_button, without_class_counter_button, without_level_counter_button)

    # Configuration pour les boutons de filtre d'écoles (7 boutons sur une ligne)
    for col in range(7):
        central_frame.grid_columnconfigure(col, weight=1)

    # Boutons de filtre pour les écoles
    school_filter_buttons = {}

    def toggle_school_filter(school_name, var):
        """Alterne l'état du filtre d'école et met à jour l'affichage."""
        current_state = var.get()
        new_state = not current_state
        var.set(new_state)

        # Afficher la fenêtre de chargement
        loading_popup = show_loading_window_for_school_filter(school_name, new_state)

        # Forcer plusieurs mises à jour pour s'assurer que la fenêtre de chargement s'affiche immédiatement
        loading_popup.update()
        loading_popup.update_idletasks()
        app.update()
        app.update_idletasks()

        # Variable pour contrôler la boucle de mise à jour de l'animation
        keep_animating = [True]  # Utiliser une liste pour permettre la modification dans les fonctions imbriquées

        def keep_animation_alive():
            """Fonction récursive pour maintenir l'animation de la fenêtre de chargement"""
            if keep_animating[0] and loading_popup and loading_popup.winfo_exists():
                loading_popup.update()
                app.update_idletasks()
                app.after(50, keep_animation_alive)  # Mettre à jour toutes les 50ms

        # Démarrer la boucle d'animation
        keep_animation_alive()

        def perform_filter_update():
            try:
                # Mettre à jour l'apparence du bouton
                button = school_filter_buttons[school_name]
                school_color = school_colors.get(school_name, "#6b7280")

                if new_state:
                    # Version douce de la couleur de l'école
                    soft_color = soften_color(school_color, 0.5)
                    hover_color = soften_color(school_color, 0.3)
                    button.configure(
                        fg_color=soft_color,
                        text_color="white",
                        hover_color=hover_color
                    )
                else:
                    button.configure(
                        fg_color="#ffffff",  # Blanc pour inactif
                        text_color="#1a365d",
                        hover_color="#f8fafc"
                    )

                # Sauvegarder les préférences utilisateur
                preferences = load_user_preferences()
                if "school_filters" not in preferences:
                    preferences["school_filters"] = {}
                preferences["school_filters"][school_name] = new_state
                save_user_preferences(preferences)

                # Rafraîchir l'affichage des classes
                refresh_school_display()

                # Fermer la fenêtre de chargement après un délai minimal
                def close_loading():
                    keep_animating[0] = False
                    try:
                        if loading_popup and loading_popup.winfo_exists():
                            loading_popup.destroy()
                    except:
                        pass

                app.after(500, close_loading)

            except Exception as e:
                print(f"Erreur lors de la mise à jour du filtre: {e}")
                keep_animating[0] = False
                try:
                    if loading_popup and loading_popup.winfo_exists():
                        loading_popup.destroy()
                except:
                    pass

        # Démarrer la mise à jour après un court délai
        app.after(10, perform_filter_update)

    def refresh_school_display():
        """Met à jour l'affichage des écoles selon les filtres actifs."""
        # Récupérer la semaine actuelle
        week_label = selected_week.get()
        if not week_label:
            return

        week_num = week_label.split()[-1]
        week_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"semaine_{week_num}")
        school_data = analyze_school_classes(week_folder)

        # Créer l'interface des classes avec les filtres
        create_classes_dashboard(content, school_data, week_folder)

    def check_matrix_modifications():
        """Vérifie périodiquement si le fichier matrix a été modifié et rafraîchit l'affichage."""
        nonlocal matrix_last_modified, matrix_watch_job
        
        week_label = selected_week.get()
        if not week_label:
            # Pas de semaine sélectionnée, arrêter la surveillance
            return
        
        matrix_path = _get_matrix_path_for_selected_week()
        
        if os.path.exists(matrix_path):
            try:
                # Obtenir le timestamp de dernière modification
                current_mtime = os.path.getmtime(matrix_path)
                
                # Vérifier si le fichier a été modifié
                if matrix_path in matrix_last_modified:
                    if current_mtime > matrix_last_modified[matrix_path]:
                        # Le fichier a été modifié, rafraîchir l'affichage
                        print(f"Detection d'une modification du fichier matrix.xlsx - Rafraichissement automatique en cours...")
                        matrix_last_modified[matrix_path] = current_mtime
                        
                        # Mettre à jour les compteurs d'élèves
                        matrix_stats = analyze_matrix_file(matrix_path)
                        update_header_counters(matrix_stats)
                        
                        # IMPORTANT : Récupérer le dossier de la semaine et analyser complètement
                        week_num = week_label.split()[-1]
                        week_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"semaine_{week_num}")
                        
                        # 1) Analyser les assignations depuis matrix.xlsx
                        matrix_assignments = analyze_matrix_assignments(matrix_path)
                        
                        # 2) Mettre à jour les fichiers Excel des écoles avec les élèves assignés
                        update_school_files_with_assignments(week_folder, matrix_assignments)
                        
                        # 3) Analyser à nouveau les données des écoles (avec les nouvelles données)
                        school_data = analyze_school_classes(week_folder)
                        
                        # 4) Recréer complètement l'interface des classes
                        create_classes_dashboard(content, school_data, week_folder)
                        
                        print(f"Rafraichissement termine - Dashboard mis a jour avec les nouveaux nombres d'eleves")
                else:
                    # Première fois qu'on surveille ce fichier
                    matrix_last_modified[matrix_path] = current_mtime
            except Exception as e:
                print(f"Erreur lors de la verification du fichier matrix: {e}")
        
        # Programmer la prochaine vérification dans 1 seconde
        matrix_watch_job = app.after(1000, check_matrix_modifications)

    def stop_matrix_watch():
        """Arrête la surveillance du fichier matrix."""
        nonlocal matrix_watch_job
        if matrix_watch_job:
            app.after_cancel(matrix_watch_job)
            matrix_watch_job = None

    # Créer les boutons de filtre pour chaque école
    # Couleurs pour chaque type d'école
    school_colors = {
        "A": "#3b82f6",      # Bleu
        "B": "#10b981",      # Vert
        "C/CS": "#f59e0b",   # Orange
        "C/CI": "#8b5cf6",   # Violet
        "Morning": "#ef4444", # Rouge
        "Premium/CS": "#06b6d4", # Cyan
        "Premium/CI": "#f97316"  # Orange foncé
    }

    # Fonction pour créer une version plus douce de la couleur
    def soften_color(color, factor=0.6):
        """Rend une couleur plus douce/claire."""
        if color.startswith('#'):
            # Convertir hex vers RGB
            r = int(color[1:3], 16)
            g = int(color[3:5], 16)
            b = int(color[5:7], 16)

            # Appliquer le facteur d'adoucissement
            r = int(r + (255 - r) * (1 - factor))
            g = int(g + (255 - g) * (1 - factor))
            b = int(b + (255 - b) * (1 - factor))

            # Reconvertir vers hex
            return f"#{r:02x}{g:02x}{b:02x}"
        return color

    for i, (school_name, var) in enumerate(school_filters.items()):
        school_color = school_colors.get(school_name, "#6b7280")

        button = ctk.CTkButton(
            central_frame,
            text=school_name,
            font=("Arial", 12, "bold"),  # Police plus petite
            height=30,  # Hauteur augmentée
            width=55,  # Largeur diminuée
            fg_color=soften_color(school_color, 0.5) if var.get() else "#ffffff",
            text_color="white" if var.get() else "#1a365d",
            hover_color=soften_color(school_color, 0.3) if var.get() else "#f8fafc",
            border_width=1,
            border_color="#cbd5e1",
            corner_radius=5,
            command=lambda s=school_name, v=var: toggle_school_filter(s, v)
        )
        button.grid(row=0, column=i, padx=0.5, pady=2, sticky="ew")  # Espacement minimal
        school_filter_buttons[school_name] = button

    # Zone "Actions préalables" à droite
    actions_btn = ctk.CTkButton(
        header_frame,
        text="Menu ▾",
        width=80,
        height=40,
        fg_color="#89B8E3",
        hover_color="#A1C9F1",
        text_color="white",
        font=("Arial", 13, "bold"),
        corner_radius=10,
        command=lambda: open_actions_menu(actions_btn),
    )
    actions_btn.grid(row=0, column=2, sticky="e", padx=(0, 8))

    # Indicateur de présence du fichier matrix pour la semaine sélectionnée
    matrix_status_pill = ctk.CTkFrame(
        header_frame,
        fg_color="#d3d3d3",
        corner_radius=999,
    )
    matrix_status_pill.grid(row=0, column=3, sticky="e", padx=(10, 0), pady=2)

    matrix_icon_label = ctk.CTkLabel(
        matrix_status_pill,
        text="●",
        font=("Arial", 18, "bold"),
        text_color="#6b6b6b",
    )
    matrix_icon_label.pack(side="left", padx=(8, 4), pady=2)

    matrix_short_label = ctk.CTkLabel(
        matrix_status_pill,
        text="MATRIX",
        font=("Arial", 11, "bold"),
        text_color="#ffffff",
    )
    matrix_short_label.pack(side="left", padx=(0, 10), pady=2)


    def _get_matrix_path_for_selected_week() -> str:
        """Retourne le chemin absolu du fichier matrix pour la semaine sélectionnée."""
        week_label = selected_week.get()  # ex: "Semaine 1"
        try:
            week_num = week_label.split()[-1]
        except Exception:
            week_num = "1"
        folder = f"semaine_{week_num}"
        # Retourne un chemin absolu basé sur le répertoire du script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        return os.path.abspath(os.path.join(script_dir, folder, "matrix.xlsx"))

    def update_matrix_status():
        """Met à jour l'indicateur design (pastille) pour la présence du fichier matrix."""
        matrix_path = _get_matrix_path_for_selected_week()
        if os.path.exists(matrix_path):
            matrix_status_pill.configure(fg_color="#2e7d32")
            matrix_icon_label.configure(text_color="#c8f7c5")
            matrix_short_label.configure(text="MATRIX", text_color="#ffffff")
        else:
            matrix_status_pill.configure(fg_color="#c62828")
            matrix_icon_label.configure(text_color="#ffcdd2")
            matrix_short_label.configure(text="MATRIX", text_color="#ffffff")

    # Logo central (affiché quand aucune semaine n'est sélectionnée)
    logo_center = ctk.CTkFrame(main_area, fg_color="#f7f9fc")
    logo_center.grid(row=0, column=0, rowspan=2, sticky="nsew", padx=20, pady=20)

    # Configuration pour centrer verticalement
    logo_center.grid_rowconfigure(0, weight=1)  # Espace au-dessus
    logo_center.grid_rowconfigure(2, weight=1)  # Espace en-dessous
    logo_center.grid_columnconfigure(0, weight=1)

    # Frame centrale pour le contenu
    center_content = ctk.CTkFrame(logo_center, fg_color="transparent")
    center_content.grid(row=1, column=0)

    # Essayer de charger le logo en grand
    try:
        logo_path = os.path.join(os.path.dirname(__file__), "logo_0.png")
        if os.path.exists(logo_path):
            logo_image = Image.open(logo_path)
            # Agrandir à 4x + 20% la taille originale
            original_width, original_height = logo_image.size
            logo_scale_factor = 1.2  # 20% plus gros (1.0 = taille normale, 1.5 = 50% plus gros, etc.)
            new_width = int(original_width * logo_scale_factor)
            new_height = int(original_height * logo_scale_factor)
            logo_image = logo_image.resize((new_width, new_height), Image.Resampling.LANCZOS)
            logo_ctk_large = ctk.CTkImage(light_image=logo_image, size=(new_width, new_height))

            # Texte de bienvenue (plus gros et plus haut)
            welcome_label = ctk.CTkLabel(
                center_content,
                text="Bienvenue dans le tableau de bord\n\nSélectionnez une semaine pour commencer",
                font=("Arial", 28, "bold"),  # Plus gros : 20 → 28
                text_color="#1f4e79",
                justify="center"
            )
            welcome_label.pack(pady=(0, 0))  # Espace augmenté à 60px sous le texte de bienvenue

            # Logo sans cadre
            logo_center_label = ctk.CTkLabel(
                center_content,
                image=logo_ctk_large,
                text="",
                fg_color="transparent"
            )
            logo_center_label.pack(pady=(100, 0))  # Espace de 20px sous le logo
        else:
            # Si pas de logo, afficher juste un message centré
            # Texte de bienvenue
            welcome_label = ctk.CTkLabel(
                center_content,
                text="Bienvenue dans le tableau de bord\n\nSélectionnez une semaine pour commencer",
                font=("Arial", 22, "bold"),
                text_color="#1f4e79",
                justify="center"
            )
            welcome_label.pack()
    except Exception as e:
        print(f"Erreur lors du chargement du logo central: {e}")
        # Texte de bienvenue en cas d'erreur
        welcome_label = ctk.CTkLabel(
            center_content,
            text="Bienvenue dans le tableau de bord\n\nSélectionnez une semaine pour commencer",
            font=("Arial", 22, "bold"),
            text_color="#1f4e79",
            justify="center"
        )
        welcome_label.pack()

    # Zone de contenu principal (masquée au départ)
    content = ctk.CTkFrame(main_area, fg_color="#f7f9fc", corner_radius=15)
    content.grid_remove()  # Masqué au départ
    content.grid_rowconfigure(0, weight=1)
    content.grid_columnconfigure(0, weight=1)

    # Créer le dashboard
    # Au départ, aucune semaine n'est sélectionnée, donc pas de mise à jour du statut matrix

    # Maximisation automatique après création de la fenêtre
    def maximize_window():
        try:
            # Windows : fenêtre maximisée
            app.state("zoomed")
        except Exception:
            # Fallback : plein écran générique
            try:
                app.attributes("-fullscreen", True)
            except Exception:
                pass

    # Effet de fondu à l'ouverture (similaire à la fenêtre de login)
    def fade_in(step: float = 0.05):
        current_alpha = app.attributes("-alpha")
        if current_alpha < 1.0:
            new_alpha = min(1.0, current_alpha + step)
            app.attributes("-alpha", new_alpha)
            app.after(20, fade_in)

    # On programme la maximisation juste après le rendu initial,
    # puis on lance l'animation de fondu
    app.after(100, maximize_window)
    fade_in()

    # Lancement de la boucle principale
    app.mainloop()


open_main_window("John Doe", 1920, 1080)