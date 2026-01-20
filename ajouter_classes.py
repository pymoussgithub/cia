import customtkinter as ctk
import pandas as pd
import os
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def open_add_class_dialog(horaire, school_key, display_name, school_color, week_folder, refresh_callback=None):
    NIVEAUX = ["A0", "A0/A0+", "Pré-A1", "Pré-A1/A1", "A1", "A1.2", "A1.2/A2",
               "A2", "A2/A2.2", "A2.2", "A2.2/B1", "B1", "B1.2", "B2", "Pitchoune", "Non Spécifié"]

    def soften_color(color, factor=0.6):
        if color.startswith('#'):
            r = int(color[1:3], 16)
            g = int(color[3:5], 16)
            b = int(color[5:7], 16)
            r = int(r + (255 - r) * (1 - factor))
            g = int(g + (255 - g) * (1 - factor))
            b = int(b + (255 - b) * (1 - factor))
            return f"#{r:02x}{g:02x}{b:02x}"
        return color

    OPTION_STYLE = {
        "height": 38,
        "corner_radius": 8,
        "fg_color": soften_color(school_color, 0.5),
        "text_color": "white",
        "button_color": soften_color(school_color, 0.3),
        "button_hover_color": soften_color(school_color, 0.3),
        "dropdown_fg_color": soften_color(school_color, 0.5),
        "dropdown_hover_color": soften_color(school_color, 0.3),
        "dropdown_text_color": "white",
        "cursor": "hand2"
    }

    school_to_file_mapping = {
        'ecole_a': 'ecole_a.xlsx',
        'ecole_b': 'ecole_b.xlsx',
        'ecole_c_cs': 'ECOLE_C_cours_standard.xlsx',
        'ecole_c_ci': 'ECOLE_C_cours_intensif.xlsx',
        'ecole_morning': 'MORNING.xlsx',
        'ecole_premium_cs': 'ECOLE_PREMIUM_cours_standard.xlsx',
        'ecole_premium_ci': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
    }

    # Extraire le numéro de semaine du week_folder
    week_name = os.path.basename(week_folder)
    if week_name.startswith("semaine_"):
        try:
            week_num = week_name.split("_")[1]
            week_display = f"Semaine {week_num}"
        except:
            week_display = week_name
    else:
        week_display = week_name

    dialog = ctk.CTkToplevel()
    dialog.title(f"Ajout d'une classe - {display_name} - {week_display}")
    dialog.geometry("450x350")
    dialog.resizable(False, False)
    dialog.configure(fg_color="white")
    dialog.attributes("-topmost", True)

    main_frame = ctk.CTkFrame(dialog, fg_color="transparent")
    main_frame.pack(fill="both", expand=True, padx=30, pady=25)

    # --- NOM ---
    ctk.CTkLabel(main_frame, text="NOM DE LA CLASSE", font=("Inter", 10, "bold"), text_color="#6B7280").pack(anchor="w")
    classe_entry = ctk.CTkEntry(main_frame, placeholder_text="Ex : CP-A", height=38, corner_radius=8, fg_color="#F3F4F6", border_width=0)
    classe_entry.pack(fill="x", pady=(0, 15))

    # --- NIVEAU ---
    ctk.CTkLabel(main_frame, text="NIVEAU", font=("Inter", 10, "bold"), text_color="#6B7280").pack(anchor="w")
    niveau_menu = ctk.CTkOptionMenu(main_frame, values=NIVEAUX, **OPTION_STYLE)
    niveau_menu.pack(fill="x", pady=(0, 15))
    niveau_menu.set("Choisir...")


    error_label = ctk.CTkLabel(main_frame, text="", font=("Inter", 11), text_color="#EF4444")
    error_label.pack(pady=10)

    def save_class():
        classe_nom = classe_entry.get().strip()
        niveau = niveau_menu.get()

        if not classe_nom or niveau == "Choisir...":
            error_label.configure(text="⚠️ Nom et niveau requis")
            return

        # Vérifier que le nom de classe n'est pas un nombre (int ou float)
        try:
            # Essayer de convertir en float pour détecter les nombres
            float(classe_nom)
            error_label.configure(text="⚠️ Le nom de la classe ne peut pas être un nombre")
            return
        except ValueError:
            # Si ça lève une exception, c'est bien une string, on continue
            pass

        # Valeurs par défaut pour l'intervenant
        intervenant_final = "Non spécifié"
        type_final = "Non spécifié"

        has_real_intervenant = False

        filename = school_to_file_mapping.get(school_key)
        file_path = os.path.join(week_folder, filename)

        try:
            wb = load_workbook(file_path)
            original_order = wb.sheetnames.copy()

            def clean(text):
                t = str(text).lower()
                for w in ["professeur", "prof", "animateur", "anim", "Animateur", "Rôle", "rôle", "role"]:
                    t = t.replace(w, "")
                return t.replace(" ", "").strip()

            target = clean(horaire)
            sheet = next((s for s in original_order if clean(s) == target or s == horaire), None)

            if not sheet:
                error_label.configure(text="⚠️ Horaire introuvable")
                return

            df = pd.read_excel(file_path, sheet_name=sheet)

            new_row = {}
            for col in df.columns:
                c = col.lower()
                if "classe" in c or "groupe" in c: new_row[col] = classe_nom
                elif "niveau" in c: new_row[col] = niveau
                elif "intervenant" in c: new_row[col] = intervenant_final
                elif "type" in c: new_row[col] = type_final
                elif "élève" in c or "effectif" in c: new_row[col] = ""
                else: new_row[col] = ""

            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

            del wb[sheet]
            ws = wb.create_sheet(sheet)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            wb._sheets.sort(key=lambda x: original_order.index(x.title))
            wb.save(file_path)
            wb.close()


            dialog.destroy()
            messagebox.showinfo("Succès", f"Classe {classe_nom} créée")
            if refresh_callback: refresh_callback()

        except Exception as e:
            print(f"Erreur de sauvegarde: {e}")
            error_label.configure(text="⚠️ Erreur lors de la sauvegarde")

    ctk.CTkButton(
        main_frame, text="Créer la classe", fg_color=school_color, hover_color=school_color,
        height=42, corner_radius=8, font=("Inter", 13, "bold"), command=save_class
    ).pack(fill="x", pady=(10, 0))

    dialog.after(200, lambda: (dialog.focus_force(), classe_entry.focus_force()))