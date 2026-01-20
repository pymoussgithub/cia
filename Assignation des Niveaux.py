import pandas as pd
import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
import ttkbootstrap as tb
import os
import sys
import re
import json
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

class AppPedagogique(ctk.CTk):
    def __init__(self, file_path=None):
        super().__init__()

        # --- DÉTECTION DU CHEMIN ---
        # Le script NE PREND QUE le chemin passé en argument depuis fenetre_principale.py
        if file_path:
            # Chemin absolu passé depuis fenetre_principale.py
            self.file_path = os.path.abspath(file_path)
        else:
            # Erreur : aucun chemin fourni
            error_msg = "Ce script doit être lancé depuis la fenêtre principale.\nUtilisez le bouton 'Assigner les niveaux aux élèves'."
            print(f"ERREUR: {error_msg}")
            messagebox.showerror("Erreur de lancement", error_msg)
            sys.exit()

        # --- CONFIGURATION DES NIVEAUX ---
        self.NIVEAUX = ["A0", "A0/A0+", "Pré-A1", "Pré-A1/A1", "A1", "A1.2", "A1.2/A2",
                        "A2", "A2/A2.2", "A2.2", "A2.2/B1", "B1", "B1.2", "B2", "Pitchoune"]
        # Couleurs très légères avec forte transparence simulée (quasi-transparentes)
        self.COULEURS = ["#E8F2FA", "#E1EDF8", "#DAE8F6", "#E8F0E8", "#E1F0E6", "#D1EDEA",
                         "#FFF4E0", "#FFF0D1", "#FFEBCC", "#FCE8E8", "#F0E6F2", "#E8DCE8",
                         "#E4F2F0", "#DDE5EA", "#FCE4EC"]
        self.map_colors = dict(zip(self.NIVEAUX, self.COULEURS))
        self.AGES = {"3-7": (3, 7), "8-11": (8, 11), "12-17+": (12, 25)}

        # Variables de filtres
        self.filter_levels = {niv: ctk.BooleanVar(value=False) for niv in self.NIVEAUX}
        self.filter_ages = {label: ctk.BooleanVar(value=False) for label in self.AGES.keys()}
        self.filter_ci = {"Avec CI": ctk.BooleanVar(value=False), "Sans CI": ctk.BooleanVar(value=False)}
        self.filter_no_level = ctk.BooleanVar(value=False)  # Filtre pour les élèves sans niveau

        # Détection automatique du numéro de semaine depuis le chemin
        week_name = os.path.basename(os.path.dirname(self.file_path))
        if week_name.startswith("semaine_"):
            try:
                week_num = week_name.split("_")[1]
                week_title = f"Semaine {week_num}"
            except:
                week_title = week_name
        else:
            week_title = "Gestion Pédagogique"

        self.title(f"Attribution des Niveaux - {week_title}")
        
        # 1. On définit le mode d'apparence avant de créer les widgets
        ctk.set_appearance_mode("light")

        # 2. On charge et on crée tout le contenu
        self.load_data()
        self.setup_styles()
        self.create_widgets()
        self.refresh_table(preserve_selection=False)

        # 3. Système de surveillance
        self.matrix_last_modified = None
        self.matrix_watch_job = None
        self.start_matrix_watch()

        # 4. EN DERNIER : On force le plein écran
        # On utilise after(200, ...) pour laisser un micro-délai au rendu graphique
        self.after(200, lambda: self.state('zoomed'))
        
    def load_data(self):
        """Charge le fichier passé en argument depuis fenetre_principale.py."""
        if not os.path.exists(self.file_path):
            error_msg = f"Le fichier matrix.xlsx n'a pas été trouvé :\n{self.file_path}\n\nVérifiez que le fichier existe dans le dossier de la semaine sélectionnée."
            print(f"ERREUR: {error_msg}")
            messagebox.showerror("Fichier matrix introuvable", error_msg)
            sys.exit()

        try:
            self.df = self.safe_read_excel(self.file_path)
            # Normalisation des colonnes
            self.cols_map = {
                "Ecole": self.find_column(["ecole", "école", "school"]),
                "Classe": self.find_column(["classe", "class", "groupe"]),
                "Stagiaire": self.find_column(["stagiaire", "nom", "élève"]),
                "Âge": self.find_column(["âge", "age"]),
                "Cours 2": self.find_column(["Cours 2"]),
                "Niveau": self.find_column(["niveau"]),
                "Classe CI": self.find_column(["classe ci", "classe_ci"]),
                "Prof": self.find_column(["prof", "professeur", "enseignant"]),
                "Prof CI": self.find_column(["prof ci", "prof_ci", "professeur ci"]),
                "Départ": self.find_column(["cours 1 du", "départ", "depart"]),
                "Arrivée": self.find_column(["cours 1 au", "arrivée", "arrivee"]),
                "Départ CI": self.find_column(["cours 2 du", "départ ci", "depart ci"]),
                "Arrivée CI": self.find_column(["cours 2 au", "arrivée ci", "arrivee ci"])
            }
            if not self.cols_map["Stagiaire"]:
                raise ValueError("Impossible de trouver la colonne 'Stagiaire'.")
            
            if not self.cols_map["Niveau"]:
                self.df["Niveau"] = None
                self.cols_map["Niveau"] = "Niveau"

            if not self.cols_map["Cours 2"]:
                self.df["Cours 2"] = None
                self.cols_map["Cours 2"] = "Cours 2"

            if not self.cols_map["Classe CI"]:
                self.df["Classe CI"] = None
                self.cols_map["Classe CI"] = "Classe CI"

            if not self.cols_map["Prof"]:
                self.df["Prof"] = None
                self.cols_map["Prof"] = "Prof"

            if not self.cols_map["Prof CI"]:
                self.df["Prof CI"] = None
                self.cols_map["Prof CI"] = "Prof CI"

            if not self.cols_map["Départ"]:
                self.df["Départ"] = None
                self.cols_map["Départ"] = "Départ"

            if not self.cols_map["Arrivée"]:
                self.df["Arrivée"] = None
                self.cols_map["Arrivée"] = "Arrivée"

            if not self.cols_map["Départ CI"]:
                self.df["Départ CI"] = None
                self.cols_map["Départ CI"] = "Départ CI"

            if not self.cols_map["Arrivée CI"]:
                self.df["Arrivée CI"] = None
                self.cols_map["Arrivée CI"] = "Arrivée CI"
        except Exception as e:
            error_msg = f"Erreur lors de l'ouverture du fichier :\n{e}"
            print(f"ERREUR: {error_msg}")
            messagebox.showerror("Erreur de lecture", error_msg)
            sys.exit()

    def find_column(self, keywords):
        for col in self.df.columns:
            if any(k in str(col).lower() for k in keywords): return col
        return None

    def safe_read_excel(self, file_path, sheet_name=None):
        """Lit un fichier Excel de manière sécurisée avec gestion d'erreur."""
        try:
            if sheet_name:
                return pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                return pd.read_excel(file_path)
        except Exception as e:
            error_msg = f"Erreur lors de la lecture du fichier Excel {file_path}: {str(e)}"
            print(f"❌ {error_msg}")
            # Essayer de libérer le fichier si c'est un problème de verrouillage
            import gc
            gc.collect()
            raise Exception(error_msg)

    def safe_load_workbook(self, file_path):
        """Charge un workbook Excel de manière sécurisée."""
        if load_workbook is None:
            raise Exception("openpyxl n'est pas disponible")

        try:
            return load_workbook(file_path)
        except Exception as e:
            error_msg = f"Erreur lors du chargement du workbook {file_path}: {str(e)}"
            print(f"❌ {error_msg}")
            raise Exception(error_msg)

    def format_date_jour_mois(self, date_str):
        """Formate une date pour afficher seulement le jour et le mois (JJ/MM)."""
        if not date_str or pd.isna(date_str) or str(date_str).strip() == "":
            return ""

        try:
            # Convertir en string et nettoyer
            date_clean = str(date_str).strip()

            # Si c'est déjà au format JJ/MM, le retourner tel quel
            if len(date_clean) == 5 and date_clean[2] == '/':
                return date_clean

            # Essayer de parser différentes formes de dates
            from datetime import datetime
            import re

            # Chercher un pattern JJ/MM/YYYY ou JJ/MM/YYYY HH:MM:SS
            match = re.search(r'(\d{1,2})/(\d{1,2})(?:/\d{4})?(?:\s+\d{1,2}:\d{1,2}(?::\d{1,2})?)?', date_clean)
            if match:
                jour = match.group(1).zfill(2)
                mois = match.group(2).zfill(2)
                return f"{jour}/{mois}"

            # Essayer de parser avec datetime
            # Essayer différents formats courants
            formats_to_try = [
                '%d/%m/%Y %H:%M:%S',
                '%d/%m/%Y %H:%M',
                '%d/%m/%Y',
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d',
                '%m/%d/%Y %H:%M:%S',
                '%m/%d/%Y'
            ]

            for fmt in formats_to_try:
                try:
                    parsed_date = datetime.strptime(date_clean, fmt)
                    return parsed_date.strftime('%d/%m')
                except ValueError:
                    continue

            # Si rien ne marche, retourner la valeur originale nettoyée
            return date_clean

        except Exception:
            return str(date_str).strip()

    def get_available_weeks(self):
        """Génère la liste des semaines disponibles pour l'import."""
        weeks = []
        current_dir = os.path.dirname(os.path.abspath(self.file_path))
        parent_dir = os.path.dirname(current_dir)

        # Déterminer la semaine actuelle
        folder_name = os.path.basename(current_dir)
        if folder_name.startswith("semaine_"):
            try:
                current_week = int(folder_name.split("_")[1])
            except:
                current_week = 1
        else:
            current_week = 1

        # Générer les semaines 1 à 9, sauf la semaine actuelle
        for i in range(1, 10):
            if i != current_week:
                week_path = os.path.join(parent_dir, f"semaine_{i}", "matrix.xlsx")
                if os.path.exists(week_path):
                    weeks.append(f"Semaine {i}")

        return weeks

    def on_week_selected_for_import(self, selected_week):
        """Gère la sélection d'une semaine dans le dropdown d'import."""
        if selected_week and selected_week != "📥 Importer élèves":
            # Extraire le numéro de semaine
            week_num = int(selected_week.split(" ")[1])

            # Afficher la fenêtre de confirmation
            self.show_import_confirmation(week_num)

            # Remettre le dropdown à son état initial
            self.import_dropdown.set("📥 Importer élèves")

    def show_import_confirmation(self, week_num):
        """Affiche une fenêtre de confirmation pour l'import avec design moderne."""
        # Créer la fenêtre moderne
        dialog = ctk.CTkToplevel(self)
        dialog.title("")
        dialog.geometry("450x200")
        dialog.resizable(False, False)

        # Centrer la fenêtre
        dialog.geometry("+{}+{}".format(
            self.winfo_rootx() + self.winfo_width()//2 - 225,
            self.winfo_rooty() + self.winfo_height()//2 - 100
        ))

        # Configuration moderne
        dialog.attributes("-topmost", True)
        dialog.overrideredirect(True)  # Pas de barre de titre

        # Frame principal avec coins arrondis
        main_frame = ctk.CTkFrame(
            dialog,
            fg_color="#ffffff",
            bg_color="transparent",
            corner_radius=20,
            border_width=2,
            border_color="#e2e8f0"
        )
        main_frame.pack(fill="both", expand=True, padx=2, pady=2)

        # Frame intérieur pour le contenu
        content_frame = ctk.CTkFrame(
            main_frame,
            fg_color="#f8fafc",
            corner_radius=12
        )
        content_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Header avec icône
        header_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        header_frame.pack(fill="x", pady=(15, 5))

        icon_label = ctk.CTkLabel(
            header_frame,
            text="📥",
            font=("Segoe UI", 24),
            text_color="#6366f1"
        )
        icon_label.pack(side="left", padx=(10, 0))

        title_label = ctk.CTkLabel(
            header_frame,
            text=f"Import depuis semaine {week_num}",
            font=("Segoe UI", 14, "bold"),
            text_color="#1e293b"
        )
        title_label.pack(side="left", padx=(8, 0))

        # Message de confirmation
        message_label = ctk.CTkLabel(
            content_frame,
            text="Importer les assignations :\n• Niveaux, écoles, horaires, classes",
            font=("Segoe UI", 12),
            text_color="#475569",
            justify="center"
        )
        message_label.pack(pady=(10, 20))

        # Boutons
        button_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        button_frame.pack(pady=(0, 15))

        def on_confirm():
            dialog.destroy()
            self.import_from_week(week_num)

        def on_cancel():
            dialog.destroy()

        cancel_btn = ctk.CTkButton(
            button_frame,
            text="❌ Annuler",
            width=100,
            height=35,
            font=("Segoe UI", 11),
            fg_color="#64748b",
            hover_color="#475569",
            corner_radius=8,
            command=on_cancel
        )
        cancel_btn.pack(side="left", padx=(10, 8))

        confirm_btn = ctk.CTkButton(
            button_frame,
            text="✅ Importer",
            width=100,
            height=35,
            font=("Segoe UI", 11, "bold"),
            fg_color="#10b981",
            hover_color="#059669",
            corner_radius=8,
            command=on_confirm
        )
        confirm_btn.pack(side="left")

    def import_from_week(self, week_num):
        """Importe les niveaux, écoles, horaires et classes depuis une semaine spécifique."""
        current_dir = os.path.dirname(os.path.abspath(self.file_path))
        parent_dir = os.path.dirname(current_dir)

        # Chemin vers le fichier source
        source_path = os.path.join(parent_dir, f"semaine_{week_num}", "matrix.xlsx")

        if not os.path.exists(source_path):
            error_msg = f"Le fichier matrix.xlsx de la semaine {week_num} est introuvable à :\n{source_path}"
            print(f"ERREUR: {error_msg}")
            messagebox.showerror("Erreur", error_msg)
            return

        try:
            # Lire le fichier source
            df_source = self.safe_read_excel(source_path)

            # Créer un mapping complet des assignations par nom d'élève
            source_cols = self.find_columns_for_df(df_source)
            if not source_cols["Stagiaire"]:
                available_cols = list(df_source.columns)
                error_msg = f"Colonne 'Stagiaire' manquante dans le fichier de la semaine {week_num}.\nColonnes disponibles : {available_cols}"
                print(f"ERREUR: {error_msg}")
                messagebox.showerror("Erreur", error_msg)
                return

            # Mapping intelligent sur les noms (normaliser les noms)
            mapping = {}
            for _, row in df_source.iterrows():
                nom = str(row[source_cols["Stagiaire"]]).strip()
                if pd.notna(nom) and nom != "":
                    student_data = {}

                    # Récupérer le niveau s'il existe
                    if "Niveau" in source_cols and source_cols["Niveau"] and pd.notna(row[source_cols["Niveau"]]):
                        student_data["niveau"] = row[source_cols["Niveau"]]

                    # Récupérer l'école si elle existe
                    if "Ecole" in source_cols and source_cols["Ecole"] and pd.notna(row[source_cols["Ecole"]]):
                        student_data["ecole"] = row[source_cols["Ecole"]]

                    # Récupérer l'horaire si il existe
                    if "Horaire" in source_cols and source_cols["Horaire"] and pd.notna(row[source_cols["Horaire"]]):
                        student_data["horaire"] = row[source_cols["Horaire"]]

                    # Récupérer la classe si elle existe
                    if "Classe" in source_cols and source_cols["Classe"] and pd.notna(row[source_cols["Classe"]]):
                        student_data["classe"] = row[source_cols["Classe"]]

                    if student_data:  # Seulement si au moins une donnée à importer
                        mapping[nom] = student_data

            if not mapping:
                error_msg = f"Aucune donnée d'assignation trouvée dans le fichier de la semaine {week_num}."
                print(f"ERREUR: {error_msg}")
                messagebox.showerror("Erreur", error_msg)
                return

            count = 0
            imported_students = []
            imported_data = []  # Liste des données importées pour chaque élève

            # Appliquer les données aux élèves actuels
            for idx, row in self.df.iterrows():
                nom = str(row[self.cols_map["Stagiaire"]]).strip()

                if nom in mapping:
                    student_data = mapping[nom]
                    imported_info = {"nom": nom, "donnees": []}

                    # Importer le niveau si la colonne existe et que l'élève n'a pas déjà un niveau
                    if "niveau" in student_data and self.cols_map["Niveau"]:
                        current_level = row[self.cols_map["Niveau"]]
                        if pd.isna(current_level) or current_level == "" or str(current_level).strip() == "":
                            self.df.at[idx, self.cols_map["Niveau"]] = student_data["niveau"]
                            imported_info["donnees"].append(f"Niveau: {student_data['niveau']}")

                    # Importer l'école si la colonne existe et que l'élève n'a pas déjà une école
                    if "ecole" in student_data and self.cols_map["Ecole"]:
                        current_ecole = row[self.cols_map["Ecole"]]
                        if pd.isna(current_ecole) or current_ecole == "" or str(current_ecole).strip() == "":
                            self.df.at[idx, self.cols_map["Ecole"]] = student_data["ecole"]
                            imported_info["donnees"].append(f"École: {student_data['ecole']}")

                    # Import de l'horaire désactivé (colonne supprimée)
                    # if "horaire" in student_data and self.cols_map["Horaire"]:
                    #     current_horaire = row[self.cols_map["Horaire"]]
                    #     if pd.isna(current_horaire) or current_horaire == "" or str(current_horaire).strip() == "":
                    #         self.df.at[idx, self.cols_map["Horaire"]] = student_data["horaire"]
                    #         imported_info["donnees"].append(f"Horaire: {student_data['horaire']}")

                    # Importer la classe si la colonne existe et que l'élève n'a pas déjà une classe
                    if "classe" in student_data and self.cols_map["Classe"]:
                        current_classe = row[self.cols_map["Classe"]]
                        if pd.isna(current_classe) or current_classe == "" or str(current_classe).strip() == "":
                            self.df.at[idx, self.cols_map["Classe"]] = student_data["classe"]
                            imported_info["donnees"].append(f"Classe: {student_data['classe']}")

                    # Si au moins une donnée a été importée
                    if imported_info["donnees"]:
                        imported_students.append((nom, " + ".join(imported_info["donnees"])))
                        imported_data.append(imported_info)
                        count += 1

            # Sauvegarder le fichier matrix
            self.df.to_excel(self.file_path, index=False)

            # Vérifier et créer les classes manquantes dans les fichiers Excel des écoles
            # Désactivé car la colonne Horaire a été supprimée
            # if imported_data:
            #     print("Vérification et création des classes manquantes...")
            #     self.create_missing_classes_from_import(imported_data)

            # Rafraîchir l'affichage
            self.refresh_table(preserve_selection=False)
            self.update_counters()

            if count == 0:
                warning_msg = f"Aucun élève de la semaine {week_num} n'a pu être mis à jour.\nVérifiez que les noms correspondent et que les élèves n'ont pas déjà des assignations."
                print(f"AVERTISSEMENT: {warning_msg}")
                messagebox.showwarning("Aucun import", warning_msg)
            else:
                success_msg = f"{count} élèves mis à jour depuis la semaine {week_num}"
                print(f"SUCCÈS: {success_msg}")
                self.show_import_success_dialog(week_num, imported_students)

        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            error_msg = f"Erreur lors de l'import : {str(e)}\n\nDétails :\n{error_details}"
            print(f"ERREUR: {error_msg}")
            messagebox.showerror("Erreur", error_msg)

    def create_missing_classes_from_import(self, imported_data):
        """Crée les classes manquantes dans les fichiers Excel des écoles après l'import."""
        if load_workbook is None:
            print("⚠️ openpyxl n'est pas disponible, impossible de créer les classes manquantes")
            return

        # Mapping des écoles
        school_file_mapping = {
            'ecole_a': 'ecole_a.xlsx',
            'ecole_b': 'ecole_b.xlsx',
            'ecole_c_cs': 'ECOLE_C_cours_standard.xlsx',
            'ecole_c_ci': 'ECOLE_C_cours_intensif.xlsx',
            'ecole_morning': 'MORNING.xlsx',
            'ecole_premium_cs': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'ecole_premium_ci': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        # Grouper les élèves par école, horaire et classe
        classes_to_create = {}  # {school_key: {horaire: {classe: [eleves]}}}

        for data in imported_data:
            nom = data["nom"]
            # Récupérer les données actuelles de l'élève dans le dataframe
            student_row = self.df[self.df[self.cols_map["Stagiaire"]] == nom]
            if not student_row.empty:
                ecole = student_row[self.cols_map["Ecole"]].values[0] if self.cols_map["Ecole"] and pd.notna(student_row[self.cols_map["Ecole"]].values[0]) else None
                # horaire = student_row[self.cols_map["Horaire"]].values[0] if self.cols_map["Horaire"] and pd.notna(student_row[self.cols_map["Horaire"]].values[0]) else None  # Désactivé
                classe = student_row[self.cols_map["Classe"]].values[0] if self.cols_map["Classe"] and pd.notna(student_row[self.cols_map["Classe"]].values[0]) else None
                niveau = student_row[self.cols_map["Niveau"]].values[0] if self.cols_map["Niveau"] and pd.notna(student_row[self.cols_map["Niveau"]].values[0]) else None

                # Convertir l'école en clé d'école
                school_key = None
                for key, display_name in {
                    'ecole_a': 'A',
                    'ecole_b': 'B',
                    'ecole_c_cs': 'C/CS',
                    'ecole_c_ci': 'C/CI',
                    'ecole_morning': 'Morning',
                    'ecole_premium_cs': 'Premium/CS',
                    'ecole_premium_ci': 'Premium/CI'
                }.items():
                    if pd.notna(ecole) and str(ecole).strip() == display_name:
                        school_key = key
                        break

                if school_key and pd.notna(horaire) and pd.notna(classe):
                    horaire_str = str(horaire).strip()
                    classe_str = str(classe).strip()

                    if school_key not in classes_to_create:
                        classes_to_create[school_key] = {}
                    if horaire_str not in classes_to_create[school_key]:
                        classes_to_create[school_key][horaire_str] = {}
                    if classe_str not in classes_to_create[school_key][horaire_str]:
                        classes_to_create[school_key][horaire_str][classe_str] = []

                    # Ajouter l'élève avec ses informations
                    classes_to_create[school_key][horaire_str][classe_str].append({
                        'nom': nom,
                        'niveau': str(niveau).strip() if pd.notna(niveau) else ""
                    })

        # Créer les classes manquantes
        week_folder = os.path.dirname(self.file_path)
        classes_created = 0

        for school_key, horaires in classes_to_create.items():
            excel_filename = school_file_mapping.get(school_key)
            if not excel_filename:
                continue

            excel_path = os.path.join(week_folder, excel_filename)
            if not os.path.exists(excel_path):
                continue

            try:
                wb = self.safe_load_workbook(excel_path)
                sheet_modified = False

                for horaire, classes in horaires.items():
                    # Nettoyer le nom d'horaire pour la recherche
                    horaire_clean = self.clean_horaire_name(horaire)

                    # Trouver la feuille correspondant à l'horaire
                    target_sheet = None
                    for sheet_name in wb.sheetnames:
                        sheet_clean = self.clean_horaire_name(sheet_name)
                        if sheet_clean == horaire_clean or horaire_clean in sheet_clean or sheet_clean in horaire_clean:
                            target_sheet = wb[sheet_name]
                            break

                    if not target_sheet:
                        print(f"⚠️ Feuille horaire '{horaire}' non trouvée dans {excel_filename}")
                        continue

                    # Vérifier chaque classe
                    for classe_nom, eleves in classes.items():
                        # Chercher si la classe existe déjà
                        classe_exists = False
                        classe_row = None

                        for row_idx in range(2, target_sheet.max_row + 1):
                            cell_value = str(target_sheet.cell(row=row_idx, column=1).value or '').strip()
                            if cell_value == classe_nom:
                                classe_exists = True
                                classe_row = row_idx
                                break

                        if not classe_exists:
                            # Trouver la première ligne vide pour ajouter la classe
                            new_row = target_sheet.max_row + 1

                            # Ajouter la classe (colonne 1)
                            target_sheet.cell(row=new_row, column=1, value=classe_nom)

                            # Ajouter les élèves dans les colonnes appropriées
                            eleves_text = ', '.join([eleve['nom'] for eleve in eleves])
                            target_sheet.cell(row=new_row, column=5, value=eleves_text)  # Colonne 5 = élèves

                            # Ajouter le niveau dans la colonne appropriée si disponible
                            if eleves and eleves[0]['niveau']:
                                # Chercher la colonne niveau (généralement colonne 4)
                                target_sheet.cell(row=new_row, column=4, value=eleves[0]['niveau'])

                            sheet_modified = True
                            classes_created += 1
                            print(f"✅ Classe '{classe_nom}' créée dans {excel_filename} (horaire: {horaire})")
                        else:
                            print(f"ℹ️ Classe '{classe_nom}' existe déjà dans {excel_filename}")

                if sheet_modified:
                    wb.save(excel_path)
                    wb.close()

            except Exception as e:
                print(f"❌ Erreur lors de la création des classes dans {excel_filename}: {e}")

        if classes_created > 0:
            print(f"✅ {classes_created} classe(s) créée(s) automatiquement")

            # Mettre à jour le dashboard des classes si on a accès à la fenêtre principale
            try:
                # Essayer de trouver la fenêtre principale et rafraîchir le dashboard
                for window in ctk.CTk._all_toplevels:
                    if hasattr(window, 'create_classes_dashboard') and hasattr(window, 'content'):
                        try:
                            # Rafraîchir le dashboard des classes
                            school_data = window.analyze_school_classes(week_folder)
                            window.create_classes_dashboard(window.content, school_data, week_folder)
                            print("✅ Dashboard des classes mis à jour")
                            break
                        except:
                            pass
            except:
                pass

    def show_import_success_dialog(self, week_num, imported_students):
        """Affiche une fenêtre de succès avec la liste des élèves importés sous forme de liste verticale."""
        dialog = ctk.CTkToplevel(self)
        dialog.title(f"Import réussi - Semaine {week_num}")
        dialog.geometry("600x500")
        dialog.resizable(True, True)
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(fg_color="white")

        # Centrer la fenêtre
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (600 // 2)
        y = (dialog.winfo_screenheight() // 2) - (500 // 2)
        dialog.geometry(f"600x500+{x}+{y}")

        # Titre
        title_label = ctk.CTkLabel(
            dialog,
            text=f"✅ {len(imported_students)} élèves mis à jour depuis la semaine {week_num}",
            font=("Inter", 16, "bold"),
            text_color="#10b981"
        )
        title_label.pack(pady=(20, 15))

        # Frame principale pour la liste
        main_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        # Label d'en-tête
        header_label = ctk.CTkLabel(
            main_frame,
            text="Élèves importés :",
            font=("Inter", 12, "bold"),
            text_color="#6B7280"
        )
        header_label.pack(anchor="w", pady=(0, 10))

        # Frame scrollable pour la liste des élèves
        scrollable_frame = ctk.CTkScrollableFrame(main_frame, fg_color="#F9FAFB", corner_radius=10)
        scrollable_frame.pack(fill="both", expand=True)

        # Créer les éléments de la liste
        for student_name, imported_data in imported_students:
            # Frame pour chaque élève
            item_frame = ctk.CTkFrame(scrollable_frame, fg_color="white", corner_radius=6)
            item_frame.pack(fill="x", padx=8, pady=(0, 2))

            # Nom de l'élève (moins gros)
            name_label = ctk.CTkLabel(
                item_frame,
                text=student_name,
                font=("Inter", 12, "bold"),
                text_color="#111827",
                anchor="w"
            )
            name_label.pack(anchor="w", padx=12, pady=(8, 4))

            # Données importées (un peu plus gros)
            data_label = ctk.CTkLabel(
                item_frame,
                text=f"Importé : {imported_data}",
                font=("Inter", 11),
                text_color="#6B7280",
                anchor="w",
                wraplength=520
            )
            data_label.pack(anchor="w", padx=12, pady=(0, 8))

        # Bouton OK en bas de la fenêtre
        ok_button = ctk.CTkButton(
            dialog,
            text="OK",
            width=140,
            height=38,
            font=("Inter", 12, "bold"),
            fg_color="#10B981",
            hover_color="#059669",
            corner_radius=8,
            command=dialog.destroy
        )
        ok_button.pack(pady=(15, 25))

    def find_columns_for_df(self, df):
        """Trouve les colonnes nécessaires dans un DataFrame."""
        cols_map = {}
        cols_map["Stagiaire"] = self.find_column_in_df(df, ["stagiaire", "nom", "name", "élève"])
        cols_map["Niveau"] = self.find_column_in_df(df, ["niveau", "level", "niveau actuel"])
        cols_map["Ecole"] = self.find_column_in_df(df, ["ecole", "école", "school"])
        # cols_map["Horaire"] = self.find_column_in_df(df, ["horaire", "horaire", "time", "schedule"])  # Désactivé
        cols_map["Classe"] = self.find_column_in_df(df, ["classe", "class", "groupe"])
        return cols_map

    def find_column_in_df(self, df, keywords):
        """Trouve une colonne dans un DataFrame selon des mots-clés."""
        # D'abord chercher une correspondance exacte
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if col_lower in keywords:
                return col

        # Ensuite chercher une correspondance partielle (mais plus stricte)
        for col in df.columns:
            col_lower = str(col).lower().strip()
            for k in keywords:
                if k in col_lower and len(k) > 2:  # Éviter les faux positifs avec des mots courts
                    return col
        return None

    def clean_horaire_name(self, sheet_name):
        """Nettoie le nom de la feuille pour extraire le nom d'horaire de manière cohérente."""
        # Cette fonction doit être identique à celle utilisée dans fenetre_principale.py et Assignation des Niveaux.py
        sheet_lower = sheet_name.lower()
        type_intervenant = "animateur" if "animateur" in sheet_lower else "professeur"

        # Liste complète des mots à supprimer (avec variations)
        words_to_remove = [
            "animateur", "Animateur", "anim", "Anim",
            "professeur", "Professeur", "prof", "Prof",
            "Rôle", "rôle", "role", "Role"
        ]

        # Nettoyer le nom en supprimant tous les mots-clés d'intervenants
        horaire = sheet_name
        for word in words_to_remove:
            horaire = horaire.replace(word, "")

        # Nettoyer les espaces multiples et supprimer les espaces au début/fin
        horaire = " ".join(horaire.split()).strip()

        # Si le résultat est vide, utiliser le nom original
        return horaire or sheet_name

    def setup_styles(self):
        self.style = tb.Style(theme="flatly")
        self.style.configure("Treeview", rowheight=45, font=("Segoe UI", 12))
        self.style.configure("Treeview.Heading", font=("Segoe UI", 12, "bold"))
        for niv, color in self.map_colors.items():
            self.style.configure(f"{niv}.Treeview", background=color)

    def create_widgets(self):
        # --- COLONNE GAUCHE (FILTRES & STATS) ---
        self.sidebar_left = ctk.CTkFrame(self, width=250, corner_radius=0, fg_color="#f8f9fa")
        self.sidebar_left.pack(side="left", fill="y")

        # En-tête avec titre et bouton reset
        header_frame = ctk.CTkFrame(self.sidebar_left, fg_color="transparent")
        header_frame.pack(fill="x", padx=15, pady=(15, 10))

        ctk.CTkLabel(header_frame, text="🔍 FILTRES", font=("Segoe UI", 16, "bold"), text_color="#2c3e50").pack(side="left")
        self.reset_btn = ctk.CTkButton(header_frame, text="⟲", width=35, height=35, font=("Segoe UI", 12, "bold"),
                                      fg_color="#e74c3c", hover_color="#c0392b", command=self.reset_filters)
        self.reset_btn.pack(side="right")


        # Zone fixe pour les filtres (sans scroll)
        filters_container = ctk.CTkFrame(self.sidebar_left, fg_color="transparent")
        filters_container.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # === CONTENEUR POUR ÂGE ET COURS INTENSIFS CÔTE À CÔTE ===
        age_ci_container = ctk.CTkFrame(filters_container, fg_color="transparent")
        age_ci_container.pack(fill="x", pady=(0, 8))
        age_ci_container.grid_columnconfigure(0, weight=1)
        age_ci_container.grid_columnconfigure(1, weight=1)

        # Créer directement les sections dans le conteneur grille

        # === FILTRE ÂGE (Colonne gauche) ===
        age_section = ctk.CTkFrame(age_ci_container, fg_color="white", corner_radius=10)
        age_section.grid(row=0, column=0, padx=(0, 4), sticky="nsew")

        age_header = ctk.CTkFrame(age_section, fg_color="#ecf0f1", corner_radius=8)
        age_header.pack(fill="x", padx=2, pady=(2, 8))
        ctk.CTkLabel(age_header, text="👶 ÂGE", font=("Segoe UI", 12, "bold"), text_color="#2c3e50").pack(pady=6)

        age_buttons = ctk.CTkFrame(age_section, fg_color="transparent")
        age_buttons.pack(fill="x", padx=8, pady=(0, 8))

        for i, (label, var) in enumerate(self.filter_ages.items()):
            color = ["#3498db", "#e67e22", "#9b59b6"][i % 3]
            btn_container = ctk.CTkFrame(age_buttons, fg_color="transparent")
            btn_container.pack(fill="x", pady=1)

            toggle_btn = ctk.CTkButton(
                btn_container, text=f"{label}", font=("Segoe UI", 10, "bold"),
                height=32, fg_color="#ffffff", text_color="#1a365d", hover_color="#f8fafc",
                border_width=1, border_color="#cbd5e1",
                corner_radius=6, command=lambda v=var, l=label, c=color, b=None: self.toggle_filter(v, l, c, b)
            )
            toggle_btn.pack(fill="x", padx=4, pady=1)
            toggle_btn.filter_var = var
            toggle_btn.original_color = "#66BB6A"  # Couleur verte douce pour l'état actif

            if not hasattr(self, 'filter_buttons'):
                self.filter_buttons = {}
            self.filter_buttons[label] = toggle_btn
            self.update_toggle_button(toggle_btn, var.get())

        # === FILTRE COURS INTENSIFS (Colonne droite) ===
        ci_section = ctk.CTkFrame(age_ci_container, fg_color="white", corner_radius=10)
        ci_section.grid(row=0, column=1, padx=(4, 0), sticky="nsew")

        ci_header = ctk.CTkFrame(ci_section, fg_color="#ecf0f1", corner_radius=8)
        ci_header.pack(fill="x", padx=2, pady=(2, 8))
        ctk.CTkLabel(ci_header, text="⚡ CI", font=("Segoe UI", 11, "bold"), text_color="#2c3e50").pack(pady=6)

        ci_buttons = ctk.CTkFrame(ci_section, fg_color="transparent")
        ci_buttons.pack(fill="x", padx=8, pady=(0, 8))

        for i, (label, var) in enumerate(self.filter_ci.items()):
            color = ["#66BB6A", "#e74c3c"][i % 2]
            btn_container = ctk.CTkFrame(ci_buttons, fg_color="transparent")
            btn_container.pack(fill="x", pady=1)

            toggle_btn = ctk.CTkButton(
                btn_container, text=f"{label}", font=("Segoe UI", 9, "bold"),
                height=32, fg_color="#ffffff", text_color="#1a365d", hover_color="#f8fafc",
                border_width=1, border_color="#cbd5e1",
                corner_radius=6, command=lambda v=var, l=label, c=color, b=None: self.toggle_filter(v, l, c, b)
            )
            toggle_btn.pack(fill="x", padx=4, pady=1)
            toggle_btn.filter_var = var
            toggle_btn.original_color = "#66BB6A"  # Couleur verte douce pour l'état actif

            if not hasattr(self, 'filter_buttons'):
                self.filter_buttons = {}
            self.filter_buttons[label] = toggle_btn
            self.update_toggle_button(toggle_btn, var.get())

        # === FILTRE NIVEAUX (avec couleurs des niveaux) ===
        self.create_levels_filter_section(filters_container)


        # --- COLONNE DROITE (NIVEAUX) ---
        self.sidebar_right = ctk.CTkFrame(self, width=180, corner_radius=0)
        self.sidebar_right.pack(side="right", fill="y")
        
        for i, niv in enumerate(self.NIVEAUX):
            pady_top = 10 if i == 0 else 2  # Plus d'espace au-dessus du premier bouton
            btn = ctk.CTkButton(self.sidebar_right, text=niv, height=30, font=("Segoe UI", 11, "bold"),
                                fg_color=self.map_colors[niv], text_color="black", hover_color="#ecf0f1",
                                command=lambda n=niv: self.assign_level(n))
            btn.pack(pady=(pady_top, 2), padx=10, fill="x")

        ctk.CTkButton(self.sidebar_right, text="EFFACER", fg_color="#34495e", height=45, text_color="white", 
                      font=("Segoe UI", 11, "bold"), command=lambda: self.assign_level(None)).pack(side="bottom", pady=20, padx=10, fill="x")

        # Bouton quitter en bas à gauche
        quit_btn = ctk.CTkButton(
            self.sidebar_left,
            text="Quitter",
            height=45,
            fg_color="#e57373",
            hover_color="#ef9a9a",
            text_color="white",
            font=("Segoe UI", 11, "bold"),
            corner_radius=8,
            command=self.on_quit
        )
        quit_btn.pack(side="bottom", padx=15, pady=(10, 15), fill="x")
        
        # Gérer la fermeture par la croix de la fenêtre
        self.protocol("WM_DELETE_WINDOW", self.on_quit)

        # --- CENTRE (TABLEAU) ---
        self.center_container = ctk.CTkFrame(self, fg_color="transparent")
        self.center_container.pack(side="left", fill="both", expand=True, padx=15, pady=10)

        # Frame pour la barre de recherche et les statistiques
        search_stats_frame = ctk.CTkFrame(self.center_container, fg_color="transparent")
        search_stats_frame.pack(fill="x", pady=(0, 10))

        # Configuration de la grille : 3 colonnes (recherche 1/3, stats 2/3)
        search_stats_frame.grid_columnconfigure(0, weight=1)  # Barre de recherche (1/3)
        search_stats_frame.grid_columnconfigure(1, weight=2)  # Statistiques (2/3)

        # Barre de recherche (colonne 0)
        self.search_var = ctk.StringVar()
        self.search_var.trace_add("write", lambda *args: self.refresh_table(preserve_selection=False))
        search_entry = ctk.CTkEntry(
            search_stats_frame,
            placeholder_text="Rechercher un nom...",
            textvariable=self.search_var,
            height=30
        )
        search_entry.grid(row=0, column=0, sticky="ew", padx=(0, 10))

        # Statistiques compactes avec design moderne (colonne 1)
        stats_frame = ctk.CTkFrame(search_stats_frame, fg_color="transparent")
        stats_frame.grid(row=0, column=1, sticky="ew")

        # Configuration pour répartir l'espace
        stats_frame.grid_columnconfigure(0, weight=1)  # Stats prennent l'espace disponible
        stats_frame.grid_columnconfigure(1, weight=0)  # Bouton fixe

        # Frame pour les statistiques avec design moderne
        stats_display_frame = ctk.CTkFrame(
            stats_frame,
            fg_color="#f8fafc",
            corner_radius=8,
            border_width=1,
            border_color="#e1e5e9"
        )
        stats_display_frame.grid(row=0, column=0, sticky="ew", padx=(40, 40))

        # Labels de statistiques avec design moderne - utilisation d'une grille pour mieux contrôler l'espace
        stats_inner_frame = ctk.CTkFrame(stats_display_frame, fg_color="transparent")
        stats_inner_frame.pack(fill="both", expand=True, padx=8, pady=6)

        # Configuration de la grille pour 8 éléments: total | sep | assigned | unassigned | sep | filtered | sep | selected
        # Donner plus d'espace aux labels principaux
        stats_inner_frame.grid_columnconfigure(0, weight=2)  # 👤 Total - plus d'espace
        stats_inner_frame.grid_columnconfigure(1, weight=0)  # Séparateur
        stats_inner_frame.grid_columnconfigure(2, weight=2)  # ✅ Avec niveau - plus d'espace
        stats_inner_frame.grid_columnconfigure(3, weight=2)  # ❌ Sans niveau - plus d'espace
        stats_inner_frame.grid_columnconfigure(4, weight=0)  # Séparateur
        stats_inner_frame.grid_columnconfigure(5, weight=1)  # 🔍 Filtrés
        stats_inner_frame.grid_columnconfigure(6, weight=0)  # Séparateur
        stats_inner_frame.grid_columnconfigure(7, weight=1)  # 🖱️ Sélectionnés

        self.stats_total_label = ctk.CTkLabel(
            stats_inner_frame,
            text="👤 0",
            font=("Segoe UI", 12, "bold"),
            text_color="#475569"
        )
        self.stats_total_label.grid(row=0, column=0, sticky="w", padx=(0, 4))

        # Petit séparateur entre total et stats avec/sans niveau
        separator_total = ctk.CTkFrame(
            stats_inner_frame,
            fg_color="#d1d5db",
            width=1,
            height=16
        )
        separator_total.grid(row=0, column=1, sticky="ns", padx=4)

        self.stats_assigned_label = ctk.CTkLabel(
            stats_inner_frame,
            text="✅ 0",
            font=("Segoe UI", 12, "bold"),
            text_color="#16a34a"
        )
        self.stats_assigned_label.grid(row=0, column=2, sticky="w", padx=(6, 8))

        self.stats_unassigned_label = ctk.CTkLabel(
            stats_inner_frame,
            text="❌ 0",
            font=("Segoe UI", 12, "bold"),
            text_color="#dc2626"
        )
        self.stats_unassigned_label.grid(row=0, column=3, sticky="w", padx=(0, 6))

        # Séparateur vertical
        separator = ctk.CTkFrame(
            stats_inner_frame,
            fg_color="#d1d5db",
            width=1,
            height=20
        )
        separator.grid(row=0, column=4, sticky="ns", padx=(0, 8))

        # Statistiques filtrées
        self.stats_filtered_label = ctk.CTkLabel(
            stats_inner_frame,
            text="🔍 0",
            font=("Segoe UI", 12, "bold"),
            text_color="#2563eb"  # Bleu plus visible
        )
        self.stats_filtered_label.grid(row=0, column=5, sticky="w")

        # Séparateur avant les sélectionnés
        separator_selected = ctk.CTkFrame(
            stats_inner_frame,
            fg_color="#d1d5db",
            width=1,
            height=20
        )
        separator_selected.grid(row=0, column=6, sticky="ns", padx=(8, 2))

        # Statistiques des élèves sélectionnés
        self.stats_selected_label = ctk.CTkLabel(
            stats_inner_frame,
            text="🖱️ 0",
            font=("Segoe UI", 12, "bold"),
            text_color="#7c3aed"  # Violet pour les sélectionnés
        )
        self.stats_selected_label.grid(row=0, column=7, sticky="w")

        # Boutons d'ajout et suppression d'élèves
        actions_frame = ctk.CTkFrame(stats_frame, fg_color="transparent")
        actions_frame.grid(row=0, column=1, sticky="e")

        # Bouton ajouter élève (+)
        add_student_btn = ctk.CTkButton(
            actions_frame,
            text="+",
            width=35,
            height=30,
            font=("Segoe UI", 14, "bold"),
            fg_color="#10b981",
            hover_color="#059669",
            corner_radius=6,
            command=self.open_add_student_dialog
        )
        add_student_btn.pack(side="left", padx=(0, 5))

        # Bouton supprimer élève (-)
        remove_student_btn = ctk.CTkButton(
            actions_frame,
            text="-",
            width=35,
            height=30,
            font=("Segoe UI", 14, "bold"),
            fg_color="#ef4444",
            hover_color="#dc2626",
            corner_radius=6,
            command=self.open_remove_student_dialog
        )
        remove_student_btn.pack(side="left", padx=(0, 10))

        # Bouton dépliable Importer élèves
        self.import_dropdown = ctk.CTkOptionMenu(
            actions_frame,
            values=self.get_available_weeks(),
            command=self.on_week_selected_for_import,
            font=("Segoe UI", 11, "bold"),
            height=30,
            width=140,
            fg_color="#6366f1",
            button_color="#4f46e5",
            button_hover_color="#4338ca",
            text_color="white",
            dropdown_fg_color="#6366f1",
            dropdown_hover_color="#4f46e5",
            dropdown_text_color="white"
        )
        self.import_dropdown.set("📥 Importer élèves")
        self.import_dropdown.pack(side="left")

        self.table_frame = ctk.CTkFrame(self.center_container)
        self.table_frame.pack(fill="both", expand=True)

        # Barre de scroll design personnalisée
        self.scrollbar = ctk.CTkScrollbar(
            self.table_frame,
            orientation="vertical",
            width=16,
            fg_color="#e8f4fd",  # Fond bleu très clair
            button_color="#4a90e2",  # Bleu pour les boutons
            button_hover_color="#357abd"  # Bleu plus foncé au hover
        )
        self.scrollbar.pack(side="right", fill="y")
        
        self.tree = ttk.Treeview(self.table_frame, columns=("Stagiaire", "Niveau", "Age", "Classe", "Prof", "Arrivée", "Départ", "sep", "CI", "Classe CI", "Prof CI", "Arr. CI", "Dép. CI"),
                                 show="headings", style="Treeview", yscrollcommand=self.scrollbar.set)
        self.scrollbar.configure(command=self.tree.yview)

        for col in [("Stagiaire", 150), ("Niveau", 80), ("Age", 40), ("Classe", 70), ("Prof", 70), ("Arrivée", 60), ("Départ", 60), ("sep", 10), ("CI", 30), ("Classe CI", 70), ("Prof CI", 70), ("Arr. CI", 60), ("Dép. CI", 60)]:
            if col[0] == "sep":
                self.tree.heading(col[0], text="")
                self.tree.column(col[0], width=col[1], anchor="center", stretch=False)
            else:
                self.tree.heading(col[0], text=col[0])
                self.tree.column(col[0], width=col[1], anchor="center" if col[0] != "Stagiaire" else "w")
        self.tree.pack(side="left", fill="both", expand=True)

        for niv, color in self.map_colors.items():
            self.tree.tag_configure(niv, background=color)

        # Style pour le séparateur vertical
        self.style.configure("Separator.Treeview", background="#d1d5db", relief="solid")

        # Binding pour mettre à jour les compteurs quand la sélection change
        self.tree.bind('<<TreeviewSelect>>', lambda e: self.update_counters())

        # Variables pour le système de sélection par glisser
        self.drag_zone = None
        self.drag_buttons = []
        self.selected_item = None
        self.hovered_button = None

        # Lier les événements pour le système de sélection par glisser (molette)
        self.tree.bind('<ButtonPress-2>', self.on_right_press)    # Clic molette enfoncé
        self.tree.bind('<B2-Motion>', self.on_right_drag)         # Glisser avec clic molette
        self.tree.bind('<ButtonRelease-2>', self.on_right_release) # Clic molette relâché

        # Lier le clic droit pour afficher le menu contextuel des classes
        self.tree.bind('<Button-3>', self.show_classes_context_menu)

    def on_right_press(self, event):
        """Gère l'enfoncement du clic droit - crée la zone de sélection."""
        region = self.tree.identify_region(event.x, event.y)

        if region == "cell" or region == "tree":
            item = self.tree.identify_row(event.y)

            if item:
                self.selected_item = item
                self.create_drag_zone(event.x_root, event.y_root)

    def on_right_drag(self, event):
        """Gère le glisser avec le clic droit enfoncé - détecte le survol."""
        if self.drag_zone:
            # Vérifier si on survole un bouton (sans déplacer la zone)
            self.check_button_hover(event.x_root, event.y_root)

    def on_right_release(self, event):
        """Gère le relâchement du clic droit - assigne le niveau."""

        if self.drag_zone and self.selected_item:
            # Vérifier si on relâche sur un bouton
            hovered = self.get_button_at_position(event.x_root, event.y_root)

            if hovered:
                # Récupérer le niveau du bouton
                level = hovered.cget('text')
                if level == " ":  # Bouton vide = effacer
                    level = None

                # Assigner le niveau
                self.assign_level_from_drag(self.selected_item, level)

        # Nettoyer
        self.destroy_drag_zone()

    def create_drag_zone(self, x, y):
        """Crée la zone flottante avec les boutons de niveaux."""
        # Détruire la zone existante si elle existe (mais préserver selected_item)
        if self.drag_zone:
            self.drag_zone.destroy()
            self.drag_zone = None
            self.drag_buttons = []
            self.hovered_button = None
            # NE PAS remettre selected_item à None ici !

        # Créer la zone flottante
        self.drag_zone = ctk.CTkToplevel(self)
        self.drag_zone.title("")
        self.drag_zone.geometry("280x90")
        self.drag_zone.overrideredirect(True)  # Pas de bordures
        self.drag_zone.attributes("-topmost", True)

        # Position initiale
        self.drag_zone.geometry(f"+{x - 50}+{y - 30}")

        # Frame principal avec style moderne
        main_frame = ctk.CTkFrame(
            self.drag_zone,
            fg_color="#ffffff",
            corner_radius=10,
            border_width=1,
            border_color="#e1e5e9"
        )
        main_frame.pack(fill="both", expand=True, padx=2, pady=2)

        # Créer les boutons (même disposition que l'ancienne popup)
        niveaux = self.NIVEAUX + [" "]  # Ajouter un bouton vide
        boutons_par_ligne = 5

        self.drag_buttons = []

        for ligne in range(3):
            ligne_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
            ligne_frame.pack(fill="x", pady=1)

            debut = ligne * boutons_par_ligne
            fin = min(debut + boutons_par_ligne, len(niveaux))

            for i in range(debut, fin):
                niveau = niveaux[i]

                if niveau == " ":
                    color = "#f8f9fa"
                    text = " "
                else:
                    color = self.map_colors[niveau]
                    text = niveau

                btn = ctk.CTkButton(
                    ligne_frame,
                    text=text,
                    font=("Segoe UI", 10, "bold"),
                    height=25,
                    width=50,
                    fg_color=color,
                    text_color="black",
                    hover_color="#ffffff",  # Sera changé dynamiquement
                    command=lambda: None  # Pas de commande, on gère avec les événements souris
                )
                btn.pack(side="left", padx=1, pady=1)
                self.drag_buttons.append(btn)

    def check_button_hover(self, mouse_x, mouse_y):
        """Vérifie si la souris survole un bouton et met à jour la surbrillance."""
        for btn in self.drag_buttons:
            # Obtenir les coordonnées absolues du bouton
            btn_x = btn.winfo_rootx()
            btn_y = btn.winfo_rooty()
            btn_width = btn.winfo_width()
            btn_height = btn.winfo_height()

            # Vérifier si la souris est sur ce bouton
            if (btn_x <= mouse_x <= btn_x + btn_width and
                btn_y <= mouse_y <= btn_y + btn_height):
                # Sur ce bouton
                if self.hovered_button != btn:
                    # Désactiver l'ancien bouton survolé
                    if self.hovered_button:
                        self.hovered_button.configure(
                            border_width=0,
                            fg_color=self.get_original_color(self.hovered_button)
                        )
                    # Activer le nouveau
                    self.hovered_button = btn
                    self.hovered_button.configure(
                        border_width=2,
                        border_color="#2563eb",
                        fg_color="#dbeafe"  # Bleu très clair pour la surbrillance
                    )
                return

        # Pas sur un bouton
        if self.hovered_button:
            self.hovered_button.configure(
                border_width=0,
                fg_color=self.get_original_color(self.hovered_button)
            )
            self.hovered_button = None

    def get_original_color(self, button):
        """Récupère la couleur originale d'un bouton."""
        text = button.cget('text')
        if text == " ":
            return "#f8f9fa"
        else:
            return self.map_colors.get(text, "#f8f9fa")

    def get_button_at_position(self, mouse_x, mouse_y):
        """Retourne le bouton aux coordonnées spécifiées, ou None."""

        for i, btn in enumerate(self.drag_buttons):
            # Obtenir les coordonnées absolues du bouton
            btn_x = btn.winfo_rootx()
            btn_y = btn.winfo_rooty()
            btn_width = btn.winfo_width()
            btn_height = btn.winfo_height()

            # Vérifier si la souris est sur ce bouton
            if (btn_x <= mouse_x <= btn_x + btn_width and
                btn_y <= mouse_y <= btn_y + btn_height):
                return btn

        return None

    def destroy_drag_zone(self):
        """Détruit la zone de glisser-déposer."""
        if self.drag_zone:
            self.drag_zone.destroy()
            self.drag_zone = None
        self.drag_buttons = []
        self.hovered_button = None
        self.selected_item = None

    def assign_level_from_drag(self, item, level):
        """Assigne un niveau depuis le système de glisser."""

        def complete_assignment():
            nom = self.tree.item(item, 'values')[0]
            niveau_value = level

            if nom in self.df[self.cols_map["Stagiaire"]].values:
                self.df.loc[self.df[self.cols_map["Stagiaire"]] == nom, self.cols_map["Niveau"]] = niveau_value
                self.df.to_excel(self.file_path, index=False)

            self.refresh_table(preserve_selection=False)
            self.update_counters()  # Mise à jour des statistiques du haut

        self.show_save_notification(complete_assignment)

    def show_level_popup(self, item, x, y):
        """Affiche une fenêtre popup avec tous les niveaux."""
        # Fermer la popup existante si elle existe
        if self.level_popup:
            self.level_popup.destroy()

        # Créer la fenêtre popup
        self.level_popup = ctk.CTkToplevel(self)
        self.level_popup.title("")
        self.level_popup.geometry("450x120")  # Format paysage
        self.level_popup.resizable(False, False)

        # Positionner la fenêtre près du curseur
        self.level_popup.geometry(f"+{x}+{y}")

        # Configuration de la popup
        self.level_popup.attributes("-topmost", True)  # Toujours au dessus
        self.level_popup.focus()  # Donner le focus

        # Frame principal pour les boutons
        main_frame = ctk.CTkFrame(self.level_popup, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=8, pady=8)

        # Créer 3 lignes de boutons
        niveaux = self.NIVEAUX + [" "]  # Ajouter un bouton vide
        boutons_par_ligne = 5

        for ligne in range(3):
            # Frame pour chaque ligne
            ligne_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
            ligne_frame.pack(fill="x", pady=1)

            # Boutons de cette ligne
            debut = ligne * boutons_par_ligne
            fin = min(debut + boutons_par_ligne, len(niveaux))

            for i in range(debut, fin):
                niveau = niveaux[i]

                if niveau == " ":  # Bouton vide pour effacer
                    color = "#f8f9fa"
                    text = " "
                else:
                    color = self.map_colors[niveau]
                    text = niveau

                btn = ctk.CTkButton(
                    ligne_frame,
                    text=text,
                    font=("Segoe UI", 11, "bold"),
                    height=30,
                    width=70,
                    fg_color=color,
                    text_color="black",  # Toujours en noir
                    hover_color="#ecf0f1",
                    command=lambda n=niveau if niveau != " " else None: self.select_level(item, n)
                )
                btn.pack(side="left", padx=1, pady=1)

        # Gérer la fermeture de la fenêtre
        self.level_popup.protocol("WM_DELETE_WINDOW", lambda: self.close_level_popup())


    def show_save_notification(self, callback):
        """Affiche une fenêtre de notification de sauvegarde moderne."""
        # Créer la fenêtre moderne
        save_popup = ctk.CTkToplevel(self)
        save_popup.title("")
        save_popup.geometry("280x100")
        save_popup.resizable(False, False)

        # Centrer la fenêtre
        save_popup.geometry("+{}+{}".format(
            self.winfo_rootx() + self.winfo_width()//2 - 140,
            self.winfo_rooty() + self.winfo_height()//2 - 50
        ))

        # Configuration moderne
        save_popup.attributes("-topmost", True)
        save_popup.overrideredirect(True)  # Pas de barre de titre

        # Frame principal avec coins arrondis
        main_frame = ctk.CTkFrame(
            save_popup,
            fg_color="#ffffff",
            bg_color="transparent",
            corner_radius=15,
            border_width=2,
            border_color="#e1e5e9"
        )
        main_frame.pack(fill="both", expand=True, padx=2, pady=2)

        # Frame intérieur pour le contenu
        content_frame = ctk.CTkFrame(
            main_frame,
            fg_color="#f8fafc",
            corner_radius=10
        )
        content_frame.pack(fill="both", expand=True, padx=8, pady=8)

        # Icône et texte
        icon_label = ctk.CTkLabel(
            content_frame,
            text="⏳",
            font=("Segoe UI", 24),
            text_color="#3b82f6"
        )
        icon_label.pack(pady=(10, 5))

        text_label = ctk.CTkLabel(
            content_frame,
            text="Sauvegarde en cours...",
            font=("Segoe UI", 13, "bold"),
            text_color="#1e293b"
        )
        text_label.pack(pady=(0, 10))

        # Programmer la disparition après 200ms et l'exécution du callback
        def complete_with_notification():
            save_popup.destroy()
            callback()

        self.after(500, complete_with_notification)

        return save_popup


    def create_filter_section(self, parent, title, filter_dict, colors=None):
        """Crée une section de filtres avec des boutons toggle stylisés."""
        # Conteneur de section
        section_frame = ctk.CTkFrame(parent, fg_color="white", corner_radius=10)
        section_frame.pack(fill="x", pady=(0, 8))

        # En-tête de section
        header = ctk.CTkFrame(section_frame, fg_color="#ecf0f1", corner_radius=8)
        header.pack(fill="x", padx=2, pady=(2, 8))

        ctk.CTkLabel(header, text=title, font=("Segoe UI", 12, "bold"), text_color="#2c3e50").pack(pady=6)

        # Conteneur pour les boutons
        buttons_frame = ctk.CTkFrame(section_frame, fg_color="transparent")
        buttons_frame.pack(fill="x", padx=8, pady=(0, 8))

        # Boutons toggle stylisés
        for i, (label, var) in enumerate(filter_dict.items()):
            color = colors[i % len(colors)] if colors else "#95a5a6"

            # Frame pour chaque bouton avec espacement
            btn_container = ctk.CTkFrame(buttons_frame, fg_color="transparent")
            btn_container.pack(fill="x", pady=1)

            # Bouton toggle personnalisé
            toggle_btn = ctk.CTkButton(
                btn_container,
                text=label,
                font=("Segoe UI", 10, "bold"),
                height=32,
                fg_color="#ffffff",
                text_color="#1a365d",
                hover_color="#f8fafc",
                border_width=1, border_color="#cbd5e1",
                corner_radius=6,
                command=lambda v=var, l=label, c=color, b=None: self.toggle_filter(v, l, c, b)
            )
            toggle_btn.pack(fill="x", padx=4, pady=1)
            toggle_btn.filter_var = var
            toggle_btn.original_color = "#66BB6A"  # Couleur verte douce pour l'état actif

            # Stocker référence pour pouvoir la modifier plus tard
            if not hasattr(self, 'filter_buttons'):
                self.filter_buttons = {}
            self.filter_buttons[label] = toggle_btn

            # Appliquer l'état initial
            self.update_toggle_button(toggle_btn, var.get())

    def create_levels_filter_section(self, parent):
        """Crée la section des niveaux avec leurs couleurs respectives."""
        section_frame = ctk.CTkFrame(parent, fg_color="white", corner_radius=10)
        section_frame.pack(fill="x", pady=(0, 8))

        # En-tête de section
        header = ctk.CTkFrame(section_frame, fg_color="#ecf0f1", corner_radius=8)
        header.pack(fill="x", padx=2, pady=(2, 8))

        ctk.CTkLabel(header, text="🎓 NIVEAUX", font=("Segoe UI", 12, "bold"), text_color="#2c3e50").pack(pady=6)

        # Conteneur pour les boutons (2 colonnes)
        buttons_frame = ctk.CTkFrame(section_frame, fg_color="transparent")
        buttons_frame.pack(fill="x", padx=8, pady=(0, 8))

        # Organiser en grille 2 colonnes
        total_levels = len(self.filter_levels)
        for i, (niv, var) in enumerate(self.filter_levels.items()):
            row = i // 2
            col = i % 2

            if col == 0:
                # Nouveau conteneur de ligne
                row_frame = ctk.CTkFrame(buttons_frame, fg_color="transparent")
                row_frame.pack(fill="x", pady=1)
                # Configurer la grille pour cette ligne
                row_frame.grid_columnconfigure(0, weight=1)
                row_frame.grid_columnconfigure(1, weight=1)

            toggle_btn = ctk.CTkButton(
                row_frame,
                text=niv,
                font=("Segoe UI", 9, "bold"),
                height=30,
                fg_color="#ffffff",  # Blanc pour meilleur contraste
                text_color="#1a365d",
                hover_color="#f8fafc",
                border_width=1, border_color="#cbd5e1",
                corner_radius=5,
                command=lambda v=var, l=niv, c="#66BB6A", b=None: self.toggle_filter(v, l, c, b)  # Vert doux pour actif
            )
            toggle_btn.grid(row=0, column=col, padx=2, pady=1, sticky="ew")
            toggle_btn.filter_var = var
            toggle_btn.original_color = "#66BB6A"  # Couleur verte douce pour l'état actif

            # Stocker référence
            if not hasattr(self, 'filter_buttons'):
                self.filter_buttons = {}
            self.filter_buttons[niv] = toggle_btn

            # Appliquer l'état initial
            self.update_toggle_button(toggle_btn, var.get())

        # Ajouter le bouton de filtre "sans niveau" sur la même ligne que Pitchoune
        self.add_no_level_filter_button_same_row(buttons_frame)

    def add_no_level_filter_button(self, parent):
        """Ajoute le bouton de filtre 'sans niveau'."""
        # Créer une nouvelle ligne pour le bouton "sans niveau"
        no_level_frame = ctk.CTkFrame(parent, fg_color="transparent")
        no_level_frame.pack(fill="x", pady=1)
        no_level_frame.grid_columnconfigure(0, weight=1)
        no_level_frame.grid_columnconfigure(1, weight=1)

        # Bouton "Sans niveau" qui prend toute la largeur
        no_level_btn = ctk.CTkButton(
            no_level_frame,
            text="🚫 SANS NIVEAU",
            font=("Segoe UI", 10, "bold"),
            height=30,
            fg_color="#ffffff",
            text_color="#1a365d",
            hover_color="#f8fafc",
            border_width=1, border_color="#cbd5e1",
            corner_radius=5,
            command=lambda: self.toggle_no_level_filter()
        )
        no_level_btn.grid(row=0, column=0, columnspan=2, padx=2, pady=1, sticky="ew")

        # Stocker la référence du bouton
        self.no_level_button = no_level_btn
        self.update_no_level_button()

    def add_no_level_filter_button_same_row(self, parent):
        """Ajoute le bouton de filtre 'sans niveau' sur la même ligne que Pitchoune."""
        # Compter le nombre de niveaux
        num_levels = len(self.filter_levels)

        # Si le nombre de niveaux est impair, Pitchoune est seul sur sa ligne
        # On peut donc ajouter "sans niveau" à côté
        if num_levels % 2 == 1:
            # Trouver tous les row_frames
            row_frames = []
            for child in parent.winfo_children():
                if isinstance(child, ctk.CTkFrame) and child.cget('fg_color') == 'transparent':
                    row_frames.append(child)

            # Le dernier row_frame contient Pitchoune
            if row_frames:
                last_row_frame = row_frames[-1]

                # Ajouter le bouton "sans niveau" dans la colonne 1
                no_level_btn = ctk.CTkButton(
                    last_row_frame,
                    text="🚫 SANS NIVEAU",
                    font=("Segoe UI", 9, "bold"),  # Même taille de police que les autres
                    height=30,
                    fg_color="#ffffff",
                    text_color="#1a365d",
                    hover_color="#f8fafc",
                    border_width=1, border_color="#cbd5e1",
                    corner_radius=5,
                    command=lambda: self.toggle_no_level_filter()
                )
                no_level_btn.grid(row=0, column=1, padx=2, pady=1, sticky="ew")

                # Stocker la référence du bouton
                self.no_level_button = no_level_btn
                self.update_no_level_button()
        else:
            # Si le nombre de niveaux est pair, créer une nouvelle ligne
            self.add_no_level_filter_button(parent)

    def toggle_no_level_filter(self):
        """Bascule l'état du filtre 'sans niveau'."""
        current_state = self.filter_no_level.get()
        new_state = not current_state
        self.filter_no_level.set(new_state)
        self.update_no_level_button()
        self.refresh_table(preserve_selection=False)

    def update_no_level_button(self):
        """Met à jour l'apparence du bouton 'sans niveau'."""
        is_active = self.filter_no_level.get()

        if is_active:
            self.no_level_button.configure(
                fg_color="#e74c3c",  # Rouge pour indiquer actif
                text_color="white",
                hover_color="#ef5350",  # Rouge plus clair au survol
                border_color="#c0392b"
            )
        else:
            self.no_level_button.configure(
                fg_color="#ffffff",
                text_color="#1a365d",
                hover_color="#f8fafc",
                border_color="#cbd5e1"
            )

    def toggle_filter(self, var, label, color, button=None):
        """Alterne l'état d'un filtre et met à jour l'interface."""
        current_state = var.get()
        new_state = not current_state
        var.set(new_state)

        # Mettre à jour le bouton correspondant
        if hasattr(self, 'filter_buttons') and label in self.filter_buttons:
            self.update_toggle_button(self.filter_buttons[label], new_state)
        elif button:
            self.update_toggle_button(button, new_state)

        # Rafraîchir le tableau
        self.refresh_table(preserve_selection=False)

    def update_toggle_button(self, button, is_active):
        """Met à jour l'apparence d'un bouton toggle."""
        base_text = button.cget('text')

        if is_active:
            # État actif : fond vert avec bordure verte
            button.configure(
                fg_color="#66BB6A",  # Vert doux pour indiquer actif
                text_color="white",
                hover_color="#81C784",  # Vert doux plus clair au hover
                border_width=1, border_color="#4CAF50"
            )
        else:
            # État inactif : blanc avec texte bleu marine et bordure grise
            button.configure(
                fg_color="#ffffff",  # Blanc pour meilleur contraste
                text_color="#1a365d",
                hover_color="#f8fafc",
                border_width=1, border_color="#cbd5e1"
            )

    def adjust_color_brightness(self, color, amount):
        """Ajuste la luminosité d'une couleur hex."""
        # Convertir hex vers RGB
        color = color.lstrip('#')
        rgb = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))

        # Ajuster la luminosité
        new_rgb = []
        for c in rgb:
            new_c = max(0, min(255, c + amount))
            new_rgb.append(new_c)

        # Reconvertir vers hex
        return f"#{new_rgb[0]:02x}{new_rgb[1]:02x}{new_rgb[2]:02x}"

    def is_dark_color(self, color):
        """Détermine si une couleur est sombre."""
        color = color.lstrip('#')
        rgb = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
        # Calcul de luminosité (formule standard)
        brightness = (rgb[0] * 299 + rgb[1] * 587 + rgb[2] * 114) / 1000
        return brightness < 128

    def reset_filters(self):
        """Remet à zéro tous les filtres."""
        for filters in [self.filter_levels, self.filter_ages, self.filter_ci]:
            for var in filters.values():
                var.set(False)

        # Remettre à zéro le filtre "sans niveau"
        self.filter_no_level.set(False)

        # Remettre à jour tous les boutons
        if hasattr(self, 'filter_buttons'):
            for button in self.filter_buttons.values():
                self.update_toggle_button(button, False)

        # Remettre à jour le bouton "sans niveau"
        self.update_no_level_button()

        self.refresh_table(preserve_selection=False)

    def refresh_table(self, preserve_selection=False):
        # Sauvegarder la sélection si demandé
        selected_names = []
        if preserve_selection:
            selected_items = self.tree.selection()
            for item in selected_items:
                nom = self.tree.item(item)['values'][0]
                selected_names.append(nom)

        for item in self.tree.get_children(): self.tree.delete(item)

        # Filtres actifs
        active_levels = [n for n, v in self.filter_levels.items() if v.get()]
        active_ages = [self.AGES[label] for label, v in self.filter_ages.items() if v.get()]
        ci_avec = self.filter_ci["Avec CI"].get()
        ci_sans = self.filter_ci["Sans CI"].get()
        no_level_filter = self.filter_no_level.get()
        search_term = self.search_var.get().lower()

        # Trier les élèves par ordre alphabétique
        df_sorted = self.df.sort_values(by=self.cols_map["Stagiaire"])

        for _, row in df_sorted.iterrows():
            classe = str(row[self.cols_map["Classe"]]) if self.cols_map["Classe"] and pd.notna(row[self.cols_map["Classe"]]) else ""
            prof = str(row[self.cols_map["Prof"]]) if self.cols_map["Prof"] and pd.notna(row[self.cols_map["Prof"]]) else ""
            arrivee = self.format_date_jour_mois(row[self.cols_map["Arrivée"]]) if self.cols_map["Arrivée"] and pd.notna(row[self.cols_map["Arrivée"]]) else ""
            depart = self.format_date_jour_mois(row[self.cols_map["Départ"]]) if self.cols_map["Départ"] and pd.notna(row[self.cols_map["Départ"]]) else ""
            # Nouvelle logique : CI = "OUI" si "Cours 2 Du" n'est pas vide
            ci_val = "OUI" if self.cols_map["Départ CI"] and pd.notna(row[self.cols_map["Départ CI"]]) and str(row[self.cols_map["Départ CI"]]).strip() != "" else ""
            classe_ci = str(row[self.cols_map["Classe CI"]]) if self.cols_map["Classe CI"] and pd.notna(row[self.cols_map["Classe CI"]]) else ""
            prof_ci = str(row[self.cols_map["Prof CI"]]) if self.cols_map["Prof CI"] and pd.notna(row[self.cols_map["Prof CI"]]) else ""
            arrivee_ci = self.format_date_jour_mois(row[self.cols_map["Arrivée CI"]]) if self.cols_map["Arrivée CI"] and pd.notna(row[self.cols_map["Arrivée CI"]]) else ""
            depart_ci = self.format_date_jour_mois(row[self.cols_map["Départ CI"]]) if self.cols_map["Départ CI"] and pd.notna(row[self.cols_map["Départ CI"]]) else ""
            nom = str(row[self.cols_map["Stagiaire"]])
            niv = str(row[self.cols_map["Niveau"]]) if pd.notna(row[self.cols_map["Niveau"]]) else ""
            age = row[self.cols_map["Âge"]]

            # Logique de filtrage
            if search_term not in nom.lower(): continue
            if active_levels and niv not in active_levels: continue
            if active_ages:
                if not pd.notna(age) or not any(low <= age <= high for low, high in active_ages): continue
            if (ci_avec or ci_sans) and not ((ci_avec and ci_val == "OUI") or (ci_sans and (ci_val == "NON" or ci_val == ""))): continue
            if no_level_filter and niv: continue  # Si filtre "sans niveau" actif, exclure les élèves qui ont un niveau

            self.tree.insert("", "end", values=(nom, niv, age, classe, prof, arrivee, depart, "", ci_val, classe_ci, prof_ci, depart_ci, arrivee_ci), tags=(niv,))

        self.update_counters()

        # Restaurer la sélection si demandé
        if preserve_selection and selected_names:
            def restore_selection():
                items_to_select = []
                for item in self.tree.get_children():
                    item_name = self.tree.item(item)['values'][0]
                    if item_name in selected_names:
                        items_to_select.append(item)

                if items_to_select:
                    self.tree.selection_set(items_to_select)

            # Différer légèrement la restauration
            self.after(10, restore_selection)

    def assign_level(self, level):
        selected = self.tree.selection()
        if not selected: return

        # Fonction de callback pour la sauvegarde
        def complete_assignment():
            for item in selected:
                nom = self.tree.item(item)['values'][0]
                self.df.loc[self.df[self.cols_map["Stagiaire"]] == nom, self.cols_map["Niveau"]] = level
            self.df.to_excel(self.file_path, index=False)

            # Rafraîchir le tableau en préservant la sélection
            self.refresh_table(preserve_selection=True)

        # Afficher la notification moderne
        self.show_save_notification(complete_assignment)

    def update_counters(self):
        total = len(self.df)
        done = self.df[self.cols_map["Niveau"]].replace("", None).dropna().count()
        unassigned = total - done

        # Calcul du nombre d'élèves selon les filtres actifs
        filtered_count = self.get_filtered_count()

        # Calcul du nombre d'élèves sélectionnés
        selected_count = len(self.tree.selection()) if hasattr(self, 'tree') else 0

        self.stats_total_label.configure(text=f"👤 {total}")
        self.stats_assigned_label.configure(text=f"✅ {done}")
        self.stats_unassigned_label.configure(text=f"❌ {unassigned}")
        self.stats_filtered_label.configure(text=f"🔍 {filtered_count}")
        self.stats_selected_label.configure(text=f"🖱️ {selected_count}")

    def get_filtered_count(self):
        """Calcule le nombre d'élèves selon les filtres actifs."""
        # Récupérer les filtres actifs (même logique que refresh_table)
        active_levels = [n for n, v in self.filter_levels.items() if v.get()]
        active_ages = [self.AGES[label] for label, v in self.filter_ages.items() if v.get()]
        ci_avec = self.filter_ci["Avec CI"].get()
        ci_sans = self.filter_ci["Sans CI"].get()
        no_level_filter = self.filter_no_level.get()
        search_term = self.search_var.get().lower()

        # Trier les élèves par ordre alphabétique
        df_sorted = self.df.sort_values(by=self.cols_map["Stagiaire"])

        count = 0
        for _, row in df_sorted.iterrows():
            classe = str(row[self.cols_map["Classe"]]) if self.cols_map["Classe"] and pd.notna(row[self.cols_map["Classe"]]) else ""
            prof = str(row[self.cols_map["Prof"]]) if self.cols_map["Prof"] and pd.notna(row[self.cols_map["Prof"]]) else ""
            arrivee = self.format_date_jour_mois(row[self.cols_map["Arrivée"]]) if self.cols_map["Arrivée"] and pd.notna(row[self.cols_map["Arrivée"]]) else ""
            depart = self.format_date_jour_mois(row[self.cols_map["Départ"]]) if self.cols_map["Départ"] and pd.notna(row[self.cols_map["Départ"]]) else ""
            # Nouvelle logique : CI = "OUI" si "Cours 2 Du" n'est pas vide
            ci_val = "OUI" if self.cols_map["Départ CI"] and pd.notna(row[self.cols_map["Départ CI"]]) and str(row[self.cols_map["Départ CI"]]).strip() != "" else ""
            classe_ci = str(row[self.cols_map["Classe CI"]]) if self.cols_map["Classe CI"] and pd.notna(row[self.cols_map["Classe CI"]]) else ""
            prof_ci = str(row[self.cols_map["Prof CI"]]) if self.cols_map["Prof CI"] and pd.notna(row[self.cols_map["Prof CI"]]) else ""
            arrivee_ci = self.format_date_jour_mois(row[self.cols_map["Arrivée CI"]]) if self.cols_map["Arrivée CI"] and pd.notna(row[self.cols_map["Arrivée CI"]]) else ""
            depart_ci = self.format_date_jour_mois(row[self.cols_map["Départ CI"]]) if self.cols_map["Départ CI"] and pd.notna(row[self.cols_map["Départ CI"]]) else ""
            nom = str(row[self.cols_map["Stagiaire"]])
            niv = str(row[self.cols_map["Niveau"]]) if pd.notna(row[self.cols_map["Niveau"]]) else ""
            age = row[self.cols_map["Âge"]]

            # Appliquer les mêmes filtres que dans refresh_table
            if search_term not in nom.lower(): continue
            if active_levels and niv not in active_levels: continue
            if active_ages:
                if not pd.notna(age) or not any(low <= age <= high for low, high in active_ages): continue
            if (ci_avec or ci_sans) and not ((ci_avec and ci_val == "OUI") or (ci_sans and (ci_val == "NON" or ci_val == ""))): continue
            if no_level_filter and niv: continue  # Si filtre "sans niveau" actif, exclure les élèves qui ont un niveau

            count += 1

        return count

    def start_matrix_watch(self):
        """Démarre la surveillance du fichier matrix."""
        self.check_matrix_modifications()

    def check_matrix_modifications(self):
        """Vérifie périodiquement si le fichier matrix a été modifié et rafraîchit l'affichage."""
        if not os.path.exists(self.file_path):
            # Fichier n'existe plus, arrêter la surveillance
            return

        try:
            # Obtenir le timestamp de dernière modification
            current_mtime = os.path.getmtime(self.file_path)

            # Vérifier si le fichier a été modifié
            if self.matrix_last_modified is not None:
                if current_mtime > self.matrix_last_modified:
                    # Le fichier a été modifié, rafraîchir l'affichage
                    print(f"✅ Détection d'une modification du fichier matrix.xlsx - Rafraîchissement automatique en cours...")
                    self.matrix_last_modified = current_mtime

                    # Recharger les données depuis le fichier
                    try:
                        self.df = self.safe_read_excel(self.file_path)
                        # Re-normaliser les colonnes au cas où la structure a changé
                        self.cols_map = {
                            "Ecole": self.find_column(["ecole", "école", "school"]),
                            # "Horaire": self.find_column(["horaire", "horaire", "time", "schedule"]),  # Désactivé
                            "Classe": self.find_column(["classe", "class", "groupe"]),
                            "Stagiaire": self.find_column(["stagiaire", "nom", "élève"]),
                            "Âge": self.find_column(["âge", "age"]),
                            "Cours 2": self.find_column(["cours 2", "intensif", "ci"]),
                            "Niveau": self.find_column(["niveau"]),
                            "Classe CI": self.find_column(["classe ci", "classe_ci"]),
                            "Prof": self.find_column(["prof", "professeur", "enseignant"]),
                            "Prof CI": self.find_column(["prof ci", "prof_ci", "professeur ci"]),
                            "Départ": self.find_column(["cours 1 du", "départ", "depart"]),
                            "Arrivée": self.find_column(["cours 1 au", "arrivée", "arrivee"]),
                            "Départ CI": self.find_column(["cours 2 du", "départ ci", "depart ci"]),
                            "Arrivée CI": self.find_column(["cours 2 au", "arrivée ci", "arrivee ci"])
                        }
                        
                        # Rafraîchir le tableau et les compteurs
                        self.refresh_table(preserve_selection=False)
                        self.update_counters()
                        
                        print(f"✅ Rafraîchissement terminé - Tableau mis à jour avec les nouvelles assignations")
                    except Exception as e:
                        print(f"❌ Erreur lors du rafraîchissement: {e}")
            else:
                # Première fois qu'on surveille ce fichier
                self.matrix_last_modified = current_mtime
        except Exception as e:
            print(f"❌ Erreur lors de la vérification du fichier matrix: {e}")

        # Programmer la prochaine vérification dans 1 seconde
        self.matrix_watch_job = self.after(1000, self.check_matrix_modifications)

    def stop_matrix_watch(self):
        """Arrête la surveillance du fichier matrix."""
        if self.matrix_watch_job:
            self.after_cancel(self.matrix_watch_job)
            self.matrix_watch_job = None

    def on_quit(self):
        """Gère la fermeture propre de l'application."""
        self.stop_matrix_watch()
        self.quit()

    def show_classes_context_menu(self, event):
        """Affiche un menu contextuel avec les classes des écoles filtrées."""
        # Fermer tout menu contextuel existant
        if hasattr(self, 'current_context_menu') and self.current_context_menu:
            try:
                if self.current_context_menu.winfo_exists():
                    self.current_context_menu.destroy()
            except:
                pass
            self.current_context_menu = None

        # Récupérer les élèves sélectionnés
        selected_items = self.tree.selection()
        if not selected_items:
            return

        # Récupérer tous les noms des élèves sélectionnés
        student_names = []
        for item in selected_items:
            student_name = self.tree.item(item)['values'][0]  # Le nom est dans la première colonne
            student_names.append(student_name)

        # Récupérer les coordonnées de la souris
        x = event.x_root
        y = 10  # Position près du haut de l'écran

        # Calculer la hauteur disponible (presque tout l'écran)
        screen_height = self.winfo_screenheight()
        available_height = screen_height - 30  # Laisser un petit margin en haut et en bas

        # Créer le menu contextuel
        menu = ctk.CTkToplevel(self)
        menu.title("")
        menu.geometry(f"700x{available_height}")
        menu.geometry(f"+{x}+{y}")
        menu.resizable(True, True)
        menu.transient(self)
        menu.attributes("-topmost", True)
        menu.overrideredirect(True)  # Pas de barre de titre

        # Stocker la référence du menu actuel
        self.current_context_menu = menu

        # Variable pour suivre si le menu est détruit
        menu_destroyed = False

        def cleanup_bindings():
            """Nettoie les bindings après fermeture."""
            try:
                if hasattr(menu, '_click_handler_id'):
                    self.unbind("<Button-1>", menu._click_handler_id)
            except:
                pass

        def close_menu():
            """Ferme le menu proprement."""
            nonlocal menu_destroyed
            if not menu_destroyed and menu.winfo_exists():
                menu_destroyed = True
                cleanup_bindings()
                try:
                    menu.destroy()
                except:
                    pass
                # Réinitialiser la référence du menu actuel
                self.current_context_menu = None

        # Frame principal avec scroll
        main_frame = ctk.CTkFrame(menu, fg_color="white", corner_radius=10)
        main_frame.pack(fill="both", expand=True, padx=2, pady=2)

        # En-tête avec titre et bouton fermer
        header_frame = ctk.CTkFrame(main_frame, fg_color="#f0f9ff", corner_radius=8)
        header_frame.pack(fill="x", padx=8, pady=(8, 4))

        title_label = ctk.CTkLabel(
            header_frame,
            text=f"📚 Assigner {len(student_names)} élève(s) à une classe",
            font=("Segoe UI", 13, "bold"),
            text_color="#1e293b"
        )
        title_label.pack(side="left", padx=12, pady=8)

        # Bouton fermer dans l'en-tête
        close_header_btn = ctk.CTkButton(
            header_frame,
            text="✕",
            width=30,
            height=30,
            font=("Segoe UI", 14, "bold"),
            fg_color="#ef4444",
            hover_color="#dc2626",
            text_color="white",
            corner_radius=15,
            command=close_menu
        )
        close_header_btn.pack(side="right", padx=8, pady=4)

        # Frame avec scrollbar
        scrollable = ctk.CTkScrollableFrame(main_frame, fg_color="#f8fafc")
        scrollable.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        # Récupérer les données des écoles filtrées
        school_data = self.get_filtered_school_data()

        if not school_data:
            no_data_label = ctk.CTkLabel(
                scrollable,
                text="Aucune donnée d'école trouvée pour cette semaine.",
                font=("Segoe UI", 12),
                text_color="#6b7280"
            )
            no_data_label.pack(pady=30)
        else:
            self.display_school_data_in_menu(scrollable, school_data, student_names, menu, close_menu)

        # Frame pour les boutons d'action en bas
        action_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        action_frame.pack(fill="x", padx=8, pady=(0, 8))

        # Bouton pour dé-assigner les élèves
        unassign_btn = ctk.CTkButton(
            action_frame,
            text="🚫 Retirer l'assignation",
            font=("Segoe UI", 11, "bold"),
            height=35,
            fg_color="#ef4444",
            hover_color="#dc2626",
            text_color="white",
            command=lambda: self.unassign_students_from_class(student_names, close_menu)
        )
        unassign_btn.pack(side="left", padx=(0, 8))

        # Bouton fermer en bas
        close_bottom_btn = ctk.CTkButton(
            action_frame,
            text="Fermer",
            font=("Segoe UI", 11, "bold"),
            height=35,
            width=100,
            fg_color="#6b7280",
            hover_color="#4b5563",
            text_color="white",
            command=close_menu
        )
        close_bottom_btn.pack(side="right")

        # Détection de clic extérieur améliorée
        def on_click_outside(event):
            """Détecte les clics en dehors du menu."""
            if not menu_destroyed and menu.winfo_exists():
                try:
                    # Vérifier si le clic est en dehors du menu
                    menu_x = menu.winfo_x()
                    menu_y = menu.winfo_y()
                    menu_width = menu.winfo_width()
                    menu_height = menu.winfo_height()
                    
                    click_x = event.x_root
                    click_y = event.y_root
                    
                    if not (menu_x <= click_x <= menu_x + menu_width and
                           menu_y <= click_y <= menu_y + menu_height):
                        close_menu()
                except:
                    pass

        # Lier les événements de clic sur la fenêtre principale
        click_handler_id = self.bind("<Button-1>", on_click_outside, add="+")
        
        # Stocker l'ID pour cleanup
        menu._click_handler_id = click_handler_id

        # Gérer la fermeture avec FocusOut (avec délai pour éviter les fermetures accidentelles)
        def on_focus_out(event):
            if not menu_destroyed:
                menu.after(150, lambda: close_menu() if not menu_destroyed and menu.winfo_exists() else None)

        menu.bind("<FocusOut>", on_focus_out)

        # Gérer la fermeture avec Escape
        def on_key_press(event):
            if event.keysym == 'Escape' and not menu_destroyed:
                close_menu()

        menu.bind("<Key>", on_key_press)
        menu.focus_set()

        # Stocker la référence du menu pour cleanup
        if not hasattr(self, 'context_menus'):
            self.context_menus = []
        self.context_menus.append(menu)

    def get_filtered_school_data(self):
        """Récupère les données des écoles filtrées selon les préférences utilisateur."""
        try:
            # Charger les préférences utilisateur
            preferences_file = os.path.join(os.path.dirname(__file__), "user_preferences.json")
            if os.path.exists(preferences_file):
                with open(preferences_file, 'r', encoding='utf-8') as f:
                    preferences = json.load(f)
            else:
                preferences = {}

            # Récupérer les filtres actifs
            school_filters = preferences.get("school_filters", {})
            active_schools = [school for school, active in school_filters.items() if active]

            if not active_schools:
                # Si aucun filtre n'est défini, activer toutes les écoles par défaut
                active_schools = ["A", "B", "C/CS", "C/CI", "Morning", "Premium/CS", "Premium/CI"]

            # Analyser les données des écoles pour cette semaine
            week_folder = os.path.dirname(self.file_path)
            school_data = self.analyze_school_classes(week_folder)

            # Filtrer selon les écoles actives
            filtered_data = {}
            school_mapping = {
                'ecole_a': 'A',
                'ecole_b': 'B',
                'ecole_c_cs': 'C/CS',
                'ecole_c_ci': 'C/CI',
                'ecole_morning': 'Morning',
                'ecole_premium_cs': 'Premium/CS',
                'ecole_premium_ci': 'Premium/CI'
            }

            for school_key, school_info in school_data.items():
                display_name = school_mapping.get(school_key, school_key)
                if display_name in active_schools:
                    filtered_data[school_key] = school_info

            return filtered_data

        except Exception as e:
            print(f"Erreur lors de la récupération des données d'écoles: {e}")
            return {}

    def analyze_school_classes(self, week_folder):
        """Analyse les fichiers Excel d'écoles pour une semaine donnée."""
        result = {
            'ecole_a': [],
            'ecole_b': [],
            'ecole_c_cs': [],
            'ecole_c_ci': [],
            'ecole_morning': [],
            'ecole_premium_cs': [],
            'ecole_premium_ci': []
        }

        # Mapping des fichiers Excel vers les clés d'écoles
        file_to_school_mapping = {
            'ecole_a.xlsx': 'ecole_a',
            'ecole_b.xlsx': 'ecole_b',
            'ECOLE_C_cours_standard.xlsx': 'ecole_c_cs',
            'ECOLE_C_cours_intensif.xlsx': 'ecole_c_ci',
            'MORNING.xlsx': 'ecole_morning',
            'ECOLE_PREMIUM_cours_standard.xlsx': 'ecole_premium_cs',
            'ECOLE_PREMIUM_cours_intensifs.xlsx': 'ecole_premium_ci'
        }

        # Fonction helper pour analyser un fichier Excel d'école
        def analyze_school_file(excel_path, school_key):
            if not os.path.exists(excel_path):
                return

            try:
                # Lire toutes les feuilles du fichier
                xl_file = pd.ExcelFile(excel_path)
                sheet_names = xl_file.sheet_names

                for sheet_name in sheet_names:
                    try:
                        df = self.safe_read_excel(excel_path, sheet_name=sheet_name)

                        if df.empty:
                            # La feuille est vide mais on garde l'horaire
                            sheet_lower = sheet_name.lower()
                            type_intervenant = "animateur" if "animateur" in sheet_lower else "professeur"
                            horaire = self.clean_horaire_name(sheet_name)

                            result[school_key].append({
                                'horaire': horaire or sheet_name,
                                'intervenant': "",
                                'type_intervenant': type_intervenant,
                                'classes': []
                            })
                            continue

                        # Analyser et nettoyer le nom de la feuille
                        sheet_lower = sheet_name.lower()
                        type_intervenant = "animateur" if "animateur" in sheet_lower else "professeur"
                        horaire = self.clean_horaire_name(sheet_name)

                        # Chercher les colonnes de classes
                        classe_cols = []
                        for col in df.columns:
                            col_lower = str(col).lower()
                            if any(keyword in col_lower for keyword in ['classe', 'groupe', 'section']):
                                classe_cols.append(col)

                        classes_info = []
                        for _, row in df.iterrows():
                            classe_nom = None
                            for col in classe_cols:
                                val = str(row.get(col, '')).strip()
                                if val and val.lower() not in ['', 'nan', 'none']:
                                    classe_nom = val
                                    break

                            if classe_nom:
                                classes_info.append({
                                    'nom_classe': classe_nom,
                                    'nb_eleves': 0,  # On ne calcule pas ici pour simplifier
                                    'niveau': '',
                                    'eleves': []
                                })

                        result[school_key].append({
                            'horaire': horaire or sheet_name,
                            'intervenant': "",
                            'type_intervenant': type_intervenant,
                            'classes': classes_info
                        })

                    except Exception as e:
                        print(f"Erreur lors de l'analyse de la feuille '{sheet_name}': {e}")
                        continue

            except Exception as e:
                print(f"Erreur lors de l'analyse du fichier {excel_path}: {e}")

        # Analyser tous les fichiers Excel du dossier semaine
        if os.path.exists(week_folder):
            for filename in os.listdir(week_folder):
                if filename.lower().endswith('.xlsx') and filename.lower() != 'matrix.xlsx':
                    file_path = os.path.join(week_folder, filename)
                    school_key = file_to_school_mapping.get(filename)
                    if school_key:
                        analyze_school_file(file_path, school_key)

        return result

    def display_school_data_in_menu(self, parent_frame, school_data, student_names, menu, close_menu_func):
        """Affiche les données des écoles dans le menu contextuel avec la structure École > Horaire > Classes cards."""
        # Mapping pour les noms d'affichage et couleurs
        school_display = {
            'ecole_a': ('A', '#3b82f6'),
            'ecole_b': ('B', '#10b981'),
            'ecole_c_cs': ('C/CS', '#f59e0b'),
            'ecole_c_ci': ('C/CI', '#8b5cf6'),
            'ecole_morning': ('Morning', '#ef4444'),
            'ecole_premium_cs': ('Premium/CS', '#06b6d4'),
            'ecole_premium_ci': ('Premium/CI', '#f97316')
        }

        for school_key, horaires in school_data.items():
            if not horaires:  # Si pas d'horaires pour cette école
                continue

            display_name, school_color = school_display.get(school_key, (school_key, '#6b7280'))

            # Frame pour l'école
            school_frame = ctk.CTkFrame(parent_frame, fg_color="white", corner_radius=8, border_width=1, border_color=school_color)
            school_frame.pack(fill="x", pady=(0, 8), padx=5)

            # En-tête de l'école
            school_header = ctk.CTkFrame(school_frame, fg_color="#f0f9ff", corner_radius=6)
            school_header.pack(fill="x", padx=8, pady=(8, 5))

            school_title = ctk.CTkLabel(
                school_header,
                text=f"🏫 École {display_name}",
                font=("Segoe UI", 12, "bold"),
                text_color=school_color
            )
            school_title.pack(pady=6)

            # Conteneur pour les horaires
            horaires_frame = ctk.CTkFrame(school_frame, fg_color="transparent")
            horaires_frame.pack(fill="x", padx=8, pady=(0, 8))

            # Calculer le nombre d'horaires par ligne (comme dans fenetre_principale.py)
            total_horaires = len(horaires)
            horaires_per_row = min(3, max(2, total_horaires))  # Entre 2 et 3 horaires par ligne

            # Configurer les colonnes dynamiquement
            for col in range(horaires_per_row):
                horaires_frame.grid_columnconfigure(col, weight=1)

            horaire_row = 0
            horaire_col = 0

            for horaire_info in horaires:
                horaire = horaire_info.get('horaire', 'Horaire inconnu')
                intervenant = horaire_info.get('intervenant', '')
                type_intervenant = horaire_info.get('type_intervenant', 'professeur')
                classes = horaire_info.get('classes', [])

                # Frame pour l'horaire
                horaire_frame = ctk.CTkFrame(horaires_frame, fg_color="#f8fafc", corner_radius=6, border_width=1, border_color="#e5e7eb")
                horaire_frame.grid(row=horaire_row, column=horaire_col, sticky="nsew", padx=(0, 4) if horaire_col < horaires_per_row - 1 else 0, pady=(0, 4))

                # En-tête de l'horaire
                horaire_header = ctk.CTkFrame(horaire_frame, fg_color="white", corner_radius=4)
                horaire_header.pack(fill="x", padx=6, pady=(6, 4))

                intervenant_icon = "👨‍🏫" if type_intervenant == "professeur" else "🎭"
                horaire_title = ctk.CTkLabel(
                    horaire_header,
                    text=f"{intervenant_icon} Horaire : {horaire} : {type_intervenant}",
                    font=("Segoe UI", 11, "bold"),
                    text_color="#374151"
                )
                horaire_title.pack(side="left", pady=4)

                # Conteneur pour les classes (cards)
                if classes:
                    classes_frame = ctk.CTkFrame(horaire_frame, fg_color="transparent")
                    classes_frame.pack(fill="x", padx=6, pady=(0, 8))

                    # Configuration pour une grille de classes
                    classes_per_row = 4  # Nombre de classes par ligne
                    for i in range(classes_per_row):
                        classes_frame.grid_columnconfigure(i, weight=1)

                    for idx, classe_info in enumerate(classes):
                        classe_nom = classe_info.get('nom_classe', 'Classe inconnue')
                        row = idx // classes_per_row
                        col = idx % classes_per_row

                        # Petite card pour chaque classe
                        class_card = ctk.CTkFrame(
                            classes_frame,
                            fg_color="white",
                            corner_radius=4,
                            width=80,
                            height=35
                        )
                        class_card.grid(row=row, column=col, padx=2, pady=2, sticky="nsew")
                        class_card.grid_propagate(False)

                        # Label du nom de classe (centré)
                        class_label = ctk.CTkLabel(
                            class_card,
                            text=classe_nom,
                            font=("Segoe UI", 9, "bold"),
                            text_color="#374151"
                        )
                        class_label.place(relx=0.5, rely=0.5, anchor="center")

                        # Rendre la card cliquable
                        class_card.configure(cursor="hand2")

                        # Fonction pour gérer l'effet hover
                        def on_enter(event, frame=class_card):
                            frame.configure(fg_color="#f3f4f6")

                        def on_leave(event, frame=class_card):
                            frame.configure(fg_color="white")

                        # Fonction pour propager les événements hover aux enfants
                        def bind_hover_to_children(widget, enter_func, leave_func):
                            """Applique les événements hover à tous les enfants"""
                            widget.bind("<Enter>", enter_func)
                            widget.bind("<Leave>", leave_func)
                            for child in widget.winfo_children():
                                bind_hover_to_children(child, enter_func, leave_func)

                        # Lier les événements
                        class_card.bind("<Enter>", on_enter)
                        class_card.bind("<Leave>", on_leave)

                        # Propager l'effet hover à tous les enfants
                        bind_hover_to_children(class_card, on_enter, on_leave)

                        # Gestion du clic sur la card
                        def on_class_click(event, s_names=student_names, s_key=school_key, h_info=horaire_info, c_info=classe_info, m=menu, close_func=close_menu_func):
                            self.assign_class_to_student(s_names, s_key, h_info, c_info)
                            # Fermer le menu après assignation
                            m.after(50, close_func)

                        class_card.bind("<Button-1>", on_class_click)
                        class_label.bind("<Button-1>", on_class_click)
                        class_label.configure(cursor="hand2")
                else:
                    no_classes_label = ctk.CTkLabel(
                        horaire_frame,
                        text="Aucune classe définie",
                        font=("Segoe UI", 10, "italic"),
                        text_color="#9ca3af"
                    )
                    no_classes_label.pack(padx=12, pady=(0, 8))

                # Gestion de la grille des horaires
                horaire_col += 1
                if horaire_col >= horaires_per_row:
                    horaire_col = 0
                    horaire_row += 1

    def get_professor_for_class(self, school_key, classe_nom):
        """Récupère le nom du professeur pour une classe donnée dans une école."""
        # Mapping pour les fichiers Excel des écoles
        school_file_mapping = {
            'ecole_a': 'ecole_a.xlsx',
            'ecole_b': 'ecole_b.xlsx',
            'ecole_c_cs': 'ECOLE_C_cours_standard.xlsx',
            'ecole_c_ci': 'ECOLE_C_cours_intensif.xlsx',
            'ecole_morning': 'MORNING.xlsx',
            'ecole_premium_cs': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'ecole_premium_ci': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        excel_filename = school_file_mapping.get(school_key)
        if not excel_filename:
            return ""

        week_folder = os.path.dirname(self.file_path)
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            return ""

        try:
            # Lire toutes les feuilles du fichier
            xl_file = pd.ExcelFile(excel_path)
            sheet_names = xl_file.sheet_names

            for sheet_name in sheet_names:
                try:
                    df_sheet = self.safe_read_excel(excel_path, sheet_name=sheet_name)

                    if df_sheet.empty:
                        continue

                    # Chercher les colonnes de classes et de professeurs
                    classe_cols = []
                    prof_cols = []

                    for col in df_sheet.columns:
                        col_lower = str(col).lower()
                        if any(keyword in col_lower for keyword in ['classe', 'groupe', 'section']):
                            classe_cols.append(col)
                        if any(keyword in col_lower for keyword in ['prof', 'professeur', 'enseignant', 'intervenant']):
                            prof_cols.append(col)

                    # Chercher la classe dans cette feuille
                    for _, row in df_sheet.iterrows():
                        classe_trouvee = None
                        for col in classe_cols:
                            val = str(row.get(col, '')).strip()
                            if val and val.lower() not in ['', 'nan', 'none'] and val == classe_nom:
                                classe_trouvee = val
                                break

                        if classe_trouvee:
                            # Chercher le professeur dans la même ligne
                            for col in prof_cols:
                                prof_val = str(row.get(col, '')).strip()
                                if prof_val and prof_val.lower() not in ['', 'nan', 'none']:
                                    return prof_val

                except Exception as e:
                    print(f"Erreur lors de la lecture de la feuille '{sheet_name}': {e}")
                    continue

        except Exception as e:
            print(f"Erreur lors de la lecture du fichier {excel_filename}: {e}")

        return ""

    def assign_class_to_student(self, student_names, school_key, horaire_info, classe_info):
        """Assigne l'école, l'horaire et la classe à un ou plusieurs élèves."""
        # Mapping pour convertir les clés d'école en noms d'affichage
        school_mapping = {
            'ecole_a': 'A',
            'ecole_b': 'B',
            'ecole_c_cs': 'C/CS',
            'ecole_c_ci': 'C/CI',
            'ecole_morning': 'Morning',
            'ecole_premium_cs': 'Premium/CS',
            'ecole_premium_ci': 'Premium/CI'
        }

        # Mapping pour les fichiers Excel des écoles
        school_file_mapping = {
            'ecole_a': 'ecole_a.xlsx',
            'ecole_b': 'ecole_b.xlsx',
            'ecole_c_cs': 'ECOLE_C_cours_standard.xlsx',
            'ecole_c_ci': 'ECOLE_C_cours_intensif.xlsx',
            'ecole_morning': 'MORNING.xlsx',
            'ecole_premium_cs': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'ecole_premium_ci': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        school_name = school_mapping.get(school_key, school_key)
        horaire = horaire_info.get('horaire', '')
        classe_nom = classe_info.get('nom_classe', '')

        # Récupérer le professeur de cette classe
        professor_name = self.get_professor_for_class(school_key, classe_nom)

        # S'assurer que student_names est une liste
        if isinstance(student_names, str):
            student_names = [student_names]

        # Fonction de callback pour la sauvegarde
        def complete_assignment():
            assigned_count = 0
            old_assignments = {}  # Pour stocker les anciennes assignations

            # Traiter chaque élève sélectionné
            for student_name in student_names:
                if student_name in self.df[self.cols_map["Stagiaire"]].values:
                    mask = self.df[self.cols_map["Stagiaire"]] == student_name
                    
                    # Récupérer les anciennes valeurs AVANT de les modifier
                    old_ecole = self.df.loc[mask, self.cols_map["Ecole"]].values[0] if self.cols_map["Ecole"] else None
                    # old_horaire = self.df.loc[mask, self.cols_map["Horaire"]].values[0] if self.cols_map["Horaire"] else None  # Désactivé
                    old_classe = self.df.loc[mask, self.cols_map["Classe"]].values[0] if self.cols_map["Classe"] else None

                    # Stocker l'ancienne assignation si elle existe
                    if pd.notna(old_ecole) and pd.notna(old_classe) and str(old_ecole).strip() and str(old_classe).strip():
                        old_assignments[student_name] = {
                            'ecole': str(old_ecole).strip(),
                            # 'horaire': str(old_horaire).strip() if pd.notna(old_horaire) else '',  # Désactivé
                            'classe': str(old_classe).strip()
                        }

                    # Assigner les nouvelles valeurs (convertir les colonnes au bon type si nécessaire)
                    if self.cols_map["Ecole"]:
                        col_name = self.cols_map["Ecole"]
                        if self.df[col_name].dtype == 'float64':
                            self.df[col_name] = self.df[col_name].astype('object')
                        self.df.loc[mask, col_name] = school_name
                    # if self.cols_map["Horaire"]:  # Désactivé
                    #     col_name = self.cols_map["Horaire"]
                    #     if self.df[col_name].dtype == 'float64':
                    #         self.df[col_name] = self.df[col_name].astype('object')
                    #     self.df.loc[mask, col_name] = horaire
                    if self.cols_map["Classe"]:
                        col_name = self.cols_map["Classe"]
                        if self.df[col_name].dtype == 'float64':
                            self.df[col_name] = self.df[col_name].astype('object')
                        self.df.loc[mask, col_name] = classe_nom

                    # Mettre à jour le professeur si on l'a trouvé
                    if self.cols_map["Prof"] and professor_name:
                        col_name = self.cols_map["Prof"]
                        if self.df[col_name].dtype == 'float64':
                            self.df[col_name] = self.df[col_name].astype('object')
                        self.df.loc[mask, col_name] = professor_name

                    assigned_count += 1

            # Sauvegarder le fichier matrix
            self.df.to_excel(self.file_path, index=False)
            prof_info = f", prof {professor_name}" if professor_name else ""
            print(f"💾 Assignation sauvegardée : {assigned_count} élève(s) assigné(s) à l'école {school_name}, classe {classe_nom}{prof_info}")

            # Retirer les élèves de leurs anciennes classes (si elles existent)
            if old_assignments:
                self.remove_students_from_old_classes(old_assignments, school_file_mapping)

            # Ajouter les élèves dans la nouvelle classe
            self.update_school_excel_file(school_key, school_file_mapping, horaire, classe_nom, student_names)

            # Rafraîchir l'affichage
            self.refresh_table(preserve_selection=False)
            self.update_counters()

        # Exécuter directement la sauvegarde sans notification popup
        complete_assignment()

    def unassign_students_from_class(self, student_names, close_menu_func):
        """Retire l'assignation (école, horaire, classe) des élèves sélectionnés."""
        # S'assurer que student_names est une liste
        if isinstance(student_names, str):
            student_names = [student_names]

        # Mapping pour les fichiers Excel des écoles
        school_file_mapping = {
            'ecole_a': 'ecole_a.xlsx',
            'ecole_b': 'ecole_b.xlsx',
            'ecole_c_cs': 'ECOLE_C_cours_standard.xlsx',
            'ecole_c_ci': 'ECOLE_C_cours_intensif.xlsx',
            'ecole_morning': 'MORNING.xlsx',
            'ecole_premium_cs': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'ecole_premium_ci': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        # Mapping pour convertir les noms d'affichage en clés d'école
        school_name_to_key = {
            'A': 'ecole_a',
            'B': 'ecole_b',
            'C/CS': 'ecole_c_cs',
            'C/CI': 'ecole_c_ci',
            'Morning': 'ecole_morning',
            'Premium/CS': 'ecole_premium_cs',
            'Premium/CI': 'ecole_premium_ci'
        }

        # Fonction de callback pour la sauvegarde
        def complete_unassignment():
            unassigned_count = 0
            old_assignments = {}  # Pour stocker les anciennes assignations

            # Traiter chaque élève sélectionné
            for student_name in student_names:
                if student_name in self.df[self.cols_map["Stagiaire"]].values:
                    mask = self.df[self.cols_map["Stagiaire"]] == student_name
                    
                    # Récupérer les anciennes valeurs AVANT de les supprimer
                    old_ecole = self.df.loc[mask, self.cols_map["Ecole"]].values[0] if self.cols_map["Ecole"] else None
                    # old_horaire = self.df.loc[mask, self.cols_map["Horaire"]].values[0] if self.cols_map["Horaire"] else None  # Désactivé
                    old_classe = self.df.loc[mask, self.cols_map["Classe"]].values[0] if self.cols_map["Classe"] else None

                    # Stocker l'ancienne assignation si elle existe
                    if pd.notna(old_ecole) and pd.notna(old_classe) and str(old_ecole).strip() and str(old_classe).strip():
                        old_assignments[student_name] = {
                            'ecole': str(old_ecole).strip(),
                            # 'horaire': str(old_horaire).strip() if pd.notna(old_horaire) else '',  # Désactivé
                            'classe': str(old_classe).strip()
                        }

                    # Supprimer les assignations (mettre à None)
                    if self.cols_map["Ecole"]:
                        self.df.loc[mask, self.cols_map["Ecole"]] = None
                    # if self.cols_map["Horaire"]:  # Désactivé
                    #     self.df.loc[mask, self.cols_map["Horaire"]] = None
                    if self.cols_map["Classe"]:
                        self.df.loc[mask, self.cols_map["Classe"]] = None

                    unassigned_count += 1

            # Sauvegarder le fichier matrix
            self.df.to_excel(self.file_path, index=False)
            print(f"🗑️  Dé-assignation sauvegardée : {unassigned_count} élève(s) retiré(s) de leur classe")

            # Retirer les élèves des fichiers Excel des écoles
            if old_assignments:
                self.remove_students_from_old_classes(old_assignments, school_file_mapping)

            # Rafraîchir l'affichage
            self.refresh_table(preserve_selection=False)
            self.update_counters()

            # Fermer le menu
            close_menu_func()       

       
        complete_unassignment()

    def remove_students_from_old_classes(self, old_assignments, school_file_mapping):
        """Retire les élèves de leurs anciennes classes dans les fichiers Excel des écoles."""
        if load_workbook is None:
            return

        # Grouper par école pour traiter fichier par fichier
        files_to_process = {}
        
        # Inverser le mapping pour trouver la clé d'école depuis le nom d'affichage
        school_name_to_key = {
            'A': 'ecole_a',
            'B': 'ecole_b',
            'C/CS': 'ecole_c_cs',
            'C/CI': 'ecole_c_ci',
            'Morning': 'ecole_morning',
            'Premium/CS': 'ecole_premium_cs',
            'Premium/CI': 'ecole_premium_ci'
        }

        for student_name, old_data in old_assignments.items():
            old_ecole = old_data['ecole']
            # old_horaire = old_data.get('horaire', '')  # Plus utilisé
            old_classe = old_data['classe']

            # Trouver la clé d'école
            school_key = school_name_to_key.get(old_ecole)
            if not school_key:
                continue

            excel_filename = school_file_mapping.get(school_key)
            if not excel_filename:
                continue

            if excel_filename not in files_to_process:
                files_to_process[excel_filename] = []

            files_to_process[excel_filename].append({
                'student': student_name,
                'classe': old_classe,
                'school_key': school_key
            })

        # Traiter chaque fichier
        week_folder = os.path.dirname(self.file_path)
        
        for excel_filename, removals in files_to_process.items():
            excel_path = os.path.join(week_folder, excel_filename)
            
            if not os.path.exists(excel_path):
                continue
            
            try:
                wb = self.safe_load_workbook(excel_path)
                
                for removal in removals:
                    student_name = removal['student']
                    old_classe = removal['classe']

                    # Chercher dans toutes les feuilles pour la classe spécifiée
                    found_in_sheet = False
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]

                        # Chercher la ligne correspondant à la classe (colonne 1 : "Nom de la classe")
                        classe_row = None
                        for row_idx in range(2, sheet.max_row + 1):  # Commencer à la ligne 2 (après header)
                            cell_value = str(sheet.cell(row=row_idx, column=1).value or '').strip()
                            if cell_value == old_classe:
                                classe_row = row_idx
                                break

                        if classe_row:
                            # Colonne des élèves (généralement colonne 5)
                            eleves_col = 5
                            current_value = str(sheet.cell(row=classe_row, column=eleves_col).value or '').strip()

                            if current_value and current_value.lower() not in ['', 'nan', 'none', 'liste des élèves...']:
                                # Séparer les élèves
                                eleves_list = [e.strip() for e in current_value.split(',') if e.strip()]

                                # Retirer l'élève
                                if student_name in eleves_list:
                                    eleves_list.remove(student_name)
                                    found_in_sheet = True

                                    # Mettre à jour la cellule des élèves
                                    if eleves_list:
                                        new_value = ', '.join(sorted(eleves_list))
                                        sheet.cell(row=classe_row, column=eleves_col, value=new_value)
                                    else:
                                        # Si la classe devient vide, vider aussi la colonne niveau
                                        sheet.cell(row=classe_row, column=eleves_col, value='')
                                        sheet.cell(row=classe_row, column=4, value='')  # Colonne niveau

                                    print(f"✅ Élève {student_name} retiré de {excel_filename} (feuille: {sheet_name}, classe: {old_classe})")

                    if not found_in_sheet:
                        print(f"ℹ️ Élève {student_name} non trouvé dans {excel_filename} pour la classe {old_classe}")
                
                # Sauvegarder le fichier
                wb.save(excel_path)
                wb.close()
                print(f"✅ Fichier {excel_filename} mis à jour (élèves retirés des anciennes classes)")
                
            except Exception as e:
                print(f"❌ Erreur lors de la suppression dans {excel_filename}: {e}")
                import traceback
                traceback.print_exc()

    def update_school_excel_file(self, school_key, school_file_mapping, horaire, classe_nom, student_names):
        """Met à jour le fichier Excel de l'école avec les élèves assignés."""
        if load_workbook is None:
            print("⚠️ openpyxl n'est pas disponible, impossible de mettre à jour le fichier Excel de l'école")
            return

        # Obtenir le fichier Excel de l'école
        excel_filename = school_file_mapping.get(school_key)
        if not excel_filename:
            print(f"⚠️ Aucun fichier Excel trouvé pour l'école {school_key}")
            return

        # Chemin complet du fichier
        week_folder = os.path.dirname(self.file_path)
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            print(f"⚠️ Fichier Excel introuvable : {excel_path}")
            return

        try:
            # Ouvrir le fichier Excel
            wb = load_workbook(excel_path)

            # Chercher la feuille correspondant à l'horaire
            target_sheet = None
            for sheet_name in wb.sheetnames:
                # Nettoyer le nom de la feuille pour la comparaison
                sheet_normalized = self.clean_horaire_name(sheet_name).lower().strip()
                horaire_normalized = horaire.lower().strip()

                # Comparaison flexible : vérifier si l'horaire contient des chiffres similaires
                # Par exemple : "8h15 - 10h15" doit matcher "8h15 à 10h15 Professeur"
                import re
                # Extraire les chiffres de l'horaire
                horaire_numbers = re.findall(r'\d+', horaire_normalized)
                sheet_numbers = re.findall(r'\d+', sheet_normalized)
                
                # Si les chiffres correspondent ou si les noms sont similaires
                if (sheet_normalized == horaire_normalized or 
                    horaire_normalized in sheet_normalized or 
                    sheet_normalized in horaire_normalized or
                    (horaire_numbers and sheet_numbers and horaire_numbers == sheet_numbers)):
                    target_sheet = wb[sheet_name]
                    print(f"✅ Feuille trouvée : '{sheet_name}' pour horaire '{horaire}'")
                    break

            if target_sheet is None:
                print(f"⚠️ Feuille horaire '{horaire}' non trouvée dans {excel_filename}")
                print(f"   Feuilles disponibles : {wb.sheetnames}")
                print(f"   Horaire recherché normalisé : '{horaire_normalized}'")
                wb.close()
                return

            # Chercher la ligne correspondant à la classe (colonne 1 : "Nom de la classe")
            classe_row = None
            for row_idx in range(2, target_sheet.max_row + 1):  # Commencer à la ligne 2 (après header)
                cell_value = str(target_sheet.cell(row=row_idx, column=1).value or '').strip()
                if cell_value == classe_nom:
                    classe_row = row_idx
                    break

            if classe_row is None:
                print(f"⚠️ Classe '{classe_nom}' non trouvée dans la feuille '{target_sheet.title}'")
                wb.close()
                return

            # Colonne 5 : Liste des élèves
            eleves_col = 5
            current_value = str(target_sheet.cell(row=classe_row, column=eleves_col).value or '').strip()

            # Récupérer la liste actuelle des élèves
            existing_eleves = []
            if current_value and current_value.lower() not in ['', 'nan', 'none', 'liste des élèves...']:
                existing_eleves = [e.strip() for e in current_value.split(',') if e.strip()]

            # Ajouter seulement les élèves qui ne sont pas déjà présents (éviter les doublons)
            new_eleves = []
            for student_name in student_names:
                if student_name not in existing_eleves:
                    new_eleves.append(student_name)

            # Si il y a de nouveaux élèves à ajouter
            if new_eleves:
                if existing_eleves:
                    all_eleves = existing_eleves + new_eleves
                else:
                    all_eleves = new_eleves

                # Trier les élèves par ordre alphabétique pour une meilleure présentation
                all_eleves_sorted = sorted(all_eleves)
                eleves_text = ', '.join(all_eleves_sorted)

                target_sheet.cell(row=classe_row, column=eleves_col, value=eleves_text)

                # Sauvegarder le fichier
                wb.save(excel_path)
                print(f"✅ Fichier {excel_filename} mis à jour : {len(new_eleves)} élève(s) ajouté(s) à la classe {classe_nom}")
            else:
                print(f"ℹ️ Élève(s) déjà présent(s) dans la classe {classe_nom}")

            wb.close()

        except Exception as e:
            print(f"❌ Erreur lors de la mise à jour du fichier {excel_filename}: {e}")
            import traceback
            traceback.print_exc()

    def open_add_student_dialog(self):
        """Ouvre une fenêtre pour ajouter un nouvel élève."""
        # Configuration des écoles disponibles
        SCHOOLS = ["A", "B", "C/CS", "C/CI", "Morning", "Premium/CS", "Premium/CI"]

        # Créer la fenêtre
        dialog = ctk.CTkToplevel(self)
        dialog.title("Ajouter un élève")
        dialog.geometry("500x550")
        dialog.resizable(False, False)
        dialog.configure(fg_color="white")
        dialog.attributes("-topmost", True)

        # Centrer la fenêtre
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (550 // 2)
        dialog.geometry(f"500x550+{x}+{y}")

        # Frame principal
        main_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=30, pady=25)

        # Titre
        title_label = ctk.CTkLabel(
            main_frame,
            text="➕ Ajouter un élève",
            font=("Inter", 16, "bold"),
            text_color="#1e293b"
        )
        title_label.pack(pady=(0, 20))

        # === SECTION NOM ET PRÉNOM ===
        name_section = ctk.CTkFrame(main_frame, fg_color="transparent")
        name_section.pack(fill="x", pady=(0, 15))

        # Configuration de la grille pour nom et prénom
        name_section.grid_columnconfigure(0, weight=1)  # NOM
        name_section.grid_columnconfigure(1, weight=1)  # PRÉNOM

        # Champ NOM (avec majuscules automatiques)
        nom_label = ctk.CTkLabel(
            name_section, text="NOM",
            font=("Inter", 10, "bold"),
            text_color="#6B7280"
        )
        nom_label.grid(row=0, column=0, sticky="ew")

        nom_entry = ctk.CTkEntry(
            name_section,
            placeholder_text="",
            height=38,
            corner_radius=8,
            fg_color="#F3F4F6",
            border_width=0
        )
        nom_entry.grid(row=1, column=0, sticky="ew", padx=(0, 8))

        # Fonction pour convertir automatiquement en majuscules
        def to_uppercase(*args):
            current_text = nom_entry.get()
            if current_text != current_text.upper():
                nom_entry.delete(0, 'end')
                nom_entry.insert(0, current_text.upper())

        # Lier l'événement de changement de texte
        nom_entry.bind('<KeyRelease>', to_uppercase)

        # Champ PRÉNOM
        prenom_label = ctk.CTkLabel(
            name_section, text="PRÉNOM",
            font=("Inter", 10, "bold"),
            text_color="#6B7280"
        )
        prenom_label.grid(row=0, column=1, sticky="ew")

        prenom_entry = ctk.CTkEntry(
            name_section,
            placeholder_text="",
            height=38,
            corner_radius=8,
            fg_color="#F3F4F6",
            border_width=0
        )
        prenom_entry.grid(row=1, column=1, sticky="ew", padx=(8, 0))

        # Fonction pour mettre automatiquement la première lettre en majuscule
        def capitalize_first_letter(*args):
            current_text = prenom_entry.get()
            if current_text:
                # Mettre la première lettre en majuscule et le reste en minuscules
                capitalized_text = current_text[0].upper() + current_text[1:].lower() if len(current_text) > 1 else current_text.upper()
                if current_text != capitalized_text:
                    prenom_entry.delete(0, 'end')
                    prenom_entry.insert(0, capitalized_text)
                    # Repositionner le curseur à la fin
                    prenom_entry.icursor('end')

        # Lier l'événement de changement de texte pour la capitalisation
        prenom_entry.bind('<KeyRelease>', capitalize_first_letter)

        # === SECTION ÂGE, NIVEAU, CI ===
        first_row = ctk.CTkFrame(main_frame, fg_color="transparent")
        first_row.pack(fill="x", pady=(0, 15))

        # Configuration de la grille pour âge, niveau, CI
        first_row.grid_columnconfigure(0, weight=0)  # ÂGE (plus étroit)
        first_row.grid_columnconfigure(1, weight=1)  # NIVEAU
        first_row.grid_columnconfigure(2, weight=0)  # CI (plus étroit)

        # Champ ÂGE
        age_label = ctk.CTkLabel(
            first_row, text="ÂGE",
            font=("Inter", 10, "bold"),
            text_color="#6B7280"
        )
        age_label.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        age_entry = ctk.CTkEntry(
            first_row,
            placeholder_text="",
            height=38,
            width=60,  # Largeur fixe plus petite
            corner_radius=8,
            fg_color="#F3F4F6",
            border_width=0
        )
        age_entry.grid(row=1, column=0, sticky="ew", padx=(0, 8))

        # Fonction pour valider que seuls des chiffres sont saisis dans l'âge
        def validate_age_input(*args):
            current_text = age_entry.get()
            # Ne garder que les chiffres
            filtered_text = ''.join(c for c in current_text if c.isdigit())
            if current_text != filtered_text:
                age_entry.delete(0, 'end')
                age_entry.insert(0, filtered_text)

        # Lier l'événement de changement de texte pour la validation des chiffres
        age_entry.bind('<KeyRelease>', validate_age_input)

        # Menu NIVEAU
        level_label = ctk.CTkLabel(
            first_row, text="NIVEAU",
            font=("Inter", 10, "bold"),
            text_color="#6B7280"
        )
        level_label.grid(row=0, column=1, sticky="ew", padx=(8, 8))

        level_menu = ctk.CTkOptionMenu(
            first_row,
            values=self.NIVEAUX,
            height=38,
            corner_radius=8,
            fg_color="#F3F4F6",
            text_color="#1e293b",
            button_color="#E5E7EB",
            button_hover_color="#D1D5DB",
            dropdown_fg_color="#F3F4F6",
            dropdown_hover_color="#E5E7EB",
            dropdown_text_color="#1e293b"
        )
        level_menu.grid(row=1, column=1, sticky="ew", padx=(8, 8))
        level_menu.set("Choisir un niveau...")

        # Case à cocher CI (alignée avec les autres)
        ci_label = ctk.CTkLabel(
            first_row, text="CI",
            font=("Inter", 10, "bold"),
            text_color="#6B7280"
        )
        ci_label.grid(row=0, column=2, sticky="ew", padx=(8, 0))

        ci_var = ctk.BooleanVar(value=False)
        ci_checkbox = ctk.CTkCheckBox(
            first_row,
            text="",
            variable=ci_var,
            fg_color="#10b981",
            hover_color="#059669",
            width=38,
            height=38
        )
        ci_checkbox.grid(row=1, column=2, sticky="ew", padx=(8, 0))

        # === SECTION ÉCOLE (ligne complète) ===
        school_row = ctk.CTkFrame(main_frame, fg_color="transparent")
        school_row.pack(fill="x", pady=(0, 15))

        # Menu ÉCOLE
        school_label = ctk.CTkLabel(
            school_row, text="ÉCOLE",
            font=("Inter", 10, "bold"),
            text_color="#6B7280"
        )
        school_label.pack(anchor="w")

        school_menu = ctk.CTkOptionMenu(
            school_row,
            values=SCHOOLS,
            height=38,
            corner_radius=8,
            fg_color="#F3F4F6",
            text_color="#1e293b",
            button_color="#E5E7EB",
            button_hover_color="#D1D5DB",
            dropdown_fg_color="#F3F4F6",
            dropdown_hover_color="#E5E7EB",
            dropdown_text_color="#1e293b"
        )
        school_menu.pack(fill="x")
        school_menu.set("Choisir une école...")

        # === SECTION HORAIRE, CLASSE ===
        schedule_row = ctk.CTkFrame(main_frame, fg_color="transparent")
        schedule_row.pack(fill="x", pady=(0, 15))

        # Configuration de la grille pour horaire, classe
        schedule_row.grid_columnconfigure(0, weight=1)  # HORAIRE
        schedule_row.grid_columnconfigure(1, weight=1)  # CLASSE

        # Menu HORAIRE
        horaire_label = ctk.CTkLabel(
            schedule_row, text="HORAIRE",
            font=("Inter", 10, "bold"),
            text_color="#6B7280"
        )
        horaire_label.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        dialog.horaire_menu = ctk.CTkOptionMenu(
            schedule_row,
            values=["Choisir d'abord une école"],
            height=38,
            corner_radius=8,
            fg_color="#E5E7EB",  # Gris au départ
            text_color="#9CA3AF",
            button_color="#D1D5DB",
            button_hover_color="#9CA3AF",
            dropdown_fg_color="#E5E7EB",
            dropdown_hover_color="#D1D5DB",
            dropdown_text_color="#6B7280",
            state="disabled"  # Désactivé au départ
        )
        dialog.horaire_menu.grid(row=1, column=0, sticky="ew", padx=(0, 8))
        dialog.horaire_menu.set("Choisir d'abord une école")

        # Menu CLASSE
        classe_label = ctk.CTkLabel(
            schedule_row, text="CLASSE",
            font=("Inter", 10, "bold"),
            text_color="#6B7280"
        )
        classe_label.grid(row=0, column=1, sticky="ew", padx=(8, 0))

        dialog.classe_menu = ctk.CTkOptionMenu(
            schedule_row,
            values=["Choisir d'abord une école"],
            height=38,
            corner_radius=8,
            fg_color="#E5E7EB",  # Gris au départ
            text_color="#9CA3AF",
            button_color="#D1D5DB",
            button_hover_color="#9CA3AF",
            dropdown_fg_color="#E5E7EB",
            dropdown_hover_color="#D1D5DB",
            dropdown_text_color="#6B7280",
            state="disabled"  # Désactivé au départ
        )
        dialog.classe_menu.grid(row=1, column=1, sticky="ew", padx=(8, 0))
        dialog.classe_menu.set("Choisir d'abord une école")

        # Fonction pour mettre à jour les classes quand l'horaire change
        def update_classe_menu(*args):
            if dialog.horaire_menu is None or dialog.classe_menu is None:
                return
            selected_horaire = dialog.horaire_menu.get()
            if hasattr(dialog, 'classes_by_horaire') and selected_horaire in dialog.classes_by_horaire:
                classes = dialog.classes_by_horaire[selected_horaire]
                if classes:
                    dialog.classe_menu.configure(state="normal", fg_color="#F3F4F6", text_color="#1e293b", values=classes)
                    dialog.classe_menu.set("Choisir une classe...")
                else:
                    dialog.classe_menu.configure(state="normal", fg_color="#F3F4F6", text_color="#9CA3AF", values=["Aucune classe disponible"])
                    dialog.classe_menu.set("Aucune classe disponible")
            else:
                dialog.classe_menu.configure(state="disabled", fg_color="#E5E7EB", text_color="#9CA3AF", values=["Choisir d'abord un horaire"])
                dialog.classe_menu.set("Choisir d'abord un horaire")

        # Fonction pour mettre à jour la couleur du texte du menu horaire
        def update_horaire_text_color(*args):
            if dialog.horaire_menu is None:
                return
            selected_value = dialog.horaire_menu.get()
            if selected_value and selected_value not in ["Choisir d'abord une école", "Aucun horaire disponible", "Choisir un horaire..."]:
                dialog.horaire_menu.configure(text_color="#1e293b")  # Noir pour valeur sélectionnée
            else:
                dialog.horaire_menu.configure(text_color="#9CA3AF")  # Gris pour message d'aide

        # Fonction pour mettre à jour la couleur du texte du menu classe
        def update_classe_text_color(*args):
            if dialog.classe_menu is None:
                return
            selected_value = dialog.classe_menu.get()
            if selected_value and selected_value not in ["Choisir d'abord une école", "Choisir d'abord un horaire", "Aucune classe disponible", "Choisir une classe..."]:
                dialog.classe_menu.configure(text_color="#1e293b")  # Noir pour valeur sélectionnée
            else:
                dialog.classe_menu.configure(text_color="#9CA3AF")  # Gris pour message d'aide

        # Fonction pour mettre à jour les menus horaire et classe
        def update_horaire_classe_menus(*args):
            selected_school = school_menu.get()
            school_key = self.get_school_key_from_display_name(selected_school) if selected_school != "Choisir une école..." else None

            if school_key:
                # Récupérer les données de l'école
                week_folder = os.path.dirname(self.file_path)
                school_data = self.analyze_school_classes(week_folder)
                school_info = school_data.get(school_key, [])

                # Extraire les horaires uniques
                horaires = []
                classes_by_horaire = {}

                for horaire_info in school_info:
                    horaire = horaire_info.get('horaire', '')
                    if horaire and str(horaire).lower() != 'nan' and horaire not in horaires:
                        horaires.append(horaire)
                        classes_by_horaire[horaire] = []

                    # Récupérer les classes pour cet horaire
                    classes = horaire_info.get('classes', [])
                    for classe_info in classes:
                        classe_nom = classe_info.get('nom_classe', '')
                        if classe_nom and classe_nom not in classes_by_horaire.get(horaire, []):
                            classes_by_horaire[horaire].append(classe_nom)

                # Trier les horaires et classes
                horaires.sort()
                for horaire in classes_by_horaire:
                    classes_by_horaire[horaire].sort()

                # Stocker les données pour utilisation ultérieure
                dialog.classes_by_horaire = classes_by_horaire

                # Activer et remplir les menus
                if dialog.horaire_menu is not None:
                    if horaires:
                        dialog.horaire_menu.configure(state="normal", fg_color="#F3F4F6", text_color="#1e293b", values=horaires)
                        dialog.horaire_menu.set("Choisir un horaire...")
                    else:
                        dialog.horaire_menu.configure(state="normal", fg_color="#F3F4F6", text_color="#9CA3AF", values=["Aucun horaire disponible"])
                        dialog.horaire_menu.set("Aucun horaire disponible")

                if dialog.classe_menu is not None:
                    dialog.classe_menu.configure(state="disabled", fg_color="#E5E7EB", text_color="#9CA3AF", values=["Choisir d'abord un horaire"])
                    dialog.classe_menu.set("Choisir d'abord un horaire")
            else:
                # Désactiver les menus si aucune école sélectionnée
                if dialog.horaire_menu is not None:
                    dialog.horaire_menu.configure(state="disabled", fg_color="#E5E7EB", text_color="#9CA3AF", values=["Choisir d'abord une école"])
                    dialog.horaire_menu.set("Choisir d'abord une école")
                if dialog.classe_menu is not None:
                    dialog.classe_menu.configure(state="disabled", fg_color="#E5E7EB", text_color="#9CA3AF", values=["Choisir d'abord une école"])
                    dialog.classe_menu.set("Choisir d'abord une école")



        # Label d'erreur
        error_label = ctk.CTkLabel(
            main_frame,
            text="",
            font=("Inter", 11),
            text_color="#EF4444"
        )
        error_label.pack(pady=(0, 10))

        # Fonction de sauvegarde
        def save_student():
            nom = nom_entry.get().strip()
            prenom = prenom_entry.get().strip()
            name = f"{nom} {prenom}".strip()  # Combiner nom et prénom
            age_text = age_entry.get().strip()
            level = level_menu.get()
            school = school_menu.get()
            horaire = dialog.horaire_menu.get() if dialog.horaire_menu and dialog.horaire_menu.get() not in ["Choisir d'abord une école", "Aucun horaire disponible", "Choisir un horaire..."] else ""
            classe = dialog.classe_menu.get() if dialog.classe_menu and dialog.classe_menu.get() not in ["Choisir d'abord une école", "Choisir d'abord un horaire", "Aucune classe disponible", "Choisir une classe..."] else ""
            is_ci = ci_var.get()

            # Validation
            if not nom:
                error_label.configure(text="⚠️ Le nom est requis")
                return

            if not prenom:
                error_label.configure(text="⚠️ Le prénom est requis")
                return

            if not age_text:
                error_label.configure(text="⚠️ L'âge est requis")
                return

            try:
                age = int(age_text)
                if age < 3 or age > 99:
                    error_label.configure(text="⚠️ L'âge doit être entre 3 et 99 ans")
                    return
            except ValueError:
                error_label.configure(text="⚠️ L'âge doit être un nombre")
                return

            if level == "Choisir un niveau...":
                error_label.configure(text="⚠️ Le niveau est requis")
                return

            if school == "Choisir une école...":
                error_label.configure(text="⚠️ L'école est requise")
                return

            # Vérifier que si un horaire est sélectionné, une classe l'est aussi (et vice versa)
            if horaire and not classe:
                error_label.configure(text="⚠️ Si un horaire est sélectionné, une classe doit l'être aussi")
                return
            if classe and not horaire:
                error_label.configure(text="⚠️ Si une classe est sélectionnée, un horaire doit l'être aussi")
                return

            # Créer la nouvelle ligne
            new_row = {
                self.cols_map["Stagiaire"]: name,
                self.cols_map["Âge"]: age,
                self.cols_map["Niveau"]: level,
                self.cols_map["Ecole"]: school
            }

            if self.cols_map["Cours 2"]:
                new_row[self.cols_map["Cours 2"]] = "OUI" if is_ci else ""
            # if self.cols_map["Horaire"]:  # Désactivé
            #     new_row[self.cols_map["Horaire"]] = horaire
            if self.cols_map["Classe"]:
                new_row[self.cols_map["Classe"]] = classe

            # Ajouter au DataFrame
            self.df = pd.concat([self.df, pd.DataFrame([new_row])], ignore_index=True)

            # Sauvegarder
            self.df.to_excel(self.file_path, index=False)

            # Créer la classe si elle n'existe pas
            if classe and horaire:
                school_key = self.get_school_key_from_display_name(school)
                if school_key:
                    self.create_missing_classes_from_import([{
                        'nom': name,
                        'donnees': [f"École: {school}", f"Horaire: {horaire}", f"Classe: {classe}"]
                    }])

            # Rafraîchir l'affichage
            self.refresh_table(preserve_selection=False)
            self.update_counters()

            dialog.destroy()
            messagebox.showinfo("Succès", f"Élève {name} ajouté avec succès!")

        # Bouton sauvegarder
        save_btn = ctk.CTkButton(
            main_frame,
            text="Ajouter l'élève",
            fg_color="#10b981",
            hover_color="#059669",
            height=42,
            corner_radius=8,
            font=("Inter", 13, "bold"),
            command=save_student
        )
        save_btn.pack(fill="x", pady=(5, 0))

        # Lier les événements après la définition de tous les widgets
        school_menu.configure(command=update_horaire_classe_menus)
        if dialog.horaire_menu is not None:
            dialog.horaire_menu.configure(command=lambda value: (update_classe_menu(), update_horaire_text_color()))
        if dialog.classe_menu is not None:
            dialog.classe_menu.configure(command=update_classe_text_color)

        # Donner le focus au champ nom
        dialog.after(200, lambda: nom_entry.focus_force())

    def get_school_key_from_display_name(self, display_name):
        """Convertit un nom d'affichage d'école en clé interne."""
        mapping = {
            'A': 'ecole_a',
            'B': 'ecole_b',
            'C/CS': 'ecole_c_cs',
            'C/CI': 'ecole_c_ci',
            'Morning': 'ecole_morning',
            'Premium/CS': 'ecole_premium_cs',
            'Premium/CI': 'ecole_premium_ci'
        }
        return mapping.get(display_name)

    def open_remove_student_dialog(self):
        """Ouvre une fenêtre pour supprimer un élève."""
        # Créer la fenêtre
        dialog = ctk.CTkToplevel(self)
        dialog.title("Supprimer un élève")
        dialog.geometry("650x650")
        dialog.resizable(True, True)
        dialog.configure(fg_color="white")
        dialog.attributes("-topmost", True)

        # Centrer la fenêtre
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (650 // 2)
        y = (dialog.winfo_screenheight() // 2) - (650 // 2)
        dialog.geometry(f"650x650+{x}+{y}")

        # Frame principal
        main_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=30, pady=25)

        # Titre
        title_label = ctk.CTkLabel(
            main_frame,
            text="➖ Supprimer un élève",
            font=("Inter", 16, "bold"),
            text_color="#1e293b"
        )
        title_label.pack(pady=(0, 20))

        # Zone de recherche
        search_label = ctk.CTkLabel(
            main_frame, text="RECHERCHER UN ÉLÈVE",
            font=("Inter", 10, "bold"),
            text_color="#6B7280"
        )
        search_label.pack(anchor="w")

        search_var = ctk.StringVar()
        search_entry = ctk.CTkEntry(
            main_frame,
            textvariable=search_var,
            placeholder_text="Tapez le nom de l'élève...",
            height=38,
            corner_radius=8,
            fg_color="#F3F4F6",
            border_width=0
        )
        search_entry.pack(fill="x", pady=(0, 15))

        # Frame scrollable pour les résultats (max 6 élèves)
        results_frame = ctk.CTkScrollableFrame(
            main_frame,
            fg_color="#f8fafc",
            corner_radius=10
        )
        results_frame.pack(fill="both", expand=True)

        # Variable pour stocker les élèves filtrés
        current_students = []

        # Fonction de recherche
        def search_students(*args):
            search_term = search_var.get().lower().strip()

            # Vider les résultats précédents
            for widget in results_frame.winfo_children():
                widget.destroy()

            if not search_term:
                display_students([])  # Afficher le message d'aide
                return

            # Chercher les élèves correspondants (maximum 6)
            current_students.clear()
            count = 0
            for _, row in self.df.iterrows():
                if count >= 6:  # Limiter à 6 élèves maximum
                    break
                nom = str(row[self.cols_map["Stagiaire"]]).lower()
                if search_term in nom:
                    student_data = {
                        'nom': str(row[self.cols_map["Stagiaire"]]),
                        'age': row[self.cols_map["Âge"]],
                        'niveau': str(row[self.cols_map["Niveau"]]) if pd.notna(row[self.cols_map["Niveau"]]) else "",
                        'classe': str(row[self.cols_map["Classe"]]) if self.cols_map["Classe"] and pd.notna(row[self.cols_map["Classe"]]) else "",
                        'prof': str(row[self.cols_map["Prof"]]) if self.cols_map["Prof"] and pd.notna(row[self.cols_map["Prof"]]) else "",
                        'arrivee': self.format_date_jour_mois(row[self.cols_map["Arrivée"]]) if self.cols_map["Arrivée"] and pd.notna(row[self.cols_map["Arrivée"]]) else "",
                        'depart': self.format_date_jour_mois(row[self.cols_map["Départ"]]) if self.cols_map["Départ"] and pd.notna(row[self.cols_map["Départ"]]) else "",
                        'ci': "OUI" if self.cols_map["Départ CI"] and pd.notna(row[self.cols_map["Départ CI"]]) and str(row[self.cols_map["Départ CI"]]).strip() != "" else "",
                        'classe_ci': str(row[self.cols_map["Classe CI"]]) if self.cols_map["Classe CI"] and pd.notna(row[self.cols_map["Classe CI"]]) else "",
                        'prof_ci': str(row[self.cols_map["Prof CI"]]) if self.cols_map["Prof CI"] and pd.notna(row[self.cols_map["Prof CI"]]) else "",
                        'arr_ci': self.format_date_jour_mois(row[self.cols_map["Départ CI"]]) if self.cols_map["Départ CI"] and pd.notna(row[self.cols_map["Départ CI"]]) else "",
                        'dep_ci': self.format_date_jour_mois(row[self.cols_map["Arrivée CI"]]) if self.cols_map["Arrivée CI"] and pd.notna(row[self.cols_map["Arrivée CI"]]) else ""
                    }
                    current_students.append(student_data)
                    count += 1

            # Afficher les résultats
            display_students(current_students)

        # Lier la recherche
        search_var.trace_add("write", search_students)

        # Fonction d'affichage des élèves
        def display_students(students):
            # Configurer la grille pour 2 colonnes
            results_frame.grid_columnconfigure(0, weight=1)
            results_frame.grid_columnconfigure(1, weight=1)

            # Si aucun élève trouvé, afficher un message d'aide
            if not students:
                help_container = ctk.CTkFrame(results_frame, fg_color="transparent")
                help_container.pack(fill="both", expand=True)
                help_label = ctk.CTkLabel(
                    help_container,
                    text="🔍 Tapez le nom d'un élève dans la barre de recherche ci-dessus\npour voir les élèves correspondants",
                    font=("Inter", 14, "italic"),
                    text_color="#6b7280",
                    justify="center"
                )
                help_label.pack(expand=True)
                return

            # Organiser les élèves en lignes de 2
            for idx, student in enumerate(students):
                row = idx // 2
                col = idx % 2

                # Frame pour chaque élève
                student_frame = ctk.CTkFrame(
                    results_frame,
                    fg_color="white",
                    corner_radius=8,
                    border_width=1,
                    border_color="#e5e7eb"
                )
                student_frame.grid(row=row, column=col, sticky="nsew", padx=(5, 2) if col == 0 else (2, 5), pady=(0, 8))

                # En-tête avec nom et âge
                header_frame = ctk.CTkFrame(student_frame, fg_color="transparent")
                header_frame.pack(fill="x", padx=12, pady=(10, 5))

                name_label = ctk.CTkLabel(
                    header_frame,
                    text=f"👤 {student['nom']}",
                    font=("Inter", 14, "bold"),
                    text_color="#1e293b"
                )
                name_label.pack(side="left")

                age_label = ctk.CTkLabel(
                    header_frame,
                    text=f"Âge: {student['age']}",
                    font=("Inter", 12),
                    text_color="#6b7280"
                )
                age_label.pack(side="right")

                # Informations détaillées
                info_frame = ctk.CTkFrame(student_frame, fg_color="transparent")
                info_frame.pack(fill="x", padx=12, pady=(0, 10))

                # Ligne 1: Niveau et École
                level_school_frame = ctk.CTkFrame(info_frame, fg_color="transparent")
                level_school_frame.pack(fill="x", pady=(0, 3))

                niveau_text = f"📚 Niveau: {student['niveau']}" if student['niveau'] else "📚 Niveau: Non défini"
                level_label = ctk.CTkLabel(
                    level_school_frame,
                    text=niveau_text,
                    font=("Inter", 11),
                    text_color="#374151"
                )
                level_label.pack(side="left", padx=(0, 15))

                ecole_text = f"🏫 École: {student['ecole']}" if student['ecole'] else "🏫 École: Non définie"
                school_label = ctk.CTkLabel(
                    level_school_frame,
                    text=ecole_text,
                    font=("Inter", 11),
                    text_color="#374151"
                )
                school_label.pack(side="left")

                # Ligne 2: Classe et Prof
                classe_prof_frame = ctk.CTkFrame(info_frame, fg_color="transparent")
                classe_prof_frame.pack(fill="x", pady=(0, 3))

                classe_text = f"📝 Classe: {student['classe']}" if student['classe'] else "📝 Classe: Non définie"
                classe_label = ctk.CTkLabel(
                    classe_prof_frame,
                    text=classe_text,
                    font=("Inter", 11),
                    text_color="#374151"
                )
                classe_label.pack(side="left", padx=(0, 15))

                prof_text = f"👨‍🏫 Prof: {student['prof']}" if student['prof'] else "👨‍🏫 Prof: Non défini"
                prof_label = ctk.CTkLabel(
                    classe_prof_frame,
                    text=prof_text,
                    font=("Inter", 11),
                    text_color="#374151"
                )
                prof_label.pack(side="left")

                # Ligne 3: Classe CI et Prof CI
                classe_ci_prof_ci_frame = ctk.CTkFrame(info_frame, fg_color="transparent")
                classe_ci_prof_ci_frame.pack(fill="x", pady=(0, 3))

                arr_ci_text = f"📅 Arr. CI: {student['arr_ci']}" if student.get('arr_ci') else "📅 Arr. CI: Non définie"
                arr_ci_label = ctk.CTkLabel(
                    classe_ci_prof_ci_frame,
                    text=arr_ci_text,
                    font=("Inter", 11),
                    text_color="#374151"
                )
                arr_ci_label.pack(side="left", padx=(0, 15))

                dep_ci_text = f"📅 Dép. CI: {student.get('dep_ci')}" if student.get('dep_ci') else "📅 Dép. CI: Non définie"
                dep_ci_label = ctk.CTkLabel(
                    classe_ci_prof_ci_frame,
                    text=dep_ci_text,
                    font=("Inter", 11),
                    text_color="#374151"
                )
                dep_ci_label.pack(side="left")

                # Ligne 3: CI
                ci_frame = ctk.CTkFrame(info_frame, fg_color="transparent")
                ci_frame.pack(fill="x", pady=(0, 8))

                ci_text = f"⚡ Cours intensifs: {student['ci']}"
                ci_color = "#10b981" if student['ci'] == "OUI" else "#6b7280"
                ci_label = ctk.CTkLabel(
                    ci_frame,
                    text=ci_text,
                    font=("Inter", 11),
                    text_color=ci_color
                )
                ci_label.pack(side="left")

                # Bouton supprimer
                def delete_student(s_name=student['nom']):
                    # Créer une boîte de dialogue de confirmation au premier plan
                    confirm_dialog = ctk.CTkToplevel(dialog)
                    confirm_dialog.title("")
                    confirm_dialog.geometry("400x150")
                    confirm_dialog.resizable(False, False)
                    confirm_dialog.attributes("-topmost", True)
                    confirm_dialog.transient(dialog)

                    # Centrer la fenêtre par rapport à la fenêtre parente
                    dialog_x = dialog.winfo_rootx()
                    dialog_y = dialog.winfo_rooty()
                    dialog_width = dialog.winfo_width()
                    dialog_height = dialog.winfo_height()

                    x = dialog_x + (dialog_width // 2) - (400 // 2)
                    y = dialog_y + (dialog_height // 2) - (150 // 2)
                    confirm_dialog.geometry(f"400x150+{x}+{y}")

                    # Configuration moderne
                    confirm_dialog.overrideredirect(True)  # Pas de barre de titre

                    # Frame principal avec coins arrondis
                    main_frame = ctk.CTkFrame(
                        confirm_dialog,
                        fg_color="#ffffff",
                        corner_radius=15,
                        border_width=2,
                        border_color="#e1e5e9"
                    )
                    main_frame.pack(fill="both", expand=True, padx=2, pady=2)

                    # Frame intérieur pour le contenu
                    content_frame = ctk.CTkFrame(
                        main_frame,
                        fg_color="#f8fafc",
                        corner_radius=10
                    )
                    content_frame.pack(fill="both", expand=True, padx=8, pady=8)

                    # Message de confirmation
                    message_label = ctk.CTkLabel(
                        content_frame,
                        text=f"Êtes-vous sûr de vouloir supprimer\nl'élève '{s_name}' ?",
                        font=("Inter", 13, "bold"),
                        text_color="#1e293b",
                        justify="center"
                    )
                    message_label.pack(pady=(15, 20))

                    # Variable pour suivre la réponse
                    confirmed = [False]

                    def on_confirm():
                        confirmed[0] = True
                        confirm_dialog.destroy()

                    def on_cancel():
                        confirmed[0] = False
                        confirm_dialog.destroy()

                    # Boutons
                    button_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
                    button_frame.pack(pady=(0, 15))

                    cancel_btn = ctk.CTkButton(
                        button_frame,
                        text="❌ Annuler",
                        width=90,
                        height=32,
                        font=("Inter", 11),
                        fg_color="#6b7280",
                        hover_color="#475569",
                        corner_radius=6,
                        command=on_cancel
                    )
                    cancel_btn.pack(side="left", padx=(10, 5))

                    confirm_btn = ctk.CTkButton(
                        button_frame,
                        text="✅ Supprimer",
                        width=90,
                        height=32,
                        font=("Inter", 11, "bold"),
                        fg_color="#ef4444",
                        hover_color="#dc2626",
                        corner_radius=6,
                        command=on_confirm
                    )
                    confirm_btn.pack(side="left", padx=(5, 10))

                    # Attendre que la boîte de dialogue se ferme
                    confirm_dialog.wait_window()

                    # Si confirmé, procéder à la suppression
                    if confirmed[0]:
                        # Récupérer les informations de l'élève AVANT suppression pour mettre à jour les fichiers Excel
                        student_info = self.get_student_info_for_excel_removal(s_name)

                        # Supprimer de la base matrix
                        self.df = self.df[self.df[self.cols_map["Stagiaire"]] != s_name]
                        self.df.to_excel(self.file_path, index=False)

                        # Supprimer des fichiers Excel des écoles si nécessaire
                        if student_info:
                            self.remove_student_from_excel_files_with_info(student_info)

                        # Rafraîchir l'affichage principal
                        self.refresh_table(preserve_selection=False)
                        self.update_counters()

                        # Mettre à jour le dashboard des classes si on a accès à la fenêtre principale
                        try:
                            week_folder = os.path.dirname(self.file_path)
                            # Essayer de trouver la fenêtre principale et rafraîchir le dashboard
                            for window in ctk.CTk._all_toplevels:
                                if hasattr(window, 'create_classes_dashboard') and hasattr(window, 'content'):
                                    try:
                                        # Rafraîchir le dashboard des classes
                                        school_data = window.analyze_school_classes(week_folder)
                                        window.create_classes_dashboard(window.content, school_data, week_folder)
                                        print("✅ Dashboard des classes mis à jour après suppression")
                                        break
                                    except:
                                        pass
                        except:
                            pass

                        # Fermer la fenêtre de suppression
                        dialog.destroy()

                delete_btn = ctk.CTkButton(
                    student_frame,
                    text="🗑️ Supprimer",
                    width=100,
                    height=32,
                    font=("Inter", 11, "bold"),
                    fg_color="#ef4444",
                    hover_color="#dc2626",
                    corner_radius=6,
                    command=delete_student
                )
                delete_btn.pack(side="right", padx=12, pady=(0, 10))

        # Afficher le message d'aide initial
        display_students([])

        # Donner le focus à la recherche
        dialog.after(200, lambda: search_entry.focus_force())

        # Lier la fermeture
        def on_closing():
            dialog.destroy()

        dialog.protocol("WM_DELETE_WINDOW", on_closing)

    def get_student_info_for_excel_removal(self, student_name):
        """Récupère les informations d'un élève avant sa suppression pour mettre à jour les fichiers Excel."""
        student_row = self.df[self.df[self.cols_map["Stagiaire"]] == student_name]
        if student_row.empty:
            return None

        ecole = student_row[self.cols_map["Ecole"]].values[0] if self.cols_map["Ecole"] and pd.notna(student_row[self.cols_map["Ecole"]].values[0]) else None
        # horaire = student_row[self.cols_map["Horaire"]].values[0] if self.cols_map["Horaire"] and pd.notna(student_row[self.cols_map["Horaire"]].values[0]) else None  # Désactivé
        classe = student_row[self.cols_map["Classe"]].values[0] if self.cols_map["Classe"] and pd.notna(student_row[self.cols_map["Classe"]].values[0]) else None

        if not ecole or not classe:  # horaire plus requis
            return None

        return {
            'nom': student_name,
            'ecole': ecole,
            # 'horaire': horaire,  # Désactivé
            'classe': classe
        }

    def remove_student_from_excel_files_with_info(self, student_info):
        """Supprime un élève des fichiers Excel des écoles en utilisant les informations fournies."""
        if not student_info or load_workbook is None:
            return

        student_name = student_info['nom']
        ecole = student_info['ecole']
        # horaire = student_info['horaire']  # Désactivé - colonne supprimée
        classe = student_info['classe']

        # Désactiver la suppression des fichiers Excel car l'horaire n'est plus disponible
        print(f"⚠️ Suppression des fichiers Excel désactivée pour {student_name} (horaire non disponible)")
        return

        # Mapping des écoles
        school_file_mapping = {
            'ecole_a': 'ecole_a.xlsx',
            'ecole_b': 'ecole_b.xlsx',
            'ecole_c_cs': 'ECOLE_C_cours_standard.xlsx',
            'ecole_c_ci': 'ECOLE_C_cours_intensif.xlsx',
            'ecole_morning': 'MORNING.xlsx',
            'ecole_premium_cs': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'ecole_premium_ci': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        week_folder = os.path.dirname(self.file_path)

        # Convertir l'école en clé
        school_key = self.get_school_key_from_display_name(ecole)
        if not school_key:
            return

        excel_filename = school_file_mapping.get(school_key)
        if not excel_filename:
            return

        excel_path = os.path.join(week_folder, excel_filename)
        if not os.path.exists(excel_path):
            return

        try:
            wb = load_workbook(excel_path)

            # Chercher la feuille horaire
            target_sheet = None
            for sheet_name in wb.sheetnames:
                sheet_clean = self.clean_horaire_name(sheet_name)
                horaire_clean = str(horaire).strip()
                if sheet_clean == horaire_clean or horaire_clean in sheet_clean:
                    target_sheet = wb[sheet_name]
                    break

            if not target_sheet:
                wb.close()
                return

            # Chercher la classe
            classe_row = None
            for row_idx in range(2, target_sheet.max_row + 1):
                cell_value = str(target_sheet.cell(row=row_idx, column=1).value or '').strip()
                if cell_value == str(classe).strip():
                    classe_row = row_idx
                    break

            if not classe_row:
                wb.close()
                return

            # Colonne des élèves (généralement colonne 5)
            eleves_col = 5
            current_value = str(target_sheet.cell(row=classe_row, column=eleves_col).value or '').strip()

            if current_value and current_value.lower() not in ['', 'nan', 'none', 'liste des élèves...']:
                # Séparer les élèves
                eleves_list = [e.strip() for e in current_value.split(',') if e.strip()]

                # Retirer l'élève
                if student_name in eleves_list:
                    eleves_list.remove(student_name)

                    # Mettre à jour la cellule des élèves
                    if eleves_list:
                        new_value = ', '.join(sorted(eleves_list))
                        target_sheet.cell(row=classe_row, column=eleves_col, value=new_value)
                    else:
                        # Si la classe devient vide, vider aussi la colonne niveau
                        target_sheet.cell(row=classe_row, column=eleves_col, value='')
                        target_sheet.cell(row=classe_row, column=4, value='')  # Colonne niveau

                    wb.save(excel_path)
                    print(f"✅ Élève {student_name} retiré de {excel_filename} (horaire: {horaire}, classe: {classe})")

            wb.close()

        except Exception as e:
            print(f"❌ Erreur lors de la suppression de {student_name} dans {excel_filename}: {e}")

    def remove_student_from_excel_files(self, student_name):
        """Supprime un élève des fichiers Excel des écoles (méthode dépréciée - utilise remove_student_from_excel_files_with_info)."""
        student_info = self.get_student_info_for_excel_removal(student_name)
        if student_info:
            self.remove_student_from_excel_files_with_info(student_info)


if __name__ == "__main__":
    # Récupérer le chemin du fichier matrix passé en argument depuis fenetre_principale.py
    file_path = sys.argv[1] if len(sys.argv) > 1 else None
    app = AppPedagogique(file_path)
    app.mainloop()