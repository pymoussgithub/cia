import customtkinter as ctk
import os
import json
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

def load_personnel_lists(week_folder):
    """
    Charge les listes de niveaux, professeurs et animateurs depuis personnel.json
    """
    # Liste par défaut des niveaux
    niveaux = ["A0", "A0/A0+", "Pré-A1", "Pré-A1/A1", "A1", "A1.2", "A1.2/A2",
               "A2", "A2/A2.2", "A2.2", "A2.2/B1", "B1", "B1.2", "B2", "Pitchoune", "Non Spécifié"]

    # Listes vides par défaut - tout est chargé dynamiquement depuis personnel.json
    profs = ["Non Spécifié"]  # Garder "Non Spécifié" comme option de secours
    anims = ["Non Spécifié"]  # Garder "Non Spécifié" comme option de secours

    if week_folder:
        personnel_path = os.path.join(week_folder, "personnel.json")
        if os.path.exists(personnel_path):
            try:
                with open(personnel_path, 'r', encoding='utf-8') as f:
                    personnel_data = json.load(f)

                # Charger les professeurs depuis personnel.json
                profs = []
                if "professeurs" in personnel_data:
                    for prof in personnel_data["professeurs"]:
                        if "nom" in prof:
                            profs.append(prof["nom"])
                profs.append("Non Spécifié")  # Ajouter l'option "Non Spécifié"

                # Charger les animateurs depuis personnel.json
                anims = []
                if "animateurs" in personnel_data:
                    for anim in personnel_data["animateurs"]:
                        if "nom" in anim:
                            anims.append(anim["nom"])
                anims.append("Non Spécifié")  # Ajouter l'option "Non Spécifié"

                return niveaux, profs, anims
            except Exception as e:
                print(f"Erreur lors du chargement de personnel.json: {e}")

    # Retourner les valeurs par défaut si le fichier n'existe pas ou erreur
    return niveaux, profs, anims

def open_classe_details(classe_info, horaire, intervenant, type_intervenant, school_color, school_name, week_folder=None, refresh_callback=None):
    print(classe_info)
    # --- IMPORTS LOCAUX ---
    from tkinter import messagebox

    # --- CONFIGURATION DES NIVEAUX ET COULEURS ---
    NIVEAUX = ["A0", "A0/A0+", "Pré-A1", "Pré-A1/A1", "A1", "A1.2", "A1.2/A2",
               "A2", "A2/A2.2", "A2.2", "A2.2/B1", "B1", "B1.2", "B2"]
    COULEURS = ["#E8F2FA", "#E1EDF8", "#DAE8F6", "#E8F0E8", "#E1F0E6", "#D1EDEA",
                "#FFF4E0", "#FFF0D1", "#FFEBCC", "#FCE8E8", "#F0E6F2", "#E8DCE8",
                "#E4F2F0", "#DDE5EA"]
    map_colors = dict(zip(NIVEAUX, COULEURS))

    # Couleurs des écoles (même que fenetre_principale.py)
    school_colors = {
        "A": "#3b82f6",      # Bleu
        "B": "#10b981",      # Vert
        "C/CS": "#f59e0b",   # Orange
        "C/CI": "#8b5cf6",   # Violet
        "Morning": "#ef4444", # Rouge
        "Premium/CS": "#06b6d4", # Cyan
        "Premium/CI": "#f97316"  # Orange foncé
    }

    # === INITIALISATION DE LA FENÊTRE ===
    # Extraire le numéro de semaine du week_folder
    if week_folder:
        week_name = os.path.basename(week_folder)
        if week_name.startswith("semaine_"):
            try:
                week_num = week_name.split("_")[1]
                week_display = f"Semaine {week_num}"
            except:
                week_display = week_name
        else:
            week_display = week_name
        window_title = f"Fiche Classe - {week_display}"
    else:
        window_title = "Fiche Classe"

    detail_window = ctk.CTkToplevel()
    detail_window.title(window_title)
    detail_window.geometry("1250x820")
    detail_window.configure(fg_color="white") # Fond de la fenêtre BLANC

    detail_window.after(200, lambda: detail_window.focus_force())
    detail_window.attributes("-topmost", True)

    # Variables pour stocker les références aux widgets modifiables
    class_label_ref = None
    niveau_dropdown_ref = None
    intervenant_dropdown_ref = None
    eleves_counter_label_ref = None  # Référence au label du compteur d'élèves

    # Variables pour la sélection des élèves
    selected_students = set()  # Ensemble des indices des élèves sélectionnés
    student_cards = []  # Liste des cartes d'élèves pour référence

    # Charger les listes depuis personnel.json
    liste_niveaux, liste_intervenants_profs, liste_intervenants_anims = load_personnel_lists(week_folder)

    # Fonctions de callback pour l'édition
    def edit_class_name():
        """Ouvre une boîte de dialogue pour modifier le nom de la classe"""
        dialog = ctk.CTkInputDialog(text="Nouveau nom de classe :", title="Modifier la classe")
        new_name = dialog.get_input()
        if new_name and new_name.strip():
            ancien_nom = classe_info['nom_classe']
            nouveau_nom = new_name.strip()

            # Mettre à jour les données locales
            classe_info['nom_classe'] = nouveau_nom
            if class_label_ref:
                class_label_ref.configure(text=nouveau_nom)

            # Mettre à jour le fichier Excel de l'école
            update_school_excel_file_class_name(week_folder, school_name, horaire, ancien_nom, nouveau_nom)

            # Mettre à jour personnel.json
            update_personnel_json_class_name(week_folder, ancien_nom, nouveau_nom)

            # Afficher un message de confirmation
            messagebox.showinfo("Modification réussie", f"Le nom de la classe a été changé de '{ancien_nom}' à '{nouveau_nom}' et personnel.json a été mis à jour.")

            # Rafraîchir le dashboard
            refresh_main_dashboard(week_folder)

    def on_niveau_selected(selected_niveau):
        """Gère la sélection d'un niveau dans le dropdown."""
        if selected_niveau and selected_niveau != f"🎓 {classe_info.get('niveau', 'N/A')}":
            # Extraire le niveau sans l'emoji
            niveau_clean = selected_niveau.replace("🎓 ", "")
            ancien_niveau = classe_info.get('niveau', 'N/A')

            # Mettre à jour les données locales
            classe_info['niveau'] = niveau_clean
            if niveau_dropdown_ref:
                niveau_dropdown_ref.set(f"🎓 {niveau_clean}")

            # Mettre à jour le fichier Excel de l'école
            update_school_excel_file_niveau(week_folder, school_name, horaire, classe_info['nom_classe'], niveau_clean)

            # Rafraîchir le dashboard
            refresh_main_dashboard(week_folder)

    def on_intervenant_selected(selected_intervenant):
        """Gère la sélection d'un intervenant dans le dropdown."""
        if selected_intervenant and selected_intervenant != f"{'👨‍🏫' if type_intervenant == 'professeur' else '🎭'} {intervenant}":
            # Extraire le nom sans l'emoji
            intervenant_clean = selected_intervenant.replace(f"{'👨‍🏫' if type_intervenant == 'professeur' else '🎭'} ", "")
            ancien_intervenant = intervenant

            # Mettre à jour les données locales
            if intervenant_dropdown_ref:
                intervenant_dropdown_ref.set(f"{'👨‍🏫' if type_intervenant == 'professeur' else '🎭'} {intervenant_clean}")

            # Mettre à jour le fichier Excel de l'école
            update_school_excel_file_intervenant(week_folder, school_name, horaire, classe_info['nom_classe'], intervenant_clean)

            # Mettre à jour personnel.json
            update_personnel_json(week_folder, ancien_intervenant, intervenant_clean, classe_info['nom_classe'])

            # Mettre à jour matrix.xlsx pour tous les élèves de cette classe
            update_matrix_professor_for_class_students(week_folder, school_name, horaire, classe_info['nom_classe'], intervenant_clean)

            # Rafraîchir le dashboard
            refresh_main_dashboard(week_folder)

    def on_animateur_selected(selected_animateur):
        """Gère la sélection d'un animateur dans le dropdown."""
        if selected_animateur and selected_animateur != "Non Spécifié":
            # Extraire le nom sans l'emoji si nécessaire
            if selected_animateur.startswith("🎭 "):
                animateur_clean = selected_animateur.replace("🎭 ", "")
            else:
                animateur_clean = selected_animateur

            # Mettre à jour le fichier Excel de l'école pour l'animateur
            update_school_excel_file_animateur(week_folder, school_name, horaire, classe_info['nom_classe'], animateur_clean)

            # Mettre à jour personnel.json pour l'animateur
            update_personnel_json_animateur(week_folder, animateur_clean, classe_info['nom_classe'])

            # Rafraîchir le dashboard
            refresh_main_dashboard(week_folder)

            # Remettre le dropdown à afficher le nouvel animateur
            # Note: Le dropdown sera mis à jour lors du rafraîchissement de la fenêtre

    def select_student(student_index, card):
        """Sélectionne ou désélectionne un élève"""
        if student_index in selected_students:
            # Désélectionner
            selected_students.remove(student_index)
            card.configure(fg_color="white", border_width=0)
        else:
            # Sélectionner
            selected_students.add(student_index)
            card.configure(fg_color="#EFF6FF", border_width=2, border_color="#3B82F6")

    def make_student_card_selectable(card, student_index):
        """Rend une carte d'élève sélectionnable"""
        card.bind("<Button-1>", lambda e, idx=student_index, c=card: select_student(idx, c))

        # Rendre tous les widgets enfants sélectionnables récursivement
        def make_children_selectable(widget):
            for child in widget.winfo_children():
                child.bind("<Button-1>", lambda e, idx=student_index, c=card: select_student(idx, c))
                make_children_selectable(child)

        make_children_selectable(card)

    def get_current_animateur_from_excel(week_folder, school_name, horaire, classe_nom):
        """Récupère l'animateur actuel assigné à une classe spécifique depuis le fichier Excel de l'école."""
        if load_workbook is None or not week_folder:
            return "Non Spécifié"

        # Mapping des écoles vers les fichiers Excel
        school_file_mapping = {
            'A': 'ecole_a.xlsx',
            'B': 'ecole_b.xlsx',
            'C/CS': 'ECOLE_C_cours_standard.xlsx',
            'C/CI': 'ECOLE_C_cours_intensif.xlsx',
            'Morning': 'MORNING.xlsx',
            'Premium/CS': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'Premium/CI': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        if school_name not in school_file_mapping:
            return "Non Spécifié"

        excel_filename = school_file_mapping[school_name]
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            return "Non Spécifié"

        try:
            wb = load_workbook(excel_path)

            # Chercher la feuille correspondant à l'horaire
            target_sheet = None
            for sheet_name in wb.sheetnames:
                sheet_normalized = clean_horaire_name(sheet_name).lower().strip()
                horaire_normalized = horaire.lower().strip()
                if sheet_normalized == horaire_normalized or horaire_normalized in sheet_normalized:
                    target_sheet = wb[sheet_name]
                    break

            if not target_sheet:
                return "Non Spécifié"

            # Chercher la ligne de la classe
            classe_row = None
            for row_idx in range(2, target_sheet.max_row + 1):
                if str(target_sheet.cell(row=row_idx, column=1).value or '').strip() == classe_nom:
                    classe_row = row_idx
                    break

            if not classe_row:
                return "Non Spécifié"

            # Chercher la colonne animateur
            animateur_col = None
            for col_idx in range(1, target_sheet.max_column + 1):
                header_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                # Chercher spécifiquement "animateur", "anim", "Animateur" ou "Rôle" mais pas "intervenant" ou "prof"
                if (('animateur' in header_value or 'anim' in header_value or 'Animateur' in header_value or 'rôle' in header_value or 'role' in header_value) and
                    'intervenant' not in header_value and 'prof' not in header_value):
                    animateur_col = col_idx
                    break

            # Si pas trouvé, utiliser la colonne 4 par défaut (comme dans update_school_excel_file_animateur)
            if animateur_col is None:
                animateur_col = 4

            # Récupérer l'animateur de la classe
            animateur_value = str(target_sheet.cell(row=classe_row, column=animateur_col).value or '').strip()
            if animateur_value and animateur_value not in ['', 'nan', 'none', 'Non spécifié']:
                return animateur_value
            else:
                return "Non Spécifié"

        except Exception as e:
            print(f"Erreur lors de la récupération de l'animateur de la classe {classe_nom}: {e}")
            return "Non Spécifié"

    def get_students_from_class_excel(week_folder, school_name, horaire, classe_nom):
        """Récupère la liste des élèves d'une classe spécifique depuis le fichier Excel de l'école."""
        if load_workbook is None or not week_folder:
            return []

        # Mapping des écoles vers les fichiers Excel
        school_file_mapping = {
            'A': 'ecole_a.xlsx',
            'B': 'ecole_b.xlsx',
            'C/CS': 'ECOLE_C_cours_standard.xlsx',
            'C/CI': 'ECOLE_C_cours_intensif.xlsx',
            'Morning': 'MORNING.xlsx',
            'Premium/CS': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'Premium/CI': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        if school_name not in school_file_mapping:
            return []

        excel_filename = school_file_mapping[school_name]
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            return []

        try:
            wb = load_workbook(excel_path)

            # Chercher la feuille correspondant à l'horaire
            target_sheet = None
            for sheet_name in wb.sheetnames:
                sheet_normalized = clean_horaire_name(sheet_name).lower().strip()
                horaire_normalized = horaire.lower().strip()
                if sheet_normalized == horaire_normalized or horaire_normalized in sheet_normalized:
                    target_sheet = wb[sheet_name]
                    break

            if not target_sheet:
                return []

            # Chercher la ligne de la classe
            classe_row = None
            for row_idx in range(2, target_sheet.max_row + 1):
                if str(target_sheet.cell(row=row_idx, column=1).value or '').strip() == classe_nom:
                    classe_row = row_idx
                    break

            if not classe_row:
                return []

            # Récupérer les élèves de cette classe
            eleves_classe = []
            eleves_col = None

            # Trouver la colonne des élèves (priorité à "liste des élèves")
            # 1ère priorité : colonne contenant "liste" ET "élèves"
            for col_idx in range(1, target_sheet.max_column + 1):
                cell_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                if 'liste' in cell_value and ('élèves' in cell_value or 'eleves' in cell_value):
                    eleves_col = col_idx
                    break

            # 2ème priorité : colonne contenant juste "élèves" ou "eleves"
            if eleves_col is None:
                for col_idx in range(1, target_sheet.max_column + 1):
                    cell_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                    if ('élèves' in cell_value or 'eleves' in cell_value) and 'liste' not in cell_value:
                        eleves_col = col_idx
                        break

            # Récupérer les élèves de la classe
            if eleves_col:
                eleves_value = str(target_sheet.cell(row=classe_row, column=eleves_col).value or '').strip()
                if eleves_value and eleves_value not in ['', 'nan', 'none']:
                    # Diviser par virgule et nettoyer
                    eleves_classe = [nom.strip() for nom in eleves_value.split(',') if nom.strip()]

            return eleves_classe

        except Exception as e:
            print(f"Erreur lors de la récupération des élèves de la classe {classe_nom}: {e}")
            return []

    def update_matrix_class_name_for_students(week_folder, student_names, nouveau_nom_classe):
        """Met à jour le nom de la classe dans matrix.xlsx pour la liste d'élèves donnée."""
        if not load_workbook or not week_folder or not student_names:
            return

        matrix_path = os.path.join(week_folder, "matrix.xlsx")

        if not os.path.exists(matrix_path):
            return

        try:
            wb = load_workbook(matrix_path)
            ws = wb.active

            # Identifier les colonnes importantes
            stagiaire_col = None
            classe_col = None

            for col in range(1, ws.max_column + 1):
                col_name = str(ws.cell(row=1, column=col).value or '').strip()
                col_name_lower = col_name.lower()

                if 'stagiaire' in col_name_lower or 'nom' in col_name_lower or 'élève' in col_name_lower or 'eleve' in col_name_lower:
                    stagiaire_col = col
                elif ('classe' in col_name_lower or 'class' in col_name_lower) and not any(exclure in col_name_lower for exclure in ['cours 1', 'cours 2', 'cours 3', 'cours 4', '.1', '.2', 'arrivée', 'départ']):
                    classe_col = col

            if not stagiaire_col or not classe_col:
                print("ERREUR: Colonnes stagiaire ou classe non trouvées dans matrix.xlsx")
                return

            # Normaliser les noms d'élèves à mettre à jour
            eleves_normalises = {eleve.lower().strip() for eleve in student_names}

            # Parcourir toutes les lignes pour trouver les élèves à mettre à jour
            updated_count = 0

            for row_idx in range(2, ws.max_row + 1):
                eleve_nom_brut = str(ws.cell(row=row_idx, column=stagiaire_col).value or '').strip()
                eleve_nom_normalise = eleve_nom_brut.lower().strip()

                # Chercher une correspondance
                correspondance_trouvee = False
                for eleve_cible in eleves_normalises:
                    if eleve_cible in eleve_nom_normalise or eleve_nom_normalise in eleve_cible:
                        correspondance_trouvee = True
                        break

                if correspondance_trouvee:
                    # Mettre à jour le nom de la classe
                    ws.cell(row=row_idx, column=classe_col, value=nouveau_nom_classe)
                    updated_count += 1

            # Sauvegarder le fichier matrix
            wb.save(matrix_path)

            if updated_count > 0:
                print(f"Mise à jour matrix.xlsx: {updated_count} élève(s) assigné(s) à la classe '{nouveau_nom_classe}'")

        except Exception as e:
            print(f"Erreur lors de la mise à jour de matrix.xlsx: {e}")

    def update_school_excel_file_class_name(week_folder, school_name, horaire, ancien_nom, nouveau_nom):
        """Met à jour le nom d'une classe dans le fichier Excel de l'école."""
        if load_workbook is None or not week_folder:
            return

        # Mapping des écoles vers les fichiers Excel
        school_file_mapping = {
            'A': 'ecole_a.xlsx',
            'B': 'ecole_b.xlsx',
            'C/CS': 'ECOLE_C_cours_standard.xlsx',
            'C/CI': 'ECOLE_C_cours_intensif.xlsx',
            'Morning': 'MORNING.xlsx',
            'Premium/CS': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'Premium/CI': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        if school_name not in school_file_mapping:
            return

        excel_filename = school_file_mapping[school_name]
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            return

        # Récupérer les élèves de l'ancienne classe avant de la renommer
        eleves_classe = get_students_from_class_excel(week_folder, school_name, horaire, ancien_nom)

        try:
            wb = load_workbook(excel_path)

            # Chercher la feuille correspondant à l'horaire
            target_sheet = None
            for sheet_name in wb.sheetnames:
                sheet_normalized = clean_horaire_name(sheet_name).lower().strip()
                horaire_normalized = horaire.lower().strip()
                if sheet_normalized == horaire_normalized or horaire_normalized in sheet_normalized:
                    target_sheet = wb[sheet_name]
                    break

            if not target_sheet:
                return

            # Chercher la ligne de l'ancienne classe et mettre à jour le nom
            for row_idx in range(2, target_sheet.max_row + 1):
                if str(target_sheet.cell(row=row_idx, column=1).value or '').strip() == ancien_nom:
                    target_sheet.cell(row=row_idx, column=1, value=nouveau_nom)
                    wb.save(excel_path)
                    break

        except Exception as e:
            print(f"Erreur lors de la mise à jour du nom de classe dans {excel_filename}: {e}")

        # Mettre à jour matrix.xlsx pour tous les élèves de cette classe
        if eleves_classe:
            update_matrix_class_name_for_students(week_folder, eleves_classe, nouveau_nom)

    def update_school_excel_file_niveau(week_folder, school_name, horaire, classe_nom, nouveau_niveau):
        """Met à jour le niveau d'une classe dans le fichier Excel de l'école."""
        if load_workbook is None or not week_folder:
            return

        # Mapping des écoles vers les fichiers Excel
        school_file_mapping = {
            'A': 'ecole_a.xlsx',
            'B': 'ecole_b.xlsx',
            'C/CS': 'ECOLE_C_cours_standard.xlsx',
            'C/CI': 'ECOLE_C_cours_intensif.xlsx',
            'Morning': 'MORNING.xlsx',
            'Premium/CS': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'Premium/CI': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        if school_name not in school_file_mapping:
            return

        excel_filename = school_file_mapping[school_name]
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            return

        try:
            wb = load_workbook(excel_path)

            # Chercher la feuille correspondant à l'horaire
            target_sheet = None
            for sheet_name in wb.sheetnames:
                sheet_normalized = clean_horaire_name(sheet_name).lower().strip()
                horaire_normalized = horaire.lower().strip()
                if sheet_normalized == horaire_normalized or horaire_normalized in sheet_normalized:
                    target_sheet = wb[sheet_name]
                    break

            if not target_sheet:
                return

            # Chercher la ligne de la classe et mettre à jour la colonne niveau
            for row_idx in range(2, target_sheet.max_row + 1):
                if str(target_sheet.cell(row=row_idx, column=1).value or '').strip() == classe_nom:
                    # Chercher la colonne niveau par nom d'en-tête
                    niveau_col = None
                    for col_idx in range(1, target_sheet.max_column + 1):
                        header_value = str(target_sheet.cell(row=1, column=col_idx).value or '').strip()
                        if 'niveau' in header_value.lower() or 'level' in header_value.lower():
                            niveau_col = col_idx
                            break

                    # Si on trouve la colonne niveau, la mettre à jour
                    if niveau_col:
                        target_sheet.cell(row=row_idx, column=niveau_col, value=nouveau_niveau)
                        wb.save(excel_path)
                    break

        except Exception as e:
            print(f"Erreur lors de la mise à jour du niveau dans {excel_filename}: {e}")

    def update_school_excel_file_intervenant(week_folder, school_name, horaire, classe_nom, nouvel_intervenant):
        """Met à jour l'intervenant d'une classe dans le fichier Excel de l'école."""
        if load_workbook is None or not week_folder:
            return

        # Mapping des écoles vers les fichiers Excel
        school_file_mapping = {
            'A': 'ecole_a.xlsx',
            'B': 'ecole_b.xlsx',
            'C/CS': 'ECOLE_C_cours_standard.xlsx',
            'C/CI': 'ECOLE_C_cours_intensif.xlsx',
            'Morning': 'MORNING.xlsx',
            'Premium/CS': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'Premium/CI': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        if school_name not in school_file_mapping:
            return

        excel_filename = school_file_mapping[school_name]
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            return

        try:
            wb = load_workbook(excel_path)

            # Chercher la feuille correspondant à l'horaire
            target_sheet = None
            for sheet_name in wb.sheetnames:
                sheet_normalized = clean_horaire_name(sheet_name).lower().strip()
                horaire_normalized = horaire.lower().strip()
                if sheet_normalized == horaire_normalized or horaire_normalized in sheet_normalized:
                    target_sheet = wb[sheet_name]
                    break

            if not target_sheet:
                return

            # Chercher la ligne de la classe et mettre à jour la colonne intervenant (colonne 2 ou 3 selon le format)
            for row_idx in range(2, target_sheet.max_row + 1):
                if str(target_sheet.cell(row=row_idx, column=1).value or '').strip() == classe_nom:
                    # Essayer de trouver la colonne intervenant (généralement colonne 2 ou 3)
                    for col_idx in [2, 3]:
                        header_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                        if ('intervenant' in header_value or 'prof' in header_value or 'animateur' in header_value or
                            'Animateur' in header_value or 'rôle' in header_value or 'role' in header_value):
                            target_sheet.cell(row=row_idx, column=col_idx, value=nouvel_intervenant)
                            wb.save(excel_path)
                            return

        except Exception as e:
            print(f"Erreur lors de la mise à jour de l'intervenant dans {excel_filename}: {e}")

    def update_school_excel_file_animateur(week_folder, school_name, horaire, classe_nom, nouvel_animateur):
        """Met à jour l'animateur d'une classe dans le fichier Excel de l'école."""
        if load_workbook is None or not week_folder:
            return

        # Mapping des écoles vers les fichiers Excel
        school_file_mapping = {
            'A': 'ecole_a.xlsx',
            'B': 'ecole_b.xlsx',
            'C/CS': 'ECOLE_C_cours_standard.xlsx',
            'C/CI': 'ECOLE_C_cours_intensif.xlsx',
            'Morning': 'MORNING.xlsx',
            'Premium/CS': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'Premium/CI': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        if school_name not in school_file_mapping:
            return

        excel_filename = school_file_mapping[school_name]
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            return

        try:
            wb = load_workbook(excel_path)

            # Chercher la feuille correspondant à l'horaire
            target_sheet = None
            for sheet_name in wb.sheetnames:
                sheet_normalized = clean_horaire_name(sheet_name).lower().strip()
                horaire_normalized = horaire.lower().strip()
                if sheet_normalized == horaire_normalized or horaire_normalized in sheet_normalized:
                    target_sheet = wb[sheet_name]
                    break

            if not target_sheet:
                return

            # Chercher la ligne de la classe
            classe_row = None
            for row_idx in range(2, target_sheet.max_row + 1):
                if str(target_sheet.cell(row=row_idx, column=1).value or '').strip() == classe_nom:
                    classe_row = row_idx
                    break

            if not classe_row:
                return

            # Chercher la colonne animateur (différente de la colonne intervenant/professeur)
            animateur_col = None
            for col_idx in range(1, target_sheet.max_column + 1):
                header_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                # Chercher spécifiquement "animateur", "anim", "Animateur" ou "Rôle" mais pas "intervenant" ou "prof"
                if (('animateur' in header_value or 'anim' in header_value or 'Animateur' in header_value or 'rôle' in header_value or 'role' in header_value) and
                    'intervenant' not in header_value and 'prof' not in header_value):
                    animateur_col = col_idx
                    break

            # Si pas trouvé, essayer de trouver une colonne vide ou créer une logique pour la colonne 4
            if animateur_col is None:
                # Pour les écoles A et B, on peut utiliser la colonne 4 pour l'animateur
                animateur_col = 4

            # Mettre à jour la colonne animateur
            target_sheet.cell(row=classe_row, column=animateur_col, value=nouvel_animateur)
            wb.save(excel_path)
            print(f"Animateur '{nouvel_animateur}' assigné à la classe '{classe_nom}' dans {excel_filename}")

        except Exception as e:
            print(f"Erreur lors de la mise à jour de l'animateur dans {excel_filename}: {e}")

    def update_personnel_json(week_folder, ancien_intervenant, nouvel_intervenant, classe_nom):
        """Met à jour personnel.json pour transférer la classe d'un intervenant à un autre."""
        if not week_folder:
            return

        personnel_path = os.path.join(week_folder, "personnel.json")

        if not os.path.exists(personnel_path):
            return

        try:
            with open(personnel_path, 'r', encoding='utf-8') as f:
                personnel_data = json.load(f)

            # Déterminer le type d'intervenant (professeur ou animateur)
            type_intervenant = "professeurs"  # Par défaut

            # Trouver l'ancien intervenant et retirer la classe
            for intervenant in personnel_data.get("professeurs", []):
                if intervenant.get("nom") == ancien_intervenant:
                    if "classes" in intervenant and classe_nom in intervenant["classes"]:
                        intervenant["classes"].remove(classe_nom)
                    break
            else:
                # Si pas trouvé dans professeurs, chercher dans animateurs
                for intervenant in personnel_data.get("animateurs", []):
                    if intervenant.get("nom") == ancien_intervenant:
                        if "classes" in intervenant and classe_nom in intervenant["classes"]:
                            intervenant["classes"].remove(classe_nom)
                        type_intervenant = "animateurs"
                        break

            # Ajouter la classe au nouvel intervenant
            found = False
            for intervenant in personnel_data.get(type_intervenant, []):
                if intervenant.get("nom") == nouvel_intervenant:
                    if "classes" not in intervenant:
                        intervenant["classes"] = []
                    if classe_nom not in intervenant["classes"]:
                        intervenant["classes"].append(classe_nom)
                    found = True
                    break

            # Si le nouvel intervenant n'existe pas, on pourrait le créer, mais pour l'instant on skip
            if not found:
                print(f"Avertissement: Nouvel intervenant '{nouvel_intervenant}' non trouve dans personnel.json")

            # Sauvegarder les modifications
            with open(personnel_path, 'w', encoding='utf-8') as f:
                json.dump(personnel_data, f, indent=4, ensure_ascii=False)

        except Exception as e:
            print(f"Erreur lors de la mise à jour de personnel.json: {e}")

    def update_personnel_json_animateur(week_folder, nouvel_animateur, classe_nom):
        """Met à jour personnel.json pour assigner un animateur à une classe."""
        if not week_folder:
            return

        personnel_path = os.path.join(week_folder, "personnel.json")

        if not os.path.exists(personnel_path):
            return

        try:
            with open(personnel_path, 'r', encoding='utf-8') as f:
                personnel_data = json.load(f)

            # Chercher l'animateur dans la liste des animateurs
            found = False
            for animateur in personnel_data.get("animateurs", []):
                if animateur.get("nom") == nouvel_animateur:
                    if "classes" not in animateur:
                        animateur["classes"] = []
                    if classe_nom not in animateur["classes"]:
                        animateur["classes"].append(classe_nom)
                    found = True
                    break

            # Si l'animateur n'existe pas, on pourrait le créer, mais pour l'instant on skip
            if not found:
                print(f"Avertissement: Animateur '{nouvel_animateur}' non trouvé dans personnel.json")

            # Sauvegarder les modifications
            with open(personnel_path, 'w', encoding='utf-8') as f:
                json.dump(personnel_data, f, indent=4, ensure_ascii=False)

        except Exception as e:
            print(f"Erreur lors de la mise à jour de personnel.json pour l'animateur: {e}")

    def get_current_selected_week():
        """Récupère le numéro de la semaine actuellement sélectionnée."""
        try:
            # Essayer d'importer la variable globale depuis fenetre_principale
            import sys
            main_module = sys.modules.get('fenetre_principale')
            if main_module and hasattr(main_module, 'selected_week'):
                week_label = main_module.selected_week.get()  # ex: "Semaine 1"
                try:
                    week_num = week_label.split()[-1]  # Extraire "1"
                    return int(week_num)
                except (ValueError, IndexError):
                    return 1  # Valeur par défaut
            return 1  # Valeur par défaut si pas trouvé
        except Exception:
            return 1  # Valeur par défaut en cas d'erreur

    def update_matrix_professor_for_class_students(week_folder, school_name, horaire, classe_nom, nouveau_prof):
        """
        Met à jour la colonne "Prof" dans matrix.xlsx pour tous les élèves d'une classe donnée.
        Utilise la semaine sélectionnée dans l'interface principale.
        """
        if not load_workbook:
            return

        # Étape 1: Récupérer la liste des élèves de cette classe depuis le fichier Excel de l'école
        eleves_classe = get_students_from_class_excel(week_folder, school_name, horaire, classe_nom)

        if not eleves_classe:
            print(f"Aucun élève trouvé dans la classe '{classe_nom}'")
            return

        # Étape 2: Construire le chemin du matrix.xlsx de la semaine sélectionnée
        week_num = get_current_selected_week()
        script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # Remonter d'un niveau depuis classes_details.py
        matrix_path = os.path.join(script_dir, f"semaine_{week_num}", "matrix.xlsx")

        if not os.path.exists(matrix_path):
            print(f"Erreur: Le fichier matrix.xlsx de la semaine {week_num} n'existe pas: {matrix_path}")
            return

        try:
            wb = load_workbook(matrix_path)
            ws = wb.active

            # Identifier les colonnes importantes
            stagiaire_col = None
            prof_col = None

            for col in range(1, ws.max_column + 1):
                col_name = str(ws.cell(row=1, column=col).value or '').strip()
                col_name_lower = col_name.lower()

                if 'stagiaire' in col_name_lower or 'nom' in col_name_lower or 'élève' in col_name_lower or 'eleve' in col_name_lower:
                    stagiaire_col = col
                elif 'prof' in col_name_lower and not any(exclure in col_name_lower for exclure in ['cours 1', 'cours 2', 'arrivée', 'départ']):
                    prof_col = col

            if not stagiaire_col:
                print("ERREUR: Aucune colonne stagiaire trouvée dans matrix.xlsx")
                return

            if not prof_col:
                print("ERREUR: Aucune colonne prof trouvée dans matrix.xlsx")
                return

            # Normaliser les noms d'élèves à mettre à jour
            eleves_normalises = {eleve.lower().strip() for eleve in eleves_classe}

            # Parcourir toutes les lignes pour trouver les élèves à mettre à jour
            updated_count = 0

            for row_idx in range(2, ws.max_row + 1):
                eleve_nom_brut = str(ws.cell(row=row_idx, column=stagiaire_col).value or '').strip()
                eleve_nom_normalise = eleve_nom_brut.lower().strip()

                # Chercher une correspondance
                if any(eleve_nom_normalise in eleve_cible or eleve_cible in eleve_nom_normalise for eleve_cible in eleves_normalises):
                    # Mettre à jour le professeur
                    ws.cell(row=row_idx, column=prof_col, value=nouveau_prof)
                    updated_count += 1

            # Sauvegarder le fichier matrix
            wb.save(matrix_path)

            if updated_count > 0:
                print(f"Matrix.xlsx (semaine {week_num}) mis à jour: {updated_count} élève(s) de la classe '{classe_nom}' assigné(s) au prof '{nouveau_prof}'")
            else:
                print(f"Aucun élève de la classe '{classe_nom}' trouvé dans matrix.xlsx semaine {week_num}")

        except Exception as e:
            print(f"Erreur lors de la mise à jour de matrix.xlsx pour les élèves de la classe: {e}")

    def update_personnel_json_class_name(week_folder, ancien_nom_classe, nouveau_nom_classe):
        """Met à jour personnel.json pour changer le nom d'une classe."""
        if not week_folder:
            return

        personnel_path = os.path.join(week_folder, "personnel.json")

        if not os.path.exists(personnel_path):
            return

        try:
            with open(personnel_path, 'r', encoding='utf-8') as f:
                personnel_data = json.load(f)

            # Mettre à jour le nom de classe dans tous les professeurs
            for intervenant in personnel_data.get("professeurs", []):
                if "classes" in intervenant:
                    if ancien_nom_classe in intervenant["classes"]:
                        intervenant["classes"].remove(ancien_nom_classe)
                        intervenant["classes"].append(nouveau_nom_classe)

            # Mettre à jour le nom de classe dans tous les animateurs
            for intervenant in personnel_data.get("animateurs", []):
                if "classes" in intervenant:
                    if ancien_nom_classe in intervenant["classes"]:
                        intervenant["classes"].remove(ancien_nom_classe)
                        intervenant["classes"].append(nouveau_nom_classe)

            # Sauvegarder les modifications
            with open(personnel_path, 'w', encoding='utf-8') as f:
                json.dump(personnel_data, f, indent=4, ensure_ascii=False)

        except Exception as e:
            print(f"Erreur lors de la mise à jour du nom de classe dans personnel.json: {e}")

    def refresh_main_dashboard(week_folder):
        """Rafraîchit le dashboard des classes dans fenetre_principale.py avec un message de chargement."""
        if not week_folder or not refresh_callback:
            return

        # Afficher un message de chargement
        loading_popup = show_loading_popup("Mise à jour du dashboard...")

        def do_refresh():
            try:
                # Utiliser le callback passé en paramètre pour rafraîchir le dashboard
                refresh_callback()
                print("Dashboard des classes mis a jour")
            finally:
                # Fermer le message de chargement
                if loading_popup and loading_popup.winfo_exists():
                    loading_popup.destroy()

        # Lancer le rafraîchissement après un court délai
        detail_window.after(100, do_refresh)

    def create_safe_toplevel(width=300, height=120, title="", resizable=False, overrideredirect=False, transient=True):
        """Crée un CTkToplevel de manière sécurisée en vérifiant si detail_window existe encore."""
        parent_window = detail_window if detail_window and detail_window.winfo_exists() else None
        popup = ctk.CTkToplevel(parent_window)
        popup.title(title)
        popup.geometry(f"{width}x{height}")
        popup.resizable(resizable, resizable)
        popup.attributes("-topmost", True)
        if overrideredirect:
            popup.overrideredirect(True)
        else:
            popup.focus_force()

        # Centrer sur la fenêtre parente si elle existe, sinon sur l'écran
        if parent_window and parent_window.winfo_exists():
            x = parent_window.winfo_rootx() + (parent_window.winfo_width() // 2) - (width // 2)
            y = parent_window.winfo_rooty() + (parent_window.winfo_height() // 2) - (height // 2)
            if transient:
                popup.transient(parent_window)
        else:
            # Centrer sur l'écran
            screen_width = popup.winfo_screenwidth()
            screen_height = popup.winfo_screenheight()
            x = (screen_width // 2) - (width // 2)
            y = (screen_height // 2) - (height // 2)
        popup.geometry(f"{width}x{height}+{x}+{y}")

        return popup

    def show_loading_popup(message):
        """Affiche un popup de chargement."""
        loading_popup = create_safe_toplevel(300, 120, "", False, True)

        # Style moderne
        main_frame = ctk.CTkFrame(
            loading_popup,
            fg_color="#FFFFFF",
            corner_radius=15,
            border_width=1,
            border_color="#CBD5E1"
        )
        main_frame.pack(fill="both", expand=True, padx=2, pady=2)

        content_frame = ctk.CTkFrame(main_frame, fg_color="#F8FAFC", corner_radius=10)
        content_frame.pack(fill="both", expand=True, padx=8, pady=8)

        # Message
        message_label = ctk.CTkLabel(
            content_frame,
            text=message,
            font=("Segoe UI", 12),
            text_color="#1E293B"
        )
        message_label.pack(pady=(15, 10))

        # Barre de progression
        progress_bar = ctk.CTkProgressBar(
            content_frame,
            width=200,
            height=4,
            corner_radius=10,
            fg_color="#F1F5F9",
            progress_color="#3B82F6",
            mode="indeterminate"
        )
        progress_bar.pack(pady=(0, 15))
        progress_bar.start()

        return loading_popup

    def clean_horaire_name(sheet_name):
        """Nettoie le nom de la feuille pour extraire le nom d'horaire de manière cohérente."""
        # Cette fonction doit être identique à celle utilisée dans fenetre_principale.py et Assignation des Niveaux.py
        sheet_lower = sheet_name.lower()
        type_intervenant = "animateur" if "animateur" in sheet_lower else "professeur"

        # Liste complète des mots à supprimer (avec variations)
        words_to_remove = [
            "animateur", "Animateur", "anim", "Anim",
            "professeur", "Professeur", "prof", "Prof",
            "Rôle", "rôle", "role", "Role"
        ]

        # Nettoyer le nom en supprimant tous les mots-clés d'intervenants
        horaire = sheet_name
        for word in words_to_remove:
            horaire = horaire.replace(word, "")

        # Nettoyer les espaces multiples et supprimer les espaces au début/fin
        horaire = " ".join(horaire.split()).strip()

        # Si le résultat est vide, utiliser le nom original
        return horaire or sheet_name

    # === 1. HEADER (BLANC) ===
    header_bar = ctk.CTkFrame(detail_window, fg_color="white", height=80, corner_radius=0)
    header_bar.pack(fill="x", side="top", padx=0, pady=0)
    header_bar.pack_propagate(False)

    header_content = ctk.CTkFrame(header_bar, fg_color="transparent")
    header_content.pack(fill="both", expand=True, padx=40)

    # Titre cliquable : École (colorée) - Classe (cliquable)
    title_frame = ctk.CTkFrame(header_content, fg_color="transparent")
    title_frame.pack(side="left")

    # Badge École avec couleur appropriée
    school_color_code = school_colors.get(school_name, "#6b7280")  # Couleur par défaut grise
    school_label = ctk.CTkLabel(title_frame, text="Ecole : " + school_name, font=("Segoe UI", 24, "bold"),
                               fg_color=school_color_code, text_color="white", corner_radius=8,
                               padx=15, pady=6)
    school_label.pack(side="left")

    # Séparateur
    ctk.CTkLabel(title_frame, text=" - ", font=("Segoe UI", 24, "bold"), text_color="#0F172A").pack(side="left", padx=5)

    # Nom de classe (non cliquable)
    class_name = classe_info.get('nom_classe', 'Sans Nom')
    class_label = ctk.CTkLabel(title_frame, text="Classe : " + class_name, font=("Segoe UI", 24, "bold"),
                              text_color="#0F172A")
    class_label.pack(side="left")
    class_label_ref = class_label

   # Design : Badge avec bordure fine
    edit_class_btn = ctk.CTkButton(
        title_frame,
        text="Modifier\nNom",          # Texte + Icône pour plus de clarté
        width=60,
        height=45,
        corner_radius=6,
        font=("Segoe UI", 11, "bold"),
        fg_color="white",
        text_color="#475569",
        border_color="#E2E8F0",
        border_width=1,
        hover_color="#F8FAFC",
        command=edit_class_name
    )
    edit_class_btn.pack(side="left", padx=(12, 0))

    # Badges d'infos à droite
    right_side = ctk.CTkFrame(header_content, fg_color="transparent")
    right_side.pack(side="right")

    '''# Badge Horaire (non cliquable)
    ctk.CTkLabel(right_side, text=f"🕒 {horaire}", font=("Segoe UI", 12, "bold"),
                 fg_color="#F0FDF4", text_color="#166534", corner_radius=8, padx=15, pady=6).pack(side="left", padx=6)
    '''

    # Section Niveau
    niveau_frame = ctk.CTkFrame(right_side, fg_color="transparent")
    niveau_frame.pack(side="left", padx=6)
    ctk.CTkLabel(niveau_frame, text="Niveau", font=("Segoe UI", 9, "bold"), text_color="#6B7280").pack(anchor="center", pady=(0, 2))
    niveau_values = [f"🎓 {niv}" for niv in liste_niveaux]
    niveau_dropdown = ctk.CTkOptionMenu(
        niveau_frame,
        values=niveau_values,
        command=lambda selected: on_niveau_selected(selected),
        font=("Segoe UI", 11, "bold"),
        height=32,
        width=120,
        fg_color="#EFF6FF",
        button_color="#EFF6FF",
        button_hover_color="#DBEAFE",
        text_color="#1E40AF",
        dropdown_fg_color="#EFF6FF",
        dropdown_hover_color="#DBEAFE",
        dropdown_text_color="#1E40AF"
    )
    niveau_dropdown.set(f"🎓 {classe_info.get('niveau', 'N/A')}")
    niveau_dropdown.pack()
    niveau_dropdown_ref = niveau_dropdown

    # Section Professeur
    prof_frame = ctk.CTkFrame(right_side, fg_color="transparent")
    prof_frame.pack(side="left", padx=6)
    ctk.CTkLabel(prof_frame, text="Professeur", font=("Segoe UI", 9, "bold"), text_color="#6B7280").pack(anchor="center", pady=(0, 2))
    if type_intervenant == "professeur":
        interv_values = [f"👨‍🏫 {interv}" for interv in liste_intervenants_profs]
    else:
        interv_values = [f"🎭 {interv}" for interv in liste_intervenants_anims]

    intervenant_dropdown = ctk.CTkOptionMenu(
        prof_frame,
        values=interv_values,
        command=lambda selected: on_intervenant_selected(selected),
        font=("Segoe UI", 11, "bold"),
        height=32,
        width=140,
        fg_color="#FAF5FF",
        button_color="#FAF5FF",
        button_hover_color="#F3E8FF",
        text_color="#6B21A8",
        dropdown_fg_color="#FAF5FF",
        dropdown_hover_color="#F3E8FF",
        dropdown_text_color="#6B21A8"
    )
    intervenant_dropdown.set(f"{'👨‍🏫' if type_intervenant == 'professeur' else '🎭'} {intervenant}")
    intervenant_dropdown.pack()
    intervenant_dropdown_ref = intervenant_dropdown

    # Section Animateur (pour les écoles A et B uniquement)
    if school_name in ["A", "B"]:
        anim_frame = ctk.CTkFrame(right_side, fg_color="transparent")
        anim_frame.pack(side="left", padx=6)
        ctk.CTkLabel(anim_frame, text="Animateur", font=("Segoe UI", 9, "bold"), text_color="#6B7280").pack(anchor="center", pady=(0, 2))

        # Récupérer l'animateur actuel assigné à la classe
        current_animateur = get_current_animateur_from_excel(week_folder, school_name, horaire, classe_info['nom_classe'])

        # Dropdown pour assigner un animateur
        anim_values = ["Non Spécifié"] + [f"🎭 {anim}" for anim in liste_intervenants_anims]
        anim_dropdown = ctk.CTkOptionMenu(
            anim_frame,
            values=anim_values,
            command=lambda selected: on_animateur_selected(selected),
            font=("Segoe UI", 10, "bold"),
            height=32,
            width=120,
            fg_color="#A7F3D0",  # Couleur plus douce (vert clair)
            button_color="#A7F3D0",
            button_hover_color="#86EFAC",
            text_color="#065F46",
            dropdown_fg_color="#F0FDF4",
            dropdown_hover_color="#DCFCE7",
            dropdown_text_color="#166534"
        )
        # Afficher l'animateur actuel ou "Non Spécifié"
        display_text = f"🎭 {current_animateur}" if current_animateur != "Non Spécifié" else "Non Spécifié"
        anim_dropdown.set(display_text)
        anim_dropdown.pack()

    # === 2. FRAME CENTRALE GRISE (Bords arrondis) ===
    # Cette frame contient uniquement la liste des élèves
    main_grey_frame = ctk.CTkFrame(detail_window, fg_color="#F1F5F9", corner_radius=25)
    main_grey_frame.pack(fill="both", expand=True, padx=30, pady=10)

    # -- Titre et Compteur à l'intérieur de la frame grise --
    eleves = classe_info.get('eleves', [])
    nb_eleves = len(eleves)

    title_container = ctk.CTkFrame(main_grey_frame, fg_color="transparent")
    title_container.pack(fill="x", padx=35, pady=(25, 0))

    ctk.CTkLabel(title_container, text="LISTE DES ÉLÈVES", font=("Segoe UI", 13, "bold"),
                 text_color="#64748B").pack(side="left")

    # Compteur avec couleur rouge si plus de 15 élèves
    compteur_color = "#DC2626" if nb_eleves > 15 else school_color  # Rouge si dépassement
    compteur_text_color = "white" if nb_eleves <= 15 else "white"  # Toujours blanc
    eleves_counter_label = ctk.CTkLabel(title_container, text=f"{nb_eleves}/15", font=("Segoe UI", 14, "bold"),
                 text_color=compteur_text_color, fg_color=compteur_color, corner_radius=6, padx=10, pady=2)
    eleves_counter_label.pack(side="left", padx=15)
    eleves_counter_label_ref = eleves_counter_label  # Garder la référence pour mise à jour

    # Boutons d'actions alignés à droite
    buttons_frame = ctk.CTkFrame(title_container, fg_color="transparent")
    buttons_frame.pack(side="right")

    # Bouton Supprimer de la classe
    delete_btn = ctk.CTkButton(
        buttons_frame,
        text="🗑️ Supprimer de la classe",
        width=160,
        height=32,
        font=("Segoe UI", 10, "bold"),
        fg_color="#EF4444",
        hover_color="#DC2626",
        text_color="white",
        corner_radius=6,
        command=lambda: remove_students_from_class()
    )
    delete_btn.pack(side="left", padx=(0, 8))

    # Bouton Assigner à une nouvelle classe
    assign_btn = ctk.CTkButton(
        buttons_frame,
        text="📚 Assigner à une nouvelle classe",
        width=180,
        height=32,
        font=("Segoe UI", 10, "bold"),
        fg_color="#3B82F6",
        hover_color="#2563EB",
        text_color="white",
        corner_radius=6,
        command=lambda: assign_students_to_new_class()
    )
    assign_btn.pack(side="left", padx=(0, 8))

    # Bouton Infos de l'élève
    info_btn = ctk.CTkButton(
        buttons_frame,
        text="ℹ️ Infos de l'élève",
        width=140,
        height=32,
        font=("Segoe UI", 10, "bold"),
        fg_color="#6B7280",
        hover_color="#4B5563",
        text_color="white",
        corner_radius=6,
        command=lambda: show_student_info()
    )
    info_btn.pack(side="left")

    # -- Grille des élèves avec scroll --
    scrollable_container = ctk.CTkScrollableFrame(main_grey_frame, fg_color="transparent")
    scrollable_container.pack(fill="both", expand=True, padx=25, pady=20)

    # Frame interne pour la grille
    grid_container = ctk.CTkFrame(scrollable_container, fg_color="transparent")
    grid_container.pack(fill="both", expand=True)
    grid_container.grid_columnconfigure((0, 1, 2, 3, 4), weight=1, pad=12)

    def generate_student_grid():
        """Génère la grille des élèves"""
        # Vider la grille existante
        for widget in grid_container.winfo_children():
            widget.destroy()

        # Vider les listes de références
        student_cards.clear()
        selected_students.clear()

        # Afficher tous les élèves (pas de limite à 15)
        current_nb_eleves = len(eleves)
        for i in range(current_nb_eleves):
            e = eleves[i]
            row, col = i // 5, i % 5

            # Carte blanche
            card = ctk.CTkFrame(grid_container, fg_color="white", corner_radius=12, border_width=0)
            card.grid(row=row, column=col, padx=8, pady=8, sticky="nsew")

            # Stocker la référence de la carte
            student_cards.append(card)

            # Nom de l'élève
            name_label = ctk.CTkLabel(card, text=e.get('nom').upper(), font=("Segoe UI", 14, "bold"),
                         text_color="#1E293B", wraplength=170)
            name_label.pack(pady=(18, 5), padx=10)

            # Sous-conteneur pour les mini-badges
            b_frame = ctk.CTkFrame(card, fg_color="transparent")
            b_frame.pack(pady=(5, 18))

            # Badge Niveau
            niv = e.get('niveau', 'A0')
            n_color = map_colors.get(niv, "#E2E8F0")
            level_badge = ctk.CTkLabel(b_frame, text=niv, font=("Segoe UI", 10, "bold"),
                         fg_color=n_color, text_color="#1E293B", corner_radius=5, padx=8, pady=4)
            level_badge.pack(side="left")

            # Badge CI (Bleu)
            if e.get('ci', False):
                ci_badge = ctk.CTkLabel(b_frame, text="CI", font=("Segoe UI", 10, "bold"),
                             fg_color="#2563EB", text_color="white", corner_radius=5, padx=8, pady=4)
                ci_badge.pack(side="left", padx=(6, 0))

            # Rendre la carte sélectionnable
            make_student_card_selectable(card, i)

    # Générer la grille initiale
    generate_student_grid()

    def remove_students_from_class():
        """Supprime les élèves sélectionnés de la classe"""
        if not selected_students:
            # Afficher un message d'erreur si aucun élève n'est sélectionné
            error_popup = create_safe_toplevel(350, 150, "Erreur")
            error_popup.lift()
            error_popup.after(50, lambda: [error_popup.focus_force(), error_popup.grab_set()])

            # Centrer sur la fenêtre principale
            x = detail_window.winfo_rootx() + (detail_window.winfo_width() // 2) - 175
            y = detail_window.winfo_rooty() + (detail_window.winfo_height() // 2) - 75

            ctk.CTkLabel(error_popup, text="Aucun élève sélectionné", font=("Segoe UI", 14, "bold")).pack(pady=(20, 10))
            ctk.CTkLabel(error_popup, text="Veuillez sélectionner au moins un élève.", font=("Segoe UI", 11)).pack(pady=(0, 20))

            ctk.CTkButton(error_popup, text="OK", command=error_popup.destroy, width=100).pack()
            return

        # Demander confirmation
        confirm_popup = create_safe_toplevel(400, 180, "Confirmation")

        # Centrer sur la fenêtre principale
        x = detail_window.winfo_rootx() + (detail_window.winfo_width() // 2) - 200
        y = detail_window.winfo_rooty() + (detail_window.winfo_height() // 2) - 90
        confirm_popup.geometry(f"400x180+{x}+{y}")

        # S'assurer que la fenêtre apparaît au premier plan
        confirm_popup.lift()
        confirm_popup.after(50, lambda: [confirm_popup.focus_force(), confirm_popup.grab_set()])

        selected_count = len(selected_students)
        ctk.CTkLabel(confirm_popup, text=f"Supprimer {selected_count} élève{'s' if selected_count > 1 else ''} de la classe ?",
                    font=("Segoe UI", 14, "bold")).pack(pady=(20, 10))
        ctk.CTkLabel(confirm_popup, text="Cette action est irréversible.", font=("Segoe UI", 11)).pack(pady=(0, 20))

        def confirm_delete():
            # Récupérer les informations des élèves avant suppression pour mettre à jour matrix.xlsx
            eleves_a_supprimer = []
            for idx in sorted(selected_students):
                if idx < len(eleves):
                    eleve = eleves[idx]
                    eleves_a_supprimer.append({
                        'nom': eleve.get('nom', ''),
                        'school_name': school_name,
                        'horaire': horaire,
                        'classe_nom': classe_info.get('nom_classe', '')
                    })

            # Supprimer les élèves sélectionnés (dans l'ordre inverse pour éviter les décalages d'indices)
            for idx in sorted(selected_students, reverse=True):
                if idx < len(eleves):
                    eleves.pop(idx)

            # Mettre à jour le nombre d'élèves dans classe_info
            classe_info['nb_eleves'] = len(eleves)

            # Vider la sélection
            selected_students.clear()

            # Mettre à jour l'affichage de la grille des élèves
            refresh_student_grid()

            # Mettre à jour matrix.xlsx pour les élèves supprimés
            if eleves_a_supprimer:
                update_matrix_after_student_removal(week_folder, eleves_a_supprimer)

            # Mettre à jour le fichier Excel de l'école (supprimer la classe si elle devient vide)
            update_school_excel_after_student_removal(week_folder, school_name, horaire, classe_info.get('nom_classe', ''), len(eleves))

            # Rafraîchir le dashboard principal
            if refresh_callback:
                refresh_callback()

            confirm_popup.destroy()

        def cancel_delete():
            confirm_popup.destroy()

        btn_frame = ctk.CTkFrame(confirm_popup, fg_color="transparent")
        btn_frame.pack(pady=(0, 20))

        ctk.CTkButton(btn_frame, text="Annuler", command=cancel_delete, fg_color="#6B7280", width=100).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="Supprimer", command=confirm_delete, fg_color="#EF4444", width=100).pack(side="left")

    def create_or_update_class_in_excel(week_folder, school_name, horaire, new_class_name, eleves_list):
        """
        Crée ou met à jour une classe dans le fichier Excel de l'école.
        """
        if load_workbook is None or not week_folder:
            return

        # Mapping des écoles vers les fichiers Excel
        school_file_mapping = {
            'A': 'ecole_a.xlsx',
            'B': 'ecole_b.xlsx',
            'C/CS': 'ECOLE_C_cours_standard.xlsx',
            'C/CI': 'ECOLE_C_cours_intensif.xlsx',
            'Morning': 'MORNING.xlsx',
            'Premium/CS': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'Premium/CI': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        if school_name not in school_file_mapping:
            return

        excel_filename = school_file_mapping[school_name]
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            return

        try:
            wb = load_workbook(excel_path)

            # Chercher la feuille correspondant à l'horaire
            target_sheet = None
            for sheet_name in wb.sheetnames:
                sheet_normalized = clean_horaire_name(sheet_name).lower().strip()
                horaire_normalized = horaire.lower().strip()
                if sheet_normalized == horaire_normalized or horaire_normalized in sheet_normalized:
                    target_sheet = wb[sheet_name]
                    break

            if not target_sheet:
                return

            # Chercher si la classe existe déjà dans cette feuille
            classe_row = None
            classe_col = None

            for row_idx in range(2, target_sheet.max_row + 1):
                for col_idx in range(1, target_sheet.max_column + 1):
                    cell_value = str(target_sheet.cell(row=row_idx, column=col_idx).value or '').strip()
                    if cell_value == new_class_name:
                        classe_row = row_idx
                        classe_col = col_idx
                        break
                if classe_row:
                    break

            # Chercher la colonne des élèves
            eleves_col = None
            for col_idx in range(1, target_sheet.max_column + 1):
                cell_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                if ('élèves' in cell_value or 'eleves' in cell_value) and 'liste' not in cell_value:
                    eleves_col = col_idx
                    break

            # Si pas trouvé, essayer "liste des élèves"
            if eleves_col is None:
                for col_idx in range(1, target_sheet.max_column + 1):
                    cell_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                    if 'liste' in cell_value and ('élèves' in cell_value or 'eleves' in cell_value):
                        eleves_col = col_idx
                        break

            # Si toujours pas trouvé, utiliser la colonne 5 comme fallback
            if eleves_col is None:
                eleves_col = 5

            if classe_row:
                # La classe existe déjà, mettre à jour la liste d'élèves
                eleves_noms = [eleve['nom'] for eleve in eleves_list if eleve['nom']]
                eleves_text = ', '.join(eleves_noms)
                target_sheet.cell(row=classe_row, column=eleves_col, value=eleves_text)
            else:
                # Créer une nouvelle classe - trouver une ligne vide
                new_row = None
                for row_idx in range(2, target_sheet.max_row + 10):  # +10 pour être sûr
                    # Vérifier si la ligne est vide
                    has_data = False
                    for col_idx in range(1, target_sheet.max_column + 1):
                        cell_value = str(target_sheet.cell(row=row_idx, column=col_idx).value or '').strip()
                        if cell_value:
                            has_data = True
                            break
                    if not has_data:
                        new_row = row_idx
                        break

                if new_row:
                    # Ajouter la nouvelle classe
                    target_sheet.cell(row=new_row, column=1, value=new_class_name)  # Colonne 1 pour le nom de classe

                    # Ajouter les élèves
                    eleves_noms = [eleve['nom'] for eleve in eleves_list if eleve['nom']]
                    eleves_text = ', '.join(eleves_noms)
                    target_sheet.cell(row=new_row, column=eleves_col, value=eleves_text)

            # Sauvegarder le fichier
            wb.save(excel_path)
            print(f"Classe '{new_class_name}' créée/mise à jour dans {excel_filename}")

        except Exception as e:
            print(f"Erreur lors de la création/mise à jour de la classe dans {excel_filename}: {e}")

    def update_matrix_for_new_class_assignment(week_folder, eleves_list, school_name, horaire, new_class_name):
        """
        Met à jour matrix.xlsx pour assigner les élèves à la nouvelle classe.
        """
        if not load_workbook or not week_folder or not eleves_list:
            return

        matrix_path = os.path.join(week_folder, "matrix.xlsx")

        if not os.path.exists(matrix_path):
            return

        try:
            wb = load_workbook(matrix_path)
            ws = wb.active

            # Identifier les colonnes importantes
            stagiaire_col = None
            ecoles_cols = {}
            horaires_cols = {}
            classes_cols = {}

            for col in range(1, ws.max_column + 1):
                col_name = str(ws.cell(row=1, column=col).value or '').strip()
                col_name_lower = col_name.lower()

                if 'stagiaire' in col_name_lower or 'nom' in col_name_lower or 'élève' in col_name_lower or 'eleve' in col_name_lower:
                    stagiaire_col = col
                elif 'école' in col_name_lower or 'ecole' in col_name_lower or 'school' in col_name_lower:
                    if not any(exclure in col_name_lower for exclure in ['cours 1', 'cours 2', 'cours 3', 'cours 4', '.1', '.2', 'arrivée', 'départ']):
                        ecoles_cols[col] = col_name
                elif 'horaire' in col_name_lower or 'heure' in col_name_lower or 'time' in col_name_lower:
                    if not any(exclure in col_name_lower for exclure in ['cours 1', 'cours 2', 'cours 3', 'cours 4', '.1', '.2', 'arrivée', 'départ']):
                        horaires_cols[col] = col_name
                elif ('classe' in col_name_lower or 'class' in col_name_lower or 'groupe' in col_name_lower):
                    if not any(exclure in col_name_lower for exclure in ['cours 1', 'cours 2', 'cours 3', 'cours 4', '.1', '.2', 'arrivée', 'départ']):
                        classes_cols[col] = col_name

            if not stagiaire_col:
                print("ERREUR: Aucune colonne eleves trouvee dans le matrix")
                return

            # Normaliser les noms d'élèves à assigner
            eleves_normalises = {eleve['nom'].lower().strip(): eleve for eleve in eleves_list}

            # Parcourir toutes les lignes pour trouver les élèves à mettre à jour
            updated_count = 0

            for row_idx in range(2, ws.max_row + 1):
                eleve_nom_brut = str(ws.cell(row=row_idx, column=stagiaire_col).value or '').strip()
                eleve_nom_normalise = eleve_nom_brut.lower().strip()

                # Chercher une correspondance
                eleve_info = None
                for search_name_norm, search_eleve in eleves_normalises.items():
                    if search_name_norm in eleve_nom_normalise or eleve_nom_normalise in search_name_norm:
                        eleve_info = search_eleve
                        break

                if eleve_info:
                    # Mettre à jour l'école, l'horaire et la classe pour cet élève
                    for col in ecoles_cols:
                        ws.cell(row=row_idx, column=col, value=school_name)
                    for col in horaires_cols:
                        ws.cell(row=row_idx, column=col, value=horaire)
                    for col in classes_cols:
                        ws.cell(row=row_idx, column=col, value=new_class_name)

                    updated_count += 1

            # Sauvegarder le fichier matrix
            wb.save(matrix_path)

            if updated_count > 0:
                print(f"Matrix.xlsx mis à jour: {updated_count} élève(s) assigné(s) à la classe '{new_class_name}'")

        except Exception as e:
            print(f"Erreur lors de la mise à jour de matrix.xlsx: {e}")

    def _show_message(msg_type, title, message):
        """Affiche un message d'information, d'avertissement ou d'erreur."""
        detail_window.attributes("-topmost", False)
        if msg_type == "warning":
            result = messagebox.showwarning(title, message, parent=detail_window)
        elif msg_type == "info":
            result = messagebox.showinfo(title, message, parent=detail_window)
        elif msg_type == "error":
            result = messagebox.showerror(title, message, parent=detail_window)
        elif msg_type == "yesno":
            result = messagebox.askyesno(title, message, parent=detail_window)
        detail_window.attributes("-topmost", True)
        return result

    def _create_student_class_assignment_menu():
        """Crée le menu d'assignation de classe pour les élèves."""
        # Créer la fenêtre du menu
        menu = create_safe_toplevel(700, 600, "", True, False, True)
        menu.overrideredirect(True)
        menu.configure(fg_color="white")

        # Centrer la fenêtre sur l'écran
        menu.update_idletasks()
        screen_width = menu.winfo_screenwidth()
        screen_height = menu.winfo_screenheight()
        x = (screen_width // 2) - (700 // 2)
        y = (screen_height // 2) - (600 // 2)
        menu.geometry(f"700x600+{x}+{y}")

        # Variable pour suivre si le menu est détruit
        menu_destroyed = False

        def close_menu():
            """Ferme le menu."""
            nonlocal menu_destroyed
            if not menu_destroyed and menu.winfo_exists():
                menu_destroyed = True
                try:
                    menu.destroy()
                except:
                    pass

        # Analyser les données des écoles pour cette semaine
        school_data = _analyze_school_classes_for_students(week_folder)

        # Frame principal avec scroll
        main_frame = ctk.CTkFrame(menu, fg_color="white", corner_radius=10)
        main_frame.pack(fill="both", expand=True, padx=2, pady=2)

        # En-tête
        header_frame = ctk.CTkFrame(main_frame, fg_color="#f0f9ff", corner_radius=8)
        header_frame.pack(fill="x", padx=8, pady=(8, 4))

        selected_count = len(selected_students)
        title_label = ctk.CTkLabel(
            header_frame,
            text=f"📚 Assigner {selected_count} élève{'s' if selected_count > 1 else ''} à une classe",
            font=("Inter", 13, "bold"),
            text_color="#1e293b"
        )
        title_label.pack(side="left", padx=12, pady=8)

        # Bouton fermer
        close_header_btn = ctk.CTkButton(
            header_frame,
            text="✕",
            width=30,
            height=30,
            font=("Inter", 14, "bold"),
            fg_color="#ef4444",
            hover_color="#dc2626",
            text_color="white",
            corner_radius=15,
            command=close_menu
        )
        close_header_btn.pack(side="right", padx=8, pady=4)

        # Frame avec scrollbar
        scrollable = ctk.CTkScrollableFrame(main_frame, fg_color="#f8fafc")
        scrollable.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        # Afficher les données des écoles
        if not school_data:
            no_data_label = ctk.CTkLabel(
                scrollable,
                text="Aucune donnée d'école trouvée pour cette semaine.",
                font=("Inter", 12),
                text_color="#6b7280"
            )
            no_data_label.pack(pady=30)
        else:
            _display_school_data_for_students(scrollable, school_data, menu, close_menu)

        # Frame pour les boutons d'action en bas
        action_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        action_frame.pack(fill="x", padx=8, pady=(0, 8))

        # Bouton fermer en bas
        close_bottom_btn = ctk.CTkButton(
            action_frame,
            text="Fermer",
            font=("Inter", 11, "bold"),
            height=35,
            width=100,
            fg_color="#6b7280",
            hover_color="#4b5563",
            text_color="white",
            command=close_menu
        )
        close_bottom_btn.pack(side="right")

        # Gestion des événements
        def on_click_outside(event):
            if not menu_destroyed and menu.winfo_exists():
                try:
                    menu_x = menu.winfo_x()
                    menu_y = menu.winfo_y()
                    menu_width = menu.winfo_width()
                    menu_height = menu.winfo_height()

                    click_x = event.x_root
                    click_y = event.y_root

                    if not (menu_x <= click_x <= menu_x + menu_width and
                           menu_y <= click_y <= menu_y + menu_height):
                        close_menu()
                except:
                    pass

        click_handler_id = detail_window.bind("<Button-1>", on_click_outside, add="+")
        menu._click_handler_id = click_handler_id

        def on_key_press(event):
            if event.keysym == 'Escape' and not menu_destroyed:
                close_menu()

        menu.bind("<Key>", on_key_press)
        menu.focus_set()

    def _analyze_school_classes_for_students(week_folder):
        """Analyse les fichiers Excel d'écoles pour les élèves."""
        try:
            import pandas as pd
        except ImportError:
            return {}

        result = {
            'ecole_a': [],
            'ecole_b': [],
            'ecole_c_cs': [],
            'ecole_c_ci': [],
            'ecole_morning': [],
            'ecole_premium_cs': [],
            'ecole_premium_ci': []
        }

        # Mapping des fichiers Excel vers les clés d'écoles
        file_to_school_mapping = {
            'ecole_a.xlsx': 'ecole_a',
            'ecole_b.xlsx': 'ecole_b',
            'ECOLE_C_cours_standard.xlsx': 'ecole_c_cs',
            'ECOLE_C_cours_intensif.xlsx': 'ecole_c_ci',
            'MORNING.xlsx': 'ecole_morning',
            'ECOLE_PREMIUM_cours_standard.xlsx': 'ecole_premium_cs',
            'ECOLE_PREMIUM_cours_intensifs.xlsx': 'ecole_premium_ci'
        }

        # Fonction helper pour analyser un fichier Excel d'école
        def analyze_school_file(excel_path, school_key):
            if not os.path.exists(excel_path):
                return

            try:
                # Lire toutes les feuilles du fichier
                xl_file = pd.ExcelFile(excel_path)
                sheet_names = xl_file.sheet_names

                for sheet_name in sheet_names:
                    try:
                        df = pd.read_excel(excel_path, sheet_name=sheet_name)

                        if df.empty:
                            # La feuille est vide mais on garde l'horaire
                            horaire = clean_horaire_name(sheet_name)
                            result[school_key].append({
                                'horaire': horaire or sheet_name,
                                'classes': []
                            })
                            continue

                        # Analyser et nettoyer le nom de la feuille
                        horaire = clean_horaire_name(sheet_name)

                        # Chercher les colonnes de classes
                        classe_cols = []
                        for col in df.columns:
                            col_lower = str(col).lower()
                            if any(keyword in col_lower for keyword in ['classe', 'groupe', 'section']):
                                classe_cols.append(col)

                        classes_info = []
                        for _, row in df.iterrows():
                            classe_nom = None
                            for col in classe_cols:
                                val = str(row.get(col, '')).strip()
                                if val and val.lower() not in ['', 'nan', 'none']:
                                    classe_nom = val
                                    break

                            if classe_nom:
                                classes_info.append({
                                    'nom_classe': classe_nom,
                                    'nb_eleves': 0,  # Sera calculé plus tard
                                    'niveau': '',
                                    'eleves': []
                                })

                        result[school_key].append({
                            'horaire': horaire or sheet_name,
                            'classes': classes_info
                        })

                    except Exception as e:
                        print(f"Erreur lors de l'analyse de la feuille '{sheet_name}': {e}")
                        continue

            except Exception as e:
                print(f"Erreur lors de l'analyse du fichier {excel_path}: {e}")

        # Analyser tous les fichiers Excel du dossier semaine
        if os.path.exists(week_folder):
            for filename in os.listdir(week_folder):
                if filename.lower().endswith('.xlsx') and filename.lower() != 'matrix.xlsx':
                    file_path = os.path.join(week_folder, filename)
                    school_key = file_to_school_mapping.get(filename)
                    if school_key:
                        analyze_school_file(file_path, school_key)

        return result

    def _display_school_data_for_students(parent_frame, school_data, menu, close_menu_func):
        """Affiche les données des écoles dans le menu d'assignation pour les élèves."""
        # Mapping pour les noms d'affichage et couleurs
        school_display = {
            'ecole_a': ('A', '#3b82f6'),
            'ecole_b': ('B', '#10b981'),
            'ecole_c_cs': ('C/CS', '#f59e0b'),
            'ecole_c_ci': ('C/CI', '#8b5cf6'),
            'ecole_morning': ('Morning', '#ef4444'),
            'ecole_premium_cs': ('Premium/CS', '#06b6d4'),
            'ecole_premium_ci': ('Premium/CI', '#f97316')
        }

        for school_key, horaires in school_data.items():
            if not horaires:  # Si pas d'horaires pour cette école
                continue

            display_name, school_color = school_display.get(school_key, (school_key, '#6b7280'))

            # Frame pour l'école
            school_frame = ctk.CTkFrame(parent_frame, fg_color="white", corner_radius=8, border_width=1, border_color=school_color)
            school_frame.pack(fill="x", pady=(0, 8), padx=5)

            # En-tête de l'école
            school_header = ctk.CTkFrame(school_frame, fg_color="#f0f9ff", corner_radius=6)
            school_header.pack(fill="x", padx=8, pady=(8, 5))

            school_title = ctk.CTkLabel(
                school_header,
                text=f"🏫 École {display_name}",
                font=("Inter", 12, "bold"),
                text_color=school_color
            )
            school_title.pack(pady=6)

            # Conteneur pour les horaires
            horaires_frame = ctk.CTkFrame(school_frame, fg_color="transparent")
            horaires_frame.pack(fill="x", padx=8, pady=(0, 8))

            # Calculer le nombre d'horaires par ligne
            total_horaires = len(horaires)
            horaires_per_row = min(3, max(2, total_horaires))  # Entre 2 et 3 horaires par ligne

            # Configurer les colonnes dynamiquement
            for col in range(horaires_per_row):
                horaires_frame.grid_columnconfigure(col, weight=1)

            horaire_row = 0
            horaire_col = 0

            for horaire_info in horaires:
                horaire = horaire_info.get('horaire', 'Horaire inconnu')
                classes = horaire_info.get('classes', [])

                # Frame pour l'horaire
                horaire_frame = ctk.CTkFrame(horaires_frame, fg_color="#f8fafc", corner_radius=6, border_width=1, border_color="#e5e7eb")
                horaire_frame.grid(row=horaire_row, column=horaire_col, sticky="nsew", padx=(0, 4) if horaire_col < horaires_per_row - 1 else 0, pady=(0, 4))

                # En-tête de l'horaire
                horaire_header = ctk.CTkFrame(horaire_frame, fg_color="white", corner_radius=4)
                horaire_header.pack(fill="x", padx=6, pady=(6, 4))

                horaire_title = ctk.CTkLabel(
                    horaire_header,
                    text=f"🕒 Horaire : {horaire}",
                    font=("Inter", 11, "bold"),
                    text_color="#374151"
                )
                horaire_title.pack(side="left", pady=4)

                # Conteneur pour les classes (cards)
                if classes:
                    classes_frame = ctk.CTkFrame(horaire_frame, fg_color="transparent")
                    classes_frame.pack(fill="x", padx=6, pady=(0, 8))

                    # Configuration pour une grille de classes
                    classes_per_row = 4  # Nombre de classes par ligne
                    for i in range(classes_per_row):
                        classes_frame.grid_columnconfigure(i, weight=1)

                    for idx, classe_info in enumerate(classes):
                        classe_nom = classe_info.get('nom_classe', 'Classe inconnue')
                        row = idx // classes_per_row
                        col = idx % classes_per_row

                        # Petite card pour chaque classe
                        class_card = ctk.CTkFrame(
                            classes_frame,
                            fg_color="white",
                            corner_radius=4,
                            width=80,
                            height=35
                        )
                        class_card.grid(row=row, column=col, padx=2, pady=2, sticky="nsew")
                        class_card.grid_propagate(False)

                        # Label du nom de classe (centré)
                        class_label = ctk.CTkLabel(
                            class_card,
                            text=classe_nom,
                            font=("Inter", 9, "bold"),
                            text_color="#374151"
                        )
                        class_label.place(relx=0.5, rely=0.5, anchor="center")

                        # Rendre la card cliquable
                        class_card.configure(cursor="hand2")

                        # Fonction pour gérer l'effet hover
                        def on_enter(event, frame=class_card):
                            frame.configure(fg_color="#f3f4f6")

                        def on_leave(event, frame=class_card):
                            frame.configure(fg_color="white")

                        # Propager l'effet hover à tous les enfants
                        def bind_hover_to_children(widget, enter_func, leave_func):
                            widget.bind("<Enter>", enter_func)
                            widget.bind("<Leave>", leave_func)
                            for child in widget.winfo_children():
                                bind_hover_to_children(child, enter_func, leave_func)

                        # Lier les événements
                        class_card.bind("<Enter>", on_enter)
                        class_card.bind("<Leave>", on_leave)
                        bind_hover_to_children(class_card, on_enter, on_leave)

                        # Gestion du clic sur la card
                        def on_class_click(event, s_key=school_key, h_info=horaire_info, c_info=classe_info, m=menu, close_func=close_menu_func):
                            # Fermer immédiatement le menu
                            close_func()
                            # Puis assigner avec un petit délai pour laisser le menu se fermer
                            m.after(10, lambda: _assign_students_to_class(s_key, h_info, c_info))

                        class_card.bind("<Button-1>", on_class_click)
                        class_label.bind("<Button-1>", on_class_click)
                        class_label.configure(cursor="hand2")
                else:
                    no_classes_label = ctk.CTkLabel(
                        horaire_frame,
                        text="Aucune classe définie",
                        font=("Inter", 10, "italic"),
                        text_color="#9ca3af"
                    )
                    no_classes_label.pack(padx=12, pady=(0, 8))

                # Gestion de la grille des horaires
                horaire_col += 1
                if horaire_col >= horaires_per_row:
                    horaire_col = 0
                    horaire_row += 1

    def _assign_students_to_class(new_school_key, new_horaire_info, new_classe_info):
        """Assigne les élèves sélectionnés à une classe existante."""
        # Mapping pour convertir les clés d'école en noms d'affichage
        school_mapping = {
            'ecole_a': 'A',
            'ecole_b': 'B',
            'ecole_c_cs': 'C/CS',
            'ecole_c_ci': 'C/CI',
            'ecole_morning': 'Morning',
            'ecole_premium_cs': 'Premium/CS',
            'ecole_premium_ci': 'Premium/CI'
        }

        # Informations de la NOUVELLE classe (cliquée)
        new_school_name = school_mapping.get(new_school_key, new_school_key)
        new_horaire = new_horaire_info.get('horaire', '')
        new_classe_nom = new_classe_info.get('nom_classe', '')

        # Récupérer les élèves sélectionnés avant de les supprimer
        eleves_a_assigner = []
        for idx in sorted(selected_students, reverse=True):
            if idx < len(eleves):
                eleve = eleves[idx]
                eleves_a_assigner.append({
                    'nom': eleve.get('nom', ''),
                    'niveau': eleve.get('niveau', 'A0'),
                    'ci': eleve.get('ci', False),
                    'age': eleve.get('age', None)
                })

        if not eleves_a_assigner:
            return

        # Informations de l'ANCIENNE classe (celle affichée dans la fiche détail)
        # On utilise les variables de la fonction parente open_classe_details
        ancienne_classe_nom = classe_info['nom_classe']  # classe_info de la fonction parente = ancienne classe

        # Supprimer les élèves de l'ancienne classe
        for idx in sorted(selected_students, reverse=True):
            if idx < len(eleves):
                eleves.pop(idx)

        # Mettre à jour le nombre d'élèves dans l'ancienne classe (variable de la fonction parente)
        classe_info['nb_eleves'] = len(eleves)  # classe_info de la fonction parente

        # Retirer les élèves de l'ancienne classe dans le fichier Excel de l'école
        # Utilise school_name et horaire de la fonction parente (ancienne classe)
        _remove_students_from_old_class(week_folder, school_name, horaire, ancienne_classe_nom, eleves_a_assigner)

        # Ajouter les élèves à la nouvelle classe dans le fichier Excel de l'école
        _add_students_to_existing_class(week_folder, new_school_name, new_horaire, new_classe_nom, eleves_a_assigner)

        # Mettre à jour matrix.xlsx pour assigner ces élèves à la nouvelle classe
        _update_matrix_for_students_assignment(week_folder, eleves_a_assigner, new_school_name, new_horaire, new_classe_nom)

        # Vider la sélection
        selected_students.clear()

        # Mettre à jour l'affichage de la grille des élèves
        refresh_student_grid()

        # Rafraîchir le dashboard principal
        if refresh_callback:
            refresh_callback()

        _show_message("info", "Succès", f"{len(eleves_a_assigner)} élève{'s' if len(eleves_a_assigner) > 1 else ''} assigné{'s' if len(eleves_a_assigner) > 1 else ''} à la classe '{new_classe_nom}'")

    def _remove_students_from_old_class(week_folder, school_name, horaire, classe_nom, eleves_list):
        """
        Retire des élèves d'une classe existante dans le fichier Excel de l'école.
        """
        if load_workbook is None or not week_folder:
            return

        # Mapping des écoles vers les fichiers Excel
        school_file_mapping = {
            'A': 'ecole_a.xlsx',
            'B': 'ecole_b.xlsx',
            'C/CS': 'ECOLE_C_cours_standard.xlsx',
            'C/CI': 'ECOLE_C_cours_intensif.xlsx',
            'Morning': 'MORNING.xlsx',
            'Premium/CS': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'Premium/CI': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        if school_name not in school_file_mapping:
            return

        excel_filename = school_file_mapping[school_name]
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            return

        try:
            wb = load_workbook(excel_path)

            # Chercher la feuille correspondant à l'horaire
            target_sheet = None
            for sheet_name in wb.sheetnames:
                sheet_normalized = clean_horaire_name(sheet_name).lower().strip()
                horaire_normalized = horaire.lower().strip()
                if sheet_normalized == horaire_normalized or horaire_normalized in sheet_normalized:
                    target_sheet = wb[sheet_name]
                    break

            if not target_sheet:
                return

            # Chercher la ligne de la classe
            classe_row = None
            classe_col = None

            for row_idx in range(2, target_sheet.max_row + 1):
                for col_idx in range(1, target_sheet.max_column + 1):
                    cell_value = str(target_sheet.cell(row=row_idx, column=col_idx).value or '').strip()
                    if cell_value == classe_nom:
                        classe_row = row_idx
                        classe_col = col_idx
                        break
                if classe_row:
                    break

            if not classe_row:
                return

            # Chercher la colonne des élèves
            eleves_col = None
            for col_idx in range(1, target_sheet.max_column + 1):
                cell_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                if ('élèves' in cell_value or 'eleves' in cell_value) and 'liste' not in cell_value:
                    eleves_col = col_idx
                    break

            # Si pas trouvé, essayer "liste des élèves"
            if eleves_col is None:
                for col_idx in range(1, target_sheet.max_column + 1):
                    cell_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                    if 'liste' in cell_value and ('élèves' in cell_value or 'eleves' in cell_value):
                        eleves_col = col_idx
                        break

            # Si toujours pas trouvé, utiliser la colonne 5 comme fallback
            if eleves_col is None:
                eleves_col = 5

            # Récupérer la liste actuelle d'élèves
            current_eleves_text = str(target_sheet.cell(row=classe_row, column=eleves_col).value or '').strip()
            current_eleves = []
            if current_eleves_text and current_eleves_text not in ['', 'nan', 'none']:
                current_eleves = [nom.strip() for nom in current_eleves_text.split(',') if nom.strip()]

            # Créer un set des noms d'élèves à retirer pour une recherche rapide
            eleves_to_remove = {eleve['nom'] for eleve in eleves_list if eleve['nom']}

            # Filtrer la liste pour retirer les élèves
            remaining_eleves = [nom for nom in current_eleves if nom not in eleves_to_remove]

            # Mettre à jour la cellule avec la nouvelle liste
            if remaining_eleves:
                updated_eleves_text = ', '.join(remaining_eleves)
                target_sheet.cell(row=classe_row, column=eleves_col, value=updated_eleves_text)
            else:
                # Si plus d'élèves, vider la cellule
                target_sheet.cell(row=classe_row, column=eleves_col, value='')

            # Sauvegarder le fichier
            wb.save(excel_path)

        except Exception as e:
            print(f"Erreur lors du retrait d'élèves de la classe dans {excel_filename}: {e}")

    def _add_students_to_existing_class(week_folder, school_name, horaire, classe_nom, eleves_list):
        """Ajoute des élèves à une classe existante dans le fichier Excel de l'école."""
        if load_workbook is None or not week_folder:
            return

        # Mapping des écoles vers les fichiers Excel
        school_file_mapping = {
            'A': 'ecole_a.xlsx',
            'B': 'ecole_b.xlsx',
            'C/CS': 'ECOLE_C_cours_standard.xlsx',
            'C/CI': 'ECOLE_C_cours_intensif.xlsx',
            'Morning': 'MORNING.xlsx',
            'Premium/CS': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'Premium/CI': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        if school_name not in school_file_mapping:
            return

        excel_filename = school_file_mapping[school_name]
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            return

        try:
            wb = load_workbook(excel_path)

            # Chercher la feuille correspondant à l'horaire
            target_sheet = None
            for sheet_name in wb.sheetnames:
                sheet_normalized = clean_horaire_name(sheet_name).lower().strip()
                horaire_normalized = horaire.lower().strip()
                if sheet_normalized == horaire_normalized or horaire_normalized in sheet_normalized:
                    target_sheet = wb[sheet_name]
                    break

            if not target_sheet:
                return

            # Chercher la ligne de la classe
            classe_row = None
            classe_col = None

            for row_idx in range(2, target_sheet.max_row + 1):
                for col_idx in range(1, target_sheet.max_column + 1):
                    cell_value = str(target_sheet.cell(row=row_idx, column=col_idx).value or '').strip()
                    if cell_value == classe_nom:
                        classe_row = row_idx
                        classe_col = col_idx
                        break
                if classe_row:
                    break

            if not classe_row:
                return

            # Chercher la colonne des élèves
            eleves_col = None
            for col_idx in range(1, target_sheet.max_column + 1):
                cell_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                if ('élèves' in cell_value or 'eleves' in cell_value) and 'liste' not in cell_value:
                    eleves_col = col_idx
                    break

            # Si pas trouvé, essayer "liste des élèves"
            if eleves_col is None:
                for col_idx in range(1, target_sheet.max_column + 1):
                    cell_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                    if 'liste' in cell_value and ('élèves' in cell_value or 'eleves' in cell_value):
                        eleves_col = col_idx
                        break

            # Si toujours pas trouvé, utiliser la colonne 5 comme fallback
            if eleves_col is None:
                eleves_col = 5

            # Récupérer la liste actuelle d'élèves
            current_eleves_text = str(target_sheet.cell(row=classe_row, column=eleves_col).value or '').strip()
            current_eleves = []
            if current_eleves_text and current_eleves_text not in ['', 'nan', 'none']:
                current_eleves = [nom.strip() for nom in current_eleves_text.split(',') if nom.strip()]

            # Ajouter les nouveaux élèves
            new_eleves_names = [eleve['nom'] for eleve in eleves_list if eleve['nom']]
            all_eleves = current_eleves + new_eleves_names

            # Mettre à jour la cellule avec la nouvelle liste
            updated_eleves_text = ', '.join(all_eleves)
            target_sheet.cell(row=classe_row, column=eleves_col, value=updated_eleves_text)

            # Sauvegarder le fichier
            wb.save(excel_path)
            print(f"Élèves ajoutés à la classe '{classe_nom}' dans {excel_filename}")

        except Exception as e:
            print(f"Erreur lors de l'ajout d'élèves à la classe dans {excel_filename}: {e}")

    def _update_matrix_for_students_assignment(week_folder, eleves_list, school_name, horaire, classe_nom):
        """Met à jour matrix.xlsx pour assigner les élèves à une nouvelle classe."""
        if not load_workbook or not week_folder or not eleves_list:
            return

        matrix_path = os.path.join(week_folder, "matrix.xlsx")

        if not os.path.exists(matrix_path):
            return

        try:
            wb = load_workbook(matrix_path)
            ws = wb.active

            # Identifier les colonnes importantes
            stagiaire_col = None
            ecoles_cols = {}
            horaires_cols = {}
            classes_cols = {}

            for col in range(1, ws.max_column + 1):
                col_name = str(ws.cell(row=1, column=col).value or '').strip()
                col_name_lower = col_name.lower()

                if 'stagiaire' in col_name_lower or 'nom' in col_name_lower or 'élève' in col_name_lower or 'eleve' in col_name_lower:
                    stagiaire_col = col
                elif 'école' in col_name_lower or 'ecole' in col_name_lower or 'school' in col_name_lower:
                    if not any(exclure in col_name_lower for exclure in ['cours 1', 'cours 2', 'cours 3', 'cours 4', '.1', '.2', 'arrivée', 'départ']):
                        ecoles_cols[col] = col_name
                elif 'horaire' in col_name_lower or 'heure' in col_name_lower or 'time' in col_name_lower:
                    if not any(exclure in col_name_lower for exclure in ['cours 1', 'cours 2', 'cours 3', 'cours 4', '.1', '.2', 'arrivée', 'départ']):
                        horaires_cols[col] = col_name
                elif ('classe' in col_name_lower or 'class' in col_name_lower or 'groupe' in col_name_lower):
                    if not any(exclure in col_name_lower for exclure in ['cours 1', 'cours 2', 'cours 3', 'cours 4', '.1', '.2', 'arrivée', 'départ']):
                        classes_cols[col] = col_name

            if not stagiaire_col:
                print("ERREUR: Aucune colonne eleves trouvee dans le matrix")
                return

            # Normaliser les noms d'élèves à assigner
            eleves_normalises = {eleve['nom'].lower().strip(): eleve for eleve in eleves_list}

            # Parcourir toutes les lignes pour trouver les élèves à mettre à jour
            updated_count = 0

            for row_idx in range(2, ws.max_row + 1):
                eleve_nom_brut = str(ws.cell(row=row_idx, column=stagiaire_col).value or '').strip()
                eleve_nom_normalise = eleve_nom_brut.lower().strip()

                # Chercher une correspondance
                eleve_info = None
                for search_name_norm, search_eleve in eleves_normalises.items():
                    if search_name_norm in eleve_nom_normalise or eleve_nom_normalise in search_name_norm:
                        eleve_info = search_eleve
                        break

                if eleve_info:
                    # Mettre à jour l'école, l'horaire et la classe pour cet élève
                    for col in ecoles_cols:
                        ws.cell(row=row_idx, column=col, value=school_name)
                    for col in horaires_cols:
                        ws.cell(row=row_idx, column=col, value=horaire)
                    for col in classes_cols:
                        ws.cell(row=row_idx, column=col, value=classe_nom)

                    updated_count += 1

            # Sauvegarder le fichier matrix
            wb.save(matrix_path)

            if updated_count > 0:
                print(f"Matrix.xlsx mis à jour: {updated_count} élève(s) assigné(s) à la classe '{classe_nom}'")

        except Exception as e:
            print(f"Erreur lors de la mise à jour de matrix.xlsx: {e}")

    def assign_students_to_new_class():
        """Assigne les élèves sélectionnés à une classe existante via un menu d'assignation."""
        if not selected_students:
            # Afficher un message d'avertissement si aucun élève n'est sélectionné
            _show_message("warning", "Aucun élève sélectionné", "Veuillez d'abord sélectionner un ou plusieurs élèves.")
            return

        # Ouvrir le menu d'assignation de classe
        _create_student_class_assignment_menu()

    def show_student_info():
        """Affiche les informations détaillées des élèves sélectionnés"""
        if not selected_students:
            # Afficher un message d'erreur si aucun élève n'est sélectionné
            error_popup = create_safe_toplevel(350, 150, "Erreur")
            error_popup.lift()
            error_popup.after(50, lambda: [error_popup.focus_force(), error_popup.grab_set()])

            x = detail_window.winfo_rootx() + (detail_window.winfo_width() // 2) - 175
            y = detail_window.winfo_rooty() + (detail_window.winfo_height() // 2) - 75

            ctk.CTkLabel(error_popup, text="Aucun élève sélectionné", font=("Segoe UI", 14, "bold")).pack(pady=(20, 10))
            ctk.CTkLabel(error_popup, text="Veuillez sélectionner au moins un élève.", font=("Segoe UI", 11)).pack(pady=(0, 20))

            ctk.CTkButton(error_popup, text="OK", command=error_popup.destroy, width=100).pack()
            return

        # Créer une popup avec les infos des élèves sélectionnés
        info_popup = ctk.CTkToplevel(detail_window)
        info_popup.title("Informations des élèves")
        info_popup.geometry("500x400")
        info_popup.resizable(False, False)
        info_popup.attributes("-topmost", True)
        info_popup.lift()
        info_popup.after(50, lambda: [info_popup.focus_force(), info_popup.grab_set()])

        x = detail_window.winfo_rootx() + (detail_window.winfo_width() // 2) - 250
        y = detail_window.winfo_rooty() + (detail_window.winfo_height() // 2) - 200
        info_popup.geometry(f"500x400+{x}+{y}")

        # Titre
        ctk.CTkLabel(info_popup, text=f"Informations de {len(selected_students)} élève{'s' if len(selected_students) > 1 else ''}",
                    font=("Segoe UI", 16, "bold")).pack(pady=(20, 15))

        # Frame scrollable pour les informations
        scroll_frame = ctk.CTkScrollableFrame(info_popup, width=450, height=280)
        scroll_frame.pack(padx=25, pady=(0, 20))

        # Afficher les infos de chaque élève sélectionné
        for idx in sorted(selected_students):
            if idx < len(eleves):
                student = eleves[idx]

                # Card pour chaque élève
                student_card = ctk.CTkFrame(scroll_frame, fg_color="#F8FAFC", corner_radius=8, border_width=1, border_color="#E2E8F0")
                student_card.pack(fill="x", pady=5, padx=5)

                # Nom de l'élève
                ctk.CTkLabel(student_card, text=student.get('nom', 'N/A').upper(),
                           font=("Segoe UI", 14, "bold"), text_color="#1E293B").pack(pady=(10, 5), padx=15)

                # Informations détaillées
                info_frame = ctk.CTkFrame(student_card, fg_color="transparent")
                info_frame.pack(fill="x", padx=15, pady=(0, 10))

                # Niveau
                niv = student.get('niveau', 'A0')
                n_color = map_colors.get(niv, "#E2E8F0")
                ctk.CTkLabel(info_frame, text=f"Niveau: {niv}", fg_color=n_color,
                           text_color="#1E293B", corner_radius=4, padx=8, pady=3,
                           font=("Segoe UI", 11, "bold")).pack(side="left", padx=(0, 8))

                # CI si applicable
                if student.get('ci', False):
                    ctk.CTkLabel(info_frame, text="CI", fg_color="#2563EB",
                               text_color="white", corner_radius=4, padx=8, pady=3,
                               font=("Segoe UI", 11, "bold")).pack(side="left", padx=(0, 8))

                # Âge si disponible
                if 'age' in student:
                    ctk.CTkLabel(info_frame, text=f"Âge: {student['age']} ans",
                               fg_color="#6B7280", text_color="white", corner_radius=4,
                               padx=8, pady=3, font=("Segoe UI", 11)).pack(side="left")

        # Bouton fermer
        ctk.CTkButton(info_popup, text="Fermer", command=info_popup.destroy,
                     fg_color="#6B7280", width=100).pack(pady=(0, 20))

    def update_eleves_counter():
        """Met à jour le compteur d'élèves dans le header"""
        if eleves_counter_label_ref:
            current_nb_eleves = len(eleves)
            compteur_color = "#DC2626" if current_nb_eleves > 15 else school_color
            compteur_text_color = "white"
            eleves_counter_label_ref.configure(
                text=f"{current_nb_eleves}/15",
                fg_color=compteur_color,
                text_color=compteur_text_color
            )

    def refresh_student_grid():
        """Rafraîchit l'affichage de la grille des élèves"""
        generate_student_grid()
        update_eleves_counter()  # Mettre à jour le compteur d'élèves

    def update_matrix_after_student_removal(week_folder, eleves_a_supprimer):
        """
        Met à jour matrix.xlsx en vidant les colonnes Ecole, Horaire, Classe
        pour les élèves qui ont été supprimés de leur classe.
        """
        if not load_workbook or not week_folder or not eleves_a_supprimer:
            return

        matrix_path = os.path.join(week_folder, "matrix.xlsx")
        if not os.path.exists(matrix_path):
            return

        try:
            wb = load_workbook(matrix_path)
            ws = wb.active

            # Identifier les colonnes importantes
            stagiaire_col = None
            colonnes_assignation = []

            for col in range(1, ws.max_column + 1):
                col_name = str(ws.cell(row=1, column=col).value or '').strip()
                col_name_lower = col_name.lower()

                if 'stagiaire' in col_name_lower or 'nom' in col_name_lower or 'élève' in col_name_lower or 'eleve' in col_name_lower:
                    stagiaire_col = col
                elif any(keyword in col_name_lower for keyword in ['ecole', 'école', 'school', 'classe', 'class', 'groupe', 'horaire', 'horaire', 'heure', 'time']):
                    if not any(exclure in col_name_lower for exclure in ['cours 1', 'cours 2', 'cours 3', 'cours 4', '.1', '.2', 'arrivée', 'départ']):
                        colonnes_assignation.append(col)

            if not stagiaire_col:
                print("ERREUR: Aucune colonne eleves trouvee dans le matrix")
                return

            # Normaliser les noms d'élèves à supprimer
            eleves_normalises = {eleve['nom'].lower().strip(): eleve for eleve in eleves_a_supprimer}

            # Parcourir toutes les lignes pour trouver les élèves à mettre à jour
            updated_count = 0

            for row_idx in range(2, ws.max_row + 1):
                eleve_nom_brut = str(ws.cell(row=row_idx, column=stagiaire_col).value or '').strip()
                eleve_nom_normalise = eleve_nom_brut.lower().strip()

                # Chercher une correspondance
                eleve_info = None
                for search_name_norm, search_eleve in eleves_normalises.items():
                    if search_name_norm in eleve_nom_normalise or eleve_nom_normalise in search_name_norm:
                        eleve_info = search_eleve
                        break

                if eleve_info:
                    # Vider TOUTES les colonnes d'assignation pour cet élève
                    for col in colonnes_assignation:
                        old_value = ws.cell(row=row_idx, column=col).value
                        if old_value is not None:
                            # Utiliser la syntaxe correcte pour modifier la cellule
                            cell = ws.cell(row=row_idx, column=col)
                            cell.value = None

                    updated_count += 1

            # Sauvegarder le fichier matrix
            wb.save(matrix_path)

            if updated_count > 0:
                print(f"Matrix.xlsx mis à jour: {updated_count} élève(s) désassigné(s)")

        except Exception as e:
            print(f"ERREUR lors de la mise à jour du matrix: {e}")

    def update_school_excel_after_student_removal(week_folder, school_name, horaire, classe_nom, nouveaux_nb_eleves):
        """
        Met à jour le fichier Excel de l'école après suppression d'élèves.
        Si la classe devient vide, elle peut être supprimée ou simplement mise à jour.
        """
        if load_workbook is None or not week_folder:
            return

        # Mapping des écoles vers les fichiers Excel
        school_file_mapping = {
            'A': 'ecole_a.xlsx',
            'B': 'ecole_b.xlsx',
            'C/CS': 'ECOLE_C_cours_standard.xlsx',
            'C/CI': 'ECOLE_C_cours_intensif.xlsx',
            'Morning': 'MORNING.xlsx',
            'Premium/CS': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'Premium/CI': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        if school_name not in school_file_mapping:
            return

        excel_filename = school_file_mapping[school_name]
        excel_path = os.path.join(week_folder, excel_filename)

        if not os.path.exists(excel_path):
            return

        try:
            wb = load_workbook(excel_path)

            # Chercher la feuille correspondant à l'horaire
            target_sheet = None
            for sheet_name in wb.sheetnames:
                sheet_normalized = clean_horaire_name(sheet_name).lower().strip()
                horaire_normalized = horaire.lower().strip()
                if sheet_normalized == horaire_normalized or horaire_normalized in sheet_normalized:
                    target_sheet = wb[sheet_name]
                    break

            if not target_sheet:
                return

            # Chercher la ligne de la classe
            classe_row = None
            classe_col = None

            for row_idx in range(2, target_sheet.max_row + 1):
                for col_idx in range(1, target_sheet.max_column + 1):
                    cell_value = str(target_sheet.cell(row=row_idx, column=col_idx).value or '').strip()
                    if cell_value == classe_nom:
                        classe_row = row_idx
                        classe_col = col_idx
                        break
                if classe_row:
                    break

            if not classe_row:
                return

            # Chercher la colonne des élèves
            eleves_col = None
            for col_idx in range(1, target_sheet.max_column + 1):
                cell_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                if ('élèves' in cell_value or 'eleves' in cell_value) and 'liste' not in cell_value:
                    eleves_col = col_idx
                    break

            # Si pas trouvé, essayer "liste des élèves"
            if eleves_col is None:
                for col_idx in range(1, target_sheet.max_column + 1):
                    cell_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                    if 'liste' in cell_value and ('élèves' in cell_value or 'eleves' in cell_value):
                        eleves_col = col_idx
                        break

            # Si toujours pas trouvé, utiliser la colonne 5 comme fallback
            if eleves_col is None:
                eleves_col = 5

            # Mettre à jour le nombre d'élèves dans la cellule appropriée
            if nouveaux_nb_eleves > 0:
                # Convertir la liste d'élèves en chaîne de caractères
                eleves_noms = [eleve.get('nom', '') for eleve in eleves if eleve.get('nom')]
                eleves_text = ', '.join(eleves_noms)
                target_sheet.cell(row=classe_row, column=eleves_col, value=eleves_text)
            else:
                # Si plus d'élèves, vider la cellule ou supprimer la ligne
                target_sheet.cell(row=classe_row, column=eleves_col, value='')

            # Sauvegarder le fichier
            wb.save(excel_path)

            print(f"Fichier {excel_filename} mis à jour: classe '{classe_nom}' contient maintenant {nouveaux_nb_eleves} élève(s)")

        except Exception as e:
            print(f"ERREUR lors de la mise à jour de {excel_filename}: {e}")

    # === 3. FOOTER (EXTÉRIEUR DE LA FRAME GRISE) ===
    # Directement sur le fond blanc de detail_window
    footer_area = ctk.CTkFrame(detail_window, fg_color="white", height=80)
    footer_area.pack(fill="x", side="bottom")
    footer_area.pack_propagate(False)

    ctk.CTkButton(footer_area, text="Fermer la fiche", fg_color="#1E293B", hover_color="#0F172A",
                  corner_radius=10, font=("Segoe UI", 14, "bold"), height=45, width=180,
                  command=detail_window.destroy).pack(side="right", padx=40)

    # Mettre à jour le compteur initial une fois que tout est chargé
    update_eleves_counter()

# === TEST ===
if __name__ == "__main__":
    ctk.set_appearance_mode("light")
    root = ctk.CTk()
    root.withdraw() 
    
    test_data = {
        'nom_classe': 'CM2-B (Arts)',
        'niveau': 'Cycle 3',
        'eleves': [{'nom': f'Élève {i+1}', 'niveau': 'A1.2', 'age': 10, 'ci': i%3==0} for i in range(13)]
    }
    
    open_classe_details(test_data, "14:00 - 16:00", "Mme Sarah", "professeur", "#2563EB", "École Jean Jaurès")
    root.mainloop()