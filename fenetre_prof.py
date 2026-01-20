import customtkinter as ctk
import json
import os
from tkinter import messagebox
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

class PersonnelManager(ctk.CTkToplevel):

    def __init__(self, parent, week_folder, personnel_type="animateurs", data_changed_callback=None):
        super().__init__(parent)

        self.week_folder = week_folder
        self.personnel_type = personnel_type
        self.personnel_path = os.path.join(week_folder, "personnel.json")
        self.current_week = os.path.basename(week_folder)
        self.data_changed_callback = data_changed_callback

        # Textes dynamiques
        self.display_name = "Animateurs" if personnel_type == "animateurs" else "Professeurs"
        self.single_name = "animateur" if personnel_type == "animateurs" else "professeur"
        
        # Configuration fenêtre
        week_display = self._format_week_display(self.current_week)
        self.title(f"Gestion des {self.display_name.lower()} - {week_display}")
        self.geometry("900x650")
        self.minsize(900, 600)
        self.resizable(True, True)
        self.configure(fg_color="white")
        
        # NOTE : On met topmost à False au début pour éviter les conflits graphiques
        # On le réactivera à la fin du chargement
        self.attributes("-topmost", False)

        # Initialisation des variables de données (vides pour l'instant)
        self.data = {self.personnel_type: []}
        self.item_labels = []
        self.selected_frame = None

        # 1. On construit l'interface STRUCTURELLE (les cadres, boutons vides) tout de suite
        self._build_ui_structure()

        # 2. On diffère le chargement des DONNÉES lourdes de 100ms
        # Cela permet à la fenêtre de s'afficher instantanément
        self.after(100, self._late_initialization)

    def _late_initialization(self):
        """Cette fonction s'exécute une fois la fenêtre affichée"""
        self._load_data()
        self._refresh_list()
        self._populate_week_dropdown()
        
        # On remet la fenêtre au premier plan une fois prête
        self.attributes("-topmost", True)
        self.lift()

    # ---------- DATA & LOGIC ----------
    def _load_data(self):
        if not os.path.exists(self.personnel_path):
            self.data = {"professeurs": [], "animateurs": []}
            return
        try:
            with open(self.personnel_path, "r", encoding="utf-8") as f:
                self.data = json.load(f)
        except:
            self.data = {"professeurs": [], "animateurs": []}

        # Migration: convertir l'ancien format (liste de strings) vers le nouveau (liste de dicts)
        self._migrate_data_format()

    def _migrate_data_format(self):
        """Convertit l'ancien format (liste de strings) vers le nouveau (liste de dicts avec classes)"""
        for personnel_type in ["professeurs", "animateurs"]:
            if personnel_type not in self.data:
                self.data[personnel_type] = []
                continue

            migrated_list = []
            for item in self.data[personnel_type]:
                if isinstance(item, str):
                    # Ancien format: convertir en nouveau format
                    migrated_list.append({"nom": item, "classes": []})
                elif isinstance(item, dict) and "nom" in item:
                    # Nouveau format: s'assurer que "classes" existe
                    if "classes" not in item:
                        item["classes"] = []
                    migrated_list.append(item)
                else:
                    # Format invalide: ignorer
                    continue

            self.data[personnel_type] = migrated_list

    def _save_data(self):
        with open(self.personnel_path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=4, ensure_ascii=False)

    def _get_available_weeks(self):
        parent_dir = os.path.dirname(self.week_folder)
        weeks = []
        if not os.path.exists(parent_dir): return weeks
        
        try:
            items = os.listdir(parent_dir)
            for item in items:
                if item.startswith("semaine_") and item != self.current_week:
                    item_path = os.path.join(parent_dir, item)
                    if os.path.isdir(item_path):
                        if os.path.exists(os.path.join(item_path, "personnel.json")):
                            weeks.append(item)
            
            weeks.sort(key=lambda x: int(x.split("_")[1]) if "_" in x and x.split("_")[1].isdigit() else x)
        except Exception as e:
            print(f"Erreur scan dossiers: {e}")
            
        return weeks
    
    def _format_week_display(self, folder_name):
        if folder_name.startswith("semaine_"):
            try:
                num = folder_name.split("_")[1]
                return f"Semaine {num}"
            except: pass
        return folder_name
    
    def _parse_week_display(self, display_name):
        if display_name.startswith("Semaine "):
            try:
                num = display_name.split(" ")[1]
                return f"semaine_{num}"
            except: pass
        return display_name

    def _show_message(self, msg_type, title, message):
        self.attributes("-topmost", False)
        if msg_type == "warning": res = messagebox.showwarning(title, message, parent=self)
        elif msg_type == "info": res = messagebox.showinfo(title, message, parent=self)
        elif msg_type == "error": res = messagebox.showerror(title, message, parent=self)
        elif msg_type == "yesno": res = messagebox.askyesno(title, message, parent=self)
        self.attributes("-topmost", True)
        return res

    def _notify_data_changed(self):
        """Notifie que les données ont changé en appelant le callback si défini."""
        if self.data_changed_callback:
            try:
                self.data_changed_callback()
            except Exception as e:
                print(f"Erreur lors de l'appel du callback de changement de données: {e}")

    def _update_excel_files_for_professor_removal(self, professor_name, removed_classes):
        """
        Met à jour les fichiers Excel des écoles pour retirer le nom du professeur
        des classes qui lui étaient assignées.

        Args:
            professor_name (str): Nom du professeur
            removed_classes (list): Liste des classes qui ont été retirées
        """
        if load_workbook is None:
            print("openpyxl n'est pas disponible, impossible de mettre à jour les fichiers Excel")
            return

        # Mapping des écoles vers les fichiers Excel
        school_file_mapping = {
            'ecole_a.xlsx': 'École A',
            'ecole_b.xlsx': 'École B',
            'ECOLE_C_cours_standard.xlsx': 'École C/CS',
            'ECOLE_C_cours_intensif.xlsx': 'École C/CI',
            'MORNING.xlsx': 'Morning',
            'ECOLE_PREMIUM_cours_standard.xlsx': 'Premium/CS',
            'ECOLE_PREMIUM_cours_intensifs.xlsx': 'Premium/CI'
        }

        # Convertir les noms de classes en set pour recherche rapide
        removed_classes_set = set(removed_classes)

        # Parcourir tous les fichiers Excel du dossier semaine
        if os.path.exists(self.week_folder):
            for filename in os.listdir(self.week_folder):
                if filename.lower().endswith('.xlsx') and filename.lower() != 'matrix.xlsx':
                    if filename not in school_file_mapping:
                        continue

                    excel_path = os.path.join(self.week_folder, filename)

                    try:
                        # Ouvrir le fichier Excel
                        wb = load_workbook(excel_path)

                        # Pour chaque feuille du fichier
                        for sheet_name in wb.sheetnames:
                            sheet = wb[sheet_name]

                            # Chercher les colonnes intervenant et classe
                            intervenant_col = None
                            classe_col = None

                            for col_idx in range(1, sheet.max_column + 1):
                                header_value = str(sheet.cell(row=1, column=col_idx).value or '').lower()
                                if any(keyword in header_value for keyword in ['intervenant', 'professeur', 'animateur', 'enseignant', 'Animateur', 'Rôle', 'rôle', 'role']):
                                    intervenant_col = col_idx
                                elif 'classe' in header_value or 'class' in header_value:
                                    classe_col = col_idx

                            # Si on a trouvé les colonnes nécessaires
                            if intervenant_col and classe_col:
                                # Parcourir toutes les lignes de données
                                for row_idx in range(2, sheet.max_row + 1):
                                    # Vérifier si la ligne contient des données
                                    has_data = False
                                    for col_idx in range(1, sheet.max_column + 1):
                                        cell_value = str(sheet.cell(row=row_idx, column=col_idx).value or '').strip()
                                        if cell_value:
                                            has_data = True
                                            break

                                    if has_data:
                                        # Récupérer le nom de la classe et de l'intervenant
                                        classe_nom = str(sheet.cell(row=row_idx, column=classe_col).value or '').strip()
                                        intervenant_nom = str(sheet.cell(row=row_idx, column=intervenant_col).value or '').strip()

                                        # Si cette classe était assignée au professeur et que le professeur est celui qu'on retire
                                        if classe_nom in removed_classes_set and intervenant_nom == professor_name:
                                            # Remplacer par "Non spécifié"
                                            sheet.cell(row=row_idx, column=intervenant_col, value="Non spécifié")

                        # Sauvegarder le fichier
                        wb.save(excel_path)
                        print(f"✅ Fichier {filename} mis à jour pour {professor_name}")

                    except Exception as e:
                        print(f"Erreur lors de la mise à jour de {filename}: {e}")

    def _update_excel_files_for_professor_assignment(self, professor_name, school_key, horaire, classe_nom):
        """
        Met à jour les fichiers Excel des écoles pour assigner un professeur
        à une classe spécifique.

        Args:
            professor_name (str): Nom du professeur
            school_key (str): Clé de l'école (ex: 'ecole_a')
            horaire (str): Horaire de la classe
            classe_nom (str): Nom de la classe
        """
        if load_workbook is None:
            print("openpyxl n'est pas disponible, impossible de mettre à jour les fichiers Excel")
            return

        # Mapping des clés d'école vers les fichiers Excel
        school_file_mapping = {
            'ecole_a': 'ecole_a.xlsx',
            'ecole_b': 'ecole_b.xlsx',
            'ecole_c_cs': 'ECOLE_C_cours_standard.xlsx',
            'ecole_c_ci': 'ECOLE_C_cours_intensif.xlsx',
            'ecole_morning': 'MORNING.xlsx',
            'ecole_premium_cs': 'ECOLE_PREMIUM_cours_standard.xlsx',
            'ecole_premium_ci': 'ECOLE_PREMIUM_cours_intensifs.xlsx'
        }

        if school_key not in school_file_mapping:
            print(f"École inconnue: {school_key}")
            return

        excel_filename = school_file_mapping[school_key]
        excel_path = os.path.join(self.week_folder, excel_filename)

        if not os.path.exists(excel_path):
            print(f"Fichier Excel non trouvé: {excel_path}")
            return

        try:
            # Ouvrir le fichier Excel
            wb = load_workbook(excel_path)

            # Chercher la feuille correspondant à l'horaire
            target_sheet = None
            for sheet_name in wb.sheetnames:
                # Normaliser les noms pour la comparaison
                sheet_normalized = sheet_name.lower().replace('animateur', '').replace('professeur', '').strip()
                horaire_normalized = horaire.lower().strip()

                if sheet_normalized == horaire_normalized or horaire_normalized in sheet_normalized:
                    target_sheet = wb[sheet_name]
                    break

            if target_sheet is None:
                print(f"Feuille horaire '{horaire}' non trouvée dans {excel_filename}")
                return

            # Chercher les colonnes intervenant et classe
            intervenant_col = None
            classe_col = None

            for col_idx in range(1, target_sheet.max_column + 1):
                header_value = str(target_sheet.cell(row=1, column=col_idx).value or '').lower()
                if any(keyword in header_value for keyword in ['intervenant', 'professeur', 'animateur', 'enseignant', 'Animateur', 'Rôle', 'rôle', 'role']):
                    intervenant_col = col_idx
                elif 'classe' in header_value or 'class' in header_value:
                    classe_col = col_idx

            # Si on a trouvé les colonnes nécessaires
            if intervenant_col and classe_col:
                # Chercher la ligne de la classe spécifique
                for row_idx in range(2, target_sheet.max_row + 1):
                    # Vérifier si la ligne contient des données
                    has_data = False
                    for col_idx in range(1, target_sheet.max_column + 1):
                        cell_value = str(target_sheet.cell(row=row_idx, column=col_idx).value or '').strip()
                        if cell_value:
                            has_data = True
                            break

                    if has_data:
                        # Récupérer le nom de la classe
                        current_classe_nom = str(target_sheet.cell(row=row_idx, column=classe_col).value or '').strip()

                        # Si c'est la classe qu'on cherche
                        if current_classe_nom == classe_nom:
                            # Assigner le professeur
                            target_sheet.cell(row=row_idx, column=intervenant_col, value=professor_name)
                            break

            # Sauvegarder le fichier
            wb.save(excel_path)
            print(f"✅ Classe '{classe_nom}' assignée à {professor_name} dans {excel_filename}")

        except Exception as e:
            print(f"Erreur lors de la mise à jour de {excel_filename}: {e}")

    def _import_from_week(self, selected_week_display):
        if selected_week_display == "Aucune semaine disponible" or selected_week_display == "Chargement...": return

        selected_week = self._parse_week_display(selected_week_display)
        parent_dir = os.path.dirname(self.week_folder)
        source_path = os.path.join(parent_dir, selected_week, "personnel.json")
        
        try:
            with open(source_path, "r", encoding="utf-8") as f:
                source_data = json.load(f)

            imported_items = source_data.get(self.personnel_type, [])

            if not imported_items:
                self._show_message("info", "Info", f"Aucun {self.single_name} trouvé.")
                return

            # Migration des données importées si nécessaire
            migrated_imported = []
            for item in imported_items:
                if isinstance(item, str):
                    migrated_imported.append({"nom": item, "classes": []})
                elif isinstance(item, dict) and "nom" in item:
                    if "classes" not in item:
                        item["classes"] = []
                    migrated_imported.append(item)

            # Trouver les nouveaux éléments
            existing_names = {item["nom"] for item in self.data[self.personnel_type]}
            new_items = [p for p in migrated_imported if p["nom"] not in existing_names]

            if not new_items:
                self._show_message("info", "Info", f"Déjà tous présents.")
                return

            if self._show_message("yesno", "Confirmation", f"Importer {len(new_items)} {self.single_name}(s) ?"):
                self.data[self.personnel_type].extend(new_items)
                self.data[self.personnel_type].sort(key=lambda x: x["nom"])
                self._save_data()
                self._refresh_list()
                self._show_message("info", "Succès", f"Import réussi.")

        except Exception as e:
            self._show_message("error", "Erreur", f"Erreur: {str(e)}")

    # ---------- UI STRUCTURE ----------
    def _build_ui_structure(self):
        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=25, pady=20)

        # HEADER
        header_frame = ctk.CTkFrame(main, fg_color="transparent")
        header_frame.pack(fill="x", pady=(0, 8))
        header_frame.grid_columnconfigure(5, weight=1)  # La colonne 5 (zone de recherche) prend l'espace restant

        ctk.CTkLabel(
            header_frame,
            text=self.display_name.upper(),
            font=("Inter", 13, "bold"),
            text_color="#111827"
        ).grid(row=0, column=0, sticky="w", padx=(0, 10))

        # Bouton Refresh
        refresh_btn = ctk.CTkButton(
            header_frame,
            text="↻",
            width=35, height=35, font=("Segoe UI", 12, "bold"),
            fg_color="#e74c3c", hover_color="#c0392b",
            text_color="white",
            corner_radius=6,
            command=lambda: self._refresh_from_file()
        )
        refresh_btn.grid(row=0, column=1, padx=(0, 15))

        # Boutons d'actions entre le label et la recherche
        actions_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        actions_frame.grid(row=0, column=2, padx=(0, 15))

        # Bouton Supprimer le prof
        delete_btn = ctk.CTkButton(
            actions_frame,
            text="🚫 Supprimer le professeur",
            width=32,
            height=32,
            font=("Inter", 12, "bold"),
            fg_color="#fca5a5",  # Rouge plus doux
            hover_color="#f87171",  # Rouge hover plus doux
            text_color="white",
            corner_radius=6,
            command=self._delete_selected_professor
        )
        delete_btn.pack(side="left", padx=(0, 5))

        # Bouton Assigner une classe
        assign_btn = ctk.CTkButton(
            actions_frame,
            text="📚 Assigner une classe",
            width=32,
            height=32,
            font=("Inter", 12, "bold"),
            fg_color="#4ade80",  # Vert un peu plus foncé
            hover_color="#22c55e",  # Vert hover plus foncé
            text_color="white",
            corner_radius=6,
            command=self._assign_class_to_selected_professor
        )
        assign_btn.pack(side="left", padx=(0, 5))

        # Bouton Supprimer les classes
        remove_classes_btn = ctk.CTkButton(
            actions_frame,
            text="🚫 Supprimer les classes",
            width=32,
            height=32,
            font=("Inter", 12, "bold"),
            fg_color="#fcd34d",  # Jaune/orange plus doux
            hover_color="#fbbf24",  # Jaune hover plus doux
            text_color="white",
            corner_radius=6,
            command=self._remove_all_classes_from_selected_professor
        )
        remove_classes_btn.pack(side="left")

        ctk.CTkLabel(header_frame, text="🔍", font=("Inter", 14), text_color="#6B7280").grid(row=0, column=3, padx=(0, 5))

        self.search_var = ctk.StringVar()
        self.search_var.trace_add("write", lambda *args: self._refresh_list())

        self.search_entry = ctk.CTkEntry(
            header_frame,
            textvariable=self.search_var,
            placeholder_text="Rechercher...",
            height=32, corner_radius=8, fg_color="#F3F4F6", border_width=0, border_color="#F3F4F6", width=150
        )
        self.search_entry.grid(row=0, column=4, sticky="ew")

        # GRID CONTAINER avec scrollbar (2 colonnes maintenant)
        scroll_container = ctk.CTkFrame(main, fg_color="transparent")
        scroll_container.pack(fill="both", expand=True, pady=(0, 15))
        
        # Canvas et scrollbar
        canvas = ctk.CTkCanvas(scroll_container, bg="#F9FAFB", highlightthickness=0)
        scrollbar = ctk.CTkScrollbar(scroll_container, command=canvas.yview)
        
        self.grid_container = ctk.CTkFrame(canvas, fg_color="#F9FAFB", corner_radius=10)
        
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas_frame = canvas.create_window((0, 0), window=self.grid_container, anchor="nw")
        
        # Fonction pour ajuster la scrollregion
        def configure_scroll_region(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        self.grid_container.bind("<Configure>", configure_scroll_region)
        
        # Ajuster la largeur du frame intérieur quand le canvas change de taille
        def configure_canvas_width(event):
            canvas.itemconfig(canvas_frame, width=event.width)
        
        canvas.bind("<Configure>", configure_canvas_width)
        
        # Permettre le scroll avec la molette
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind_all("<MouseWheel>", on_mousewheel)

        # Configuration de la grille 4 colonnes
        self.grid_container.grid_columnconfigure(0, weight=1)
        self.grid_container.grid_columnconfigure(1, weight=1)
        self.grid_container.grid_columnconfigure(2, weight=1)
        self.grid_container.grid_columnconfigure(3, weight=1)
        
        # Forcer la hauteur du grid_container à 3 lignes complètes (environ 600px)
        self.grid_container.configure(height=600)

        # ADD SECTION ET IMPORT SUR LA MÊME LIGNE
        # Labels au-dessus des colonnes
        labels_frame = ctk.CTkFrame(main, fg_color="transparent")
        labels_frame.pack(fill="x", pady=(0, 4))
        labels_frame.grid_columnconfigure(0, weight=1)
        labels_frame.grid_columnconfigure(1, weight=2)
        labels_frame.grid_columnconfigure(2, weight=1)

        ctk.CTkLabel(labels_frame, text=f"Ajouter un {self.single_name}", font=("Inter", 11, "bold"), text_color="#6B7280").grid(row=0, column=0, sticky="n")
        ctk.CTkLabel(labels_frame, text="", font=("Inter", 11, "bold"), text_color="#6B7280").grid(row=0, column=1, sticky="n")
        ctk.CTkLabel(labels_frame, text=f"Importer une liste de {self.display_name.lower()}", font=("Inter", 11, "bold"), text_color="#6B7280").grid(row=0, column=2, sticky="n")

        combined_frame = ctk.CTkFrame(main, fg_color="transparent")
        combined_frame.pack(fill="x", pady=(0, 10))
        combined_frame.grid_columnconfigure(0, weight=1)  # Add section prend 20%
        combined_frame.grid_columnconfigure(1, weight=2)  # Colonne centrale vide 40%
        combined_frame.grid_columnconfigure(2, weight=1)  # Import section prend 20%
        
        # Section AJOUT (gauche)
        add_frame = ctk.CTkFrame(combined_frame, fg_color="transparent")
        add_frame.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        add_frame.grid_columnconfigure(0, weight=1)

        self.add_entry = ctk.CTkEntry(
            add_frame, placeholder_text=f"Ex : Marie", height=38, corner_radius=8,
            fg_color="#F3F4F6", border_width=0, border_color="#F3F4F6"
        )
        self.add_entry.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        self.add_entry.bind("<Return>", lambda e: self._add_item())

        ctk.CTkButton(
            add_frame, text="➕", width=38, height=38, fg_color="#2563EB", hover_color="#1D4ED8",
            command=self._add_item
        ).grid(row=0, column=1)

        # Section IMPORT (droite)
        import_frame = ctk.CTkFrame(combined_frame, fg_color="transparent")
        import_frame.grid(row=0, column=2, sticky="ew")
        import_frame.grid_columnconfigure(0, weight=1)

        self.week_dropdown = ctk.CTkOptionMenu(
            import_frame,
            values=["Chargement..."],
            height=38, corner_radius=8, fg_color="#F3F4F6", text_color="#111827",
            button_color="#2563EB", button_hover_color="#1D4ED8",
            dropdown_fg_color="#FFFFFF", dropdown_text_color="#111827", dropdown_hover_color="#DBEAFE",
            dynamic_resizing=False
        )
        self.week_dropdown.grid(row=0, column=0, sticky="ew", padx=(0, 6))

        ctk.CTkButton(
            import_frame, text="📥", width=38, height=38, fg_color="#10B981", hover_color="#059669",
            text_color="white", font=("Arial", 16),
            command=lambda: self._import_from_week(self.week_dropdown.get())
        ).grid(row=0, column=1)

        # BUTTONS
        btn_frame = ctk.CTkFrame(main, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(10, 0))
        btn_frame.grid_columnconfigure((0, 1), weight=1)
        ctk.CTkButton(btn_frame, text=f"🗑 Supprimer", height=38, fg_color="#EF4444", hover_color="#DC2626", command=self._remove_item).grid(row=0, column=0, sticky="ew", padx=(0, 6))
        ctk.CTkButton(btn_frame, text="❌ Fermer", height=38, fg_color="#9CA3AF", hover_color="#6B7280", command=self.destroy).grid(row=0, column=1, sticky="ew")

    # ---------- POPULATION ----------
    def _populate_week_dropdown(self):
        """Remplit le menu déroulant après coup pour ne pas bloquer le démarrage"""
        available_weeks = self._get_available_weeks()
        week_options = [self._format_week_display(w) for w in available_weeks] if available_weeks else ["Aucune semaine disponible"]
        self.week_dropdown.configure(values=week_options)
        self.week_dropdown.set(week_options[0])

    def _detect_class_conflicts(self, items):
        """Détecte les professeurs assignés à des classes partagées"""
        from collections import defaultdict

        # Dictionnaire classe -> liste des professeurs
        class_to_professors = defaultdict(list)

        # Remplir le dictionnaire
        for item in items:
            nom = item["nom"]
            classes = item.get("classes", [])
            for classe in classes:
                class_to_professors[classe].append(nom)

        # Trouver les professeurs en conflit (assignés à des classes avec plusieurs professeurs)
        conflicted_professors = set()
        for professors in class_to_professors.values():
            if len(professors) > 1:
                conflicted_professors.update(professors)

        return conflicted_professors

    def _refresh_from_file(self):
        """Recharge les données depuis personnel.json puis rafraîchit l'affichage"""
        self._load_data()
        self._refresh_list()

    def _refresh_list(self):
        # Nettoyage
        for w in self.item_labels: w.destroy()
        self.item_labels.clear()
        self.selected_item = None
        self.selected_frame = None

        # Filtre
        search_text = self.search_var.get().strip().lower()
        items = self.data.get(self.personnel_type, [])
        if search_text:
            items = [i for i in items if search_text in i["nom"].lower() or
                    any(search_text in classe.lower() for classe in i.get("classes", []))]

        # Tri par nom
        items.sort(key=lambda x: x["nom"])

        # Détecter les conflits de classes (professeurs assignés à la même classe)
        conflicted_professors = self._detect_class_conflicts(items)

        # Création de la grille 4 colonnes
        num_columns = 4
        for idx, item in enumerate(items):
            row = idx // num_columns
            col = idx % num_columns

            nom = item["nom"]
            classes = item.get("classes", [])

            # Déterminer la couleur de la carte (rouge si conflit)
            is_conflicted = nom in conflicted_professors
            card_bg_color = "#FEE2E2" if is_conflicted else "white"  # Rouge clair pour les conflits
            card_border_color = "#F87171" if is_conflicted else "#E5E7EB"  # Bordure rouge pour les conflits

            # Conteneur principal (card)
            person_frame = ctk.CTkFrame(
                self.grid_container,
                fg_color=card_bg_color,
                corner_radius=12,
                border_width=2,
                border_color=card_border_color
            )
            person_frame.grid(row=row, column=col, padx=8, pady=8, sticky="nsew")

            # Entête avec nom et nombre de classes
            header_frame = ctk.CTkFrame(person_frame, fg_color="transparent")
            header_frame.pack(fill="x", pady=(12, 8), padx=12)

            # Nom de l'intervenant
            name_label = ctk.CTkLabel(
                header_frame,
                text=nom,
                font=("Inter", 14, "bold"),
                anchor="w",
                text_color="#1F2937"
            )
            name_label.pack(side="left", fill="x", expand=True)
            self._make_selectable(name_label, nom, person_frame)

            # Badge avec nombre de classes
            if classes:
                count_badge = ctk.CTkFrame(
                    header_frame,
                    fg_color="#DBEAFE",
                    corner_radius=12,
                    height=24
                )
                count_badge.pack(side="right")
                self._make_selectable(count_badge, nom, person_frame)

                count_label = ctk.CTkLabel(
                    count_badge,
                    text=f"{len(classes)} classe{'s' if len(classes) > 1 else ''}",
                    font=("Inter", 10, "bold"),
                    text_color="#1E40AF"
                )
                count_label.pack(padx=8, pady=2)
                self._make_selectable(count_label, nom, person_frame)

            # Séparateur
            separator = ctk.CTkFrame(person_frame, fg_color="#F3F4F6", height=1)
            separator.pack(fill="x", padx=12, pady=(0, 8))
            self._make_selectable(separator, nom, person_frame)

            # Zone des classes avec wrapping
            if classes:
                classes_container = ctk.CTkFrame(person_frame, fg_color="transparent")
                classes_container.pack(fill="both", expand=True, padx=12, pady=(0, 12))
                self._make_selectable(classes_container, nom, person_frame)

                # Frame interne pour le wrapping
                classes_wrapper = ctk.CTkFrame(classes_container, fg_color="transparent")
                classes_wrapper.pack(fill="both", expand=True)
                self._make_selectable(classes_wrapper, nom, person_frame)

                for classe in classes:
                    # Badge pour chaque classe
                    class_badge = ctk.CTkFrame(
                        classes_wrapper,
                        fg_color="#F3F4F6",
                        corner_radius=8,
                        border_width=1,
                        border_color="#E5E7EB"
                    )
                    class_badge.pack(side="left", padx=(0, 6), pady=3)
                    self._make_selectable(class_badge, nom, person_frame)

                    # Contenu du badge
                    badge_content = ctk.CTkFrame(class_badge, fg_color="transparent")
                    badge_content.pack(padx=8, pady=4)
                    self._make_selectable(badge_content, nom, person_frame)

                    # Nom de la classe
                    class_label = ctk.CTkLabel(
                        badge_content,
                        text=classe,
                        font=("Inter", 10, "bold"),
                        text_color="#374151"
                    )
                    class_label.pack(padx=8, pady=4)
                    self._make_selectable(class_label, nom, person_frame)
            else:
                # Message si aucune classe
                empty_frame = ctk.CTkFrame(person_frame, fg_color="#F9FAFB", corner_radius=8)
                empty_frame.pack(fill="x", padx=12, pady=(0, 12))
                self._make_selectable(empty_frame, nom, person_frame)

                no_classes_label = ctk.CTkLabel(
                    empty_frame,
                    text="📚 Aucune classe assignée",
                    font=("Inter", 10, "italic"),
                    text_color="#9CA3AF"
                )
                no_classes_label.pack(pady=8)
                self._make_selectable(no_classes_label, nom, person_frame)

            # Rendre toute la card cliquable
            person_frame.bind("<Button-1>", lambda e, n=nom, f=person_frame: self._select_person_frame(n, f))

            # Stocker le nom du professeur dans le frame pour une utilisation ultérieure
            person_frame.professor_name = nom

            # Stocker la référence
            self.item_labels.append(person_frame)

    def _select_person_frame(self, name, frame):
        """Sélectionne un intervenant et met en surbrillance son frame"""
        self.selected_item = name

        # Détecter les conflits pour déterminer les couleurs de base
        items = self.data.get(self.personnel_type, [])
        conflicted_professors = self._detect_class_conflicts(items)

        # Réinitialiser tous les frames avec leur couleur de base (blanc ou rouge selon conflit)
        for f in self.item_labels:
            # Le nom du professeur est stocké dans l'attribut 'professor_name' du frame
            professor_name = getattr(f, 'professor_name', '')
            is_conflicted = professor_name in conflicted_professors
            base_bg_color = "#FEE2E2" if is_conflicted else "white"
            base_border_color = "#F87171" if is_conflicted else "#E5E7EB"

            f.configure(fg_color=base_bg_color, border_color=base_border_color)

        # Mettre en surbrillance le frame sélectionné
        frame.configure(fg_color="#EFF6FF", border_color="#3B82F6")

        self.selected_frame = frame

    def _make_selectable(self, widget, professor_name, frame):
        """Rend un widget et tous ses enfants cliquables pour la sélection"""
        widget.bind("<Button-1>", lambda e, n=professor_name, f=frame: self._select_person_frame(n, f))
        # Propager aux enfants de manière récursive
        for child in widget.winfo_children():
            self._make_selectable(child, professor_name, frame)

    def _add_item(self):
        name = self.add_entry.get().strip()
        if not name: return

        # Vérifier si le nom existe déjà
        if any(item["nom"] == name for item in self.data[self.personnel_type]):
            self._show_message("warning", "Erreur", "Existe déjà")
            return

        self.data[self.personnel_type].append({"nom": name, "classes": []})
        self.data[self.personnel_type].sort(key=lambda x: x["nom"])
        self._save_data()
        self.add_entry.delete(0, "end")
        self._refresh_list()
        self._notify_data_changed()

    def _remove_item(self):
        if not self.selected_item: return
        if self._show_message("yesno", "Confirmation", f"Supprimer {self.selected_item} ?"):
            self.data[self.personnel_type] = [
                item for item in self.data[self.personnel_type]
                if item["nom"] != self.selected_item
            ]
            self._save_data()
            self._refresh_list()
            self._notify_data_changed()

    def _remove_class_from_person(self, person_name, class_name):
        """Retire une classe spécifique d'un intervenant"""
        if self._show_message("yesno", "Confirmation", f"Retirer la classe '{class_name}' de {person_name} ?"):
            for person in self.data[self.personnel_type]:
                if person["nom"] == person_name:
                    if class_name in person.get("classes", []):
                        person["classes"].remove(class_name)
                        self._save_data()
                        self._refresh_list()

                        # Mettre à jour les fichiers Excel des écoles
                        self._update_excel_files_for_professor_removal(person_name, [class_name])

                        self._show_message("info", "Succès", f"Classe '{class_name}' retirée de {person_name}")
                        self._notify_data_changed()
                    break


    def _delete_selected_professor(self):
        """Supprime le professeur actuellement sélectionné."""
        if not self.selected_item:
            self._show_message("warning", "Aucun professeur sélectionné", "Veuillez d'abord sélectionner un professeur.")
            return
        self._delete_professor(self.selected_item, lambda: None)

    def _assign_class_to_selected_professor(self):
        """Ouvre le menu d'assignation pour le professeur sélectionné."""
        if not self.selected_item:
            self._show_message("warning", "Aucun professeur sélectionné", "Veuillez d'abord sélectionner un professeur.")
            return
        self._show_assign_class_menu(self.selected_item, lambda: None)

    def _remove_all_classes_from_selected_professor(self):
        """Supprime toutes les classes du professeur sélectionné."""
        if not self.selected_item:
            self._show_message("warning", "Aucun professeur sélectionné", "Veuillez d'abord sélectionner un professeur.")
            return
        self._remove_all_classes_from_professor(self.selected_item, lambda: None)

    def _delete_professor(self, professor_name, close_menu_func):
        """Supprime un professeur."""
        if self._show_message("yesno", "Confirmation", f"Êtes-vous sûr de vouloir supprimer {professor_name} ?"):
            self.data[self.personnel_type] = [
                item for item in self.data[self.personnel_type]
                if item["nom"] != professor_name
            ]
            self._save_data()
            self._refresh_list()
            self._notify_data_changed()
            close_menu_func()

    def _show_assign_class_menu(self, professor_name, close_menu_func):
        """Affiche le menu d'assignation de classe (adapté d'Assignation des Niveaux.py)."""
        close_menu_func()  # Fermer le menu contextuel actuel

        # Créer le menu d'assignation de classe (même structure qu'Assignation des Niveaux.py)
        self._create_class_assignment_menu(professor_name)

    def _remove_all_classes_from_professor(self, professor_name, close_menu_func):
        """Supprime toutes les classes d'un professeur."""
        # Trouver le professeur
        professor_data = None
        for person in self.data[self.personnel_type]:
            if person["nom"] == professor_name:
                professor_data = person
                break

        if not professor_data or not professor_data.get("classes"):
            self._show_message("info", "Info", f"{professor_name} n'a aucune classe assignée.")
            close_menu_func()
            return

        classes_count = len(professor_data["classes"])
        if self._show_message("yesno", "Confirmation", f"Retirer les {classes_count} classe(s) de {professor_name} ?"):
            # Sauvegarder les classes qui vont être retirées pour mettre à jour les fichiers Excel
            removed_classes = professor_data["classes"].copy()

            professor_data["classes"] = []
            self._save_data()
            self._refresh_list()

            # Mettre à jour les fichiers Excel des écoles
            if removed_classes:
                self._update_excel_files_for_professor_removal(professor_name, removed_classes)

            self._show_message("info", "Succès", f"Toutes les classes ont été retirées de {professor_name}")
            self._notify_data_changed()
            close_menu_func()

    def _create_class_assignment_menu(self, professor_name):
        """Crée le menu d'assignation de classe pour les professeurs (basé sur Assignation des Niveaux.py)."""
        # Créer la fenêtre du menu
        menu = ctk.CTkToplevel(self)
        menu.title("")
        menu.geometry("700x600")
        menu.resizable(True, True)
        menu.transient(self)
        menu.attributes("-topmost", True)
        menu.overrideredirect(True)
        menu.configure(fg_color="white")

        # Centrer la fenêtre sur l'écran
        menu.update_idletasks()
        screen_width = menu.winfo_screenwidth()
        screen_height = menu.winfo_screenheight()
        x = (screen_width // 2) - (700 // 2)
        y = (screen_height // 2) - (600 // 2)
        menu.geometry(f"700x600+{x}+{y}")

        # Stocker la référence
        if not hasattr(self, 'class_assignment_menus'):
            self.class_assignment_menus = []
        self.class_assignment_menus.append(menu)

        # Variable pour suivre si le menu est détruit
        menu_destroyed = False

        def close_menu():
            """Ferme le menu."""
            nonlocal menu_destroyed
            if not menu_destroyed and menu.winfo_exists():
                menu_destroyed = True
                try:
                    menu.destroy()
                except:
                    pass

        # Analyser les données des écoles pour cette semaine
        week_folder = os.path.dirname(self.personnel_path)
        school_data = self._analyze_school_classes(week_folder)

        # Frame principal avec scroll
        main_frame = ctk.CTkFrame(menu, fg_color="white", corner_radius=10)
        main_frame.pack(fill="both", expand=True, padx=2, pady=2)

        # En-tête
        header_frame = ctk.CTkFrame(main_frame, fg_color="#f0f9ff", corner_radius=8)
        header_frame.pack(fill="x", padx=8, pady=(8, 4))

        title_label = ctk.CTkLabel(
            header_frame,
            text=f"📚 Assigner une classe à {professor_name}",
            font=("Inter", 13, "bold"),
            text_color="#1e293b"
        )
        title_label.pack(side="left", padx=12, pady=8)

        # Bouton fermer
        close_header_btn = ctk.CTkButton(
            header_frame,
            text="✕",
            width=30,
            height=30,
            font=("Inter", 14, "bold"),
            fg_color="#ef4444",
            hover_color="#dc2626",
            text_color="white",
            corner_radius=15,
            command=close_menu
        )
        close_header_btn.pack(side="right", padx=8, pady=4)

        # Frame avec scrollbar
        scrollable = ctk.CTkScrollableFrame(main_frame, fg_color="#f8fafc")
        scrollable.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        # Afficher les données des écoles
        if not school_data:
            no_data_label = ctk.CTkLabel(
                scrollable,
                text="Aucune donnée d'école trouvée pour cette semaine.",
                font=("Inter", 12),
                text_color="#6b7280"
            )
            no_data_label.pack(pady=30)
        else:
            self._display_school_data_in_prof_menu(scrollable, school_data, professor_name, menu, close_menu)

        # Frame pour les boutons d'action en bas
        action_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        action_frame.pack(fill="x", padx=8, pady=(0, 8))

        # Bouton fermer en bas
        close_bottom_btn = ctk.CTkButton(
            action_frame,
            text="Fermer",
            font=("Inter", 11, "bold"),
            height=35,
            width=100,
            fg_color="#6b7280",
            hover_color="#4b5563",
            text_color="white",
            command=close_menu
        )
        close_bottom_btn.pack(side="right")

        # Gestion des événements
        def on_click_outside(event):
            if not menu_destroyed and menu.winfo_exists():
                try:
                    menu_x = menu.winfo_x()
                    menu_y = menu.winfo_y()
                    menu_width = menu.winfo_width()
                    menu_height = menu.winfo_height()

                    click_x = event.x_root
                    click_y = event.y_root

                    if not (menu_x <= click_x <= menu_x + menu_width and
                           menu_y <= click_y <= menu_y + menu_height):
                        close_menu()
                except:
                    pass

        click_handler_id = self.bind("<Button-1>", on_click_outside, add="+")
        menu._click_handler_id = click_handler_id

        def on_key_press(event):
            if event.keysym == 'Escape' and not menu_destroyed:
                close_menu()

        menu.bind("<Key>", on_key_press)
        menu.focus_set()

    def _analyze_school_classes(self, week_folder):
        """Analyse les fichiers Excel d'écoles pour une semaine donnée (version adaptée pour professeurs)."""
        import os
        try:
            import pandas as pd
        except ImportError:
            return {}

        result = {
            'ecole_a': [],
            'ecole_b': [],
            'ecole_c_cs': [],
            'ecole_c_ci': [],
            'ecole_morning': [],
            'ecole_premium_cs': [],
            'ecole_premium_ci': []
        }

        # Mapping des fichiers Excel vers les clés d'écoles
        file_to_school_mapping = {
            'ecole_a.xlsx': 'ecole_a',
            'ecole_b.xlsx': 'ecole_b',
            'ECOLE_C_cours_standard.xlsx': 'ecole_c_cs',
            'ECOLE_C_cours_intensif.xlsx': 'ecole_c_ci',
            'MORNING.xlsx': 'ecole_morning',
            'ECOLE_PREMIUM_cours_standard.xlsx': 'ecole_premium_cs',
            'ECOLE_PREMIUM_cours_intensifs.xlsx': 'ecole_premium_ci'
        }

        # Fonction helper pour analyser un fichier Excel d'école
        def analyze_school_file(excel_path, school_key):
            if not os.path.exists(excel_path):
                return

            try:
                # Lire toutes les feuilles du fichier
                xl_file = pd.ExcelFile(excel_path)
                sheet_names = xl_file.sheet_names

                for sheet_name in sheet_names:
                    try:
                        df = pd.read_excel(excel_path, sheet_name=sheet_name)

                        if df.empty:
                            # La feuille est vide mais on garde l'horaire
                            horaire = self._clean_horaire_name(sheet_name)
                            result[school_key].append({
                                'horaire': horaire or sheet_name,
                                'classes': []
                            })
                            continue

                        # Analyser et nettoyer le nom de la feuille
                        horaire = self._clean_horaire_name(sheet_name)

                        # Chercher les colonnes de classes
                        classe_cols = []
                        for col in df.columns:
                            col_lower = str(col).lower()
                            if any(keyword in col_lower for keyword in ['classe', 'groupe', 'section']):
                                classe_cols.append(col)

                        classes_info = []
                        for _, row in df.iterrows():
                            classe_nom = None
                            for col in classe_cols:
                                val = str(row.get(col, '')).strip()
                                if val and val.lower() not in ['', 'nan', 'none']:
                                    classe_nom = val
                                    break

                            if classe_nom:
                                classes_info.append({
                                    'nom_classe': classe_nom,
                                    'nb_eleves': 0,  # On ne calcule pas ici pour simplifier
                                    'niveau': '',
                                    'eleves': []
                                })

                        result[school_key].append({
                            'horaire': horaire or sheet_name,
                            'classes': classes_info
                        })

                    except Exception as e:
                        print(f"Erreur lors de l'analyse de la feuille '{sheet_name}': {e}")
                        continue

            except Exception as e:
                print(f"Erreur lors de l'analyse du fichier {excel_path}: {e}")

        # Analyser tous les fichiers Excel du dossier semaine
        if os.path.exists(week_folder):
            for filename in os.listdir(week_folder):
                if filename.lower().endswith('.xlsx') and filename.lower() != 'matrix.xlsx':
                    file_path = os.path.join(week_folder, filename)
                    school_key = file_to_school_mapping.get(filename)
                    if school_key:
                        analyze_school_file(file_path, school_key)

        return result

    def _clean_horaire_name(self, sheet_name):
        """Nettoie le nom de la feuille pour extraire le nom d'horaire de manière cohérente."""
        # Liste complète des mots à supprimer
        words_to_remove = [
            "animateur", "Animateur", "anim", "Anim",
            "professeur", "Professeur", "prof", "Prof",
            "Rôle", "rôle", "role", "Role"
        ]

        # Nettoyer le nom en supprimant tous les mots-clés d'intervenants
        horaire = sheet_name
        for word in words_to_remove:
            horaire = horaire.replace(word, "")

        # Nettoyer les espaces multiples et supprimer les espaces au début/fin
        horaire = " ".join(horaire.split()).strip()

        # Si le résultat est vide, utiliser le nom original
        return horaire or sheet_name

    def _display_school_data_in_prof_menu(self, parent_frame, school_data, professor_name, menu, close_menu_func):
        """Affiche les données des écoles dans le menu d'assignation pour professeurs."""
        # Mapping pour les noms d'affichage et couleurs
        school_display = {
            'ecole_a': ('A', '#3b82f6'),
            'ecole_b': ('B', '#10b981'),
            'ecole_c_cs': ('C/CS', '#f59e0b'),
            'ecole_c_ci': ('C/CI', '#8b5cf6'),
            'ecole_morning': ('Morning', '#ef4444'),
            'ecole_premium_cs': ('Premium/CS', '#06b6d4'),
            'ecole_premium_ci': ('Premium/CI', '#f97316')
        }

        for school_key, horaires in school_data.items():
            if not horaires:  # Si pas d'horaires pour cette école
                continue

            display_name, school_color = school_display.get(school_key, (school_key, '#6b7280'))

            # Frame pour l'école
            school_frame = ctk.CTkFrame(parent_frame, fg_color="white", corner_radius=8, border_width=1, border_color=school_color)
            school_frame.pack(fill="x", pady=(0, 8), padx=5)

            # En-tête de l'école
            school_header = ctk.CTkFrame(school_frame, fg_color="#f0f9ff", corner_radius=6)
            school_header.pack(fill="x", padx=8, pady=(8, 5))

            school_title = ctk.CTkLabel(
                school_header,
                text=f"🏫 École {display_name}",
                font=("Inter", 12, "bold"),
                text_color=school_color
            )
            school_title.pack(pady=6)

            # Conteneur pour les horaires
            horaires_frame = ctk.CTkFrame(school_frame, fg_color="transparent")
            horaires_frame.pack(fill="x", padx=8, pady=(0, 8))

            # Calculer le nombre d'horaires par ligne
            total_horaires = len(horaires)
            horaires_per_row = min(3, max(2, total_horaires))  # Entre 2 et 3 horaires par ligne

            # Configurer les colonnes dynamiquement
            for col in range(horaires_per_row):
                horaires_frame.grid_columnconfigure(col, weight=1)

            horaire_row = 0
            horaire_col = 0

            for horaire_info in horaires:
                horaire = horaire_info.get('horaire', 'Horaire inconnu')
                classes = horaire_info.get('classes', [])

                # Frame pour l'horaire
                horaire_frame = ctk.CTkFrame(horaires_frame, fg_color="#f8fafc", corner_radius=6, border_width=1, border_color="#e5e7eb")
                horaire_frame.grid(row=horaire_row, column=horaire_col, sticky="nsew", padx=(0, 4) if horaire_col < horaires_per_row - 1 else 0, pady=(0, 4))

                # En-tête de l'horaire
                horaire_header = ctk.CTkFrame(horaire_frame, fg_color="white", corner_radius=4)
                horaire_header.pack(fill="x", padx=6, pady=(6, 4))

                horaire_title = ctk.CTkLabel(
                    horaire_header,
                    text=f"🕒 Horaire : {horaire}",
                    font=("Inter", 11, "bold"),
                    text_color="#374151"
                )
                horaire_title.pack(side="left", pady=4)

                # Conteneur pour les classes (cards)
                if classes:
                    classes_frame = ctk.CTkFrame(horaire_frame, fg_color="transparent")
                    classes_frame.pack(fill="x", padx=6, pady=(0, 8))

                    # Configuration pour une grille de classes
                    classes_per_row = 4  # Nombre de classes par ligne
                    for i in range(classes_per_row):
                        classes_frame.grid_columnconfigure(i, weight=1)

                    for idx, classe_info in enumerate(classes):
                        classe_nom = classe_info.get('nom_classe', 'Classe inconnue')
                        row = idx // classes_per_row
                        col = idx % classes_per_row

                        # Petite card pour chaque classe
                        class_card = ctk.CTkFrame(
                            classes_frame,
                            fg_color="white",
                            corner_radius=4,
                            width=80,
                            height=35
                        )
                        class_card.grid(row=row, column=col, padx=2, pady=2, sticky="nsew")
                        class_card.grid_propagate(False)

                        # Label du nom de classe (centré)
                        class_label = ctk.CTkLabel(
                            class_card,
                            text=classe_nom,
                            font=("Inter", 9, "bold"),
                            text_color="#374151"
                        )
                        class_label.place(relx=0.5, rely=0.5, anchor="center")

                        # Rendre la card cliquable
                        class_card.configure(cursor="hand2")

                        # Fonction pour gérer l'effet hover
                        def on_enter(event, frame=class_card):
                            frame.configure(fg_color="#f3f4f6")

                        def on_leave(event, frame=class_card):
                            frame.configure(fg_color="white")

                        # Propager l'effet hover à tous les enfants
                        def bind_hover_to_children(widget, enter_func, leave_func):
                            widget.bind("<Enter>", enter_func)
                            widget.bind("<Leave>", leave_func)
                            for child in widget.winfo_children():
                                bind_hover_to_children(child, enter_func, leave_func)

                        # Lier les événements
                        class_card.bind("<Enter>", on_enter)
                        class_card.bind("<Leave>", on_leave)
                        bind_hover_to_children(class_card, on_enter, on_leave)

                        # Gestion du clic sur la card
                        def on_class_click(event, p_name=professor_name, s_key=school_key, h_info=horaire_info, c_info=classe_info, m=menu, close_func=close_menu_func):
                            # Fermer immédiatement le menu
                            close_func()
                            # Puis assigner avec un petit délai pour laisser le menu se fermer
                            m.after(10, lambda: self._assign_class_to_professor(p_name, s_key, h_info, c_info))

                        class_card.bind("<Button-1>", on_class_click)
                        class_label.bind("<Button-1>", on_class_click)
                        class_label.configure(cursor="hand2")
                else:
                    no_classes_label = ctk.CTkLabel(
                        horaire_frame,
                        text="Aucune classe définie",
                        font=("Inter", 10, "italic"),
                        text_color="#9ca3af"
                    )
                    no_classes_label.pack(padx=12, pady=(0, 8))

                # Gestion de la grille des horaires
                horaire_col += 1
                if horaire_col >= horaires_per_row:
                    horaire_col = 0
                    horaire_row += 1

    def _assign_class_to_professor(self, professor_name, school_key, horaire_info, classe_info):
        """Assigne une classe à un professeur."""
        # Mapping pour convertir les clés d'école en noms d'affichage
        school_mapping = {
            'ecole_a': 'A',
            'ecole_b': 'B',
            'ecole_c_cs': 'C/CS',
            'ecole_c_ci': 'C/CI',
            'ecole_morning': 'Morning',
            'ecole_premium_cs': 'Premium/CS',
            'ecole_premium_ci': 'Premium/CI'
        }

        school_name = school_mapping.get(school_key, school_key)
        horaire = horaire_info.get('horaire', '')
        classe_nom = classe_info.get('nom_classe', '')

        # Trouver le professeur dans les données
        for person in self.data[self.personnel_type]:
            if person["nom"] == professor_name:
                # Vérifier si la classe n'est pas déjà assignée
                if classe_nom not in person.get("classes", []):
                    person["classes"].append(classe_nom)
                    person["classes"].sort()  # Trier les classes
                    self._save_data()
                    self._refresh_list()

                    # Mettre à jour les fichiers Excel des écoles
                    self._update_excel_files_for_professor_assignment(professor_name, school_key, horaire, classe_nom)

                    self._show_message("info", "Succès", f"Classe '{classe_nom}' assignée à {professor_name}")
                    self._notify_data_changed()
                else:
                    self._show_message("info", "Info", f"La classe '{classe_nom}' est déjà assignée à {professor_name}")
                break